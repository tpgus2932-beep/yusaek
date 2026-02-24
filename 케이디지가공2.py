import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import load_workbook

BASE_FILENAME = "케이디지원가베이스.xlsx"

BASE_XLSX_PATH = r"C:\Users\ksh29\OneDrive\Desktop\원베\케이디지원가베이스.xlsx"

NUM_LINE_RE = re.compile(r"^\s*([\d,]+)\s+([\d,]+)\s+([\d,]+)\s*$")  # 단가 수량 금액

ITEM_RE = re.compile(
    r"^\s*(?:[A-Za-z]\.)*([A-Za-z])\.(\d+)-(\d+)(SH|LO)?/([^/]+)/([^/\s]+)\s*$"
)

def suffix_to_kor(sfx):
    if sfx == "SH":
        return "숏"
    if sfx == "LO":
        return "롱"
    return "기본"


def parse_input_text(raw: str) -> pd.DataFrame:
    lines = [ln.rstrip() for ln in raw.splitlines() if ln.strip()]
    rows = []
    i = 0

    while i < len(lines):
        ln = lines[i].strip()

        if ln.startswith("품명"):
            i += 1
            continue

        m_item = ITEM_RE.match(ln)
        if not m_item:
            i += 1
            continue

        _, model_no, opt_no, sfx, color, size = m_item.groups()
        suffix = suffix_to_kor(sfx)              # 숏/롱/기본
        a_value = f"{model_no} {color} {size} {suffix}"

        qty = ""
        amount = ""

        # 바로 다음 줄에서 단가/수량/금액 읽기
        if i + 1 < len(lines):
            m_num = NUM_LINE_RE.match(lines[i + 1].strip())
            if m_num:
                unit_price, qty, amount = m_num.groups()
                qty = qty.replace(",", "")
                amount = amount.replace(",", "")
                i += 2
            else:
                i += 1
        else:
            i += 1

        rows.append({
            "A(변환품명)": a_value,

            # ✅ 기존 호환 위해 컬럼명은 그대로 두고, 값은 '수량' 넣기
            "B(옵션번호)": qty,

            "원본품명": ln,

            # ✅ 날짜별 복사용 필드(숨김 컬럼처럼 사용)
            "모델번호": model_no,
            "색상": color,
            "사이즈": size,
            "기장": suffix,
            "수량": int(qty) if str(qty).strip() else 0,
            "금액": int(amount) if str(amount).strip() else 0,
        })

    if not rows:
        raise ValueError("변환할 품명 라인을 찾지 못했음.")

    df = pd.DataFrame(rows)
    if "원베_B" not in df.columns:
        df["원베_B"] = ""
    return df


def load_base_map(base_path: str) -> dict:
    if not os.path.exists(base_path):
        raise FileNotFoundError(f"파일이 없음: {base_path}")

    wb = load_workbook(base_path, data_only=True)
    ws = wb.active

    mapping = {}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None:
            continue
        key = str(a).strip()
        if key:
            mapping[key] = b
    return mapping

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("품명 변환 + 원베 매칭")
        self.geometry("980x620")

        self.df = None
        self.base_folder = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)
        ttk.Label(top, text="여기에 그대로 붙여넣기").pack(anchor="w")


        self.txt = tk.Text(self, height=16, font=("Consolas", 10))
        self.txt.pack(fill="both", expand=False, padx=10)

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=8)
        ttk.Button(btns, text="초기화", command=self.on_reset).pack(side="left", padx=8)
        ttk.Button(btns, text="날짜별 복사", command=self.on_copy_by_date).pack(side="left", padx=8)


        ttk.Button(btns, text="변환", command=self.on_convert).pack(side="left")
        ttk.Button(btns, text="이지어드민 변환(원베 매칭)", command=self.on_match_base).pack(side="left", padx=8)
        ttk.Button(btns, text="CSV로 저장(A=원베,B=옵션)", command=self.on_save_csv).pack(side="left", padx=8)
        ttk.Button(btns, text="원베 폴더 설정", command=self.on_set_base_folder)\
            .pack(side="left", padx=8)
        ttk.Label(btns, text=f"원베 파일: {BASE_XLSX_PATH}").pack(side="right")

        ttk.Label(self, text="결과").pack(anchor="w", padx=10)

        self.tree = ttk.Treeview(self, columns=("A", "B", "BASE", "SRC"), show="headings", height=12)
        self.tree.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.tree.heading("A", text="A(변환품명)")
        self.tree.heading("B", text="B(옵션번호)")
        self.tree.heading("BASE", text="원베_B(매칭값)")
        self.tree.heading("SRC", text="원본품명")

        self.tree.column("A", width=260)
        self.tree.column("B", width=90, anchor="center")
        self.tree.column("BASE", width=160)
        self.tree.column("SRC", width=420)



    def on_reset(self):
        self.df = None
        self.render_df()
        self.txt.delete("1.0", "end")
        messagebox.showinfo("초기화", "입력/결과 초기화했음.")

    def on_set_base_folder(self):
        folder = filedialog.askdirectory(title="원베 폴더 선택")
        if not folder:
            return

        base_path = os.path.join(folder, BASE_FILENAME)

        if not os.path.exists(base_path):
            messagebox.showerror(
                "오류",
                f"선택한 폴더에 파일이 없습니다.\n\n"
                f"필요한 파일명:\n{BASE_FILENAME}"
            )
            return

        self.base_folder = folder
        messagebox.showinfo(
            "설정 완료",
            f"원베 폴더가 설정되었습니다.\n\n{base_path}"
        )
    

    def render_df(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        if self.df is None or self.df.empty:
            return

        for _, r in self.df.iterrows():
            self.tree.insert(
                "",
                "end",
                values=(
                    r.get("A(변환품명)", ""),
                    r.get("B(옵션번호)", ""),
                    r.get("원베_B", ""),
                    r.get("원본품명", ""),
                ),
            )
    def on_copy_by_date(self):
        if self.df is None or self.df.empty:
            messagebox.showwarning("안내", "복사할 데이터가 없음. 먼저 [변환]을 해줘.")
            return

        # 필요한 컬럼 없으면(구버전 df) 방어
        need = ["모델번호", "색상", "사이즈", "기장", "수량", "금액"]
        for c in need:
            if c not in self.df.columns:
                messagebox.showerror("오류", f"날짜별 복사에 필요한 컬럼이 없음: {c}\nparse_input_text 업데이트 필요")
                return

        # 같은 모델+색상+사이즈+기장 단위로 수량/금액 합산
        g = (
            self.df.groupby(["모델번호", "색상", "사이즈", "기장"], as_index=False)[["수량", "금액"]]
            .sum()
        )

        # A~F 구성
        out = pd.DataFrame({
            "A": ["케이디지"] * len(g),
            "B": g["모델번호"].astype(str),
            "C": g["금액"].astype(int).astype(str),                 # 필요하면 콤마 포맷도 가능
            "D": g["색상"].astype(str),
            "E": (g["사이즈"].astype(str) + " " + g["기장"].astype(str)),
            "F": g["수량"].astype(int).astype(str),
        })

        # 엑셀에 붙여넣기 쉬운 탭 구분(헤더 없이)
        tsv = out.to_csv(
            sep="\t",
            index=False,
            header=False,
            lineterminator="\n"   # ⭐ 핵심
        ).rstrip("\n")            # ⭐ 마지막 개행 제거



        self.clipboard_clear()
        self.clipboard_append(tsv)
        messagebox.showinfo("복사 완료", "날짜별 복사 데이터(A~F)가 클립보드에 복사됐음.\n엑셀에 바로 붙여넣기 하면 됨.")


    # ✅ 반드시 class 안(들여쓰기) 있어야 함
    def on_convert(self):
        raw = self.txt.get("1.0", "end").strip()
        try:
            new_df = parse_input_text(raw)

            # ✅ 누적 추가 (기존 df가 있으면 아래로 append)
            if self.df is None or self.df.empty:
                self.df = new_df
            else:
                self.df = pd.concat([self.df, new_df], ignore_index=True)

            # ✅ (선택) 완전 동일 행 중복 제거: 변환품명+옵션번호+원본품명 기준
            self.df = self.df.drop_duplicates(
                subset=["A(변환품명)", "B(옵션번호)", "원본품명"],
                keep="first"
            ).reset_index(drop=True)

            self.render_df()
            messagebox.showinfo("완료", f"누적 변환 완료: 현재 {len(self.df)}개")
        except Exception as e:
            messagebox.showerror("오류", str(e))


    def on_match_base(self):
        # 1) 변환 데이터 있는지 확인
        if self.df is None or self.df.empty:
            messagebox.showwarning("안내", "먼저 [변환]을 눌러 결과를 만든 뒤 진행해줘.")
            return

        # 2) 원베 폴더 설정 확인
        if not getattr(self, "base_folder", None):
            messagebox.showwarning("안내", "먼저 [원베 폴더 설정]을 해주세요.")
            return

        # 3) 고정 파일명으로 전체 경로 만들기
        base_path = os.path.join(self.base_folder, BASE_FILENAME)

        # 4) 파일 존재 확인
        if not os.path.exists(base_path):
            messagebox.showerror(
                "오류",
                f"원베 파일을 찾을 수 없음.\n\n"
                f"폴더: {self.base_folder}\n"
                f"파일명: {BASE_FILENAME}\n\n"
                f"현재 경로:\n{base_path}"
            )
            return

        # 5) 매칭 실행
        try:
            base_map = load_base_map(base_path)

            # A(변환품명) 기준으로 원베 A열 매칭 -> 원베 B열 값 불러오기
            self.df["원베_B"] = self.df["A(변환품명)"].map(
                lambda k: base_map.get(str(k).strip(), "")
            )

            missing = (self.df["원베_B"].astype(str).str.strip() == "").sum()

            self.render_df()

            if missing:
                messagebox.showwarning("완료(일부 누락)", f"매칭 완료. {missing}개는 원베에서 못 찾았음(A열 키 불일치).")
            else:
                messagebox.showinfo("완료", "매칭 완료. 전부 원베에서 찾았음.")

        except Exception as e:
            messagebox.showerror("오류", str(e))


    # ✅ CSV 저장: A열=원베_B, B열=옵션번호
    def on_save_csv(self):
        if self.df is None or self.df.empty:
            messagebox.showwarning("안내", "저장할 결과가 없음. 먼저 [변환]을 해줘.")
            return

        # 필수 컬럼 확인
        need_cols = ["원베_B", "B(옵션번호)"]
        for c in need_cols:
            if c not in self.df.columns:
                messagebox.showerror("오류", f"CSV 저장에 필요한 컬럼이 없음: {c}")
                return

        # ✅ CSV 데이터 구성
        out = pd.DataFrame({
            "상품코드": self.df["원베_B"].astype(str),
            "요청수량": ["0"] * len(self.df),              # 전부 0
            "입고수량": self.df["B(옵션번호)"].astype(str) # 수량
        })

        save_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile="입고업로드.csv"
        )
        if not save_path:
            return

        # ✅ 엑셀 호환 + 줄 안 띄어지게
        out.to_csv(
            save_path,
            index=False,
            encoding="utf-8-sig",
            lineterminator="\n"
        )

        messagebox.showinfo("저장 완료", f"CSV 저장했음:\n{save_path}")

if __name__ == "__main__":
    app = App()
    app.mainloop()
