# backend/barcode_core.py
import re
import sys
import tempfile
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import Workbook

QTY_COL = 11

PATTERN_S5 = re.compile(r"S(\d{5})")
PATTERN_YUSAS5 = re.compile(r"YUSAS(\d{5})")


def _to_int(x, default=0):
    try:
        if x is None or (isinstance(x, str) and not str(x).strip()):
            return default
        return int(float(str(x).strip()))
    except Exception:
        return default


def _to_str(x):
    if x is None:
        return ""
    return str(x).strip()


def normalize_to_yusas(text: str | None) -> str:
    if text is None:
        return ""
    s = str(text)
    m2 = PATTERN_YUSAS5.search(s)
    if m2:
        return f"YUSAS{m2.group(1)}"
    m1 = PATTERN_S5.search(s)
    if m1:
        return f"YUSAS{m1.group(1)}"
    return ""


def _bytes_head(path: Path, n=4096) -> bytes:
    with open(path, "rb") as f:
        return f.read(n)


def _excel_com_convert_to_xlsx(path: Path) -> Path:
    # Excel 설치된 윈도우에서만 동작(로컬 PC면 OK)
    import win32com.client as win32  # pip install pywin32

    # Use a new Excel instance so we don't close the user's existing Excel
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(str(path))
    tmp = Path(tempfile.gettempdir()) / (path.stem + "_auto_tmp.xlsx")
    wb.SaveAs(str(tmp), FileFormat=51)  # 51 = xlsx
    wb.Close(False)
    excel.Quit()
    return tmp


def _df_to_ws(df: "pd.DataFrame"):
    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(df.fillna("").values.tolist(), start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    return wb, ws


def load_excel_any(path: Path):
    suf = path.suffix.lower()
    head = _bytes_head(path)

    # .xlsx
    if suf == ".xlsx" or head[:2] == b"PK":
        wb = openpyxl.load_workbook(path)
        return wb, wb.active

    # .xls
    if suf == ".xls":
        # 진짜 BIFF
        if head.startswith(b"\xD0\xCF\x11\xE0"):
            df = pd.read_excel(path, header=None, engine="xlrd")  # xlrd 필요
            return _df_to_ws(df)

        # 가짜 .xls (html 등) -> Excel COM 변환 시도
        try:
            tmp_xlsx = _excel_com_convert_to_xlsx(path)
            wb = openpyxl.load_workbook(tmp_xlsx)
            return wb, wb.active
        except Exception:
            # html table
            text = head.lstrip(b"\xef\xbb\xbf")
            if text[:10].lower().startswith(b"<html") or b"<table" in text.lower():
                dfs = pd.read_html(path, header=None)
                return _df_to_ws(dfs[0])
            # csv fallback
            df = pd.read_csv(path, header=None, encoding="utf-8-sig")
            return _df_to_ws(df)

    raise ValueError("지원 확장자: .xls / .xlsx")


def fill_merged_in_column(ws, col_idx=13, header_row=1):
    targets = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col <= col_idx <= mr.max_col:
            top_val = ws.cell(mr.min_row, mr.min_col).value
            targets.append({"ref": str(mr), "min_row": mr.min_row, "max_row": mr.max_row, "top_val": top_val})
    for t in targets:
        ws.unmerge_cells(t["ref"])
    for t in targets:
        for r in range(t["min_row"], t["max_row"] + 1):
            ws.cell(r, col_idx).value = t["top_val"]

    last = None
    for r in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(r, col_idx).value
        if (val is None or val == "") and last:
            ws.cell(r, col_idx).value = last
        else:
            last = val


def process_and_load_any(path: Path):
    wb, ws = load_excel_any(path)
    fill_merged_in_column(ws, col_idx=13, header_row=1)

    # H열 코드 정규화
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        cell = row[0]
        cell.value = normalize_to_yusas(cell.value)

    # N열 기준 정렬
    rows = []
    for r in range(2, ws.max_row + 1):
        nval = ws.cell(r, 14).value  # N
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rows.append([nval, row_values])

    def parse_time(v):
        if isinstance(v, datetime):
            return v
        if v is None:
            return datetime.max
        s = str(v).strip()
        try:
            return datetime.fromisoformat(s)
        except Exception:
            try:
                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except Exception:
                return datetime.max

    rows.sort(key=lambda x: parse_time(x[0]))

    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
    for c, v in enumerate(header, start=1):
        ws.cell(1, c).value = v
    row_idx = 2
    for _, row_vals in rows:
        for c, v in enumerate(row_vals, start=1):
            ws.cell(row_idx, c).value = v
        row_idx += 1

    mapping_counts = defaultdict(Counter)
    mapping_details = defaultdict(dict)
    invoice_order = defaultdict(list)
    invoice_seq = []
    seen_invoice = set()
    code_o_text = {}

    def _parse_dt(v):
        if isinstance(v, datetime):
            return v
        if v is None or str(v).strip() == "":
            return None
        s = str(v).strip()
        try:
            return datetime.fromisoformat(s)
        except Exception:
            try:
                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except Exception:
                return None

    last_time_code = {}
    cur_run_len_code = defaultdict(int)
    cur_run_members = defaultdict(list)
    run_len_per_invcode = defaultdict(int)
    prev_code_row = None

    def flush_run(code: str):
        L = cur_run_len_code.get(code, 0)
        if L > 1 and cur_run_members[code]:
            for inv_row, code_row in cur_run_members[code]:
                key = (inv_row, code_row)
                if L > run_len_per_invcode[key]:
                    run_len_per_invcode[key] = L
        cur_run_len_code[code] = 0
        cur_run_members[code] = []

    for r in range(2, ws.max_row + 1):
        code = _to_str(ws.cell(r, 8).value)     # H
        name = _to_str(ws.cell(r, 9).value)     # I
        option = _to_str(ws.cell(r, 10).value)  # J
        inv = _to_str(ws.cell(r, 13).value)     # M
        t = _parse_dt(ws.cell(r, 14).value)           # N
        qty = _to_int(ws.cell(r, QTY_COL).value, default=1)
        o_val = ws.cell(r, 15).value                  # O

        if not (inv and code) or qty <= 0:
            prev_code_row = code
            continue

        if code not in code_o_text and o_val is not None:
            code_o_text[code] = _to_str(o_val)

        if inv not in seen_invoice:
            seen_invoice.add(inv)
            invoice_seq.append(inv)

        if code not in invoice_order[inv]:
            invoice_order[inv].append(code)

        mapping_counts[inv][code] += qty
        if code not in mapping_details[inv]:
            mapping_details[inv][code] = {"name": name, "option": option}

        same_run = False
        if prev_code_row == code:
            same_run = True
        else:
            lt = last_time_code.get(code)
            if lt and t and abs((t - lt).total_seconds()) <= 2:
                same_run = True

        if same_run:
            cur_run_len_code[code] += 1
            cur_run_members[code].append((inv, code))
        else:
            if cur_run_len_code[code] > 0:
                flush_run(code)
            cur_run_len_code[code] = 1
            cur_run_members[code] = [(inv, code)]

        if t:
            last_time_code[code] = t
        prev_code_row = code

    for code in list(cur_run_len_code.keys()):
        if cur_run_len_code[code] > 0:
            flush_run(code)

    mapping_runs = defaultdict(lambda: defaultdict(int))
    for (inv, code), L in run_len_per_invcode.items():
        if L > 1:
            mapping_runs[inv][code] = L

    return mapping_counts, mapping_details, mapping_runs, invoice_order, invoice_seq, code_o_text
