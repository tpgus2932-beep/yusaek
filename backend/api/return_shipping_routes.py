from __future__ import annotations

import io
import json
import os
import re
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import httpx
import xlwt
from openpyxl import load_workbook

from fastapi import APIRouter, Body, Depends, File, HTTPException, Response, UploadFile

ABLY_BASE = "https://api.a-bly.com"
LLOGIS_LOGIN_URL = "https://partner.alps.llogis.com/auth/login"
LLOGIS_BASE = "https://pid.alps.llogis.com:18210"

ABLY_EMAIL = "eostm1997@naver.com"
ABLY_PASSWORD = "!Glqgkqdldi1126"

LLOGIS_PRINCIPAL = "348867"
LLOGIS_CREDENTIAL = "1q2w3e4r5t"
LLOGIS_EMP_NO = "348867"

_EZADMIN_BASE = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_KST = timezone(timedelta(hours=9))
_BROWSER_WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_BROWSER_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_CANCEL_REASON = {
    30: "단순변심",
    31: "사이즈/색상 불만족",
    32: "상품 하자/오배송",
    1: "셀러 변경",
}

# 상품 하자/오배송, 셀러 변경 = 판매자 사유 - exchange_return_routes.py의 process_exchange_pickup이
# reason_code==2(상품 하자, 판매자 부담)를 회수신청/문자에서 제외하는 것과 같은 취지로,
# 신규반품 회수신청도 판매자 사유 건은 자동으로 넘기지 않는다.
_SELLER_FAULT_CANCEL_REASONS = {32, 1}


def build_return_shipping_router(*, get_current_user, get_db, get_setting, enqueue_sms=None, get_shared_db=None, set_setting=None):
    router = APIRouter(prefix="/return-shipping")

    def _browser_time_flag(now: datetime) -> str:
        return (
            f"{_BROWSER_WEEKDAYS[now.weekday()]} "
            f"{_BROWSER_MONTHS[now.month - 1]} "
            f"{now.day:02d} {now.year} "
            f"{now:%H:%M:%S} GMT+0900 (한국 표준시)"
        )

    def _looks_like_ezadmin_session_error(response, body: str) -> bool:
        lowered = (body or "").lower()
        if response.url and "login" in str(response.url).lower():
            return True
        if "<html" in lowered or "<!doctype html" in lowered:
            return True
        return any(t in lowered for t in ("login", "phpsessid", "session", "로그인"))

    async def _ably_login() -> str:
        async with httpx.AsyncClient(timeout=15.0) as client:
            res = await client.post(
                f"{ABLY_BASE}/seller/login/",
                json={"email": ABLY_EMAIL, "password": ABLY_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": "Mozilla/5.0",
                    "Referer": "https://my.a-bly.com/",
                    "Origin": "https://my.a-bly.com",
                },
            )
            res.raise_for_status()
        token = res.json().get("token")
        if not token:
            raise HTTPException(status_code=502, detail="에이블리 로그인 실패: 토큰 없음")
        return token

    async def _llogis_login() -> str:
        async with httpx.AsyncClient(timeout=15.0, verify=False) as client:
            res = await client.post(
                LLOGIS_LOGIN_URL,
                json={
                    "principal": LLOGIS_PRINCIPAL,
                    "credential": LLOGIS_CREDENTIAL,
                    "macAddress": "normal-browser",
                },
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": "Mozilla/5.0",
                    "Referer": "https://partner.alps.llogis.com/",
                    "Origin": "https://partner.alps.llogis.com",
                },
            )
            res.raise_for_status()
        token = res.json().get("accessToken")
        if not token:
            raise HTTPException(status_code=502, detail="llogis 로그인 실패: 토큰 없음")
        return token

    async def _llogis_query(inv_no: str, token: str) -> dict:
        url = f"{LLOGIS_BASE}/pid/ftr/pacltrc/inner/bcraiinvinfo"
        params = {
            "filter": json.dumps(
                {
                    "srchInvNo": inv_no,
                    "blngBrshCd": None,
                    "empno": LLOGIS_EMP_NO,
                    "usrId": LLOGIS_EMP_NO,
                    "currPageId": "PIDFTR001U",
                    "crdFarePrntStat": "N",
                    "srchOrgInvNo": "",
                },
                ensure_ascii=False,
            ),
            "_": str(int(time.time() * 1000)),
        }
        headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "ko-KR,ko;q=0.9",
            "Authorization": token,
            "Content-Type": "application/json",
            "Host": "pid.alps.llogis.com:18210",
            "Menulink": json.dumps(
                {
                    "menuId": "21966",
                    "pgmId": "100001378",
                    "pgmUrl": f"{LLOGIS_BASE}/pid/pages/ftr/PIDFTR051U",
                }
            ),
            "Referer": f"{LLOGIS_BASE}/pid/pages/ftr/PIDFTR051U",
            "X-Requested-With": "XMLHttpRequest",
            "User-Agent": "Mozilla/5.0",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
        }
        async with httpx.AsyncClient(timeout=20.0, verify=False) as client:
            res = await client.get(url, params=params, headers=headers)
            res.raise_for_status()
        return res.json()

    def _content_disposition(filename: str) -> str:
        fallback = filename.encode("ascii", "ignore").decode() or "download.xls"
        quoted = "".join(f"%{b:02X}" for b in filename.encode("utf-8"))
        return f"attachment; filename=\"{fallback}\"; filename*=UTF-8''{quoted}"

    def _clean_invoice(value) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return ""
        if text.endswith(".0") and text[:-2].isdigit():
            text = text[:-2]
        return re.sub(r"\D", "", text)

    def _read_column_g_invoices(file_bytes: bytes, filename: str) -> list[dict]:
        ext = Path(filename or "").suffix.lower()
        rows: list[dict] = []
        if ext in {".xlsx", ".xlsm"}:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            try:
                for row_no, row in enumerate(ws.iter_rows(min_col=7, max_col=7, values_only=True), start=1):
                    invoice = _clean_invoice(row[0] if row else "")
                    if invoice:
                        rows.append({"row_no": row_no, "invoice_no": invoice})
            finally:
                wb.close()
            return rows

        if ext == ".xls":
            try:
                import pandas as pd
            except Exception as exc:
                raise HTTPException(status_code=400, detail=".xls 파일 처리를 위해 pandas/xlrd 설치가 필요합니다.") from exc
            df = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
            if df.shape[1] < 7:
                raise HTTPException(status_code=400, detail="엑셀 파일에 G열이 없습니다.")
            for idx, value in enumerate(df.iloc[:, 6].tolist(), start=1):
                invoice = _clean_invoice(value)
                if invoice:
                    rows.append({"row_no": idx, "invoice_no": invoice})
            return rows

        raise HTTPException(status_code=400, detail="xlsx/xlsm/xls 파일만 업로드 가능합니다.")

    def _build_stuck_invoice_xls(rows: list[dict]) -> bytes:
        book = xlwt.Workbook()
        sheet = book.add_sheet("stuck_invoices")
        headers = ["원본행", "송장번호", "배송상태", "위치", "최종스캔일", "사유"]
        for col, header in enumerate(headers):
            sheet.write(0, col, header)
        for row_idx, row in enumerate(rows, start=1):
            sheet.write(row_idx, 0, row.get("row_no", ""))
            sheet.write(row_idx, 1, row.get("invoice_no", ""))
            sheet.write(row_idx, 2, row.get("status", ""))
            sheet.write(row_idx, 3, row.get("location", ""))
            sheet.write(row_idx, 4, row.get("scan_date", ""))
            sheet.write(row_idx, 5, row.get("reason", ""))
        buf = io.BytesIO()
        book.save(buf)
        return buf.getvalue()

    def _is_stuck_llogis_result(data: dict) -> tuple[bool, dict]:
        inv_info_list = data.get("invInfoList") or []
        mvm_list = data.get("mvmList") or []
        if not inv_info_list:
            return True, {
                "status": "-",
                "location": "-",
                "scan_date": "-",
                "reason": "llogis에서 송장을 찾을 수 없음",
            }
        if not mvm_list:
            return True, {
                "status": "이동이력 없음",
                "location": "-",
                "scan_date": "-",
                "reason": "이동이력 없음",
            }

        latest = mvm_list[-1]
        status = latest.get("paclStatNm") or "-"
        location = latest.get("scanBrshNm") or "-"
        scan_date = latest.get("rgstYmd") or "-"
        if status == "운송장등록":
            return True, {
                "status": status,
                "location": location,
                "scan_date": scan_date,
                "reason": "운송장등록 이후 이동 없음",
            }
        return False, {
            "status": status,
            "location": location,
            "scan_date": scan_date,
            "reason": "",
        }

    @router.get("/ably-returns")
    async def get_ably_returns(
        start_date: str = None,
        end_date: str = None,
        user=Depends(get_current_user),
    ):
        if not end_date:
            end_date = datetime.today().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (datetime.today() - timedelta(days=30)).strftime("%Y-%m-%d")

        try:
            token = await _ably_login()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://my.a-bly.com/",
            "Origin": "https://my.a-bly.com",
        }

        all_items = []
        page = 1
        async with httpx.AsyncClient(timeout=30.0) as client:
            while True:
                res = await client.get(
                    f"{ABLY_BASE}/seller/order_cancels/",
                    headers=headers,
                    params={
                        "cancel_type": "return",
                        "processing_sub_status[]": ["41", "42"],
                        "delivery_type[]": ["standard", "today", "combine", "reserved"],
                        "order": "cancel_received_at",
                        "page": page,
                        "per_page": 30,
                        "start_date": start_date,
                        "end_date": end_date,
                    },
                )
                res.raise_for_status()
                data = res.json()

                cancels = data.get("order_cancels", [])
                max_page = data.get("max_page_number", 1)

                if not cancels:
                    break

                for cancel in cancels:
                    return_delivery = cancel.get("return_delivery") or {}
                    for item in cancel.get("order_items", []):
                        reason_code = item.get("cancel_reason")
                        all_items.append(
                            {
                                "상품명": item.get("goods_name"),
                                "옵션": item.get("option_info"),
                                "주문번호": item.get("order_sno"),
                                "전화번호": item.get("buyer_tel") or item.get("receiver_tel") or "",
                                "송장번호": item.get("invoice"),
                                "반품신청일시": item.get("cancel_received_at"),
                                "수취인명": item.get("receiver_name"),
                                "반품사유": _CANCEL_REASON.get(reason_code, f"기타({reason_code})"),
                                "고객메모": item.get("user_comment"),
                                "수거택배사": return_delivery.get("courier_name"),
                                "반품송장번호": return_delivery.get("invoice_number"),
                                "수거주소": return_delivery.get("address"),
                                "환불금액": item.get("refund_amount"),
                                "반품배송비": item.get("return_delivery_fee"),
                                "수량": item.get("ea") or 1,
                            }
                        )

                if page >= max_page:
                    break
                page += 1

        return {"items": all_items, "total": len(all_items)}

    @router.post("/llogis-check")
    async def check_llogis(
        invoice_nos: list[str] = Body(..., embed=True),
        user=Depends(get_current_user),
    ):
        if not invoice_nos:
            return {"results": {}}

        token = await _llogis_login()
        results = {}

        for inv_no in invoice_nos:
            inv_no = str(inv_no).strip()
            if not inv_no:
                continue
            try:
                data = await _llogis_query(inv_no, token)
                inv_info_list = data.get("invInfoList") or []
                mvm_list = data.get("mvmList") or []
                if not inv_info_list:
                    results[inv_no] = {
                        "status": "-",
                        "location": "-",
                        "scan_date": "-",
                        "error": "llogis에서 송장을 찾을 수 없음 (다른 택배사이거나 미등록 송장)",
                    }
                elif not mvm_list:
                    results[inv_no] = {
                        "status": "이동이력 없음",
                        "location": "-",
                        "scan_date": "-",
                        "error": None,
                    }
                else:
                    latest = mvm_list[-1]
                    results[inv_no] = {
                        "status": latest.get("paclStatNm") or "-",
                        "location": latest.get("scanBrshNm") or "-",
                        "scan_date": latest.get("rgstYmd") or "-",
                        "error": None,
                    }
            except Exception as e:
                results[inv_no] = {
                    "status": "-",
                    "location": "-",
                    "scan_date": "-",
                    "error": str(e),
                }

        return {"results": results}

    @router.post("/llogis-stuck-from-excel")
    async def llogis_stuck_from_excel(
        file: UploadFile = File(...),
        user=Depends(get_current_user),
    ):
        name = file.filename or "delivery_status.xlsx"
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="업로드된 파일이 비어 있습니다.")

        invoice_rows = _read_column_g_invoices(content, name)
        if not invoice_rows:
            raise HTTPException(status_code=400, detail="G열에서 송장번호를 찾지 못했습니다.")

        seen = set()
        deduped = []
        for row in invoice_rows:
            inv = row["invoice_no"]
            if inv in seen:
                continue
            seen.add(inv)
            deduped.append(row)

        token = await _llogis_login()
        stuck_rows = []
        for row in deduped:
            inv_no = row["invoice_no"]
            try:
                data = await _llogis_query(inv_no, token)
                is_stuck, info = _is_stuck_llogis_result(data)
                if is_stuck:
                    stuck_rows.append({**row, **info})
            except Exception as exc:
                stuck_rows.append({
                    **row,
                    "status": "-",
                    "location": "-",
                    "scan_date": "-",
                    "reason": f"조회 실패: {str(exc)[:120]}",
                })

        xls_bytes = _build_stuck_invoice_xls(stuck_rows)
        filename = f"{Path(name).stem}_안움직이는송장.xls"
        return Response(
            content=xls_bytes,
            media_type="application/vnd.ms-excel",
            headers={
                "Content-Disposition": _content_disposition(filename),
                "X-Total-Invoices": str(len(deduped)),
                "X-Stuck-Invoices": str(len(stuck_rows)),
            },
        )

    @router.post("/llogis-check-by-origin")
    async def check_llogis_by_origin(
        invoice_nos: list[str] = Body(..., embed=True),
        user=Depends(get_current_user),
    ):
        if not invoice_nos:
            return {"results": {}}

        try:
            token = await _llogis_login()
        except Exception as e:
            return {"results": {}, "error": f"llogis 로그인 실패: {e}"}

        results = {}

        for inv_no in invoice_nos:
            inv_no = str(inv_no).strip()
            if not inv_no:
                continue
            try:
                data = await _llogis_query(inv_no, token)
                if not (data.get("invInfoList") or []):
                    results[inv_no] = {"return_invoices": [], "error": "llogis에서 송장을 찾을 수 없음"}
                    continue

                return_raws = [
                    r for r in (data.get("rltnInvList") or [])
                    if r.get("wkSctCd") == "02"
                ]

                returns = []
                for r in return_raws:
                    rtn_no = r.get("rltnInvNo")
                    rtn_no_view = r.get("rltnInvNoView") or rtn_no
                    try:
                        rtn_data = await _llogis_query(rtn_no, token)
                        rtn_inv_info = (rtn_data.get("invInfoList") or [{}])[0]
                        rtn_mvm = rtn_data.get("mvmList") or []
                        rtn_latest = rtn_mvm[-1] if rtn_mvm else {}
                        status = (
                            rtn_inv_info.get("wkSctNm")
                            or rtn_latest.get("paclStatNm")
                            or "이동이력 없음"
                        )
                        returns.append({
                            "invoice_no": rtn_no_view,
                            "status": status,
                            "location": rtn_latest.get("scanBrshNm") or "-",
                            "scan_date": rtn_latest.get("rgstYmd") or "-",
                            "error": None,
                            "_inv_info_keys": list(rtn_inv_info.keys()),
                        })
                    except Exception as e:
                        returns.append({
                            "invoice_no": rtn_no_view,
                            "status": "-",
                            "location": "-",
                            "scan_date": "-",
                            "error": str(e),
                        })

                results[inv_no] = {"return_invoices": returns, "error": None}
            except Exception as e:
                results[inv_no] = {"return_invoices": [], "error": str(e)}

        return {"results": results}

    @router.get("/ably-shipping")
    async def get_ably_shipping(user=Depends(get_current_user)):
        try:
            token = await _ably_login()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Origin": "https://my.a-bly.com",
            "Referer": "https://my.a-bly.com/",
        }

        all_items = []
        async with httpx.AsyncClient(timeout=30.0) as client:
            res = await client.get(
                f"{ABLY_BASE}/seller/order_items/",
                headers=headers,
                params={
                    "processing_status[]": 3,
                    "processing_sub_status[]": 0,
                    "order": "goods_sent_at",
                    "delivery_type[]": ["standard", "today", "combine", "reserved"],
                    "per_page": 100,
                    "sponsorship_type": -1,
                    "page": 1,
                },
            )
            if res.status_code == 200:
                for item in res.json().get("order_items", []):
                    all_items.append({
                        "상품명": item.get("goods_name"),
                        "옵션": item.get("option_info"),
                        "주문번호": item.get("order_sno") or item.get("sno"),
                        "송장번호": item.get("invoice"),
                        "전화번호": item.get("receiver_tel") or "",
                        "발송일": item.get("goods_sent_at"),
                        "수량": item.get("ea") or 1,
                    })

        return {"items": all_items, "total": len(all_items)}

    @router.get("/llogis-detail")
    async def llogis_detail(inv_no: str, user=Depends(get_current_user)):
        token = await _llogis_login()
        data = await _llogis_query(inv_no, token)

        inv_info_list = data.get("invInfoList") or []
        if not inv_info_list:
            raise HTTPException(status_code=404, detail="llogis에서 송장을 찾을 수 없습니다.")

        inv_info = inv_info_list[0]
        mvm_list = data.get("mvmList") or []
        latest = mvm_list[-1] if mvm_list else {}

        return_invoices_raw = [
            r for r in (data.get("rltnInvList") or [])
            if r.get("wkSctCd") == "02"
        ]

        returns = []
        for r in return_invoices_raw:
            rtn_no = r.get("rltnInvNo")
            rtn_no_view = r.get("rltnInvNoView") or rtn_no
            try:
                rtn_data = await _llogis_query(rtn_no, token)
                rtn_inv_info = (rtn_data.get("invInfoList") or [{}])[0]
                rtn_mvm = rtn_data.get("mvmList") or []
                rtn_latest = rtn_mvm[-1] if rtn_mvm else {}
                latest_status = (
                    rtn_inv_info.get("wkSctNm")
                    or rtn_latest.get("paclStatNm")
                    or "-"
                )
                returns.append({
                    "invoice_no": rtn_no_view,
                    "status_name": r.get("paclStatNm"),
                    "latest_status": latest_status,
                    "location": rtn_latest.get("scanBrshNm") or "-",
                    "scan_date": rtn_latest.get("rgstYmd") or "-",
                    "error": None,
                    "_inv_info_keys": list(rtn_inv_info.keys()),
                })
            except Exception as e:
                returns.append({
                    "invoice_no": rtn_no_view,
                    "status_name": r.get("paclStatNm"),
                    "latest_status": "-",
                    "location": "-",
                    "scan_date": "-",
                    "error": str(e),
                })

        return {
            "inv_no": inv_no,
            "inv_info": {
                "receiver": inv_info.get("acperNm"),
                "product": inv_info.get("artcNm"),
                "sent_date": inv_info.get("acptRgstYmd"),
                "delivered_date": inv_info.get("dlvYmd"),
            },
            "latest_status": latest.get("paclStatNm") or "-",
            "location": latest.get("scanBrshNm") or "-",
            "scan_date": latest.get("rgstYmd") or "-",
            "returns": returns,
        }

    @router.get("/memos")
    async def get_memos(user=Depends(get_current_user)):
        conn = get_db()
        rows = conn.execute("SELECT invoice_no, memo, updated_at FROM delivery_memos").fetchall()
        conn.close()
        return {row["invoice_no"]: {"memo": row["memo"], "updated_at": row["updated_at"]} for row in rows}

    @router.post("/memo")
    async def upsert_memo(payload: dict = Body(...), user=Depends(get_current_user)):
        invoice_no = (payload.get("invoice_no") or "").strip()
        memo = (payload.get("memo") or "").strip()
        if not invoice_no:
            raise HTTPException(status_code=400, detail="invoice_no 필요")
        conn = get_db()
        if memo:
            conn.execute(
                "INSERT INTO delivery_memos (invoice_no, memo, updated_at) VALUES (?, ?, ?) "
                "ON CONFLICT(invoice_no) DO UPDATE SET memo=excluded.memo, updated_at=excluded.updated_at",
                (invoice_no, memo, datetime.now(timezone.utc).isoformat()),
            )
        else:
            conn.execute("DELETE FROM delivery_memos WHERE invoice_no = ?", (invoice_no,))
        conn.commit()
        conn.close()
        return {"ok": True}

    @router.post("/memos/cleanup")
    async def cleanup_memos(payload: dict = Body(...), user=Depends(get_current_user)):
        active_invoices: list[str] = payload.get("invoice_nos") or []
        if not active_invoices:
            return {"deleted": 0}
        conn = get_db()
        placeholders = ",".join("?" * len(active_invoices))
        cur = conn.execute(
            f"DELETE FROM delivery_memos WHERE invoice_no NOT IN ({placeholders}) AND invoice_no NOT LIKE '%:%'",
            active_invoices,
        )
        conn.commit()
        conn.close()
        return {"deleted": cur.rowcount}

    @router.post("/new-return-pickup")
    async def new_return_pickup(
        user=Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        # 실행 상태를 설정에 남겨 프론트가 새로고침/탭이동 후에도 "진행 중"임을
        # 알 수 있게 한다 - 정상/예외 종료 어느 경우든 finally에서 반드시 지운다.
        running_key = "daily_check_new_return_pickup_running_at"
        if set_setting:
            set_setting(running_key, datetime.now(_KST).isoformat())
        try:
            data = await _run_new_return_pickup(phpsessid)
            if set_setting and not data.get("need_session"):
                if data.get("ok"):
                    excluded = data.get("seller_fault_excluded") or 0
                    note = f" (판매자 부담 {excluded}건 제외)" if excluded > 0 else ""
                    message = f"송장 {data.get('invoice_count') or 0}건 처리{note}"
                else:
                    message = data.get("error") or data.get("detail") or "실패"
                set_setting("daily_check_new_return_pickup_last_result", message)
            return data
        finally:
            if set_setting:
                set_setting(running_key, None)

    async def _run_new_return_pickup(phpsessid: str):
        end_date = datetime.now(_KST).strftime("%Y-%m-%d")
        start_date = (datetime.now(_KST) - timedelta(days=30)).strftime("%Y-%m-%d")

        try:
            token = await _ably_login()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        ably_headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://my.a-bly.com/",
            "Origin": "https://my.a-bly.com",
        }

        inv_to_sno: dict[str, int] = {}
        sms_recipients: list[dict] = []  # [{tel, name, goods_name}]
        seller_fault_count = 0
        page = 1
        async with httpx.AsyncClient(timeout=30.0) as client:
            while True:
                res = await client.get(
                    f"{ABLY_BASE}/seller/order_cancels/",
                    headers=ably_headers,
                    params={
                        "cancel_type": "return",
                        "processing_sub_status[]": ["41"],
                        "delivery_type[]": ["standard", "today", "combine", "reserved"],
                        "order": "cancel_received_at",
                        "date_type": "cancel_received_at",
                        "page": page,
                        "per_page": 30,
                        "start_date": start_date,
                        "end_date": end_date,
                    },
                )
                res.raise_for_status()
                data = res.json()
                cancels = data.get("order_cancels", [])
                if not cancels:
                    break
                for cancel in cancels:
                    items = cancel.get("order_items", [])
                    # 판매자 사유(상품 하자/오배송, 셀러 변경) 건은 회수신청/문자 대상에서 제외
                    eligible_items = [
                        item for item in items
                        if item.get("cancel_reason") not in _SELLER_FAULT_CANCEL_REASONS
                    ]
                    seller_fault_count += len(items) - len(eligible_items)
                    for item in eligible_items:
                        inv = str(item.get("invoice") or "").strip()
                        sno = item.get("sno")
                        if inv and inv not in inv_to_sno:
                            inv_to_sno[inv] = sno
                    if eligible_items:
                        first = eligible_items[0]
                        tel_raw = (
                            cancel.get("buyer_tel") or cancel.get("receiver_tel") or
                            first.get("buyer_tel") or first.get("receiver_tel") or ""
                        )
                        tel = "".join(ch for ch in str(tel_raw) if ch.isdigit())
                        name_raw = (
                            cancel.get("receiver_name") or cancel.get("buyer_name") or
                            first.get("receiver_name") or first.get("buyer_name") or ""
                        )
                        if tel:
                            sms_recipients.append({
                                "tel": tel,
                                "name": str(name_raw).strip(),
                                "goods_name": str(first.get("goods_name") or "").strip(),
                            })
                if page >= data.get("max_page_number", 1):
                    break
                page += 1

        if not inv_to_sno:
            note = " (전체가 판매자 부담 사유로 제외됨)" if seller_fault_count else ""
            return {"ok": False, "error": f"처리할 송장번호가 없습니다 (sub_status=41 건 없음){note}"}

        invoices = list(inv_to_sno.keys())

        # XLS 생성 (BIFF .xls)
        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet1")
        sheet.write(0, 0, "송장번호")
        for i, inv in enumerate(invoices, start=1):
            sheet.write(i, 0, inv)
        buf = io.BytesIO()
        book.save(buf)
        xls_bytes = buf.getvalue()

        ez_headers_base = {
            "User-Agent": "Mozilla/5.0",
            "Referer": f"{_EZADMIN_BASE}/popup35.htm?template=DS05&set_batch_cs=1",
        }

        async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as ez_client:
            # EZAdmin DS05 XLS 업로드
            upload_res = await ez_client.post(
                f"{_EZADMIN_BASE}/popup35.htm",
                data={"template": "DS05", "action": "update_batch_cs", "set_batch_cs": "1", "set_order_label": ""},
                files={"_file": ("ably_return_invoice.xls", xls_bytes, "application/vnd.ms-excel")},
                cookies={"PHPSESSID": phpsessid},
                headers=ez_headers_base,
            )
            upload_body = (upload_res.text or "").strip()
            m = re.search(r"batch_cs_\w+", upload_body)
            if not m:
                if _looks_like_ezadmin_session_error(upload_res, upload_body):
                    return {"ok": False, "need_session": True}
                return {"ok": False, "error": f"table_name 추출 실패: {upload_body[:300]}"}
            table_name = m.group(0)

            # EZAdmin 회수 접수 (DS00 set_batch_cs takeback)
            now_kst = datetime.now(_KST)
            today_str = now_kst.strftime("%Y-%m-%d")
            time_flag = _browser_time_flag(now_kst)
            set_payload = {
                "template": "DS00",
                "action": "set_batch_cs",
                "work": "takeback",
                "table_name": table_name,
                "cs_reason": "일반",
                "arr_product": "[]",
                "receiver_seq": "8",
                "receiver_name": "유색",
                "receiver_tel1": "010",
                "receiver_tel2": "25466058",
                "receiver_mobile1": "010",
                "receiver_mobile2": "25466058",
                "receiver_zip1": "122",
                "receiver_zip2": "47",
                "receiver_address": "경기 남양주시 진건읍 진관로303번길 9-1 (배양리) JH대리점",
                "trans_who": "04",
                "trans_due_date": today_str,
                "timeFlag": time_flag,
                "cs_content": "",
                "seq": "",
                "cancel_pack": "0",
                "recover_pack": "0",
                "delete_pack": "0",
                "priority": "0",
                "auto_restockin_all": "0",
                "auto_restockin_all_bad": "0",
                "restockin_ex": "0",
                "update_unhold": "0",
                "unhold": "0",
                "set_cs_top_fix": "0",
            }
            set_res = await ez_client.post(
                f"{_EZADMIN_BASE}/function.htm",
                data=set_payload,
                cookies={"PHPSESSID": phpsessid},
                headers={
                    "User-Agent": "Mozilla/5.0",
                    "X-Requested-With": "XMLHttpRequest",
                    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
                    "Referer": f"{_EZADMIN_BASE}/popup35.htm?template=DS05",
                },
            )
            set_body = (set_res.text or "").strip()
            if _looks_like_ezadmin_session_error(set_res, set_body):
                return {"ok": False, "need_session": True}

        # 에이블리 반품요청접수
        sno_list = [v for v in inv_to_sno.values() if v]
        ably_status = None
        if sno_list:
            async with httpx.AsyncClient(timeout=15.0) as ably_client:
                ably_res = await ably_client.put(
                    f"{ABLY_BASE}/seller/order_items/request_return/",
                    headers={**ably_headers, "Content-Type": "application/json"},
                    json={"sno_list": sno_list},
                )
                ably_status = ably_res.status_code

        # SMS 발송 — 반품 최초 접수 템플릿
        sms_queued = 0
        if enqueue_sms and get_shared_db and sms_recipients:
            sender = os.environ.get("ALIGO_SENDER", "").strip()
            if sender:
                sconn = get_shared_db()
                try:
                    tmpl = sconn.execute(
                        "SELECT msg, title, msg_type FROM sms_templates WHERE name = ?",
                        ("반품 최초 접수",),
                    ).fetchone()
                finally:
                    sconn.close()
                if tmpl and tmpl["msg"]:
                    for r in sms_recipients:
                        msg = tmpl["msg"]
                        msg = msg.replace("{이름}", r["name"])
                        msg = msg.replace("{수령인}", r["name"])
                        msg = msg.replace("{상품명}", r["goods_name"])
                        try:
                            enqueue_sms(
                                {
                                    "receiver": r["tel"],
                                    "msg": msg,
                                    "msg_type": tmpl["msg_type"] or "LMS",
                                    "title": tmpl["title"] or "",
                                    "sender": sender,
                                },
                                "auto-return-pickup",
                            )
                            sms_queued += 1
                        except Exception:
                            pass

        # 체크리스트의 "오늘 실행됨" 표시는 실제로 회수신청/문자발송까지 끝난
        # 뒤에만 남긴다 - 도중에 브라우저가 새로고침되거나 예외가 나면 이
        # 지점까지 오지 못하므로 done_today가 그대로 false로 남는다.
        if set_setting:
            set_setting("daily_check_new_return_pickup", datetime.now(_KST).isoformat())

        return {
            "ok": True,
            "invoice_count": len(invoices),
            "table_name": table_name,
            "ably_status": ably_status,
            "sno_count": len(sno_list) if sno_list else 0,
            "sms_queued": sms_queued,
            "seller_fault_excluded": seller_fault_count,
        }

    return router
