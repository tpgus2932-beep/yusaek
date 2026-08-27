import asyncio
import tempfile
import traceback
import uuid
import warnings
warnings.filterwarnings("ignore", message="Unverified HTTPS request")
from collections import Counter
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
import json
import re
from time import perf_counter

import httpx
import pandas as pd
import xlwt
from fastapi import APIRouter, Body, Depends, File, HTTPException, Response, UploadFile
from openpyxl import load_workbook

from sdk.ably import AblyClient
from services.easyadmin_product import process_easyadmin_product_from_api
from services.order_history_store import init_order_history_table
from services.pastelco_utils import pastelco_login

_ABLY_BASE     = "https://api.a-bly.com"
_ABLY_EMAIL    = "eostm1997@naver.com"
_ABLY_PASSWORD = "!Glqgkqdldi1126"

_EZADMIN_BASE        = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_KST = ZoneInfo("Asia/Seoul")
_INCOMING_VERIFY_EXCLUDED_CLIENTS = {"케이디지", "리자드스탠다드", "리마인드", "계란속노른자", "도매킴"}
_SCHEDULE_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

from api.wonbe_routes import _get_wonbe_db as _get_wonbe_db, record_defect_process_logs


def build_barcode_router(
    *,
    get_current_user,
    get_barcode_state,
    to_int,
    process_and_load_any,
    load_excel_any,
    normalize_to_yusas,
    process_easyadmin_product_upload,
    content_disposition,
    get_shared_incoming_counts,
    set_shared_incoming_counts,
    get_shared_defect_counts,
    set_shared_defect_counts,
    get_shared_kimsungil_counts,
    set_shared_kimsungil_counts,
    set_shared_barcode_data,
    get_setting,
    set_setting,
    get_user_display,
    get_shared_db,
    get_db,
):
    router = APIRouter()
    _DEFECT_BASE_HEADERS = ["상품코드", "상품명", "공급처", "공급처상품명", "색상 사이즈", "주소", "표시형 상품명"]
    hapbae_target_shop = "에이블리(유색)"
    hapbae_checked_rows_key = "test_hapbae_checked_rows"
    hapbae_checked_rows_today_key = "test_hapbae_checked_rows_today"
    hapbae_registered_products_key = "test_hapbae_registered_products"

    def _normalize_text(value) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _normalize_shop_name(value) -> str:
        text = str(value or "")
        text = text.replace("（", "(").replace("）", ")")
        text = re.sub(r"[\s\u00a0\u200b-\u200d\ufeff]+", "", text)
        return text.strip().casefold()

    def _normalize_header(value) -> str:
        return re.sub(r"\s+", "", str(value or "")).strip().casefold()

    def _find_header_col(headers: list, aliases: list[str], fallback: int) -> int:
        normalized_aliases = [_normalize_header(alias) for alias in aliases]
        for idx, header in enumerate(headers, start=1):
            normalized_header = _normalize_header(header)
            if normalized_header and any(alias and alias in normalized_header for alias in normalized_aliases):
                return idx
        return fallback

    def _get_hapbae_checked_rows_shared() -> dict[str, bool]:
        raw = get_setting(hapbae_checked_rows_key) or "{}"
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = {}
        if not isinstance(parsed, dict):
            return {}
        clean = {}
        for key, value in parsed.items():
            if isinstance(key, str) and key.strip() and value and not key.strip().startswith("today::"):
                clean[key.strip()] = True
        return clean

    def _set_hapbae_checked_rows_shared(checked_rows: dict[str, bool]):
        clean = {
            key.strip(): True
            for key, value in checked_rows.items()
            if isinstance(key, str) and key.strip() and value and not key.strip().startswith("today::")
        }
        set_setting(hapbae_checked_rows_key, json.dumps(clean, ensure_ascii=False))
        return clean

    # TODAY 대량 섹션은 계정마다 진행 상황이 달라서, 다른 섹션(공용)과 분리해
    # 사용자별로 체크 상태를 저장한다: { username: { "today::...": true } }
    def _get_hapbae_checked_rows_today_all() -> dict[str, dict[str, bool]]:
        raw = get_setting(hapbae_checked_rows_today_key) or "{}"
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = {}
        if not isinstance(parsed, dict):
            return {}
        clean = {}
        for user_key, rows in parsed.items():
            if not isinstance(user_key, str) or not user_key.strip() or not isinstance(rows, dict):
                continue
            clean_rows = {
                key.strip(): True
                for key, value in rows.items()
                if isinstance(key, str) and key.strip() and value
            }
            if clean_rows:
                clean[user_key.strip()] = clean_rows
        return clean

    def _set_hapbae_checked_rows_today_all(all_rows: dict[str, dict[str, bool]]):
        set_setting(hapbae_checked_rows_today_key, json.dumps(all_rows, ensure_ascii=False))
        return all_rows

    def _get_hapbae_checked_rows(user: str) -> dict[str, bool]:
        merged = _get_hapbae_checked_rows_shared()
        merged.update(_get_hapbae_checked_rows_today_all().get(user, {}))
        return merged

    def _set_hapbae_checked_row(user: str, key: str, checked: bool) -> dict[str, bool]:
        if key.startswith("today::"):
            all_today = _get_hapbae_checked_rows_today_all()
            user_rows = dict(all_today.get(user, {}))
            if checked:
                user_rows[key] = True
            else:
                user_rows.pop(key, None)
            if user_rows:
                all_today[user] = user_rows
            else:
                all_today.pop(user, None)
            _set_hapbae_checked_rows_today_all(all_today)
        else:
            shared = _get_hapbae_checked_rows_shared()
            if checked:
                shared[key] = True
            else:
                shared.pop(key, None)
            _set_hapbae_checked_rows_shared(shared)
        return _get_hapbae_checked_rows(user)

    def _clear_hapbae_checked_rows():
        _set_hapbae_checked_rows_shared({})
        _set_hapbae_checked_rows_today_all({})

    def _get_registered_products() -> list[dict]:
        raw = get_setting(hapbae_registered_products_key) or "[]"
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = []
        if not isinstance(parsed, list):
            return []
        clean = []
        seen_codes = set()
        for item in parsed:
            if not isinstance(item, dict):
                continue
            code = str(item.get("code") or "").strip()
            label = str(item.get("label") or "").strip()
            if not code or code in seen_codes:
                continue
            seen_codes.add(code)
            clean.append({"code": code, "label": label})
        return clean

    def _set_registered_products(items: list[dict]) -> list[dict]:
        clean = []
        seen_codes = set()
        for item in items:
            code = str(item.get("code") or "").strip()
            label = str(item.get("label") or "").strip()
            if not code or code in seen_codes:
                continue
            seen_codes.add(code)
            clean.append({"code": code, "label": label})
        set_setting(hapbae_registered_products_key, json.dumps(clean, ensure_ascii=False))
        return clean

    def _lookup_wonbe_labels(codes: set[str]) -> dict[str, str]:
        if not codes:
            return {}
        conn = _get_wonbe_db()
        try:
            rows = conn.execute("SELECT 상품코드, 거래처, 상품명합 FROM wonbe").fetchall()
            lookup: dict[str, str] = {}
            for row in rows:
                raw_code = str(row["상품코드"] or "").strip()
                normalized = normalize_to_yusas(raw_code) or raw_code
                if normalized in codes and normalized not in lookup:
                    label = " / ".join(
                        p for p in [str(row["거래처"] or "").strip(), str(row["상품명합"] or "").strip()] if p
                    )
                    lookup[normalized] = label
            return lookup
        finally:
            conn.close()

    def _parse_dt_hapbae(v):
        if isinstance(v, datetime):
            return v
        if v is None or str(v).strip() == "":
            return None
        s = str(v).strip()
        try:
            return datetime.fromisoformat(s)
        except Exception:
            try:
                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except Exception:
                return None

    def _extract_hapbae_pre_match_rows(path: Path) -> list[dict]:
        wb, ws = load_excel_any(path)
        try:
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            code_col = _find_header_col(headers, ["상품코드", "바코드", "barcode", "code", "sku", "품번"], 8)
            name_col = _find_header_col(headers, ["상품명", "품명", "product", "name"], 9)
            option_col = _find_header_col(headers, ["옵션", "option", "옵션명"], 10)
            qty_col = _find_header_col(headers, ["수량", "주문수량", "qty", "quantity", "개수"], 11)
            rows = []
            for row_idx in range(2, ws.max_row + 1):
                raw_code = ws.cell(row_idx, code_col).value
                code = normalize_to_yusas(raw_code)
                duplicate_key = _normalize_text(ws.cell(row_idx, 13).value)
                shop = _normalize_text(ws.cell(row_idx, 4).value)
                if not duplicate_key and not shop and not code:
                    continue
                raw_trans_date = ws.cell(row_idx, 14).value
                rows.append({
                    "rowNumber": row_idx,
                    "shop": shop,
                    "duplicateKey": duplicate_key,
                    "code": code,
                    "productName": _normalize_text(ws.cell(row_idx, name_col).value),
                    "optionName": _normalize_text(ws.cell(row_idx, option_col).value),
                    "orderQty": _normalize_text(ws.cell(row_idx, qty_col).value),
                    "transDate": _normalize_text(raw_trans_date),
                    "_dt": _parse_dt_hapbae(raw_trans_date),
                    "runLen": 0,
                })

            # 스캔 화면(barcode_core.py process_and_load_any)과 동일한 연속 판정:
            # 시간순 정렬(동률이면 원본 행 순서 유지) 후, 송장번호·상품코드·수량이 모두
            # 유효한 행만 대상으로 바로 앞 유효 행이 같은 코드이거나 같은 코드가 2초 이내에
            # 다시 나오면 하나의 연속으로 간주한다 (날짜 문자열 완전일치 요구 X)
            rows.sort(key=lambda r: r["_dt"] or datetime.max)
            last_time_code: dict[str, datetime] = {}
            cur_run_len_code: dict[str, int] = {}
            cur_run_members: dict[str, list[int]] = {}
            prev_code_row = None

            def _flush_run(code: str):
                length = cur_run_len_code.get(code, 0)
                if length > 1:
                    for idx in cur_run_members.get(code, []):
                        rows[idx]["runLen"] = length
                cur_run_len_code[code] = 0
                cur_run_members[code] = []

            for idx, row in enumerate(rows):
                code = row["code"]
                qty = to_int(row.get("orderQty"), default=0)
                if not (row.get("duplicateKey") and code) or qty <= 0:
                    prev_code_row = code
                    continue
                t = row["_dt"]
                same_run = False
                if prev_code_row == code:
                    same_run = True
                else:
                    lt = last_time_code.get(code)
                    if lt and t and abs((t - lt).total_seconds()) <= 2:
                        same_run = True
                if same_run:
                    cur_run_len_code[code] = cur_run_len_code.get(code, 0) + 1
                    cur_run_members.setdefault(code, []).append(idx)
                else:
                    if cur_run_len_code.get(code, 0) > 0:
                        _flush_run(code)
                    cur_run_len_code[code] = 1
                    cur_run_members[code] = [idx]
                if t:
                    last_time_code[code] = t
                prev_code_row = code

            for code in list(cur_run_len_code.keys()):
                if cur_run_len_code.get(code, 0) > 0:
                    _flush_run(code)

            for row in rows:
                row.pop("_dt", None)
            return rows
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _build_ezadmin_order_state(ez_rows: list[dict]):
        """Build barcode state directly from the DS00 JSON response.

        This mirrors the former temporary-XLS parsing path without writing or
        reading an intermediate file.
        """
        html_re = re.compile(r"<[^>]+>")

        def clean(value):
            return html_re.sub("", str(value or "")).strip()

        def parse_qty(value):
            try:
                return int(float(clean(value) or "1"))
            except (TypeError, ValueError):
                return 1

        prepared = []
        for index, source_row in enumerate(ez_rows, start=2):
            cell = source_row.get("cell", {}) or {}
            raw_code = clean(cell.get("product_id", ""))
            code = normalize_to_yusas(raw_code)
            invoice = clean(cell.get("trans_no", ""))
            trans_date = clean(cell.get("trans_date", "") or cell.get("collect_date", ""))
            prepared.append({
                "rowNumber": index,
                "shop": clean(cell.get("shop_id", "")),
                "duplicateKey": invoice,
                "code": code,
                "productName": clean(cell.get("name", "")),
                "optionName": re.sub(r"^\[|\]$", "", clean(cell.get("p_options", ""))).strip(),
                "orderQty": parse_qty(cell.get("qty", "1")),
                "transDate": trans_date,
                "_dt": _parse_dt_hapbae(trans_date),
            })

        prepared.sort(key=lambda row: row["_dt"] or datetime.max)
        mapping = {}
        details = {}
        invoice_order = {}
        invoice_seq = []
        code_o_text = {}
        seen_invoices = set()

        last_time_code = {}
        current_run_length = {}
        current_run_members = {}
        run_length_by_invoice_code = {}
        prev_code = None

        def flush_run(code):
            length = current_run_length.get(code, 0)
            if length > 1:
                for invoice, product_code in current_run_members.get(code, []):
                    key = (invoice, product_code)
                    run_length_by_invoice_code[key] = max(run_length_by_invoice_code.get(key, 0), length)
            current_run_length[code] = 0
            current_run_members[code] = []

        for row in prepared:
            code = row["code"]
            invoice = row["duplicateKey"]
            qty = row["orderQty"]
            if not (invoice and code) or qty <= 0:
                prev_code = code
                continue

            if invoice not in seen_invoices:
                seen_invoices.add(invoice)
                invoice_seq.append(invoice)
                mapping[invoice] = {}
                details[invoice] = {}
                invoice_order[invoice] = []
            if code not in invoice_order[invoice]:
                invoice_order[invoice].append(code)
            mapping[invoice][code] = mapping[invoice].get(code, 0) + qty
            details[invoice].setdefault(code, {"name": row["productName"], "option": row["optionName"]})

            timestamp = row["_dt"]
            same_run = prev_code == code
            if not same_run:
                last_seen = last_time_code.get(code)
                same_run = bool(last_seen and timestamp and abs((timestamp - last_seen).total_seconds()) <= 2)
            if same_run:
                current_run_length[code] = current_run_length.get(code, 0) + 1
                current_run_members.setdefault(code, []).append((invoice, code))
            else:
                if current_run_length.get(code, 0):
                    flush_run(code)
                current_run_length[code] = 1
                current_run_members[code] = [(invoice, code)]
            if timestamp:
                last_time_code[code] = timestamp
            prev_code = code

        for code in list(current_run_length):
            flush_run(code)

        runs = {}
        for (invoice, code), length in run_length_by_invoice_code.items():
            runs.setdefault(invoice, {})[code] = length

        pre_match_rows = []
        for row in prepared:
            pre_match_rows.append({
                key: value for key, value in row.items() if key != "_dt"
            } | {"runLen": 0})
        for index, row in enumerate(prepared):
            code = row["code"]
            invoice = row["duplicateKey"]
            pre_match_rows[index]["runLen"] = run_length_by_invoice_code.get((invoice, code), 0)

        return mapping, details, runs, invoice_order, invoice_seq, code_o_text, pre_match_rows

    def _csv_escape(value) -> str:
        return '"' + str(value or "").replace('"', '""') + '"'

    def _defect_db_row_to_values(row) -> list[str]:
        color_size = " ".join(p for p in [str(row["색상"] or ""), str(row["사이즈"] or "")] if p.strip())
        return [
            str(row["상품코드"] or ""),
            str(row["상품명"] or ""),
            str(row["거래처"] or ""),
            str(row["거래처상품명"] or ""),
            color_size,
            str(row["거래처주소"] or ""),
            str(row["상품명합"] or ""),
        ]

    def _ws_text(ws, row_idx: int, col_idx: int) -> str:
        return str(ws.cell(row_idx, col_idx).value or "").strip()

    def _combine_color_size(color: str, size: str) -> str:
        return " ".join(part for part in [str(color or "").strip(), str(size or "").strip()] if part).strip()

    def _split_color_size(value: str) -> tuple[str, str]:
        text = str(value or "").strip()
        if not text:
            return "", ""
        parts = text.split(maxsplit=1)
        if len(parts) == 1:
            return parts[0], ""
        return parts[0], parts[1]

    def _read_defect_base_table():
        conn = _get_wonbe_db()
        try:
            db_rows = conn.execute(
                "SELECT 상품코드, 상품명, 거래처, 거래처상품명, 색상, 사이즈, 거래처주소, 상품명합"
                " FROM wonbe ORDER BY rowid ASC"
            ).fetchall()
            rows = [{"row_index": i, "values": _defect_db_row_to_values(r)} for i, r in enumerate(db_rows)]
            return _DEFECT_BASE_HEADERS, rows
        finally:
            conn.close()

    def _build_defect_base_lookup():
        conn = _get_wonbe_db()
        try:
            db_rows = conn.execute(
                "SELECT 상품코드, 상품명, 거래처, 거래처상품명, 색상, 사이즈, 거래처주소, 상품명합 FROM wonbe"
            ).fetchall()
            lookup = {}
            for row in db_rows:
                raw_code = str(row["상품코드"] or "").strip()
                normalized_code = normalize_to_yusas(raw_code) or raw_code
                if normalized_code:
                    color_size = _combine_color_size(str(row["색상"] or ""), str(row["사이즈"] or ""))
                    lookup[normalized_code] = {
                        "a": raw_code,
                        "b": str(row["상품명"] or ""),
                        "c": str(row["거래처"] or ""),
                        "d": str(row["거래처상품명"] or ""),
                        "e": color_size,
                        "f": str(row["거래처주소"] or ""),
                        "g": str(row["상품명합"] or ""),
                    }
            return lookup
        finally:
            conn.close()

    def _get_item_detail_lookup(state) -> dict[str, dict]:
        cached = state.get("_detail_lookup")
        if cached is not None:
            return cached

        lookup = {}
        details = state.get("details") or {}
        for codes in details.values():
            for code, det in codes.items():
                if code not in lookup:
                    lookup[code] = {
                        "name": det.get("name", "") or "",
                        "option": det.get("option", "") or "",
                    }
        state["_detail_lookup"] = lookup
        return lookup

    def _normalize_invoice_value(value: str | None) -> str:
        raw = str(value or "").strip().upper()
        if not raw:
            return ""
        raw = re.sub(r"\s+", "", raw)
        raw = re.sub(r"\.0+$", "", raw)
        if raw.isdigit():
            return raw
        m = re.fullmatch(r"(SB\d+)\.0+", raw)
        if m:
            return m.group(1)
        return raw

    def _get_invoice_lookup(state) -> dict[str, str]:
        cached = state.get("_invoice_lookup")
        if cached is not None:
            return cached

        lookup = {}
        for invoice in (state.get("mapping") or {}).keys():
            normalized = _normalize_invoice_value(invoice)
            if normalized and normalized not in lookup:
                lookup[normalized] = invoice
        state["_invoice_lookup"] = lookup
        return lookup

    def _resolve_invoice_key(state, invoice: str) -> str | None:
        if invoice in (state.get("mapping") or {}):
            return invoice
        normalized = _normalize_invoice_value(invoice)
        if not normalized:
            return None
        return _get_invoice_lookup(state).get(normalized)

    def _get_all_items(state, inv: str):
        mapping = state["mapping"]
        if inv not in mapping:
            return []

        codes = state["invoice_order"].get(inv) if state["invoice_order"] and inv in state["invoice_order"] else None
        if not codes:
            codes = sorted(mapping[inv].keys())

        defect_counts = get_shared_defect_counts()
        kimsungil_counts = get_shared_kimsungil_counts()
        incoming_counts = get_shared_incoming_counts() or {}
        hapbae_rows = state.get("hapbae_pre_match_rows") or []
        order_qty_by_code: dict[str, int] = {}
        _target_shop_key = _normalize_shop_name(hapbae_target_shop)
        for _row in hapbae_rows:
            if _normalize_shop_name(_row.get("shop")) != _target_shop_key:
                continue
            _code = _row.get("code") or ""
            _qty = to_int(_row.get("orderQty"), default=0)
            if _code and _qty > 0:
                order_qty_by_code[_code] = order_qty_by_code.get(_code, 0) + _qty
        details = (state.get("details") or {}).get(inv, {})
        runs = (state.get("runs") or {}).get(inv, {})
        items = []
        for code in codes:
            det = details.get(code, {})
            incoming_qty = incoming_counts.get(code, 0)
            order_qty = order_qty_by_code.get(code, 0)
            items.append(
                {
                    "code": code,
                    "name": det.get("name", "") or "",
                    "option": det.get("option", "") or "",
                    "remain": mapping[inv].get(code, 0),
                    "run_len": runs.get(code, 0),
                    "defect": defect_counts.get(code, 0),
                    "incoming": incoming_qty,
                    "stock": max(0, order_qty - incoming_qty) if order_qty > incoming_qty else 0,
                    "kimsungil_incoming": kimsungil_counts.get(code, 0),
                }
            )
        return items

    def _get_first_remaining_item(state, inv: str | None):
        if not inv:
            return None
        for item in _get_all_items(state, inv):
            if item.get("remain", 0) > 0:
                return item
        return None

    def _invoice_has_defect(state, inv: str | None):
        if not inv:
            return False
        mapping = state.get("mapping") or {}
        if inv not in mapping:
            return False
        defect_counts = get_shared_defect_counts()
        return any(defect_counts.get(code, 0) > 0 for code in mapping[inv].keys())

    def _find_item_detail_by_code(state, code: str):
        return _get_item_detail_lookup(state).get(code, {"name": "", "option": ""})

    def _get_defect_list(state):
        defect_counts = get_shared_defect_counts()
        defect_base_lookup = _build_defect_base_lookup()
        rows = []
        for code, count in sorted(defect_counts.items()):
            det = _find_item_detail_by_code(state, code)
            base_row = defect_base_lookup.get(code, {})
            rows.append(
                {
                    "code": code,
                    "count": count,
                    "name": det.get("name", ""),
                    "option": det.get("option", ""),
                    "base_vendor": base_row.get("c", ""),
                    "base_product": base_row.get("d", ""),
                    "base_color": base_row.get("e", ""),
                    "base_addr": base_row.get("f", ""),
                    "base_name": base_row.get("g", ""),
                    "base_option": " " if base_row.get("g", "") else "",
                }
            )
        return rows

    def _get_kimsungil_list(state):
        kimsungil_counts = get_shared_kimsungil_counts()
        incoming_counts = get_shared_incoming_counts() or {}
        defect_base_lookup = _build_defect_base_lookup()
        rows = []
        for code, count in sorted(kimsungil_counts.items()):
            det = _find_item_detail_by_code(state, code)
            base_row = defect_base_lookup.get(code, {})
            rows.append(
                {
                    "code": code,
                    "count": count,
                    "incoming_qty": int(incoming_counts.get(code, 0) or 0),
                    "name": det.get("name", ""),
                    "option": det.get("option", ""),
                    "base_vendor": base_row.get("c", ""),
                    "base_product": base_row.get("d", ""),
                    "base_color": base_row.get("e", ""),
                    "base_addr": base_row.get("f", ""),
                    "base_name": base_row.get("g", ""),
                    "base_option": " " if base_row.get("g", "") else "",
                }
            )
        return rows

    def _kimsungil_display_name(state, code: str) -> str:
        det = _find_item_detail_by_code(state, code)
        base_row = _build_defect_base_lookup().get(code, {})
        return base_row.get("g") or det.get("name") or ""

    def _log_kimsungil_event(*, code: str, name: str, action: str, method: str, count_after: int, user: str):
        now = datetime.now(_KST).isoformat()
        conn = get_shared_db()
        try:
            conn.execute(
                """
                INSERT INTO kimsungil_log
                    (created_at, code, name, action, method, count_after, username, display_name)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (now, code, name, action, method, count_after, user, get_user_display(user)),
            )
            conn.commit()
        finally:
            conn.close()

    def _build_defect_csv(state) -> str:
        defect_counts = get_shared_defect_counts()
        defect_base_lookup = _build_defect_base_lookup()
        lines = ["vendor,product,color,count,address"]
        for code, count in sorted(defect_counts.items()):
            det = _find_item_detail_by_code(state, code)
            base_row = defect_base_lookup.get(code, {})
            row = [
                base_row.get("c") or det.get("name", "") or "",
                base_row.get("d") or "",
                base_row.get("e") or det.get("option", "") or "",
                str(count),
                base_row.get("f") or "",
            ]
            lines.append(",".join(_csv_escape(value) for value in row))
        return "\n".join(lines) + "\n"

    def _to_s_code(code: str) -> str:
        value = str(code or "").strip()
        return f"S{value[5:]}" if value.startswith("YUSAS") else value

    def _build_defect_xls_bytes() -> bytes:
        defect_counts = get_shared_defect_counts()
        book = xlwt.Workbook()
        sheet = book.add_sheet("defects")
        header_style = xlwt.easyxf("font: bold on; align: horiz center;")
        sheet.write(0, 0, "상품코드", header_style)
        sheet.write(0, 1, "작업수량", header_style)
        sheet.write(0, 2, "메모", header_style)
        for row_idx, (code, count) in enumerate(sorted(defect_counts.items()), start=1):
            sheet.write(row_idx, 0, _to_s_code(code))
            sheet.write(row_idx, 1, int(count or 0))
            sheet.write(row_idx, 2, "불량")
        sheet.col(0).width = 20 * 256
        sheet.col(1).width = 12 * 256
        sheet.col(2).width = 16 * 256
        buf = tempfile.SpooledTemporaryFile()
        book.save(buf)
        buf.seek(0)
        data = buf.read()
        buf.close()
        return data

    def _copy_original_mapping(mapping) -> dict:
        return {
            invoice: {code: int(qty or 0) for code, qty in codes.items()}
            for invoice, codes in (mapping or {}).items()
        }

    def _is_invoice_completed(state, invoice: str) -> bool:
        codes = (state.get("mapping") or {}).get(invoice) or {}
        return bool(codes) and all(int(remain or 0) == 0 for remain in codes.values())

    def _build_completed_xls_bytes(state) -> tuple[bytes, int]:
        original_mapping = state.get("original_mapping") or {}
        details = state.get("details") or {}
        invoice_order = state.get("invoice_order") or {}
        invoice_seq = state.get("invoice_seq") or list(original_mapping.keys())

        book = xlwt.Workbook()
        sheet = book.add_sheet("completed")
        header_style = xlwt.easyxf("font: bold on; align: horiz center;")
        headers = ["송장번호", "상품코드", "상품명", "옵션", "수량"]
        for col_idx, header in enumerate(headers):
            sheet.write(0, col_idx, header, header_style)

        row_idx = 1
        completed_count = 0
        seen = set()
        invoice_seq_set = set(invoice_seq)
        ordered_invoices = list(invoice_seq) + [
            invoice for invoice in original_mapping.keys() if invoice not in invoice_seq_set
        ]
        for invoice in ordered_invoices:
            if invoice in seen:
                continue
            seen.add(invoice)
            if not _is_invoice_completed(state, invoice):
                continue
            completed_count += 1
            codes = invoice_order.get(invoice) or list((original_mapping.get(invoice) or {}).keys())
            written_codes = set()
            for code in codes:
                if code in written_codes:
                    continue
                written_codes.add(code)
                qty = int((original_mapping.get(invoice) or {}).get(code, 0) or 0)
                if qty <= 0:
                    continue
                det = (details.get(invoice) or {}).get(code, {})
                sheet.write(row_idx, 0, str(invoice))
                sheet.write(row_idx, 1, str(code))
                sheet.write(row_idx, 2, det.get("name", "") or "")
                sheet.write(row_idx, 3, det.get("option", "") or "")
                sheet.write(row_idx, 4, qty)
                row_idx += 1

        sheet.col(0).width = 22 * 256
        sheet.col(1).width = 18 * 256
        sheet.col(2).width = 36 * 256
        sheet.col(3).width = 36 * 256
        sheet.col(4).width = 10 * 256
        buf = tempfile.SpooledTemporaryFile()
        book.save(buf)
        buf.seek(0)
        data = buf.read()
        buf.close()
        return data, completed_count

    def _build_orders_xls_bytes(state) -> tuple[bytes, int]:
        original_mapping = state.get("original_mapping") or {}
        details = state.get("details") or {}
        invoice_order = state.get("invoice_order") or {}
        invoice_seq = state.get("invoice_seq") or list(original_mapping.keys())

        book = xlwt.Workbook()
        sheet = book.add_sheet("orders")
        header_style = xlwt.easyxf("font: bold on; align: horiz center;")
        headers = ["송장번호", "상품코드", "상품명", "옵션", "수량"]
        for col_idx, header in enumerate(headers):
            sheet.write(0, col_idx, header, header_style)

        row_idx = 1
        seen = set()
        invoice_seq_set = set(invoice_seq)
        ordered_invoices = list(invoice_seq) + [
            invoice for invoice in original_mapping.keys() if invoice not in invoice_seq_set
        ]
        for invoice in ordered_invoices:
            if invoice in seen:
                continue
            seen.add(invoice)
            codes = invoice_order.get(invoice) or list((original_mapping.get(invoice) or {}).keys())
            written_codes = set()
            for code in codes:
                if code in written_codes:
                    continue
                written_codes.add(code)
                qty = int((original_mapping.get(invoice) or {}).get(code, 0) or 0)
                if qty <= 0:
                    continue
                det = (details.get(invoice) or {}).get(code, {})
                sheet.write(row_idx, 0, str(invoice))
                sheet.write(row_idx, 1, str(code))
                sheet.write(row_idx, 2, det.get("name", "") or "")
                sheet.write(row_idx, 3, det.get("option", "") or "")
                sheet.write(row_idx, 4, qty)
                row_idx += 1

        sheet.col(0).width = 22 * 256
        sheet.col(1).width = 18 * 256
        sheet.col(2).width = 36 * 256
        sheet.col(3).width = 36 * 256
        sheet.col(4).width = 10 * 256
        buf = tempfile.SpooledTemporaryFile()
        book.save(buf)
        buf.seek(0)
        data = buf.read()
        buf.close()
        return data, len(seen)

    @router.post("/barcode/upload")
    async def barcode_upload(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        name = (file.filename or "").lower()
        if not (name.endswith(".xls") or name.endswith(".xlsx")):
            raise HTTPException(status_code=400, detail="Only xls/xlsx files are allowed")

        suffix = ".xlsx" if name.endswith(".xlsx") else ".xls"
        tmp_path = Path(tempfile.gettempdir()) / f"yusaek_upload_{uuid.uuid4().hex}{suffix}"
        data = await file.read()
        tmp_path.write_bytes(data)

        try:
            hapbae_pre_match_rows = _extract_hapbae_pre_match_rows(tmp_path)
            result = process_and_load_any(tmp_path)
            if len(result) == 7:
                processed_path, mapping, details, runs, invoice_order, invoice_seq, code_o_text = result
            elif len(result) == 6:
                mapping, details, runs, invoice_order, invoice_seq, code_o_text = result
                processed_path = None
            else:
                raise RuntimeError(f"unexpected return count: {len(result)}")
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Barcode upload processing failed: {exc}")

        set_shared_barcode_data(
            {
                "loaded": True,
                "processed_path": str(processed_path) if processed_path else None,
                "mapping": mapping,
                "original_mapping": _copy_original_mapping(mapping),
                "details": details,
                "runs": runs,
                "invoice_order": invoice_order,
                "invoice_seq": invoice_seq,
                "code_o_text": code_o_text,
                "hapbae_pre_match_rows": hapbae_pre_match_rows,
                "_detail_lookup": None,
                "_invoice_lookup": None,
            }
        )
        _clear_hapbae_checked_rows()
        return {"ok": True, "invoices": len(mapping), "codes_total": sum(len(v) for v in mapping.values())}

    @router.post("/barcode/upload-from-ezadmin-orders")
    async def upload_from_ezadmin_orders(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today = str(payload.get("date") or datetime.now().strftime("%Y-%m-%d"))
        cookies = {"PHPSESSID": phpsessid}
        _kw = {"timeout": 600.0, "verify": False, "follow_redirects": True}

        par = (
            f"template=DS00&action=&search=1&page=1&_sort=&sort_order=&panel_open=true"
            f"&field_change=&bck_search=0&recover_delete=&date_type=trans_date"
            f"&start_date={today}&start_hour=00&end_date={today}&end_hour=23"
            f"&date_period_sel=0&option%5B%5D=seq&query_str%5B%5D="
            f"&multi_supply_group=&multi_supply=&str_supply_code=0"
            f"&status_sel=4&pack_sel=0&check_set_match=0&order_cs_sel=0&work_type=0"
            f"&checkbox_options_string=&trans_corp=99&user_area="
            f"&multi_shop_group=&multi_shop=&str_shop_code=0"
            f"&tags_string=&product_tag_include_type=1&labels_string=&order_label_include_type=1"
            f"&date_type2=0&start_date2={today}&start_hour2=00&end_date2={today}&end_hour2=23"
            f"&date_period_sel2=0&category=0&option%5B%5D=&query_str%5B%5D="
            f"&c_cs=blink&order_copy=0&create_order=0&print_enable=0&product_expect=0"
            f"&return_money_expect_price=&return_money_return_price="
            f"&trans_who=0&cs_reason=&multi_user_cs_type=&user_cs_type=0"
            f"&special_option%5B%5D=%EC%82%AC%EC%9D%80%ED%92%88%EC%84%A0%ED%83%9D"
            f"&select_field=DS00_1&download_field=DS00_file&download_type=0&include_sum=on"
        )

        try:
            request_started = perf_counter()
            async with httpx.AsyncClient(**_kw) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false",
                        "nd": str(int(datetime.now().timestamp() * 1000)),
                        "rows": "5000",
                        "page": "1",
                        "sidx": "",
                        "sord": "asc",
                        "template": "DS00",
                        "action": "grid_DS00",
                        "bck_search": "0",
                        "par": par,
                    },
                    cookies=cookies,
                    headers={
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": f"{_EZADMIN_BASE}/template40.htm?template=DS00",
                    },
                )
            response_received = perf_counter()
        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        processing_started = perf_counter()
        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}

        if "rows" not in obj:
            return {"ok": False, "need_session": True}

        ez_rows = obj.get("rows", [])
        if not ez_rows:
            return {"ok": False, "error": f"{today} 조회된 주문이 없습니다."}

        try:
            mapping, details, runs, invoice_order, invoice_seq, code_o_text, hapbae_pre_match_rows = _build_ezadmin_order_state(ez_rows)
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"주문 데이터 처리 실패: {exc}")

        set_shared_barcode_data({
            "loaded": True,
            "processed_path": None,
            "mapping": mapping,
            "original_mapping": _copy_original_mapping(mapping),
            "details": details,
            "runs": runs,
            "invoice_order": invoice_order,
            "invoice_seq": invoice_seq,
            "code_o_text": code_o_text,
            "hapbae_pre_match_rows": hapbae_pre_match_rows,
            "_detail_lookup": None,
            "_invoice_lookup": None,
        })
        _clear_hapbae_checked_rows()
        processing_finished = perf_counter()
        return {
            "ok": True,
            "invoices": len(mapping),
            "codes_total": sum(len(v) for v in mapping.values()),
            "total_rows": len(ez_rows),
            "date": today,
            "timings": {
                "api_ms": round((response_received - request_started) * 1000),
                "processing_ms": round((processing_finished - processing_started) * 1000),
                "total_ms": round((processing_finished - request_started) * 1000),
            },
        }

        _html_re = re.compile(r"<[^>]+>")

        def _strip(val):
            return _html_re.sub("", str(val or "")).strip()

        book = xlwt.Workbook()
        ws_xls = book.add_sheet("Sheet1")
        for ci, h in enumerate([
            "쇼핑몰", "주문번호", "수취인", "쇼핑몰", "", "", "",
            "상품코드", "상품명", "옵션", "수량", "", "송장번호", "주문일시",
        ]):
            ws_xls.write(0, ci, h)

        for ri, row in enumerate(ez_rows, 1):
            cell = row.get("cell", {}) or {}
            shop  = _strip(cell.get("shop_id", ""))
            pid   = _strip(cell.get("product_id", ""))
            name  = _strip(cell.get("name", ""))
            opts  = re.sub(r"^\[|\]$", "", _strip(cell.get("p_options", ""))).strip()
            trans = _strip(cell.get("trans_no", ""))
            tdate = _strip(cell.get("trans_date", "") or cell.get("collect_date", ""))
            try:
                qty_val = int(float(_strip(cell.get("qty", "1")) or "1"))
            except Exception:
                qty_val = 1

            ws_xls.write(ri, 0, shop)
            ws_xls.write(ri, 1, _strip(cell.get("order_id", "")))
            ws_xls.write(ri, 2, _strip(cell.get("recv_name", "")))
            ws_xls.write(ri, 3, shop)    # col 4: 쇼핑몰 (hapbae 고정)
            ws_xls.write(ri, 4, "")
            ws_xls.write(ri, 5, "")
            ws_xls.write(ri, 6, "")
            ws_xls.write(ri, 7, pid)     # col 8: 상품코드
            ws_xls.write(ri, 8, name)    # col 9: 상품명
            ws_xls.write(ri, 9, opts)    # col 10: 옵션
            ws_xls.write(ri, 10, qty_val) # col 11: 수량
            ws_xls.write(ri, 11, "")
            ws_xls.write(ri, 12, trans)  # col 13: 송장번호
            ws_xls.write(ri, 13, tdate)  # col 14: 주문일시

        import io as _io
        buf = _io.BytesIO()
        book.save(buf)
        xls_bytes = buf.getvalue()

        tmp_path = Path(tempfile.gettempdir()) / f"yusaek_ezadmin_orders_{uuid.uuid4().hex}.xls"
        tmp_path.write_bytes(xls_bytes)

        try:
            hapbae_pre_match_rows = _extract_hapbae_pre_match_rows(tmp_path)
            result = process_and_load_any(tmp_path)
            if len(result) == 7:
                processed_path, mapping, details, runs, invoice_order, invoice_seq, code_o_text = result
            elif len(result) == 6:
                mapping, details, runs, invoice_order, invoice_seq, code_o_text = result
                processed_path = None
            else:
                raise RuntimeError(f"unexpected return count: {len(result)}")
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"주문 데이터 처리 실패: {exc}")
        finally:
            tmp_path.unlink(missing_ok=True)

        set_shared_barcode_data({
            "loaded": True,
            "processed_path": str(processed_path) if processed_path else None,
            "mapping": mapping,
            "original_mapping": _copy_original_mapping(mapping),
            "details": details,
            "runs": runs,
            "invoice_order": invoice_order,
            "invoice_seq": invoice_seq,
            "code_o_text": code_o_text,
            "hapbae_pre_match_rows": hapbae_pre_match_rows,
            "_detail_lookup": None,
            "_invoice_lookup": None,
        })
        _clear_hapbae_checked_rows()
        return {
            "ok": True,
            "invoices": len(mapping),
            "codes_total": sum(len(v) for v in mapping.values()),
            "total_rows": len(ez_rows),
            "date": today,
        }

    @router.post("/barcode/incoming/upload")
    async def incoming_upload(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        name = (file.filename or "").lower()
        if not (name.endswith(".xls") or name.endswith(".xlsx")):
            raise HTTPException(status_code=400, detail="Only xls/xlsx files are allowed")

        suffix = ".xlsx" if name.endswith(".xlsx") else ".xls"
        tmp_path = Path(tempfile.gettempdir()) / f"yusaek_incoming_{uuid.uuid4().hex}{suffix}"
        data = await file.read()
        tmp_path.write_bytes(data)

        try:
            wb, ws = load_excel_any(tmp_path)
            counts = Counter()
            for r in range(1, ws.max_row + 1):
                code = normalize_to_yusas(ws.cell(r, 1).value)
                qty = to_int(ws.cell(r, 2).value, default=0)
                if code and qty > 0:
                    counts[code] += qty
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Incoming file load failed: {exc}")

        set_shared_incoming_counts(dict(counts))
        return {"ok": True, "codes": len(counts), "total_qty": sum(counts.values())}

    @router.post("/barcode/product/upload")
    async def easyadmin_product_upload(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        name = (file.filename or "").lower()
        if not (name.endswith(".xls") or name.endswith(".xlsx") or name.endswith(".csv")):
            raise HTTPException(status_code=400, detail="Only xls/xlsx/csv files are allowed")

        suffix = Path(name).suffix or ".xlsx"
        tmp_path = Path(tempfile.gettempdir()) / f"yusaek_easyadmin_{uuid.uuid4().hex}{suffix}"
        data = await file.read()
        tmp_path.write_bytes(data)

        try:
            xls_bytes = process_easyadmin_product_upload(tmp_path)
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"Product upload processing failed: {exc}")

        filename = f"easyadmin_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        headers = {"Content-Disposition": content_disposition(filename)}
        return Response(content=xls_bytes, media_type="application/vnd.ms-excel", headers=headers)

    @router.post("/barcode/product/upload-from-api")
    async def product_upload_from_api(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        start_date = payload.get("start_date", "")
        end_date   = payload.get("end_date", "")

        async with httpx.AsyncClient(timeout=15.0) as client:
            res = await client.post(
                f"{_ABLY_BASE}/seller/login/",
                json={"email": _ABLY_EMAIL, "password": _ABLY_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "Origin": "https://seller-admin.a-bly.com",
                    "Referer": "https://seller-admin.a-bly.com/",
                    "User-Agent": "Mozilla/5.0",
                },
            )
            if not res.is_success:
                raise HTTPException(status_code=502, detail="에이블리 로그인 실패")
        token = res.json().get("access_token") or res.json().get("token")
        if not token:
            raise HTTPException(status_code=502, detail="에이블리 로그인 실패: 토큰 없음")

        ably_headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Origin": "https://seller-admin.a-bly.com",
            "Referer": "https://seller-admin.a-bly.com/",
            "User-Agent": "Mozilla/5.0",
        }

        all_goods = []
        page = 1
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                while True:
                    res = await client.post(
                        f"{_ABLY_BASE}/seller/goods/search/",
                        headers=ably_headers,
                        json={"page": page, "per_page": 30},
                    )
                    res.raise_for_status()
                    data = res.json()
                    goods = data.get("goods", [])
                    if not goods:
                        break
                    all_goods.extend(goods)
                    if page >= data.get("max_page_number", 1):
                        break
                    page += 1
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"상품 목록 조회 실패: {exc}")

        if start_date or end_date:
            filtered = []
            for g in all_goods:
                date_str = (g.get("registered_at") or g.get("created_at") or "")[:10]
                if start_date and date_str < start_date:
                    continue
                if end_date and date_str > end_date:
                    continue
                filtered.append(g)
            all_goods = filtered

        # search API returns option_groups=null; fetch per-goods detail to get option names
        if all_goods:
            ably_client = AblyClient()
            ably_client.set_token(token)

            async def _fetch_detail(sno):
                try:
                    return sno, await ably_client.get_goods_detail(sno)
                except Exception:
                    return sno, {}

            results = await asyncio.gather(*[_fetch_detail(g["sno"]) for g in all_goods])
            detail_map = {sno: detail for sno, detail in results}
            for i, g in enumerate(all_goods):
                detail = detail_map.get(g["sno"])
                if detail:
                    all_goods[i] = {
                        **g,
                        "option_groups": detail.get("option_groups") or g.get("option_groups"),
                        "options": detail.get("options") or g.get("options"),
                    }

        try:
            xls_bytes = process_easyadmin_product_from_api(all_goods)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"XLS 생성 실패: {exc}")

        filename = f"easyadmin_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        resp_headers = {"Content-Disposition": content_disposition(filename)}
        return Response(content=xls_bytes, media_type="application/vnd.ms-excel", headers=resp_headers)

    @router.get("/barcode/status")
    def barcode_status(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            return {"loaded": False}
        invoice_keys = list((state.get("mapping") or {}).keys())
        return {
            "loaded": True,
            "current_invoice": state["current_invoice"],
            "invoices": len(state["mapping"]),
            "invoice_samples": invoice_keys[:20],
            "invoice_sample_normalized": [
                {"raw": invoice, "normalized": _normalize_invoice_value(invoice)}
                for invoice in invoice_keys[:20]
            ],
            "processed_path": state["processed_path"],
            "items": _get_all_items(state, state["current_invoice"]) if state["current_invoice"] else [],
            "current_next": _get_first_remaining_item(state, state["current_invoice"]),
            "defects": _get_defect_list(state),
            "kimsungil": _get_kimsungil_list(state),
            "invoice_has_defect": _invoice_has_defect(state, state["current_invoice"]),
            "incoming_codes": len(get_shared_incoming_counts() or {}),
            "incoming_total": sum((get_shared_incoming_counts() or {}).values()),
            "kimsungil_codes": len(get_shared_kimsungil_counts() or {}),
            "kimsungil_total": sum((get_shared_kimsungil_counts() or {}).values()),
        }

    def _ez_time_flag(now: datetime) -> str:
        """I100(set_stock_data)류 템플릿이 쓰는 브라우저 포맷 timeFlag."""
        weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        return f"{weekdays[now.weekday()]} {months[now.month - 1]} {now.day:02d} {now.year} {now:%H:%M:%S} GMT+0900 (한국 표준시)"

    def _ez_time_flag_ms(now: datetime) -> str:
        """E900(packlist_json/cancel_trans/delete_trans_no)류 템플릿이 쓰는 epoch-ms timeFlag."""
        return str(int(now.timestamp() * 1000))

    def _looks_like_ez_session_error(resp, body: str) -> bool:
        lowered = (body or "").lower()
        if resp.url and "login" in str(resp.url).lower():
            return True
        if "<html" in lowered or "<!doctype html" in lowered:
            return True
        return any(t in lowered for t in ("phpsessid", "session_error", "로그인이 필요"))

    def _extract_ably_order_items(packlist_data: dict) -> tuple[list[dict], "Counter[str]"]:
        """packlist_json 응답에서 (에이블리 주문상품, 상품코드별 출고수량)을 추출."""
        product_qty: Counter = Counter()
        items: list[dict] = []
        seen_sno: set[int] = set()
        for row in (packlist_data.get("rows") or []):
            cell = row.get("cell") or {}
            product_id = str(cell.get("product_id") or "").strip()
            if product_id:
                product_qty[product_id] += 1

            data_row_raw = cell.get("data_row")
            if not data_row_raw:
                continue
            try:
                data_row = json.loads(data_row_raw)
            except (TypeError, ValueError):
                continue

            order_id_seq = data_row.get("order_id_seq")
            if order_id_seq in (None, ""):
                continue
            try:
                sno = int(order_id_seq)
            except (TypeError, ValueError):
                continue
            if sno in seen_sno:
                continue
            seen_sno.add(sno)
            items.append({"product_id": product_id, "sno": sno})
        return items, product_qty

    def _ez_invoice_headers() -> dict:
        return {
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://ga80.ezadmin.co.kr/popup25.htm?template=E900",
        }

    @router.post("/barcode/invoice/preview")
    async def barcode_invoice_preview(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        invoice_no = str(payload.get("invoice_no") or "").strip()
        if not invoice_no:
            raise HTTPException(status_code=400, detail="송장번호가 필요합니다.")

        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        cookies = {"PHPSESSID": phpsessid}
        ez_headers = _ez_invoice_headers()

        today = datetime.now()
        start_date = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
        search_par = (
            f"pack=&history_seq=&date_type=collect_date"
            f"&start_date={start_date}&end_date={end_date}"
            f"&date_period_sel=0&search_type=0&keyword="
            f"&keyword1=&keyword2=&keyword3=&keyword4=&keyword5="
            f"&super_keyword={invoice_no}&order_status=-1&order_cs=0"
            f"&query_trans_who=0&is_gift=0&work_type=0"
            f"&labels_string=&checkbox_options_string="
        )

        try:
            async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
                # Step 0: 송장번호(super_keyword)로 주문 검색 → 내부 pack 값 추출
                r0 = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false", "nd": _ez_time_flag_ms(datetime.now()),
                        "rows": "10", "page": "1", "sidx": "", "sord": "desc",
                        "readonly": "T", "template": "E900", "action": "query_json",
                        "par": search_par,
                    },
                    cookies=cookies, headers=ez_headers,
                )
                body0 = (r0.text or "").strip()
                if _looks_like_ez_session_error(r0, body0):
                    return {"ok": False, "need_session": True}
                try:
                    search_data = r0.json()
                except Exception:
                    return {"ok": False, "need_session": True}

                pack = None
                for row in (search_data.get("rows") or []):
                    cell = row.get("cell") or {}
                    if isinstance(cell, dict) and cell.get("pack"):
                        pack = cell.get("pack")
                        break

                if not pack:
                    return {
                        "ok": True,
                        "invoice_no": invoice_no,
                        "pack": None,
                        "ezadmin_found": False,
                        "ably_found": False,
                        "products": [],
                        "sno_list": [],
                    }

                # Step 1: packlist_json (읽기전용) - 상품코드 + 에이블리 주문상품 sno 추출
                r1 = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false", "nd": _ez_time_flag_ms(datetime.now()),
                        "rows": "500", "page": "1", "sidx": "", "sord": "",
                        "readonly": "T", "template": "E900", "action": "packlist_json",
                        "pack": pack, "stock": "0", "is_masking": "0",
                        "timeFlag": _ez_time_flag_ms(datetime.now()),
                    },
                    cookies=cookies, headers=ez_headers,
                )
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"패킹리스트 조회 실패: {exc}")

        body1 = (r1.text or "").strip()
        if _looks_like_ez_session_error(r1, body1):
            return {"ok": False, "need_session": True}
        try:
            packlist_data = r1.json()
        except Exception:
            return {"ok": False, "need_session": True}

        ezadmin_found = bool(packlist_data.get("rows"))
        items, product_qty = _extract_ably_order_items(packlist_data)
        sno_list = [item["sno"] for item in items]
        ably_found = bool(sno_list)

        codes = list(product_qty.keys())
        info_map: dict[str, dict] = {}
        if codes:
            conn = _get_wonbe_db()
            try:
                placeholders = ",".join(["?"] * len(codes))
                rows = conn.execute(
                    f"SELECT 상품코드, 상품명, 색상, 사이즈 FROM wonbe WHERE 상품코드 IN ({placeholders})",
                    codes,
                ).fetchall()
                info_map = {r["상품코드"]: dict(r) for r in rows}
            finally:
                conn.close()

        products = [
            {
                "product_id": code,
                "qty": qty,
                "name": info_map.get(code, {}).get("상품명", ""),
                "color": info_map.get(code, {}).get("색상", ""),
                "size": info_map.get(code, {}).get("사이즈", ""),
            }
            for code, qty in product_qty.items()
        ]

        return {
            "ok": True,
            "invoice_no": invoice_no,
            "pack": pack,
            "ezadmin_found": ezadmin_found,
            "ably_found": ably_found,
            "products": products,
            "sno_list": sno_list,
        }

    @router.post("/barcode/invoice/cancel")
    async def barcode_invoice_cancel(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        seq = str(payload.get("pack") or payload.get("seq") or "").strip()
        if not seq:
            raise HTTPException(status_code=400, detail="pack(seq) 값이 필요합니다.")

        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        cookies = {"PHPSESSID": phpsessid}
        ez_headers = _ez_invoice_headers()

        async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
            try:
                r2 = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "template": "E900", "action": "cancel_trans",
                        "seq": seq, "content": "",
                        "timeFlag": _ez_time_flag_ms(datetime.now()),
                    },
                    cookies=cookies, headers=ez_headers,
                )
            except Exception as exc:
                raise HTTPException(status_code=502, detail=f"거래 취소 실패: {exc}")
            if _looks_like_ez_session_error(r2, (r2.text or "").strip()):
                return {"ok": False, "need_session": True}

            try:
                r3 = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "template": "E900", "action": "delete_trans_no",
                        "seq": seq, "content": "",
                        "timeFlag": _ez_time_flag_ms(datetime.now()),
                    },
                    cookies=cookies, headers=ez_headers,
                )
            except Exception as exc:
                raise HTTPException(status_code=502, detail=f"송장번호 삭제 실패: {exc}")
            if _looks_like_ez_session_error(r3, (r3.text or "").strip()):
                return {"ok": False, "need_session": True}

        return {"ok": True, "pack": seq}

    @router.post("/barcode/invoice/stock-out")
    async def barcode_invoice_stock_out(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        products = payload.get("products") or []
        memo = str(payload.get("memo") or "").strip()
        if not products:
            raise HTTPException(status_code=400, detail="출고 처리할 상품이 없습니다.")

        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        cookies = {"PHPSESSID": phpsessid}
        ez_headers = _ez_invoice_headers()

        results = []
        async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
            for p in products:
                product_id = str((p or {}).get("product_id") or "").strip()
                try:
                    qty = int((p or {}).get("qty") or 0)
                except (TypeError, ValueError):
                    qty = 0
                if not product_id or qty <= 0:
                    continue
                try:
                    sr = await client.post(
                        f"{_EZADMIN_BASE}/function.htm",
                        data={
                            "template": "I100", "action": "set_stock_data",
                            "product_id": product_id, "bad": "0", "type": "out",
                            "stock_label": "", "move_warehouse": "0",
                            "stock_unit": "stock_unit_ea", "qty": str(qty), "memo": memo,
                            "timeFlag": _ez_time_flag(datetime.now()),
                        },
                        cookies=cookies, headers=ez_headers,
                    )
                    if _looks_like_ez_session_error(sr, (sr.text or "").strip()):
                        return {"ok": False, "need_session": True}
                    results.append({"product_id": product_id, "qty": qty, "ok": sr.status_code < 400})
                except Exception as exc:
                    results.append({"product_id": product_id, "qty": qty, "ok": False, "error": str(exc)})

        return {"ok": all(r["ok"] for r in results), "results": results}

    @router.post("/barcode/invoice/rollback")
    async def barcode_invoice_rollback(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        raw_sno_list = payload.get("sno_list") or []
        sno_list = []
        for s in raw_sno_list:
            try:
                sno_list.append(int(s))
            except (TypeError, ValueError):
                continue
        if not sno_list:
            raise HTTPException(status_code=400, detail="sno_list가 비어 있습니다.")

        async with httpx.AsyncClient(timeout=15.0) as login_client:
            login_res = await login_client.post(
                f"{_ABLY_BASE}/seller/login/",
                json={"email": _ABLY_EMAIL, "password": _ABLY_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "Origin": "https://seller-admin.a-bly.com",
                    "Referer": "https://seller-admin.a-bly.com/",
                    "User-Agent": "Mozilla/5.0",
                },
            )
            if not login_res.is_success:
                raise HTTPException(status_code=502, detail="에이블리 로그인 실패")
        token = login_res.json().get("access_token") or login_res.json().get("token")
        if not token:
            raise HTTPException(status_code=502, detail="에이블리 로그인 실패: 토큰 없음")

        try:
            async with httpx.AsyncClient(timeout=20.0) as ably_client:
                rollback_res = await ably_client.put(
                    f"{_ABLY_BASE}/seller/order_items/rollback_to_prepare/",
                    headers={
                        "Authorization": f"JWT {token}",
                        "Accept": "application/json",
                        "Content-Type": "application/json",
                        "Origin": "https://my.a-bly.com",
                        "Referer": "https://my.a-bly.com/",
                        "User-Agent": "Mozilla/5.0",
                    },
                    json={"sno_list": sno_list},
                )
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"에이블리 발송관리 롤백 실패: {exc}")

        rollback_ok = rollback_res.status_code in (200, 201, 204)
        return {
            "ok": rollback_ok,
            "rollback_status": rollback_res.status_code,
            "rollback_detail": None if rollback_ok else rollback_res.text[:200],
        }

    @router.get("/barcode/hapbae-pre-match")
    def hapbae_pre_match(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state.get("loaded"):
            return {
                "ok": True,
                "loaded": False,
                "incoming_loaded": bool(get_shared_incoming_counts()),
                "rows": [],
                "stock_rows": [],
                "registered_rows": [],
                "stats": {"totalRows": 0, "targetRows": 0, "duplicateRows": 0, "incomingRows": 0, "stockRows": 0},
            }

        source_rows = state.get("hapbae_pre_match_rows") or []
        incoming_counts = get_shared_incoming_counts() or {}
        target_shop_key = _normalize_shop_name(hapbae_target_shop)
        target_rows = [row for row in source_rows if _normalize_shop_name(row.get("shop")) == target_shop_key]
        counts_by_key = Counter(
            _normalize_text(row.get("duplicateKey"))
            for row in target_rows
            if _normalize_text(row.get("duplicateKey"))
        )
        duplicate_rows = [
            row
            for row in target_rows
            if _normalize_text(row.get("duplicateKey")) and counts_by_key[_normalize_text(row.get("duplicateKey"))] >= 2
        ]
        result_rows = []
        for row in duplicate_rows:
            code = row.get("code") or ""
            order_qty = to_int(row.get("orderQty"), default=0)
            if order_qty <= 0:
                continue
            result_rows.append({
                "rowNumber": row.get("rowNumber"),
                "duplicateKey": row.get("duplicateKey", ""),
                "code": code,
                "productName": row.get("productName", ""),
                "optionName": row.get("optionName", ""),
                "orderQty": order_qty,
                "incomingQty": int(incoming_counts.get(code, 0) or 0),
            })

        def _group_hapbae_rows(target: list[dict]) -> list[dict]:
            grouped_rows = []
            grouped_lookup = {}
            for row in target:
                incoming_qty = int(row.get("incomingQty") or 0)
                if incoming_qty >= 10:
                    incoming_bucket = "high"
                elif incoming_qty > 0:
                    incoming_bucket = "normal"
                else:
                    incoming_bucket = "none"
                key = (
                    _normalize_text(row.get("productName")),
                    _normalize_text(row.get("optionName")),
                    incoming_bucket,
                )
                qty = to_int(row.get("orderQty"), default=0)
                if key not in grouped_lookup:
                    grouped_lookup[key] = {
                        "productName": row.get("productName", ""),
                        "optionName": row.get("optionName", ""),
                        "orderQty": qty,
                        "incomingQty": incoming_qty,
                    }
                    grouped_rows.append(grouped_lookup[key])
                else:
                    grouped_lookup[key]["orderQty"] += qty
            grouped_rows.sort(
                key=lambda row: (
                    _normalize_text(row.get("productName")),
                    _normalize_text(row.get("optionName")),
                )
            )
            return grouped_rows

        grouped_rows = _group_hapbae_rows(result_rows)
        grouped_stock_rows = []

        # TODAY 대량: 송장번호가 딱 1번만 나온 단건 주문 중 같은 상품이 10건 이상인 것
        unique_invoice_rows = [
            row for row in target_rows
            if _normalize_text(row.get("duplicateKey"))
            and counts_by_key[_normalize_text(row.get("duplicateKey"))] == 1
        ]
        today_bulk_lookup: dict[tuple, dict] = {}
        for row in unique_invoice_rows:
            key = (
                _normalize_text(row.get("productName")),
                _normalize_text(row.get("optionName")),
            )
            if key not in today_bulk_lookup:
                today_bulk_lookup[key] = {
                    "productName": row.get("productName", ""),
                    "optionName": row.get("optionName", ""),
                    "orderCount": 0,
                    "orderQty": 0,
                    "_codes": set(),
                    "_max_run_len": 0,
                }
            today_bulk_lookup[key]["orderCount"] += 1
            today_bulk_lookup[key]["orderQty"] += to_int(row.get("orderQty"), default=0)
            if row.get("code"):
                today_bulk_lookup[key]["_codes"].add(row["code"])
            run_len = int(row.get("runLen") or 0)
            if run_len > today_bulk_lookup[key]["_max_run_len"]:
                today_bulk_lookup[key]["_max_run_len"] = run_len
        today_bulk_rows = sorted(
            [
                {
                    **{k: v for k, v in entry.items() if k not in ("_codes", "_max_run_len")},
                    "runLen": entry["_max_run_len"],
                    "incomingQty": sum(int(incoming_counts.get(code, 0) or 0) for code in entry["_codes"]),
                }
                for entry in today_bulk_lookup.values()
                if entry["_max_run_len"] >= 10
                and any(int(incoming_counts.get(code, 0) or 0) >= 10 for code in entry["_codes"])
            ],
            key=lambda r: (-r["orderCount"], _normalize_text(r.get("productName")), _normalize_text(r.get("optionName"))),
        )

        registered_rows = []
        for item in _get_registered_products():
            code = item.get("code") or ""
            if not code:
                continue
            matches = [row for row in source_rows if (row.get("code") or "") == code]
            order_qty = sum(to_int(row.get("orderQty"), default=0) for row in matches)
            incoming_qty = int(incoming_counts.get(code, 0) or 0)
            if order_qty <= 0 or incoming_qty <= 0:
                continue
            sample = matches[0]
            registered_rows.append({
                "code": code,
                "productName": sample.get("productName", ""),
                "optionName": sample.get("optionName", ""),
                "orderQty": min(order_qty, incoming_qty),
                "incomingQty": incoming_qty,
            })

        return {
            "ok": True,
            "loaded": True,
            "incoming_loaded": bool(incoming_counts),
            "rows": grouped_rows,
            "stock_rows": grouped_stock_rows,
            "today_bulk_rows": today_bulk_rows,
            "registered_rows": registered_rows,
            "stats": {
                "totalRows": len(source_rows),
                "targetRows": len(target_rows),
                "duplicateRows": len(duplicate_rows),
                "incomingRows": len(result_rows),
                "groupedRows": len(grouped_rows),
                "stockRows": 0,
                "groupedStockRows": len(grouped_stock_rows),
            },
        }

    @router.get("/barcode/hapbae-pre-match/checked")
    def get_hapbae_pre_match_checked(user: str = Depends(get_current_user)):
        return {"ok": True, "checked_rows": _get_hapbae_checked_rows(user)}

    @router.patch("/barcode/hapbae-pre-match/checked")
    def set_hapbae_pre_match_checked(payload: dict = Body(...), user: str = Depends(get_current_user)):
        key = str(payload.get("key") or "").strip()
        if not key:
            raise HTTPException(status_code=400, detail="key required")
        checked = bool(payload.get("checked"))
        return {"ok": True, "checked_rows": _set_hapbae_checked_row(user, key, checked)}

    @router.get("/barcode/hapbae-pre-match/registered")
    def get_hapbae_registered_products(user: str = Depends(get_current_user)):
        return {"ok": True, "registered": _get_registered_products()}

    @router.post("/barcode/hapbae-pre-match/registered")
    def add_hapbae_registered_product(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw_code = str(payload.get("code") or "").strip()
        if not raw_code:
            raise HTTPException(status_code=400, detail="code required")
        code = normalize_to_yusas(raw_code) or raw_code
        label = str(payload.get("label") or "").strip()
        current = [item for item in _get_registered_products() if item["code"] != code]
        current.append({"code": code, "label": label})
        return {"ok": True, "registered": _set_registered_products(current)}

    @router.delete("/barcode/hapbae-pre-match/registered")
    def remove_hapbae_registered_product(payload: dict = Body(...), user: str = Depends(get_current_user)):
        code = str(payload.get("code") or "").strip()
        if not code:
            raise HTTPException(status_code=400, detail="code required")
        current = [item for item in _get_registered_products() if item["code"] != code]
        return {"ok": True, "registered": _set_registered_products(current)}

    @router.post("/barcode/hapbae-pre-match/registered/bulk")
    def bulk_add_hapbae_registered_products(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw_codes = payload.get("codes")
        if not isinstance(raw_codes, list):
            raise HTTPException(status_code=400, detail="codes must be a list")

        normalized_codes = []
        seen = set()
        for raw in raw_codes:
            raw_str = str(raw or "").strip()
            if not raw_str:
                continue
            code = normalize_to_yusas(raw_str) or raw_str
            if code in seen:
                continue
            seen.add(code)
            normalized_codes.append(code)
        if not normalized_codes:
            raise HTTPException(status_code=400, detail="no valid codes")

        labels = _lookup_wonbe_labels(set(normalized_codes))
        current = _get_registered_products()
        by_code = {item["code"]: item for item in current}
        added = 0
        for code in normalized_codes:
            label = labels.get(code, code)
            if code in by_code:
                by_code[code] = {"code": code, "label": label}
            else:
                by_code[code] = {"code": code, "label": label}
                added += 1
        return {
            "ok": True,
            "registered": _set_registered_products(list(by_code.values())),
            "added": added,
            "total_input": len(normalized_codes),
        }

    @router.post("/barcode/stock-bulk-fetch")
    async def stock_bulk_fetch(user: str = Depends(get_current_user)):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        now = datetime.now()
        weekday_kr = ["월", "화", "수", "목", "금", "토", "일"][now.weekday()]
        number_date = f"{now.year}-{now.month}-{now.day}"
        start_date3 = f"{now.year}년 {now.month:02d}월 {now.day:02d}일 {weekday_kr}요일"

        par = (
            f"orderReserve=0&priorityDeliv=0"
            f"&numberDate={number_date}"
            f"&start_date3={start_date3}"
            f"&number_no=1&is_download=0&orderType=2"
            f"&agencyTemplate=20004&check_address=2"
            f"&teamf_transType=tonight&kakaoT_type=0&deliv_today_type=0&hanjin_deliv_type=GN"
        )
        form_data = {
            "template": "S700",
            "action": "get_success_order",
            "_search": "false",
            "nd": str(int(now.timestamp() * 1000)),
            "rows": "999999",
            "page": "1",
            "sidx": "",
            "sord": "asc",
            "readonly": "T",
            "par": par,
        }
        ez_headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": _EZADMIN_BASE,
            "Referer": f"{_EZADMIN_BASE}/template35.htm?template=S700",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
            "X-Requested-With": "XMLHttpRequest",
        }
        cookies = {"PHPSESSID": phpsessid}

        try:
            async with httpx.AsyncClient(timeout=60.0, verify=False, follow_redirects=True) as client:
                res = await client.post(f"{_EZADMIN_BASE}/function.htm", headers=ez_headers, cookies=cookies, data=form_data)
            result = res.json()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"이지어드민 응답 오류: {e}")

        api_rows = result.get("rows", [])
        if not api_rows:
            return {"ok": True, "rows": []}

        _product_re = re.compile(r'\[([A-Z]\d+)\](.*?)(?:-\[([^\]]*)\])?\s+(\d+)개')
        counts: Counter = Counter()
        for order in api_rows:
            cell = order.get("cell", {})
            products_str = str(cell.get("products", ""))
            for m in _product_re.finditer(products_str):
                code = m.group(1).strip()
                name = m.group(2).strip().rstrip("-").strip()
                option = m.group(3).strip() if m.group(3) else ""
                qty = int(m.group(4))
                counts[(code, name, option)] += qty

        bulk_rows = [
            {"code": k[0], "productName": k[1], "optionName": k[2], "qty": v}
            for k, v in sorted(counts.items(), key=lambda x: -x[1])
            if v >= 10
        ]
        return {"ok": True, "rows": bulk_rows, "total": len(api_rows)}

    @router.post("/barcode/scan/invoice")
    def scan_invoice(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        invoice = (payload.get("invoice") or "").strip()
        if not invoice:
            raise HTTPException(status_code=400, detail="invoice is required")
        resolved_invoice = _resolve_invoice_key(state, invoice)
        if not resolved_invoice:
            return {"ok": False, "type": "invoice", "result": "NOT_FOUND", "invoice": invoice}

        state["current_invoice"] = resolved_invoice
        first_item = _get_first_remaining_item(state, resolved_invoice)
        if first_item:
            state["last_scanned_code"] = first_item.get("code")

        return {
            "ok": True,
            "type": "invoice",
            "result": "SET",
            "invoice": resolved_invoice,
            "items": _get_all_items(state, resolved_invoice),
            "current_next": first_item,
            "defects": _get_defect_list(state),
            "kimsungil": _get_kimsungil_list(state),
            "invoice_has_defect": _invoice_has_defect(state, resolved_invoice),
            "post_shipment_cancelled": _is_post_shipment_cancelled(resolved_invoice),
        }

    @router.post("/barcode/scan/item")
    def scan_item(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        inv = state["current_invoice"]
        if not inv:
            return {"ok": False, "type": "item", "result": "NO_INVOICE"}

        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        if inv not in state["mapping"]:
            return {"ok": False, "type": "item", "result": "BAD_INVOICE", "invoice": inv}

        remain = state["mapping"][inv].get(code, 0)
        det = (state["details"] or {}).get(inv, {}).get(code, {})
        name = det.get("name", "") or ""
        option = det.get("option", "") or ""

        if remain <= 0:
            return {
                "ok": True,
                "type": "item",
                "result": "FALSE",
                "invoice": inv,
                "raw": raw,
                "code": code,
                "name": name,
                "option": option,
                "remain": remain,
                "items": _get_all_items(state, inv),
                "current_next": _get_first_remaining_item(state, inv),
                "defects": _get_defect_list(state),
                "kimsungil": _get_kimsungil_list(state),
            }

        state["mapping"][inv][code] = remain - 1
        state["last_scanned_code"] = code
        all_done = all(v == 0 for v in state["mapping"][inv].values())
        return {
            "ok": True,
            "type": "item",
            "result": "TRUE",
            "invoice": inv,
            "code": code,
            "name": name,
            "option": option,
            "remain": state["mapping"][inv][code],
            "invoice_done": all_done,
            "items": _get_all_items(state, inv),
            "current_next": _get_first_remaining_item(state, inv),
            "defects": _get_defect_list(state),
            "kimsungil": _get_kimsungil_list(state),
        }

    @router.post("/barcode/defect/add")
    def add_defect(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        defect_counts = dict(get_shared_defect_counts())
        defect_counts[code] = defect_counts.get(code, 0) + 1
        set_shared_defect_counts(defect_counts)

        inv = state.get("current_invoice")
        return {
            "ok": True,
            "code": code,
            "defect_count": defect_counts[code],
            "items": _get_all_items(state, inv) if inv else [],
            "current_next": _get_first_remaining_item(state, inv),
            "defects": _get_defect_list(state),
        }

    @router.get("/barcode/defect/list")
    def list_defects(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")
        return {"ok": True, "defects": _get_defect_list(state)}

    @router.get("/barcode/defect/search")
    def search_defects(q: str = "", user: str = Depends(get_current_user)):
        query = _normalize_text(q).casefold()
        if not query:
            return {"ok": True, "rows": []}
        keywords = [kw for kw in query.split() if kw]
        conn = _get_wonbe_db()
        try:
            db_rows = conn.execute(
                "SELECT 상품코드, 상품명, 거래처, 거래처상품명, 색상, 사이즈, 거래처주소, 상품명합"
                " FROM wonbe WHERE 상품명합 IS NOT NULL AND 상품명합 != '' ORDER BY rowid ASC"
            ).fetchall()
        finally:
            conn.close()
        matches = []
        for row in db_rows:
            code = str(row["상품코드"] or "").strip()
            display_name = str(row["상품명합"] or "")
            if not code or not display_name:
                continue
            if not all(kw in _normalize_text(display_name).casefold() for kw in keywords):
                continue
            matches.append({
                "code": normalize_to_yusas(code) or code,
                "base_code": code,
                "base_name": display_name,
                "base_vendor": str(row["거래처"] or ""),
                "base_product": str(row["거래처상품명"] or ""),
                "base_color": _combine_color_size(str(row["색상"] or ""), str(row["사이즈"] or "")),
                "base_addr": str(row["거래처주소"] or ""),
            })
            if len(matches) >= 50:
                break
        return {"ok": True, "rows": matches}

    @router.get("/barcode/kimsungil/list")
    def list_kimsungil(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        return {"ok": True, "kimsungil": _get_kimsungil_list(state)}

    @router.get("/barcode/kimsungil/search")
    def search_kimsungil(q: str = "", user: str = Depends(get_current_user)):
        return search_defects(q=q, user=user)

    @router.post("/barcode/kimsungil/add")
    def add_kimsungil(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        kimsungil_counts = dict(get_shared_kimsungil_counts())
        kimsungil_counts[code] = kimsungil_counts.get(code, 0) + 1
        set_shared_kimsungil_counts(kimsungil_counts)
        _log_kimsungil_event(
            code=code,
            name=_kimsungil_display_name(state, code),
            action="add",
            method="검색 추가",
            count_after=kimsungil_counts[code],
            user=user,
        )

        inv = state.get("current_invoice")
        return {
            "ok": True,
            "code": code,
            "kimsungil_count": kimsungil_counts[code],
            "items": _get_all_items(state, inv) if inv else [],
            "current_next": _get_first_remaining_item(state, inv),
            "kimsungil": _get_kimsungil_list(state),
        }

    @router.post("/barcode/kimsungil/dec")
    def decrement_kimsungil(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        kimsungil_counts = dict(get_shared_kimsungil_counts())
        if code in kimsungil_counts:
            kimsungil_counts[code] -= 1
            if kimsungil_counts[code] <= 0:
                del kimsungil_counts[code]
        set_shared_kimsungil_counts(kimsungil_counts)
        _log_kimsungil_event(
            code=code,
            name=_kimsungil_display_name(state, code),
            action="dec",
            method="수량 차감",
            count_after=kimsungil_counts.get(code, 0),
            user=user,
        )

        inv = state.get("current_invoice")
        return {
            "ok": True,
            "kimsungil": _get_kimsungil_list(state),
            "items": _get_all_items(state, inv) if inv else [],
            "current_next": _get_first_remaining_item(state, inv),
        }

    @router.post("/barcode/kimsungil/remove")
    def remove_kimsungil(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        kimsungil_counts = dict(get_shared_kimsungil_counts())
        kimsungil_counts.pop(code, None)
        set_shared_kimsungil_counts(kimsungil_counts)
        _log_kimsungil_event(
            code=code,
            name=_kimsungil_display_name(state, code),
            action="remove",
            method="삭제",
            count_after=0,
            user=user,
        )

        inv = state.get("current_invoice")
        return {
            "ok": True,
            "kimsungil": _get_kimsungil_list(state),
            "items": _get_all_items(state, inv) if inv else [],
            "current_next": _get_first_remaining_item(state, inv),
        }

    @router.post("/barcode/kimsungil/summon-to-defect")
    def summon_kimsungil_to_defect(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        kimsungil_counts = dict(get_shared_kimsungil_counts())
        incoming_counts = get_shared_incoming_counts() or {}

        moved_codes = [code for code in kimsungil_counts if int(incoming_counts.get(code, 0) or 0) > 0]
        defect_counts = dict(get_shared_defect_counts())
        moved_total = 0
        for code in moved_codes:
            qty = kimsungil_counts.pop(code, 0)
            defect_counts[code] = defect_counts.get(code, 0) + qty
            moved_total += qty
            _log_kimsungil_event(
                code=code,
                name=_kimsungil_display_name(state, code),
                action="summon",
                method="김승일 소환술(불량 이동)",
                count_after=0,
                user=user,
            )

        set_shared_kimsungil_counts(kimsungil_counts)
        set_shared_defect_counts(defect_counts)

        return {
            "ok": True,
            "moved_codes": moved_codes,
            "moved_count": len(moved_codes),
            "moved_total": moved_total,
            "defects": _get_defect_list(state),
            "kimsungil": _get_kimsungil_list(state),
        }

    @router.get("/barcode/kimsungil/log")
    def list_kimsungil_log(code: str = "", limit: int = 200, user: str = Depends(get_current_user)):
        conditions = []
        params: list = []
        if code:
            normalized = normalize_to_yusas(code) or code
            conditions.append("code = ?")
            params.append(normalized)
        where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
        conn = get_shared_db()
        try:
            rows = conn.execute(
                f"SELECT * FROM kimsungil_log {where} ORDER BY created_at DESC, id DESC LIMIT ?",
                (*params, limit),
            ).fetchall()
        finally:
            conn.close()
        return {"ok": True, "items": [dict(row) for row in rows]}

    @router.get("/barcode/defect/export")
    def export_defects(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")
        if not get_shared_defect_counts():
            raise HTTPException(status_code=400, detail="Defect list is empty")

        filename = f"defects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        headers = {"Content-Disposition": content_disposition(filename)}
        return Response(
            content=_build_defect_csv(state).encode("utf-8-sig"),
            media_type="text/csv; charset=utf-8",
            headers=headers,
        )

    @router.get("/barcode/defect/export-xls")
    def export_defects_xls(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")
        if not get_shared_defect_counts():
            raise HTTPException(status_code=400, detail="Defect list is empty")

        filename = f"defects_work_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        headers = {"Content-Disposition": content_disposition(filename)}
        return Response(
            content=_build_defect_xls_bytes(),
            media_type="application/vnd.ms-excel",
            headers=headers,
        )

    @router.post("/barcode/defect/purchase-manager-handoff")
    def defect_purchase_manager_handoff(user: str = Depends(get_current_user)):
        defect_counts = get_shared_defect_counts()
        if not defect_counts:
            raise HTTPException(status_code=400, detail="불량 목록이 비어 있습니다.")

        def normalize_s_code(value) -> str:
            code = str(value or "").strip()
            upper = code.upper()
            if upper.startswith("YUSAS"):
                return f"S{code[5:]}"
            if upper.startswith("S"):
                return f"S{code[1:]}"
            return code

        conn = _get_wonbe_db()
        try:
            base_rows = conn.execute(
                "SELECT 상품코드, 거래처, 거래처상품명, 색상, 사이즈, 원가 FROM wonbe"
            ).fetchall()
        finally:
            conn.close()

        wonbe_by_s_code = {}
        for row in base_rows:
            code = normalize_s_code(row["상품코드"])
            if code and code not in wonbe_by_s_code:
                wonbe_by_s_code[code] = row

        rows = []
        unmatched_codes = []
        for raw_code, qty in sorted(defect_counts.items()):
            code = normalize_s_code(raw_code)
            matched = wonbe_by_s_code.get(code)
            if not matched:
                unmatched_codes.append(code or str(raw_code or ""))
                continue
            rows.append({
                "code": code,
                "vendor": str(matched["거래처"] or "").strip(),
                "productName": str(matched["거래처상품명"] or "").strip(),
                "color": str(matched["색상"] or "").strip(),
                "size": str(matched["사이즈"] or "").strip(),
                "cost": matched["원가"],
                "qty": int(qty or 0),
            })

        return {"ok": True, "rows": rows, "unmatched_codes": unmatched_codes}

    @router.post("/barcode/defect/export-to-ezadmin")
    async def defect_export_to_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        defect_counts = get_shared_defect_counts()
        if not defect_counts:
            return {"ok": False, "error": "불량 목록이 비어 있습니다."}

        defect_rows = _get_defect_list(get_barcode_state(user))
        xls_bytes = _build_defect_xls_bytes()
        cookies = {"PHPSESSID": phpsessid}
        ez_headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://ga80.ezadmin.co.kr/template40.htm?template=I210",
            "X-Requested-With": "XMLHttpRequest",
        }
        base_url = f"{_EZADMIN_BASE}/function.htm"
        ts_ms = str(int(datetime.now().timestamp() * 1000))

        try:
            async with httpx.AsyncClient(timeout=120.0, verify=False, follow_redirects=True) as client:
                # Step 1: XLS 업로드 (응답이 JSON이 아닐 수 있으므로 상태코드만 확인)
                upload_r = await client.post(
                    base_url,
                    data={"template": "I200", "action": "upload_new"},
                    files={"_file": (f"defects_work_{ts_ms}.xls", xls_bytes, "application/vnd.ms-excel")},
                    cookies=cookies,
                    headers=ez_headers,
                )
                if upload_r.status_code >= 400:
                    return {"ok": False, "error": f"업로드 실패 (HTTP {upload_r.status_code})"}

                # Step 2: 미리보기 확인
                preview_r = await client.post(
                    base_url,
                    data={
                        "_search": "false", "nd": ts_ms,
                        "rows": "99999", "page": "1", "sidx": "", "sord": "asc",
                        "template": "I200", "action": "load_template_data_new",
                    },
                    cookies=cookies,
                    headers=ez_headers,
                )
                try:
                    preview_r.json()
                except Exception:
                    return {"ok": False, "need_session": True}

                # Step 3: 출고처리 실행
                time_flag = datetime.now().strftime("%a %b %d %Y %H:%M:%S GMT+0900 (한국 표준시)")
                apply_r = await client.post(
                    base_url,
                    data={
                        "template": "I200", "action": "apply_new",
                        "bad": "0", "type": "out",
                        "move_warehouse": "0", "save_stock": "0",
                        "stock_tag": "", "timeFlag": time_flag,
                    },
                    cookies=cookies,
                    headers=ez_headers,
                )
                try:
                    apply_r.json()
                except Exception:
                    return {"ok": False, "error": "출고처리 응답 파싱 실패"}

        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        execution_id = str(uuid.uuid4())
        record_defect_process_logs(
            execution_id=execution_id,
            processed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            process_type="ezadmin_stock_out",
            rows=defect_rows,
            username=user,
            display_name=get_user_display(user),
        )
        return {"ok": True, "count": len(defect_counts), "execution_id": execution_id}

    @router.get("/barcode/completed/export-xls")
    def export_completed_xls(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        content, completed_count = _build_completed_xls_bytes(state)
        if completed_count <= 0:
            raise HTTPException(status_code=400, detail="다운로드할 완료목록이 없습니다")

        filename = f"completed_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        headers = {"Content-Disposition": content_disposition(filename)}
        return Response(content=content, media_type="application/vnd.ms-excel", headers=headers)

    @router.get("/barcode/orders/export-xls")
    def export_orders_xls(user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        content, invoice_count = _build_orders_xls_bytes(state)
        if invoice_count <= 0:
            raise HTTPException(status_code=400, detail="다운로드할 주문 목록이 없습니다")

        filename = f"extended_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        headers = {"Content-Disposition": content_disposition(filename)}
        return Response(content=content, media_type="application/vnd.ms-excel", headers=headers)

    @router.get("/barcode/defect/base")
    def get_defect_base(user: str = Depends(get_current_user)):
        headers, rows = _read_defect_base_table()
        return {"ok": True, "headers": headers, "rows": rows}

    @router.post("/barcode/defect/base")
    def save_defect_base(payload: dict = Body(...), user: str = Depends(get_current_user)):
        rows = payload.get("rows")
        if not isinstance(rows, list):
            raise HTTPException(status_code=400, detail="rows must be a list")

        conn = _get_wonbe_db()
        try:
            for row in rows:
                values = row.get("values") if isinstance(row, dict) else row
                cells = [str(v or "").strip() for v in (values or [])]
                while len(cells) < 7:
                    cells.append("")
                code = normalize_to_yusas(cells[0]) or cells[0]
                if not code:
                    continue
                color, size = _split_color_size(cells[4])
                conn.execute(
                    """UPDATE wonbe SET 상품명=?, 거래처=?, 거래처상품명=?, 색상=?, 사이즈=?,
                       거래처주소=?, 상품명합=? WHERE 상품코드=?""",
                    (cells[1], cells[2], cells[3], color, size, cells[5], cells[6], code),
                )
            conn.commit()
            headers, saved_rows = _read_defect_base_table()
            return {"ok": True, "headers": headers, "rows": saved_rows}
        finally:
            conn.close()

    @router.post("/barcode/defect/dec")
    def decrement_defect(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        defect_counts = dict(get_shared_defect_counts())
        if code in defect_counts:
            defect_counts[code] -= 1
            if defect_counts[code] <= 0:
                del defect_counts[code]
        set_shared_defect_counts(defect_counts)

        inv = state.get("current_invoice")
        return {
            "ok": True,
            "defects": _get_defect_list(state),
            "items": _get_all_items(state, inv) if inv else [],
            "current_next": _get_first_remaining_item(state, inv),
        }

    @router.post("/barcode/defect/remove")
    def remove_defect(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_barcode_state(user)
        if not state["loaded"]:
            raise HTTPException(status_code=400, detail="Upload barcode data first")

        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code is required")

        code = normalize_to_yusas(raw) or raw
        defect_counts = dict(get_shared_defect_counts())
        defect_counts.pop(code, None)
        set_shared_defect_counts(defect_counts)

        inv = state.get("current_invoice")
        return {
            "ok": True,
            "defects": _get_defect_list(state),
            "items": _get_all_items(state, inv) if inv else [],
            "current_next": _get_first_remaining_item(state, inv),
        }

    @router.post("/barcode/defect/ochuul-minus")
    async def defect_ochuul_minus(user: str = Depends(get_current_user)):
        """불량 목록 기준으로 Ably 오출 재고 차감."""
        defect_counts = get_shared_defect_counts()
        if not defect_counts:
            return {"ok": True, "matched": 0, "details": [], "message": "불량 목록이 없습니다."}

        defect_rows = _get_defect_list(get_barcode_state(user))
        code_to_sno: dict[str, int] = {}
        _wconn = _get_wonbe_db()
        try:
            for _row in _wconn.execute(
                "SELECT 상품코드, 옵션번호 FROM wonbe WHERE 옵션번호 IS NOT NULL AND 옵션번호 != ''"
            ).fetchall():
                a_val = str(_row["상품코드"] or "").strip()
                k_val = _row["옵션번호"]
                if not a_val or not k_val:
                    continue
                normalized = normalize_to_yusas(a_val) or a_val
                try:
                    code_to_sno[normalized] = int(k_val)
                except (ValueError, TypeError):
                    pass
        finally:
            _wconn.close()

        sno_to_ea: dict[int, int] = {}
        missing_option_codes: list[str] = []
        for code, count in defect_counts.items():
            sno = code_to_sno.get(code)
            if sno and count > 0:
                sno_to_ea[sno] = sno_to_ea.get(sno, 0) + count
            elif count > 0:
                missing_option_codes.append(code)

        if not sno_to_ea:
            execution_id = str(uuid.uuid4())
            early_outcomes = {
                str(row.get("code") or ""): {"applied_qty": 0, "result_status": "missing_option_number"}
                for row in defect_rows
            }
            record_defect_process_logs(
                execution_id=execution_id,
                processed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                process_type="ably_ochuul_minus",
                rows=defect_rows,
                username=user,
                display_name=get_user_display(user),
                outcomes=early_outcomes,
            )
            return {
                "ok": True,
                "matched": 0,
                "details": [],
                "missing_option_codes": missing_option_codes,
                "ably_missing_codes": [],
                "message": "옵션번호가 매칭되지 않았습니다. 원가베이스 K열을 확인하세요.",
            }
            return {"ok": True, "matched": 0, "details": [],
                    "message": "매칭된 옵션이 없습니다. 원가베이스 K열을 확인하세요."}

        try:
            token = await pastelco_login()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        ably_base = "https://api.a-bly.com"
        my_headers = {
            "Authorization": f"JWT {token}", "Accept": "application/json",
            "Origin": "https://my.a-bly.com", "Referer": "https://my.a-bly.com/",
            "User-Agent": "Mozilla/5.0",
        }
        admin_headers = {
            "Authorization": f"JWT {token}", "Accept": "application/json",
            "Content-Type": "application/json",
            "Origin": "https://seller-admin.a-bly.com",
            "Referer": "https://seller-admin.a-bly.com/",
            "User-Agent": "Mozilla/5.0",
        }

        all_opts = []
        page = 1
        async with httpx.AsyncClient(timeout=30.0) as client:
            while True:
                res = await client.get(
                    f"{ably_base}/seller/today-delivery-goods-options/",
                    headers=my_headers,
                    params={"keyword_type": "goods_name", "current_page": page, "per_page": 50},
                )
                if res.status_code != 200:
                    break
                data = res.json()
                opts = data.get("data", [])
                if not opts:
                    break
                all_opts.extend(opts)
                if page >= data.get("max_page_number", 1):
                    break
                page += 1

        updates = []
        ably_option_snos = set()
        for opt in all_opts:
            sno = opt.get("sno")
            if sno is not None:
                try:
                    ably_option_snos.add(int(sno))
                except (ValueError, TypeError):
                    pass
            if sno not in sno_to_ea:
                continue
            ea = sno_to_ea[sno]
            current = int(opt.get("stock") or 0)
            new_stock = max(0, current - ea)
            updates.append({
                "sno": sno,
                "stock_sync_code": str(opt.get("stock_sync_code") or ""),
                "delivery_type": opt.get("delivery_type", "today"),
                "stock": new_stock,
                "safety_stock": int(opt.get("safety_stock") or 0),
                "use_stock": bool(opt.get("use_stock", False)),
                "is_display": bool(opt.get("is_display", True)),
                "_goods_name": opt.get("goods_name", ""),
                "_option_name": opt.get("option_name", ""),
                "_ea_minus": ea,
                "_prev_stock": current,
            })

        if updates:
            patch_payload = [{k: v for k, v in u.items() if not k.startswith("_")} for u in updates]
            async with httpx.AsyncClient(timeout=30.0) as client:
                res = await client.patch(
                    f"{ably_base}/seller/today-delivery-goods-options/bulk-update/",
                    headers=admin_headers,
                    json={"options": patch_payload},
                )
            if res.status_code not in (200, 201, 204):
                raise HTTPException(status_code=502,
                    detail=f"bulk-update 실패 (HTTP {res.status_code}): {res.text[:300]}")

        updates_by_sno = {item["sno"]: item for item in updates}
        ably_missing_codes = [
            code for code, count in defect_counts.items()
            if count > 0 and code_to_sno.get(code) and code_to_sno[code] not in ably_option_snos
        ]
        outcomes = {}
        for row in defect_rows:
            code = str(row.get("code") or "")
            item = updates_by_sno.get(code_to_sno.get(code))
            if item:
                outcomes[code] = {
                    "applied_qty": int(row.get("count") or 0),
                    "result_status": "matched",
                    "option_sno": item["sno"],
                    "stock_before": item["_prev_stock"],
                    "stock_after": item["stock"],
                }
        for row in defect_rows:
            outcomes.setdefault(str(row.get("code") or ""), {"applied_qty": 0, "result_status": "unmatched"})
        for code in missing_option_codes:
            outcomes[code] = {"applied_qty": 0, "result_status": "missing_option_number"}
        for code in ably_missing_codes:
            outcomes[code] = {"applied_qty": 0, "result_status": "not_in_ably_today"}
        execution_id = str(uuid.uuid4())
        record_defect_process_logs(
            execution_id=execution_id,
            processed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            process_type="ably_ochuul_minus",
            rows=defect_rows,
            username=user,
            display_name=get_user_display(user),
            outcomes=outcomes,
        )
        return {
            "ok": True,
            "matched": len(updates),
            "details": updates,
            "missing_option_codes": missing_option_codes,
            "ably_missing_codes": ably_missing_codes,
            "execution_id": execution_id,
        }

    @router.get("/ezadmin/session")
    def ezadmin_session_status(user: str = Depends(get_current_user)):
        phpsessid = get_setting(_EZADMIN_SESSION_KEY) or ""
        return {"ok": True, "has_session": bool(phpsessid.strip()), "phpsessid": phpsessid}

    @router.post("/ezadmin/session")
    def save_ezadmin_session(payload: dict = Body(...), user: str = Depends(get_current_user)):
        phpsessid = str(payload.get("phpsessid") or "").strip()
        if not phpsessid:
            raise HTTPException(status_code=400, detail="phpsessid is required")
        set_setting(_EZADMIN_SESSION_KEY, phpsessid)
        return {"ok": True}

    @router.get("/barcode/incoming/ezadmin-voucher-list")
    async def incoming_ezadmin_voucher_list(user: str = Depends(get_current_user)):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}
        today = datetime.now().strftime("%Y-%m-%d")
        _ez_client_kwargs = {"timeout": 20.0, "verify": False, "follow_redirects": True}
        cookies = {"PHPSESSID": phpsessid}
        try:
            async with httpx.AsyncClient(**_ez_client_kwargs) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false",
                        "nd": str(int(datetime.now().timestamp() * 1000)),
                        "rows": "9999",
                        "page": "1",
                        "sidx": "",
                        "sord": "asc",
                        "template": "IM00",
                        "action": "get_IM00_grid",
                        "par": (
                            "template=IM00&action=&page_code=IM00&search=1"
                            "&_sort=&sort_order=&date_type=crdate"
                            f"&start_date={today}&end_date={today}"
                            "&date_period_sel=0&query_option=title&query_str=&req_status=0"
                        ),
                    },
                    cookies=cookies,
                )
        except Exception:
            return {"ok": False, "need_session": True}
        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}
        if "rows" not in obj:
            return {"ok": False, "need_session": True}
        _html_tag = re.compile(r"<[^>]+>")
        vouchers = []
        for row in obj.get("rows", []):
            cell = row.get("cell", {}) or {}
            raw_sheet = str(cell.get("sheet") or "").strip()
            sheet_no = raw_sheet
            # HTML <a> 태그에서 sheet 번호 추출
            for val in cell.values():
                if not isinstance(val, str) or "<a" not in val:
                    continue
                m = re.search(r"sheet=['\"]?(\w+)['\"]?", val, re.IGNORECASE)
                if m:
                    sheet_no = m.group(1)
                break
            if not sheet_no:
                continue
            # HTML 제거한 cell 필드 정리
            clean = {}
            for k, v in cell.items():
                if isinstance(v, str):
                    stripped = _html_tag.sub("", v).strip()
                    if stripped:
                        clean[k] = stripped
                elif v is not None and v != "":
                    clean[k] = v
            vouchers.append({"sheet": sheet_no, "cell": clean})
        return {"ok": True, "date": today, "vouchers": vouchers}

    @router.post("/barcode/incoming/upload-from-ezadmin")
    async def incoming_upload_from_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        from io import BytesIO

        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today = str(payload.get("date") or datetime.now().strftime("%Y-%m-%d"))
        cookies = {"PHPSESSID": phpsessid}

        _ez_client_kwargs = {"timeout": 20.0, "verify": False, "follow_redirects": True}

        # 1) 전표 목록 조회
        try:
            async with httpx.AsyncClient(**_ez_client_kwargs) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false",
                        "nd": str(int(datetime.now().timestamp() * 1000)),
                        "rows": "9999",
                        "page": "1",
                        "sidx": "",
                        "sord": "asc",
                        "template": "IM00",
                        "action": "get_IM00_grid",
                        "par": (
                            "template=IM00&action=&page_code=IM00&search=1"
                            "&_sort=&sort_order=&date_type=crdate"
                            f"&start_date={today}&end_date={today}"
                            "&date_period_sel=0&query_option=title&query_str=&req_status=0"
                        ),
                    },
                    cookies=cookies,
                )
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"EZAdmin 연결 실패: {exc}")

        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}

        if "rows" not in obj:
            return {"ok": False, "need_session": True}

        sheet_list = [
            str(row["cell"]["sheet"])
            for row in obj.get("rows", [])
            if row.get("cell", {}).get("sheet")
        ]
        if not sheet_list:
            return {"ok": False, "need_session": False, "detail": f"{today} 전표가 없습니다."}

        # 2) 다운로드 작업 등록
        par = (
            "template=IM00&action=save_file_IM00&filename=&page_code=IM10_file_2"
            f"&sheet_list={','.join(sheet_list)}&download_type=1&select_code=IM00_file"
            f"&date_type=crdate&start_date={today}&end_date={today}&date_period_sel="
            "&multi_supply_group=undefined&multi_supply=undefined&str_supply_code=undefined"
            "&sub_domain_seq=undefined&req_status=0&query_option=title&query_str=&readonly=T"
        )
        try:
            async with httpx.AsyncClient(**_ez_client_kwargs) as client:
                await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "template": "download",
                        "action": "ins_download_worklist",
                        "work_template": "IM00",
                        "work_func": "save_file_IM00",
                        "par": par,
                    },
                    headers={"X-Requested-With": "XMLHttpRequest"},
                    cookies=cookies,
                )
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"다운로드 작업 등록 실패: {exc}")

        # 3) BL30 폴링 (최대 30회 × 2초)
        file_url = None
        for _ in range(30):
            await asyncio.sleep(2)
            try:
                async with httpx.AsyncClient(**_ez_client_kwargs) as client:
                    r = await client.post(
                        f"{_EZADMIN_BASE}/function.htm",
                        data={
                            "_search": "false",
                            "nd": str(int(datetime.now().timestamp() * 1000)),
                            "rows": "300",
                            "page": "1",
                            "sidx": "",
                            "sord": "asc",
                            "template": "BL30",
                            "action": "grid_BL30",
                            "par": (
                                "template=BL30&action=&bck_search="
                                f"&start_date={today}&start_hour=00%3A00%3A00"
                                f"&end_date={today}&end_hour=23%3A59%3A59"
                                "&date_period_sel=0"
                            ),
                        },
                        cookies=cookies,
                    )
                try:
                    bl = r.json()
                except Exception:
                    return {"ok": False, "need_session": True}
                if "rows" not in bl:
                    return {"ok": False, "need_session": True}
                for row in bl.get("rows", []):
                    cell = row.get("cell", {})
                    if cell.get("template") == "입고요청전표2" and cell.get("status") == "완료":
                        file_url = cell.get("file_name")
                        break
                if file_url:
                    break
            except Exception:
                continue

        if not file_url:
            return {"ok": False, "need_session": False, "detail": "완료된 다운로드 파일을 찾지 못했습니다."}

        # 4) 파일 다운로드 & 파싱
        try:
            async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
                r = await client.get(file_url, cookies=cookies)
            r.raise_for_status()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"파일 다운로드 실패: {exc}")

        try:
            try:
                dataframe = pd.read_excel(BytesIO(r.content), header=None, engine="xlrd")
            except Exception:
                try:
                    dataframe = pd.read_html(BytesIO(r.content), header=None)[0]
                except Exception:
                    dataframe = pd.read_csv(BytesIO(r.content), header=None, encoding="utf-8-sig")

            counts = Counter()
            for values in dataframe.itertuples(index=False, name=None):
                if len(values) < 2:
                    continue
                code = normalize_to_yusas(values[0])
                qty = to_int(values[1], default=0)
                if code and qty > 0:
                    counts[code] += qty
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"파일 파싱 실패: {exc}")

        set_shared_incoming_counts(dict(counts))
        return {"ok": True, "codes": len(counts), "total_qty": sum(counts.values())}

        suffix = ".xls"
        tmp_path = Path(tempfile.gettempdir()) / f"yusaek_ezadmin_incoming_{uuid.uuid4().hex}{suffix}"
        tmp_path.write_bytes(r.content)
        try:
            wb, ws = load_excel_any(tmp_path)
            counts = Counter()
            for row_num in range(1, ws.max_row + 1):
                code = normalize_to_yusas(ws.cell(row_num, 1).value)
                qty = to_int(ws.cell(row_num, 2).value, default=0)
                if code and qty > 0:
                    counts[code] += qty
        except Exception as exc:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"파일 파싱 실패: {exc}")
        finally:
            tmp_path.unlink(missing_ok=True)

        set_shared_incoming_counts(dict(counts))
        return {"ok": True, "codes": len(counts), "total_qty": sum(counts.values())}

    @router.post("/barcode/incoming/raw-file-from-ezadmin")
    async def incoming_raw_file_from_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today = str(payload.get("date") or datetime.now().strftime("%Y-%m-%d"))
        cookies = {"PHPSESSID": phpsessid}
        _kw = {"timeout": 20.0, "verify": False, "follow_redirects": True}

        sheet_list_override = [str(s).strip() for s in (payload.get("sheet_list") or []) if str(s).strip()]
        if sheet_list_override:
            sheet_list = sheet_list_override
        else:
            try:
                async with httpx.AsyncClient(**_kw) as client:
                    r = await client.post(
                        f"{_EZADMIN_BASE}/function.htm",
                        data={
                            "_search": "false",
                            "nd": str(int(datetime.now().timestamp() * 1000)),
                            "rows": "9999", "page": "1", "sidx": "", "sord": "asc",
                            "template": "IM00", "action": "get_IM00_grid",
                            "par": (
                                "template=IM00&action=&page_code=IM00&search=1"
                                "&_sort=&sort_order=&date_type=crdate"
                                f"&start_date={today}&end_date={today}"
                                "&date_period_sel=0&query_option=title&query_str=&req_status=0"
                            ),
                        },
                        cookies=cookies,
                    )
            except Exception as exc:
                return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

            try:
                obj = r.json()
            except Exception:
                return {"ok": False, "need_session": True}

            if "rows" not in obj:
                return {"ok": False, "need_session": True}

            sheet_list = [
                str(row["cell"]["sheet"])
                for row in obj.get("rows", [])
                if row.get("cell", {}).get("sheet")
            ]
            if not sheet_list:
                return {"ok": False, "error": f"{today} 전표가 없습니다."}

        page_code = str(payload.get("page_code") or "IM10_file_2")
        par = (
            f"template=IM00&action=save_file_IM00&filename=&page_code={page_code}"
            f"&sheet_list={','.join(sheet_list)}&download_type=1&select_code=IM00_file"
            f"&date_type=crdate&start_date={today}&end_date={today}&date_period_sel="
            "&multi_supply_group=undefined&multi_supply=undefined&str_supply_code=undefined"
            "&sub_domain_seq=undefined&req_status=0&query_option=title&query_str=&readonly=T"
        )
        try:
            async with httpx.AsyncClient(**_kw) as client:
                await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "template": "download", "action": "ins_download_worklist",
                        "work_template": "IM00", "work_func": "save_file_IM00", "par": par,
                    },
                    headers={"X-Requested-With": "XMLHttpRequest"},
                    cookies=cookies,
                )
        except Exception as exc:
            return {"ok": False, "error": f"다운로드 작업 등록 실패: {type(exc).__name__}: {exc}"}

        file_url = None
        for _ in range(30):
            await asyncio.sleep(2)
            try:
                async with httpx.AsyncClient(**_kw) as client:
                    r = await client.post(
                        f"{_EZADMIN_BASE}/function.htm",
                        data={
                            "_search": "false",
                            "nd": str(int(datetime.now().timestamp() * 1000)),
                            "rows": "300", "page": "1", "sidx": "", "sord": "desc",
                            "template": "BL30", "action": "grid_BL30",
                            "par": (
                                "template=BL30&action=&bck_search="
                                f"&start_date={today}&start_hour=00%3A00%3A00"
                                f"&end_date={today}&end_hour=23%3A59%3A59"
                                "&date_period_sel=0"
                            ),
                        },
                        cookies=cookies,
                    )
                try:
                    bl = r.json()
                except Exception:
                    return {"ok": False, "need_session": True}
                if "rows" not in bl:
                    return {"ok": False, "need_session": True}
                for row in bl.get("rows", []):
                    cell = row.get("cell", {})
                    if cell.get("template") == "입고요청전표2" and cell.get("status") == "완료":
                        file_url = cell.get("file_name")
                        break
                if file_url:
                    break
            except Exception:
                continue

        if not file_url:
            return {"ok": False, "error": "완료된 다운로드 파일을 찾지 못했습니다."}

        try:
            async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
                r = await client.get(file_url, cookies=cookies)
            r.raise_for_status()
        except Exception as exc:
            return {"ok": False, "error": f"파일 다운로드 실패: {type(exc).__name__}: {exc}"}

        from fastapi.responses import Response as FastAPIResponse
        return FastAPIResponse(
            content=r.content,
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": 'attachment; filename="incoming.xls"'},
        )

    @router.post("/barcode/base-file-from-ezadmin")
    async def base_file_from_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        from datetime import timedelta
        import io as _io
        today = datetime.now()
        start = str(payload.get("start_date") or (today - timedelta(days=90)).strftime("%Y-%m-%d"))
        end = str(payload.get("end_date") or today.strftime("%Y-%m-%d"))
        nd = str(int(today.timestamp() * 1000))

        par = (
            f"template=IO30&action=&page_code=IO00&search=1&now_page=&is_sort=&"
            f"_sort=supply_options&sort_order=1&product_qty_list=&bill_seq=&"
            f"offset_top=&work_no=&location_str=&date_type=collect_date&"
            f"start_date={start}&start_hour=00%3A00%3A00&"
            f"end_date={end}&end_hour=23%3A59%3A59&"
            f"date_period_sel=9&multi_shop_group=&multi_shop=&str_shop_code=0&"
            f"multi_supply_group=&multi_supply=&str_supply_code=0&"
            f"supply_name_search=&brand=&supply_options=&tags_string=&"
            f"product_tag_include_type=1&product_id=&name=&options=&"
            f"search_keyword_type=origin&search_keyword=&enable_stock_type=2&"
            f"order_status=3&except_soldout=1&sel_reserve_qty=none&"
            f"sel_return_qty=none&sel_lack_qty=none&sel_req_qty=none&category=0"
        )

        cookies = {"PHPSESSID": phpsessid}
        _kw = {"timeout": 30.0, "verify": False, "follow_redirects": True}

        try:
            async with httpx.AsyncClient(**_kw) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false", "nd": nd,
                        "rows": "1000", "page": "1", "sidx": "", "sord": "asc",
                        "template": "IO30", "action": "search_IO30", "par": par,
                    },
                    cookies=cookies,
                    headers={
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": "https://ga80.ezadmin.co.kr/template40.htm?template=IO30",
                    },
                )
        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}

        if "rows" not in obj:
            return {"ok": False, "need_session": True}

        def _ez_val(html_str):
            s = str(html_str or "")
            # <input ... value='X' ...> → X
            m = re.search(r"<input[^>]+\bvalue=['\"]([^'\"]*)['\"]", s, re.IGNORECASE)
            if m:
                return m.group(1)
            # <a ...>TEXT</a> → TEXT
            m = re.search(r">([^<]+)</a>", s)
            if m:
                return m.group(1).strip()
            # 태그 없으면 그대로
            return re.sub(r"<[^>]+>", "", s).strip()

        rows = [["product_name", "supply_product_name", "", "", "", "options", "request_qty", "", "product_id", "", "", "reserve_qty"]]
        for row in obj.get("rows", []):
            cell = row.get("cell", row)
            rows.append([
                _ez_val(cell.get("product_name")), _ez_val(cell.get("supply_product_name")), "", "", "",
                _ez_val(cell.get("options")), _ez_val(cell.get("request_qty")), "",
                _ez_val(cell.get("product_id")), "", "", _ez_val(cell.get("reserve_qty")),
            ])
        return {"ok": True, "rows": rows, "count": len(rows) - 1}

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        for ci, h in enumerate(["상품명", "공급처상품명", "옵션추가항목5", "옵션추가항목6", "옵션추가항목7",
                                 "옵션", "요청수량", "입고수량", "상품코드", "옵션추가항목8", "원가", "입고대기"]):
            ws.write(0, ci, h)

        for ri, row in enumerate(obj.get("rows", []), 1):
            cell = row.get("cell", row)
            ws.write(ri, 0, _ez_val(cell.get("product_name")))
            ws.write(ri, 1, _ez_val(cell.get("supply_product_name")))
            ws.write(ri, 2, ""); ws.write(ri, 3, ""); ws.write(ri, 4, "")
            ws.write(ri, 5, _ez_val(cell.get("options")))
            ws.write(ri, 6, _ez_val(cell.get("request_qty")))
            ws.write(ri, 7, "")
            ws.write(ri, 8, _ez_val(cell.get("product_id")))
            ws.write(ri, 9, ""); ws.write(ri, 10, "")
            ws.write(ri, 11, _ez_val(cell.get("reserve_qty")))

        buf = _io.BytesIO()
        wb.save(buf)

        from fastapi.responses import Response as FastAPIResponse
        return FastAPIResponse(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": 'attachment; filename="base_file.xls"'},
        )

    @router.post("/barcode/client-schedule/export-to-ezadmin")
    async def client_schedule_export_to_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        import io as _io
        today = datetime.now(_KST).date()

        def _classify(note_text: str):
            """Returns (output_note, is_tracked_date)."""
            if _SCHEDULE_DATE_RE.match(note_text):
                parsed = date.fromisoformat(note_text)
                return ("" if parsed <= today else note_text), True
            return note_text, False

        scheduled_rows = []
        for row in payload.get("rows") or []:
            code = str(row.get("productCode", ""))
            output_note, is_tracked = _classify(str(row.get("note", "")))
            scheduled_rows.append({"productCode": code, "note": output_note, "is_tracked": is_tracked})

        current_codes = {row["productCode"] for row in scheduled_rows}

        conn = get_shared_db()
        try:
            stale = conn.execute(
                "SELECT DISTINCT product_code FROM client_schedule_export_log WHERE note_date <= ?",
                (today.isoformat(),),
            ).fetchall()
        finally:
            conn.close()

        cleanup_rows = [
            {"productCode": r["product_code"], "note": "", "is_tracked": True}
            for r in stale
            if r["product_code"] not in current_codes
        ]

        all_rows = scheduled_rows + cleanup_rows
        if not all_rows:
            return {"ok": True, "count": 0}

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Sheet1")
        ws.write(0, 0, "상품코드")
        ws.write(0, 1, "상품메모")
        for ri, row in enumerate(all_rows, 1):
            ws.write(ri, 0, row["productCode"])
            ws.write(ri, 1, row["note"])

        buf = _io.BytesIO()
        wb.save(buf)
        xls_bytes = buf.getvalue()

        c620_url = f"{_EZADMIN_BASE}/template40.htm?template=C620"
        ts_ms = str(int(datetime.now().timestamp() * 1000))
        ez_headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Origin": _EZADMIN_BASE,
            "Referer": c620_url,
        }
        cookies = {"PHPSESSID": phpsessid}

        try:
            async with httpx.AsyncClient(timeout=60.0, verify=False, follow_redirects=True) as client:
                r = await client.post(
                    c620_url,
                    data={"page": "1", "action": "update2", "template": "C620", "total": "0", "status": "6"},
                    files={"_file": (f"client_schedule_{ts_ms}.xls", xls_bytes, "application/vnd.ms-excel")},
                    cookies=cookies,
                    headers=ez_headers,
                )
        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        html = r.text
        m = re.search(r'alert\("(\d+)\s*개 변경 완료 되었습니다\."\)', html)
        if not m:
            return {"ok": False, "error": "응답에서 변경 완료 문구를 찾지 못했습니다", "raw_snippet": html[:300]}

        now_iso = datetime.now(_KST).isoformat()
        conn = get_shared_db()
        try:
            for row in all_rows:
                if not row["is_tracked"]:
                    continue
                conn.execute(
                    "DELETE FROM client_schedule_export_log WHERE product_code = ?",
                    (row["productCode"],),
                )
                if row["note"]:
                    conn.execute(
                        "INSERT INTO client_schedule_export_log (product_code, note_date, exported_at) "
                        "VALUES (?, ?, ?)",
                        (row["productCode"], row["note"], now_iso),
                    )
            conn.commit()
        finally:
            conn.close()

        return {"ok": True, "count": int(m.group(1))}

    def _post_shipment_cancel_rows() -> list[dict]:
        conn = get_db()
        try:
            rows = conn.execute(
                "SELECT date, seq, invoice_no, shop, manager, ship_time, cancel_time, "
                "product_name, carrier, fetched_at FROM post_shipment_cancel ORDER BY id"
            ).fetchall()
        finally:
            conn.close()
        return [dict(row) for row in rows]

    def _is_post_shipment_cancelled(invoice_no: str) -> bool:
        value = str(invoice_no or "").strip()
        if not value:
            return False
        conn = get_db()
        try:
            row = conn.execute(
                "SELECT 1 FROM post_shipment_cancel WHERE invoice_no = ? LIMIT 1", (value,)
            ).fetchone()
        finally:
            conn.close()
        return row is not None

    @router.get("/barcode/post-shipment-cancel/list")
    def post_shipment_cancel_list(user: str = Depends(get_current_user)):
        """저장된(새로고침으로 불러온) 배송후취소 목록을 로컬 DB에서 그대로 반환.

        EZAdmin을 호출하지 않는다 - 실제 조회는 refresh 엔드포인트에서만 일어난다."""
        rows = _post_shipment_cancel_rows()
        date = rows[0]["date"] if rows else ""
        fetched_at = rows[0]["fetched_at"] if rows else ""
        return {"ok": True, "date": date, "fetched_at": fetched_at, "count": len(rows), "items": rows}

    @router.post("/barcode/post-shipment-cancel/refresh")
    async def post_shipment_cancel_refresh(user: str = Depends(get_current_user)):
        """E807(배송후취소) - 오늘 발송 후 취소된 주문의 송장번호 목록을 EZAdmin에서 다시 조회해 로컬 DB에 저장.

        실캡처(브라우저 요청) 그대로: template=E800으로 grid_E807을 호출하고,
        실제 조회조건(E807/status=8/type=cancel/오늘 날짜)은 par에 담는다."""
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today = datetime.now(_KST).strftime("%Y-%m-%d")
        par = (
            f"template=E807&action=&start_date={today}&end_date={today}"
            "&date_period_sel=0&multi_shop_group=&multi_shop=&str_shop_code=0"
            "&trans_corp=99&cancel_qty=&qty=&status=8&order_cs=0&type=cancel&view=1"
        )
        headers = {
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": f"{_EZADMIN_BASE}/template40.htm?winmode=none&template=E807&act=action&view=1",
        }
        try:
            async with httpx.AsyncClient(timeout=20.0, verify=False, follow_redirects=True) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false",
                        "nd": str(int(datetime.now().timestamp() * 1000)),
                        "rows": "9999",
                        "page": "1",
                        "sidx": "",
                        "sord": "asc",
                        "readonly": "T",
                        "template": "E800",
                        "action": "grid_E807",
                        "par": par,
                        "index": "",
                        "sort_order": "",
                    },
                    cookies={"PHPSESSID": phpsessid},
                    headers=headers,
                )
        except Exception:
            return {"ok": False, "need_session": True}
        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}
        if "rows" not in obj:
            # 오늘 배송후취소 건이 0건이면 EZAdmin이 "rows" 키 자체를 안 내려줄 때가 있다 -
            # 이걸 세션 만료로 오판하면 세션이 멀쩡한데도 계속 재입력을 요구하게 된다.
            # 진짜 세션 만료 응답에만 있는 로그인/세션 관련 문구가 있을 때만 need_session 처리.
            body_lower = (r.text or "").lower()
            if any(t in body_lower for t in ("login", "phpsessid", "session", "로그인")):
                return {"ok": False, "need_session": True}
            obj = {"rows": []}

        _html_tag = re.compile(r"<[^>]+>")
        items = []
        for row in obj.get("rows", []):
            cell = row.get("cell") or {}
            invoice_no = str(cell.get("col6") or "").strip()
            if not invoice_no:
                continue
            seq_match = re.search(r"popupcs\(\s*(\d+)", str(cell.get("col3") or ""))
            items.append({
                "seq": seq_match.group(1) if seq_match else None,
                "invoice_no": invoice_no,
                "shop": _html_tag.sub("", str(cell.get("col4") or "")).strip(),
                "manager": _html_tag.sub("", str(cell.get("col18") or "")).strip(),
                "ship_time": str(cell.get("col7") or "").strip(),
                "cancel_time": str(cell.get("col10") or "").strip(),
                "product_name": _html_tag.sub("", str(cell.get("col20") or cell.get("col15") or "")).strip(),
                "carrier": str(cell.get("col21") or "").strip(),
            })

        fetched_at = datetime.now(_KST).isoformat()
        conn = get_db()
        try:
            conn.execute("DELETE FROM post_shipment_cancel")
            conn.executemany(
                "INSERT INTO post_shipment_cancel "
                "(date, seq, invoice_no, shop, manager, ship_time, cancel_time, product_name, carrier, fetched_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [
                    (today, item["seq"], item["invoice_no"], item["shop"], item["manager"],
                     item["ship_time"], item["cancel_time"], item["product_name"], item["carrier"], fetched_at)
                    for item in items
                ],
            )
            conn.commit()
        finally:
            conn.close()

        return {"ok": True, "date": today, "fetched_at": fetched_at, "count": len(items), "items": items}

    @router.post("/barcode/incoming/verify-trans-in")
    async def barcode_incoming_verify_trans_in(user: str = Depends(get_current_user)):
        """바코드탭에서 불러온 입고 파일의 상품코드를, 이지어드민 I100(오늘 입고 거래발생) 목록과
        대조해 이지어드민 쪽에 안 잡히는 상품코드만 골라낸다.

        조회 자체는 이지어드민 실캡처(work_type=trans&work_start=1, 오늘 날짜) 그대로 -
        backend/api/inventory_dashboard_routes.py의 today_stock_check와 같은 I100/search
        호출 형태를 쓰되 par 조건(재고 stockin이 아니라 거래발생 입고)과 결과(일치 대신
        불일치)만 다르다."""
        incoming_counts = get_shared_incoming_counts() or {}
        if not incoming_counts:
            return {"ok": False, "detail": "불러온 입고 파일이 없습니다. 바코드 탭에서 입고 파일을 먼저 불러오세요."}

        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today = datetime.now(_KST).strftime("%Y-%m-%d")
        nd = str(int(datetime.now().timestamp() * 1000))
        par = (
            "auto_search=&search_all_product=&multi_supply_group=&multi_supply=&str_supply_code=0"
            "&tags_string=&product_tag_include_type=1&query_type=name&query_str="
            "&stock_type=0&stock_start=&stock_end=&notrans_day=&notrans_cnt=&notrans_status=0&stock_status=0"
            f"&start_date={today}&start_hour=00%3A00%3A00&end_date={today}&end_hour=23%3A59%3A59"
            "&date_period_sel=1&work_type=trans&work_start=1&work_end=&inout_type=0&product_date="
            f"&start_date2={today}&end_date2={today}&date_period_sel2=1"
            "&products_sort=1&category=0&except_soldout=0&temp_soldout=0&location=0"
        )

        try:
            async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false", "nd": nd,
                        "rows": "9999", "page": "1", "sidx": "", "sord": "asc",
                        "template": "I100", "action": "search", "page_code": "I100",
                        "par": par,
                    },
                    cookies={"PHPSESSID": phpsessid},
                    headers={
                        "User-Agent": "Mozilla/5.0",
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": f"{_EZADMIN_BASE}/template40.htm?template=I100",
                    },
                )
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"이지어드민 연결 실패: {exc}")

        body = (r.text or "").strip()
        if _looks_like_ez_session_error(r, body):
            return {"ok": False, "need_session": True}
        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}
        if not isinstance(obj, dict) or "rows" not in obj:
            return {"ok": False, "unexpected_response": True, "raw": obj}

        found_codes: set[str] = set()
        for row in obj.get("rows", []):
            cell = row.get("cell") or {}
            raw_code = cell.get("key") or cell.get("product_id") or cell.get("code") or ""
            code = normalize_to_yusas(raw_code) or str(raw_code).strip()
            if code:
                found_codes.add(code)

        missing_codes = sorted(code for code in incoming_counts if code not in found_codes)

        info_map: dict[str, dict] = {}
        if missing_codes:
            s_code_map = {_to_s_code(code): code for code in missing_codes}
            conn = _get_wonbe_db()
            try:
                placeholders = ",".join(["?"] * len(s_code_map))
                rows = conn.execute(
                    f"SELECT 상품코드, 상품명, 색상, 사이즈 FROM wonbe WHERE 상품코드 IN ({placeholders})",
                    list(s_code_map.keys()),
                ).fetchall()
                info_map = {
                    s_code_map[row["상품코드"]]: dict(row)
                    for row in rows
                    if row["상품코드"] in s_code_map
                }
            finally:
                conn.close()

        missing = [
            {
                "code": code,
                "incomingQty": incoming_counts.get(code, 0),
                "productName": info_map.get(code, {}).get("상품명", ""),
                "color": info_map.get(code, {}).get("색상", ""),
                "size": info_map.get(code, {}).get("사이즈", ""),
            }
            for code in missing_codes
        ]

        return {
            "ok": True,
            "date": today,
            "total_rows": len(obj.get("rows", [])),
            "incoming_codes": len(incoming_counts),
            "missing": missing,
        }

    @router.post("/barcode/incoming/verify-order-history")
    async def barcode_incoming_verify_order_history(user: str = Depends(get_current_user)):
        """바코드탭에서 불러온 입고 파일의 상품코드를, DB관리 > 발주내역의 전날(어제) 등록분과
        대조해 전날 발주내역에 없는 상품코드만 골라낸다 - verify_trans_in과 같은 패턴이지만
        비교 대상이 이지어드민 I100이 아니라 우리 쪽 order_history 테이블.

        incoming_counts의 키는 normalize_to_yusas가 만든 "YUSAS00000" 형식인데,
        order_history.product_code/wonbe.상품코드는 top90 발주 형식 그대로인 "S00000"
        형식이라 형식이 다르다 - _to_s_code로 맞춰서 비교해야 한다(안 그러면 형식이
        달라서 전부 미확인으로 잘못 뜬다)."""
        incoming_counts = get_shared_incoming_counts() or {}
        if not incoming_counts:
            return {"ok": False, "detail": "불러온 입고 파일이 없습니다. 바코드 탭에서 입고 파일을 먼저 불러오세요."}

        yesterday = (datetime.now(_KST) - timedelta(days=1)).strftime("%Y-%m-%d")

        init_order_history_table(get_db)
        conn = get_db()
        try:
            rows = conn.execute(
                "SELECT DISTINCT product_code FROM order_history "
                "WHERE recorded_at >= ? AND recorded_at <= ? AND product_code != ''",
                (f"{yesterday} 00:00:00", f"{yesterday} 23:59:59"),
            ).fetchall()
        finally:
            conn.close()
        ordered_codes = {row["product_code"] for row in rows}

        # S코드 기준으로 비교/표시한다 - 여러 YUSAS코드가 같은 S코드로 겹치면 수량은 합산.
        s_incoming_counts: dict[str, int] = {}
        for code, qty in incoming_counts.items():
            s_incoming_counts[_to_s_code(code)] = s_incoming_counts.get(_to_s_code(code), 0) + int(qty or 0)

        candidate_codes = sorted(s_code for s_code in s_incoming_counts if s_code not in ordered_codes)

        info_map: dict[str, dict] = {}
        if candidate_codes:
            wonbe_conn = _get_wonbe_db()
            try:
                placeholders = ",".join(["?"] * len(candidate_codes))
                wrows = wonbe_conn.execute(
                    f"SELECT 상품코드, 상품명, 색상, 사이즈, 거래처 FROM wonbe WHERE 상품코드 IN ({placeholders})",
                    candidate_codes,
                ).fetchall()
                info_map = {row["상품코드"]: dict(row) for row in wrows}
            finally:
                wonbe_conn.close()

        # 케이디지/리자드스탠다드/리마인드/계란속노른자/도매킴 거래처 상품은 조회 대상에서 제외.
        missing_codes = [
            s_code for s_code in candidate_codes
            if info_map.get(s_code, {}).get("거래처", "") not in _INCOMING_VERIFY_EXCLUDED_CLIENTS
        ]

        missing = [
            {
                "code": s_code,
                "incomingQty": s_incoming_counts.get(s_code, 0),
                "productName": info_map.get(s_code, {}).get("상품명", ""),
                "color": info_map.get(s_code, {}).get("색상", ""),
                "size": info_map.get(s_code, {}).get("사이즈", ""),
            }
            for s_code in missing_codes
        ]

        return {
            "ok": True,
            "date": yesterday,
            "incoming_codes": len(incoming_counts),
            "missing": missing,
        }

    return router
