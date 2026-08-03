from __future__ import annotations

import asyncio
import io
import os
import re
import sqlite3
import urllib.parse
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
import pandas as pd
import xlwt
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response

from sdk.ably import AblyClient

try:
    from bs4 import BeautifulSoup as _BS4
    _BS4_OK = True
except ImportError:
    _BS4_OK = False

_EZADMIN_BASE = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"

_TOP90_BASE = "https://top90.sosolution.net"
_TOP90_EMAIL = "values0208@naver.com"
_TOP90_PASSWORD = "!Glqgkqdldi1126"

_ABLY_BASE = "https://api.a-bly.com"
_ABLY_EMAIL = "eostm1997@naver.com"
_ABLY_PASSWORD = "!Glqgkqdldi1126"


def _parse_ably_datetime(value) -> datetime | None:
    s = str(value or "").strip()
    if not s:
        return None
    s = s.replace("T", " ")
    if s.endswith("Z"):
        s = s[:-1]
    s = re.sub(r"[+-]\d{2}:?\d{2}$", "", s).strip()
    s = s[:19]
    if len(s) == 10:
        s = s + " 00:00:00"
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None

_BACKEND_DIR = Path(__file__).resolve().parent.parent
_wonbe_db_setting = os.environ.get("WONBE_DB_PATH", "").strip()
if _wonbe_db_setting:
    WONBE_DB_PATH = Path(_wonbe_db_setting).expanduser()
    if not WONBE_DB_PATH.is_absolute():
        WONBE_DB_PATH = _BACKEND_DIR.parent / WONBE_DB_PATH
else:
    WONBE_DB_PATH = _BACKEND_DIR / "data" / "원가베이스유.db"
WONBE_XLSX_PATH = Path(r"C:\Users\ksh29\OneDrive\Desktop\원베\원가베이스유.xlsx")
_janggi_db_setting = os.environ.get("JANGGI_DB_PATH", "").strip()
if _janggi_db_setting:
    JANGGI_DB_PATH = Path(_janggi_db_setting).expanduser()
    if not JANGGI_DB_PATH.is_absolute():
        JANGGI_DB_PATH = _BACKEND_DIR.parent / JANGGI_DB_PATH
else:
    JANGGI_DB_PATH = Path(r"C:\Users\ksh29\OneDrive\Desktop\원베\날짜별장끼정리.db")
INGODAEGI_XLSX_PATH = Path(r"C:\Users\ksh29\OneDrive\Desktop\원베\입고대기.xlsx")

COLUMNS = ["상품코드", "상품명", "색상", "사이즈", "원가", "거래처", "거래처상품명", "거래처합", "상품명합", "거래처주소", "옵션번호", "에이블리상품번호", "등록일", "진열상태", "품절상태", "제조국"]

# username → {"running": bool, "total": int, "done": int, "matched": int} — 제조국 채우기 진행상황 폴링용
_country_sync_progress: dict[str, dict] = {}
EDITABLE = {"상품명합", "거래처합", "원가", "거래처주소", "등록일"}

INGODAEGI_COLUMNS = ["상품코드", "입고수량"]
ABLY_STOCK_COLUMNS = ["옵션번호", "수량"]

JANGGI_COLUMNS = ["거래처", "거래처상품명", "가격", "옵션", "사이즈", "개수", "날짜", "미송체크", "상품코드", "메모", "거래처합산"]

ACCOUNT_COLS = ["A", "B", "C", "D", "E", "F"]
ICHAE_COLS = ["날짜", "A", "B", "C", "D", "E", "F", "G", "H", "상태", "생성일시"]

# 은행명 → 금융결제원 은행코드 (정규화된 축약형 키 기준)
_BANK_CODE_MAP: dict[str, str] = {
    "산업": "002",
    "기업": "003", "ibk": "003", "ibk기업": "003",
    "국민": "004", "kb": "004", "kb국민": "004",
    "수협": "007",
    "수출입": "008",
    "농협": "011", "nh": "011", "nh농협": "011",
    "지역농협": "012", "단위농협": "012",
    "우리": "020",
    "sc제일": "023", "제일": "023", "sc": "023",
    "한국씨티": "027", "씨티": "027",
    "대구": "031", "im": "031", "아이엠": "031",
    "부산": "032",
    "광주": "034",
    "제주": "035",
    "전북": "037",
    "경남": "039",
    "새마을금고": "045", "새마을": "045",
    "신협": "048", "신용협동조합": "048",
    "저축": "050", "상호저축": "050",
    "우체국": "071",
    "하나": "081", "keb하나": "081",
    "신한": "088",
    "케이뱅크": "089", "k뱅크": "089", "케이": "089",
    "카카오뱅크": "090", "카카오": "090",
    "토스뱅크": "092", "토스": "092",
}


def _normalize_bank_key(name: str) -> str:
    s = str(name or "").strip().lower()
    s = re.sub(r"[\s()㈜]+", "", s)
    s = s.replace("주식회사", "")
    if s.endswith("은행"):
        s = s[:-2]
    return s


def _resolve_bank_code(name: str) -> str:
    key = _normalize_bank_key(name)
    return _BANK_CODE_MAP.get(key, "") if key else ""


def _get_wonbe_db() -> sqlite3.Connection:
    WONBE_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(WONBE_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _init_wonbe_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS wonbe (
            상품코드  TEXT PRIMARY KEY,
            상품명    TEXT,
            색상      TEXT,
            사이즈    TEXT,
            원가      TEXT,
            거래처    TEXT,
            거래처상품명 TEXT,
            거래처합  TEXT,
            상품명합  TEXT,
            거래처주소 TEXT,
            옵션번호  TEXT
        )
    """)
    wonbe_cols = {row["name"] for row in conn.execute("PRAGMA table_info(wonbe)").fetchall()}
    if "에이블리상품번호" not in wonbe_cols:
        conn.execute("ALTER TABLE wonbe ADD COLUMN 에이블리상품번호 TEXT")
    if "등록일" not in wonbe_cols:
        conn.execute("ALTER TABLE wonbe ADD COLUMN 등록일 TEXT")
    if "진열상태" not in wonbe_cols:
        conn.execute("ALTER TABLE wonbe ADD COLUMN 진열상태 TEXT")
    if "품절상태" not in wonbe_cols:
        conn.execute("ALTER TABLE wonbe ADD COLUMN 품절상태 TEXT")
    if "제조국" not in wonbe_cols:
        conn.execute("ALTER TABLE wonbe ADD COLUMN 제조국 TEXT")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS wonbe_meta (
            key   TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_wonbe_상품명합 ON wonbe(상품명합)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_wonbe_거래처합 ON wonbe(거래처합)")
    conn.commit()


COST_BASE_CODE_COL = 0
COST_BASE_MATCH_COL = 8


def _normalize_cost_base_key(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def load_wonbe_cost_base_df() -> pd.DataFrame:
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute(
            f"SELECT {', '.join(COLUMNS)} FROM wonbe ORDER BY rowid ASC"
        ).fetchall()
    finally:
        conn.close()
    return pd.DataFrame([dict(r) for r in rows], columns=COLUMNS)


def save_wonbe_cost_base_df(df: pd.DataFrame) -> int:
    """상품코드 기준 upsert. wonbe 테이블을 절대 비우지 않는다.
    상품코드가 빈 행은 PRIMARY KEY 충돌(빈 문자열끼리 덮어씀)을 피하기 위해 스킵하고,
    스킵된 행 수를 반환한다."""
    df = df.reindex(columns=COLUMNS, fill_value="").fillna("")
    codes = df[COLUMNS[COST_BASE_CODE_COL]].astype(str).str.strip()
    skipped = int((codes == "").sum())
    df = df[codes != ""]
    rows = [tuple(r) for r in df[COLUMNS].itertuples(index=False, name=None)]
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        if rows:
            conn.executemany(
                f"INSERT OR REPLACE INTO wonbe ({', '.join(COLUMNS)}) VALUES ({', '.join(['?'] * len(COLUMNS))})",
                rows,
            )
            conn.commit()
    finally:
        conn.close()
    return skipped


def load_wonbe_cost_base_map() -> dict[str, str]:
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute("SELECT 상품코드, 상품명합 FROM wonbe").fetchall()
    finally:
        conn.close()
    cost_map: dict[str, str] = {}
    for r in rows:
        key = _normalize_cost_base_key(r["상품명합"])
        if not key or key in cost_map:
            continue
        cost_map[key] = r["상품코드"]
    return cost_map


def load_wonbe_option_sno_map() -> dict[str, str]:
    """옵션번호(에이블리 옵션 sno) → 상품코드 매핑.

    옵션번호는 EZAdmin 상품마스타의 옵션추가1 필드에서 채워지지만, 실제로는
    에이블리 옵션 sno로 관리된다 (barcode_routes.py의 defect_ochuul_minus가
    같은 컬럼 값을 에이블리 today-delivery-goods-options 응답의 sno와
    대조해서 쓰고 있음 - 검증된 매핑).
    """
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute("SELECT 상품코드, 옵션번호 FROM wonbe WHERE 옵션번호 != ''").fetchall()
    finally:
        conn.close()
    sno_map: dict[str, str] = {}
    for r in rows:
        sno = str(r["옵션번호"]).strip()
        if sno and sno not in sno_map:
            sno_map[sno] = r["상품코드"]
    return sno_map


def load_wonbe_goods_sno_map() -> dict[str, list[tuple[str, str]]]:
    """에이블리상품번호(goods_sno) → [(옵션번호, 상품코드), ...] 매핑.

    같은 goods_sno 아래 여러 옵션(색상/사이즈 등)이 서로 다른 상품코드로
    관리되는 경우를 그룹으로 묶어, 통계 API를 goods_sno 단위로 한 번만
    호출하면 되도록 한다."""
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute(
            "SELECT 상품코드, 옵션번호, 에이블리상품번호 FROM wonbe "
            "WHERE 옵션번호 != '' AND 에이블리상품번호 != ''"
        ).fetchall()
    finally:
        conn.close()
    goods_map: dict[str, list[tuple[str, str]]] = {}
    for r in rows:
        goods_sno = str(r["에이블리상품번호"]).strip()
        option_sno = str(r["옵션번호"]).strip()
        code = r["상품코드"]
        if not goods_sno or not option_sno:
            continue
        goods_map.setdefault(goods_sno, []).append((option_sno, code))
    return goods_map


def load_wonbe_registered_at_map() -> dict[str, str]:
    """상품코드 → 등록일("YYYY-MM-DD") 매핑. "등록일 채우기"로 이미 채워진 값을 재사용한다.

    발주추천 판매량 백필이 상품 등록 이전 날짜를 실제 판매 0건으로 착각해
    채워버리지 않도록, 백필 대상 날짜를 이 등록일로 잘라내는 데 쓴다."""
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute(
            "SELECT 상품코드, 등록일 FROM wonbe WHERE 등록일 IS NOT NULL AND 등록일 != ''"
        ).fetchall()
    finally:
        conn.close()
    result: dict[str, str] = {}
    for r in rows:
        code = r["상품코드"]
        registered_at = str(r["등록일"]).strip()
        if not code or not registered_at:
            continue
        result[code] = registered_at[:10]  # "YYYY-MM-DD HH:MM:SS" -> 날짜만
    return result


def load_wonbe_product_name_map() -> dict[str, str]:
    """상품코드 → 상품명합(상품명+색상+사이즈) 매핑. 발주 대시보드 일별 데이터 테이블 표시용.

    같은 상품이라도 색상/사이즈별로 상품코드가 다른데 순수 상품명은 전부 동일해서
    구분이 안 된다 — 상품명합을 쓰면 옵션까지 포함된 이름이 나온다."""
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute(
            "SELECT 상품코드, 상품명합 FROM wonbe WHERE 상품코드 != ''"
        ).fetchall()
    finally:
        conn.close()
    return {r["상품코드"]: r["상품명합"] or "" for r in rows if r["상품코드"]}


def load_wonbe_product_cost_map() -> dict[str, int]:
    """상품명합 → 원가(int) 매핑. 마진계산(아무드/에이블리 정산) 등에서 원가 조회용."""
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        rows = conn.execute("SELECT 상품명합, 원가 FROM wonbe").fetchall()
    finally:
        conn.close()
    cost_map: dict[str, int] = {}
    for r in rows:
        key = _normalize_cost_base_key(r["상품명합"])
        if not key or key in cost_map:
            continue
        try:
            cost_map[key] = int(float(r["원가"]))
        except (TypeError, ValueError):
            continue
    return cost_map


def wonbe_cost_base_status() -> dict:
    exists = WONBE_DB_PATH.exists()
    mtime = None
    if exists:
        try:
            mtime = datetime.fromtimestamp(WONBE_DB_PATH.stat().st_mtime).isoformat()
        except Exception:
            mtime = None
    conn = _get_wonbe_db()
    try:
        _init_wonbe_table(conn)
        count = conn.execute("SELECT COUNT(*) FROM wonbe").fetchone()[0]
    finally:
        conn.close()
    return {"path": str(WONBE_DB_PATH), "exists": exists, "mtime": mtime, "rows": count}


def _init_kdg_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 케이디지원가베이스 (
            상품명합  TEXT,
            상품코드  TEXT PRIMARY KEY
        )
    """)
    conn.commit()


def _init_ingodaegi_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 입고대기 (
            상품코드  TEXT PRIMARY KEY,
            입고수량  TEXT NOT NULL DEFAULT 'ZERO'
        )
    """)
    conn.commit()


def _sync_ingodaegi_from_wonbe(conn: sqlite3.Connection) -> int:
    _init_wonbe_table(conn)
    _init_ingodaegi_table(conn)
    rows = conn.execute(
        """SELECT 상품코드 FROM wonbe
           WHERE TRIM(상품코드) != ''
             AND 상품코드 NOT IN (SELECT 상품코드 FROM 입고대기)"""
    ).fetchall()
    new_codes = [(r["상품코드"],) for r in rows]
    if new_codes:
        conn.executemany(
            "INSERT OR IGNORE INTO 입고대기 (상품코드, 입고수량) VALUES (?, 'ZERO')",
            new_codes,
        )
        conn.commit()
    return len(new_codes)


def _init_ably_stock_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 에이블리재고변경 (
            옵션번호  TEXT PRIMARY KEY,
            수량      TEXT NOT NULL DEFAULT '0'
        )
    """)
    conn.commit()


def _init_defect_process_log_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS defect_process_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            execution_id TEXT NOT NULL,
            processed_at TEXT NOT NULL,
            process_type TEXT NOT NULL,
            product_code TEXT NOT NULL,
            product_name TEXT NOT NULL DEFAULT '',
            product_option TEXT NOT NULL DEFAULT '',
            vendor TEXT NOT NULL DEFAULT '',
            vendor_product TEXT NOT NULL DEFAULT '',
            requested_qty INTEGER NOT NULL,
            applied_qty INTEGER,
            result_status TEXT NOT NULL DEFAULT 'completed',
            option_sno TEXT NOT NULL DEFAULT '',
            stock_before INTEGER,
            stock_after INTEGER,
            processed_by_username TEXT NOT NULL,
            processed_by_display_name TEXT NOT NULL DEFAULT ''
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_defect_process_logs_processed_at ON defect_process_logs(processed_at DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_defect_process_logs_execution_id ON defect_process_logs(execution_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_defect_process_logs_product_code ON defect_process_logs(product_code)")
    conn.commit()


def record_defect_process_logs(
    *,
    execution_id: str,
    processed_at: str,
    process_type: str,
    rows: list[dict],
    username: str,
    display_name: str,
    outcomes: dict[str, dict] | None = None,
) -> None:
    """Persist an immutable, per-product snapshot after a defect process succeeds."""
    outcomes = outcomes or {}
    records = []
    for row in rows:
        code = str(row.get("code") or "").strip()
        qty = int(row.get("count") or 0)
        if not code or qty <= 0:
            continue
        outcome = outcomes.get(code) or {}
        records.append((
            execution_id,
            processed_at,
            process_type,
            code,
            str(row.get("base_name") or row.get("name") or ""),
            str(row.get("base_color") or row.get("option") or ""),
            str(row.get("base_vendor") or ""),
            str(row.get("base_product") or ""),
            qty,
            outcome.get("applied_qty"),
            str(outcome.get("result_status") or "completed"),
            str(outcome.get("option_sno") or ""),
            outcome.get("stock_before"),
            outcome.get("stock_after"),
            username,
            display_name,
        ))
    if not records:
        return
    conn = _get_wonbe_db()
    try:
        _init_defect_process_log_table(conn)
        conn.executemany(
            """INSERT INTO defect_process_logs (
                execution_id, processed_at, process_type, product_code, product_name,
                product_option, vendor, vendor_product, requested_qty, applied_qty,
                result_status, option_sno, stock_before, stock_after,
                processed_by_username, processed_by_display_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            records,
        )
        conn.commit()
    finally:
        conn.close()


def _sync_ably_stock_from_wonbe(conn: sqlite3.Connection) -> int:
    _init_wonbe_table(conn)
    _init_ably_stock_table(conn)
    rows = conn.execute(
        """SELECT 옵션번호 FROM wonbe
           WHERE TRIM(옵션번호) != ''
             AND 옵션번호 NOT IN (SELECT 옵션번호 FROM 에이블리재고변경)"""
    ).fetchall()
    new_options = [(r["옵션번호"],) for r in rows]
    if new_options:
        conn.executemany(
            "INSERT OR IGNORE INTO 에이블리재고변경 (옵션번호, 수량) VALUES (?, '0')",
            new_options,
        )
        conn.commit()
    return len(new_options)


def _get_janggi_db() -> sqlite3.Connection:
    JANGGI_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(JANGGI_DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _init_account_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 거래처계좌데이터 (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            A TEXT NOT NULL DEFAULT '',
            B TEXT NOT NULL DEFAULT '',
            C TEXT NOT NULL DEFAULT '',
            D TEXT NOT NULL DEFAULT '',
            E TEXT NOT NULL DEFAULT '',
            F TEXT NOT NULL DEFAULT ''
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_account_A ON 거래처계좌데이터(A)")
    conn.commit()


def _init_ichae_table(conn: sqlite3.Connection):
    # 구 스키마(G열 또는 상품코드 컬럼 포함) 감지 시 재생성
    try:
        existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(이체파일)").fetchall()]
        if existing_cols and ("G" in existing_cols or "상품코드" in existing_cols):
            conn.execute("DROP TABLE 이체파일")
            conn.commit()
    except Exception:
        pass
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 이체파일 (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            날짜 TEXT NOT NULL DEFAULT '',
            A TEXT NOT NULL DEFAULT '',
            B TEXT NOT NULL DEFAULT '',
            C REAL NOT NULL DEFAULT 0,
            D TEXT NOT NULL DEFAULT '',
            E TEXT NOT NULL DEFAULT '주식회사 유색',
            F TEXT NOT NULL DEFAULT '',
            상태 TEXT NOT NULL DEFAULT '',
            생성일시 TEXT NOT NULL DEFAULT ''
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ichae_날짜 ON 이체파일(날짜)")
    conn.commit()


def _parse_amount_janggi(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return 0.0
    normalized = re.sub(r"[,원\s]", "", raw)
    m = re.search(r"[+-]?\d+(?:\.\d+)?", normalized)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except Exception:
        return 0.0


def _init_janggi_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS 날짜별장끼정리 (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            거래처      TEXT,
            거래처상품명 TEXT,
            가격        TEXT,
            옵션        TEXT,
            사이즈      TEXT,
            개수        TEXT,
            날짜        TEXT,
            미송체크    TEXT,
            상품코드    TEXT,
            메모        TEXT,
            거래처합산  TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_janggi_날짜 ON 날짜별장끼정리(날짜)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_janggi_거래처 ON 날짜별장끼정리(거래처)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS janggi_aliases (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS janggi_top_col_widths (
            username TEXT NOT NULL,
            col_key  TEXT NOT NULL,
            width    INTEGER NOT NULL,
            PRIMARY KEY (username, col_key)
        )
    """)
    try:
        conn.execute("ALTER TABLE 날짜별장끼정리 ADD COLUMN 일괄이체 TEXT NOT NULL DEFAULT ''")
    except Exception:
        pass
    conn.commit()


def _content_disposition(filename: str) -> str:
    ascii_name = "".join(ch if ord(ch) < 128 else "_" for ch in filename).strip("_") or "download"
    quoted = urllib.parse.quote(filename)
    return f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quoted}'


def _parse_options(options_str: str) -> tuple[str, str]:
    """'[크림-FREE-SHORT]' → ('크림', 'FREE SHORT')"""
    s = str(options_str or "").strip().lstrip("[").rstrip("]")
    parts = [p.strip() for p in s.split("-")]
    color = parts[0] if parts else ""
    size = " ".join(parts[1:]) if len(parts) > 1 else ""
    return color, size


def _strip_html(html_str: str) -> str:
    return re.sub(r"<[^>]+>", "", str(html_str or "")).strip()


def build_wonbe_router(*, get_current_user, get_setting=None):
    router = APIRouter(prefix="/wonbe")

    @router.get("/defect-process-logs")
    def get_defect_process_logs(
        offset: int = 0,
        limit: int = 50,
        date_from: str = "",
        date_to: str = "",
        process_type: str = "",
        q: str = "",
        user: str = Depends(get_current_user),
    ):
        offset = max(0, offset)
        limit = min(max(1, limit), 200)
        clauses = []
        params = []
        if date_from:
            clauses.append("processed_at >= ?")
            params.append(f"{date_from} 00:00:00")
        if date_to:
            clauses.append("processed_at <= ?")
            params.append(f"{date_to} 23:59:59")
        if process_type:
            clauses.append("process_type = ?")
            params.append(process_type)
        if q.strip():
            like = f"%{q.strip()}%"
            clauses.append("(product_code LIKE ? OR product_name LIKE ? OR vendor LIKE ? OR processed_by_username LIKE ? OR processed_by_display_name LIKE ?)")
            params.extend([like, like, like, like, like])
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        conn = _get_wonbe_db()
        try:
            _init_defect_process_log_table(conn)
            total = conn.execute(f"SELECT COUNT(*) FROM defect_process_logs{where}", params).fetchone()[0]
            rows = conn.execute(
                f"SELECT * FROM defect_process_logs{where} ORDER BY processed_at DESC, id DESC LIMIT ? OFFSET ?",
                [*params, limit, offset],
            ).fetchall()
            return {"ok": True, "total": total, "rows": [dict(row) for row in rows]}
        finally:
            conn.close()

    @router.post("/import")
    async def wonbe_import(
        file: UploadFile = File(...),
        user: str = Depends(get_current_user),
    ):
        content = await file.read()
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                raise HTTPException(status_code=400, detail="빈 파일입니다.")

            data_rows = []
            for row in rows_iter:
                vals = [str(v).strip() if v is not None else "" for v in row]
                while len(vals) < len(COLUMNS):
                    vals.append("")
                vals = vals[:len(COLUMNS)]
                if not any(vals):
                    continue
                data_rows.append(vals)
            wb.close()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"엑셀 파싱 오류: {e}")

        # 엑셀 맨 아래 행이 가장 나중에 삽입되어 rowid가 가장 커짐 → 기본 정렬(rowid DESC)에서 맨 위에 표시

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            conn.execute("DELETE FROM wonbe")
            conn.executemany(
                f"INSERT OR REPLACE INTO wonbe ({', '.join(COLUMNS)}) VALUES ({', '.join(['?']*len(COLUMNS))})",
                data_rows,
            )
            conn.commit()
            return {"ok": True, "count": len(data_rows)}
        finally:
            conn.close()

    @router.post("/init-from-default")
    def wonbe_init_from_default(user: str = Depends(get_current_user)):
        if not WONBE_XLSX_PATH.exists():
            raise HTTPException(status_code=404, detail=f"파일 없음: {WONBE_XLSX_PATH}")
        try:
            wb = openpyxl.load_workbook(str(WONBE_XLSX_PATH), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter, None)
            data_rows = []
            for row in rows_iter:
                vals = [str(v).strip() if v is not None else "" for v in row]
                while len(vals) < len(COLUMNS):
                    vals.append("")
                vals = vals[:len(COLUMNS)]
                if not any(vals):
                    continue
                data_rows.append(vals)
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"파일 읽기 오류: {e}")

        # 엑셀 맨 아래 행이 가장 나중에 삽입되어 rowid가 가장 커짐 → 기본 정렬(rowid DESC)에서 맨 위에 표시

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            conn.execute("DELETE FROM wonbe")
            conn.executemany(
                f"INSERT OR REPLACE INTO wonbe ({', '.join(COLUMNS)}) VALUES ({', '.join(['?']*len(COLUMNS))})",
                data_rows,
            )
            conn.commit()
            return {"ok": True, "count": len(data_rows)}
        finally:
            conn.close()

    # ── 입고대기 CRUD ─────────────────────────────────────────────

    @router.post("/ingodaegi/init-from-default")
    def ingodaegi_init_from_default(user: str = Depends(get_current_user)):
        if not INGODAEGI_XLSX_PATH.exists():
            raise HTTPException(status_code=404, detail=f"파일 없음: {INGODAEGI_XLSX_PATH}")
        try:
            wb = openpyxl.load_workbook(str(INGODAEGI_XLSX_PATH), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter, None)  # 헤더 스킵
            data_rows = []
            for row in rows_iter:
                code = str(row[0]).strip() if row and row[0] is not None else ""
                if not code:
                    continue
                qty = str(row[1]).strip() if len(row) > 1 and row[1] is not None else "ZERO"
                data_rows.append((code, qty or "ZERO"))
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"파일 읽기 오류: {e}")

        conn = _get_wonbe_db()
        try:
            _init_ingodaegi_table(conn)
            conn.execute("DELETE FROM 입고대기")
            conn.executemany(
                "INSERT OR REPLACE INTO 입고대기 (상품코드, 입고수량) VALUES (?, ?)",
                data_rows,
            )
            conn.commit()
            return {"ok": True, "count": len(data_rows)}
        finally:
            conn.close()

    @router.post("/ingodaegi/sync-from-wonbe")
    def ingodaegi_sync_from_wonbe(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            added = _sync_ingodaegi_from_wonbe(conn)
            return {"ok": True, "added": added}
        finally:
            conn.close()

    @router.get("/ingodaegi/search")
    def ingodaegi_search(
        q: str = "",
        offset: int = 0,
        limit: int = 50,
        user: str = Depends(get_current_user),
    ):
        conn = _get_wonbe_db()
        try:
            _init_ingodaegi_table(conn)
            q = q.strip()
            if not q:
                rows = conn.execute(
                    "SELECT * FROM 입고대기 ORDER BY 상품코드 LIMIT ? OFFSET ?",
                    (limit, offset),
                ).fetchall()
                total = conn.execute("SELECT COUNT(*) FROM 입고대기").fetchone()[0]
            else:
                like = f"%{q}%"
                rows = conn.execute(
                    "SELECT * FROM 입고대기 WHERE 상품코드 LIKE ? ORDER BY 상품코드 LIMIT ? OFFSET ?",
                    (like, limit, offset),
                ).fetchall()
                total = conn.execute(
                    "SELECT COUNT(*) FROM 입고대기 WHERE 상품코드 LIKE ?", (like,)
                ).fetchone()[0]
            return {
                "ok": True,
                "rows": [dict(r) for r in rows],
                "total": total,
                "offset": offset,
                "limit": limit,
            }
        finally:
            conn.close()

    @router.post("/ingodaegi/append")
    def ingodaegi_append(payload: dict = Body(...), user: str = Depends(get_current_user)):
        text = str(payload.get("text") or "")
        codes = []
        for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            first_cell = line.split("\t", 1)[0].strip()
            if first_cell:
                codes.append(first_cell)
        if not codes:
            raise HTTPException(status_code=400, detail="추가할 상품코드가 없습니다.")

        conn = _get_wonbe_db()
        try:
            _init_ingodaegi_table(conn)
            before = conn.execute("SELECT COUNT(*) FROM 입고대기").fetchone()[0]
            conn.executemany(
                "INSERT OR IGNORE INTO 입고대기 (상품코드, 입고수량) VALUES (?, 'ZERO')",
                [(c,) for c in codes],
            )
            conn.commit()
            after = conn.execute("SELECT COUNT(*) FROM 입고대기").fetchone()[0]
            return {"ok": True, "requested": len(codes), "inserted": after - before}
        finally:
            conn.close()

    @router.delete("/ingodaegi/row")
    def ingodaegi_delete_row(payload: dict = Body(...), user: str = Depends(get_current_user)):
        code = str(payload.get("상품코드") or "").strip()
        if not code:
            raise HTTPException(status_code=400, detail="상품코드 필요")
        conn = _get_wonbe_db()
        try:
            _init_ingodaegi_table(conn)
            cur = conn.execute("DELETE FROM 입고대기 WHERE 상품코드 = ?", (code,))
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 상품코드 없음")
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    @router.get("/ingodaegi/export")
    def ingodaegi_export(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            _init_ingodaegi_table(conn)
            rows = conn.execute("SELECT * FROM 입고대기 ORDER BY 상품코드").fetchall()
        finally:
            conn.close()

        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet1")
        for ci, h in enumerate(INGODAEGI_COLUMNS):
            sheet.write(0, ci, h)
        for ri, row in enumerate(rows, start=1):
            for ci, col in enumerate(INGODAEGI_COLUMNS):
                sheet.write(ri, ci, row[col] or "")

        buf = io.BytesIO()
        book.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": _content_disposition("입고대기.xls")},
        )

    # ── 에이블리재고변경 CRUD ─────────────────────────────────────

    @router.post("/ably-stock/sync-from-wonbe")
    def ably_stock_sync_from_wonbe(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            added = _sync_ably_stock_from_wonbe(conn)
            return {"ok": True, "added": added}
        finally:
            conn.close()

    @router.get("/ably-stock/search")
    def ably_stock_search(
        q: str = "",
        offset: int = 0,
        limit: int = 50,
        user: str = Depends(get_current_user),
    ):
        conn = _get_wonbe_db()
        try:
            _init_ably_stock_table(conn)
            q = q.strip()
            if not q:
                rows = conn.execute(
                    "SELECT * FROM 에이블리재고변경 ORDER BY 옵션번호 LIMIT ? OFFSET ?",
                    (limit, offset),
                ).fetchall()
                total = conn.execute("SELECT COUNT(*) FROM 에이블리재고변경").fetchone()[0]
            else:
                like = f"%{q}%"
                rows = conn.execute(
                    "SELECT * FROM 에이블리재고변경 WHERE 옵션번호 LIKE ? ORDER BY 옵션번호 LIMIT ? OFFSET ?",
                    (like, limit, offset),
                ).fetchall()
                total = conn.execute(
                    "SELECT COUNT(*) FROM 에이블리재고변경 WHERE 옵션번호 LIKE ?", (like,)
                ).fetchone()[0]
            return {
                "ok": True,
                "rows": [dict(r) for r in rows],
                "total": total,
                "offset": offset,
                "limit": limit,
            }
        finally:
            conn.close()

    @router.delete("/ably-stock/row")
    def ably_stock_delete_row(payload: dict = Body(...), user: str = Depends(get_current_user)):
        code = str(payload.get("옵션번호") or "").strip()
        if not code:
            raise HTTPException(status_code=400, detail="옵션번호 필요")
        conn = _get_wonbe_db()
        try:
            _init_ably_stock_table(conn)
            cur = conn.execute("DELETE FROM 에이블리재고변경 WHERE 옵션번호 = ?", (code,))
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 옵션번호 없음")
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    @router.get("/ably-stock/export")
    def ably_stock_export(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            _init_ably_stock_table(conn)
            rows = conn.execute("SELECT * FROM 에이블리재고변경 ORDER BY 옵션번호").fetchall()
        finally:
            conn.close()

        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet1")
        for ci, h in enumerate(ABLY_STOCK_COLUMNS):
            sheet.write(0, ci, h)
        for ri, row in enumerate(rows, start=1):
            for ci, col in enumerate(ABLY_STOCK_COLUMNS):
                sheet.write(ri, ci, row[col] or "")

        buf = io.BytesIO()
        book.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": _content_disposition("에이블리재고변경.xls")},
        )

    @router.get("/search")
    def wonbe_search(
        q: str = "",
        offset: int = 0,
        limit: int = 50,
        user: str = Depends(get_current_user),
    ):
        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            q = q.strip()
            if not q:
                rows = conn.execute(
                    "SELECT * FROM wonbe ORDER BY rowid DESC LIMIT ? OFFSET ?",
                    (limit, offset),
                ).fetchall()
                total = conn.execute("SELECT COUNT(*) FROM wonbe").fetchone()[0]
            else:
                like = f"%{q}%"
                rows = conn.execute(
                    """SELECT * FROM wonbe
                       WHERE 상품코드 LIKE ? OR 상품명합 LIKE ? OR 거래처합 LIKE ? OR 거래처 LIKE ?
                       ORDER BY CASE WHEN 상품코드 = ? THEN 0
                                     WHEN 상품코드 LIKE ? THEN 1 ELSE 2 END, 상품코드
                       LIMIT ? OFFSET ?""",
                    (like, like, like, like, q, f"{q}%", limit, offset),
                ).fetchall()
                total = conn.execute(
                    "SELECT COUNT(*) FROM wonbe WHERE 상품코드 LIKE ? OR 상품명합 LIKE ? OR 거래처합 LIKE ? OR 거래처 LIKE ?",
                    (like, like, like, like),
                ).fetchone()[0]
            return {
                "ok": True,
                "rows": [dict(r) for r in rows],
                "total": total,
                "offset": offset,
                "limit": limit,
            }
        finally:
            conn.close()

    @router.patch("/row")
    def wonbe_update_row(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        code = str(payload.get("상품코드") or "").strip()
        if not code:
            raise HTTPException(status_code=400, detail="상품코드 필요")

        updates = {k: str(v).strip() for k, v in payload.items() if k in EDITABLE}
        if not updates:
            raise HTTPException(status_code=400, detail="수정할 필드 없음")

        set_clause = ", ".join(f"{col} = ?" for col in updates)
        values = list(updates.values()) + [code]

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            cur = conn.execute(f"UPDATE wonbe SET {set_clause} WHERE 상품코드 = ?", values)
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 상품코드 없음")
            row = conn.execute("SELECT * FROM wonbe WHERE 상품코드 = ?", (code,)).fetchone()
            return {"ok": True, "row": dict(row)}
        finally:
            conn.close()

    @router.post("/bulk-update")
    def wonbe_bulk_update(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        q = str(payload.get("q") or "").strip()
        col = str(payload.get("col") or "").strip()
        value = str(payload.get("value") or "").strip()
        if col not in EDITABLE:
            raise HTTPException(status_code=400, detail=f"수정 불가 컬럼: {col}")

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            if q:
                like = f"%{q}%"
                cur = conn.execute(
                    f"UPDATE wonbe SET {col} = ? WHERE 상품코드 LIKE ? OR 상품명합 LIKE ? OR 거래처합 LIKE ? OR 거래처 LIKE ?",
                    (value, like, like, like, like),
                )
            else:
                cur = conn.execute(f"UPDATE wonbe SET {col} = ?", (value,))
            conn.commit()
            return {"ok": True, "count": cur.rowcount}
        finally:
            conn.close()

    @router.get("/export")
    def wonbe_export(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            rows = conn.execute("SELECT * FROM wonbe ORDER BY 상품코드").fetchall()
        finally:
            conn.close()

        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet1")
        for ci, h in enumerate(COLUMNS):
            sheet.write(0, ci, h)
        for ri, row in enumerate(rows, start=1):
            for ci, col in enumerate(COLUMNS):
                sheet.write(ri, ci, row[col] or "")

        buf = io.BytesIO()
        book.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": _content_disposition("원가베이스유.xls")},
        )

    @router.post("/sync-from-ezadmin")
    async def wonbe_sync_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        if not get_setting:
            raise HTTPException(status_code=500, detail="get_setting 미설정")
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today_str = datetime.now().strftime("%Y-%m-%d")
        start_date2 = str(payload.get("start_date") or today_str)
        end_date2 = str(payload.get("end_date") or today_str)

        nd = str(int(datetime.now().timestamp() * 1000))
        par = (
            f"auto_search=&search_all_product=&multi_supply_group=&multi_supply="
            f"&str_supply_code=0&tags_string=&product_tag_include_type=1"
            f"&query_type=name&query_str=&stock_type=0&stock_start=&stock_end="
            f"&notrans_day=&notrans_cnt=&notrans_status=0&stock_status=0"
            f"&start_date={today_str}&start_hour=00%3A00%3A00"
            f"&end_date={today_str}&end_hour=23%3A59%3A59&date_period_sel=0"
            f"&work_type=stockin&work_start=&work_end=&inout_type=0&product_date=reg_date"
            f"&start_date2={start_date2}&end_date2={end_date2}&date_period_sel2=0"
            f"&products_sort=1&category=0&except_soldout=0&temp_soldout=0&location=0"
        )
        try:
            async with httpx.AsyncClient(timeout=120.0, verify=False, follow_redirects=True) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false", "nd": nd,
                        "rows": "9999", "page": "1", "sidx": "", "sord": "asc",
                        "template": "I100", "action": "search", "page_code": "I100",
                        "par": par,
                    },
                    cookies={"PHPSESSID": phpsessid},
                    headers={
                        "User-Agent": "Mozilla/5.0",
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": f"{_EZADMIN_BASE}/template40.htm?template=I100",
                    },
                )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"EZAdmin 요청 실패: {exc}")

        raw_text = r.text
        try:
            obj = r.json()
        except Exception:
            # 세션이 만료되면 JSON이 아니라 로그인 HTML 페이지가 내려온다 — 이 경우만 진짜 need_session.
            return {"ok": False, "need_session": True, "raw_preview": raw_text[:1000]}

        if not isinstance(obj, dict) or "rows" not in obj:
            # JSON은 정상 수신했지만 rows가 없는 경우 — 세션 문제가 아닐 수 있다.
            # 세션 재입력을 유도하지 않고, 실제 받은 응답을 그대로 보여준다.
            return {"ok": False, "unexpected_response": True, "raw": obj}

        data_rows = []
        for row in obj.get("rows", []):
            c = row.get("cell", {})
            code = str(c.get("key") or "").strip()
            if not code:
                continue
            product_name = str(c.get("product_name") or "").strip()
            color, size = _parse_options(str(c.get("options") or ""))
            org_price = str(c.get("org_price") or "").strip()
            brand_raw = str(c.get("brand") or "").strip()
            brand_parts = brand_raw.split(" ", 1)
            supplier = brand_parts[0] if brand_parts else ""
            supplier_product = brand_parts[1] if len(brand_parts) > 1 else ""
            option_no = str(c.get("option_extra_column1") or "").strip()
            supplier_combined = " ".join(p for p in [supplier_product, color, size] if p)
            name_combined = " ".join(p for p in [product_name, color, size] if p)
            data_rows.append((code, product_name, color, size, org_price, supplier, supplier_product, supplier_combined, name_combined, option_no))

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)

            # 기존 거래처 → 거래처주소 매핑
            addr_rows = conn.execute(
                "SELECT 거래처, 거래처주소 FROM wonbe WHERE 거래처주소 IS NOT NULL AND 거래처주소 != ''"
            ).fetchall()
            addr_map = {r["거래처"]: r["거래처주소"] for r in addr_rows}

            # 거래처주소 채운 최종 튜플
            final_rows = [
                (code, pname, color, size, price, sup, sup_prod, sup_comb, name_comb, addr_map.get(sup, ""), opt)
                for code, pname, color, size, price, sup, sup_prod, sup_comb, name_comb, opt in data_rows
            ]

            before = conn.execute("SELECT COUNT(*) FROM wonbe").fetchone()[0]
            conn.executemany(
                """INSERT OR IGNORE INTO wonbe
                   (상품코드, 상품명, 색상, 사이즈, 원가, 거래처, 거래처상품명, 거래처합, 상품명합, 거래처주소, 옵션번호)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                final_rows,
            )
            after = conn.execute("SELECT COUNT(*) FROM wonbe").fetchone()[0]
            inserted = after - before

            # 거래처가 케이디지인 행을 케이디지원가베이스 테이블에도 추가
            _init_kdg_table(conn)
            kdg_rows = [
                (row[8], row[0])  # (상품명합, 상품코드)
                for row in final_rows
                if row[5] == "케이디지"
            ]
            if kdg_rows:
                conn.executemany(
                    """INSERT OR IGNORE INTO 케이디지원가베이스 (상품명합, 상품코드) VALUES (?, ?)""",
                    kdg_rows,
                )
            synced_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(
                "INSERT OR REPLACE INTO wonbe_meta (key, value) VALUES ('last_sync_at', ?)",
                (synced_at,),
            )
            conn.execute(
                "INSERT OR REPLACE INTO wonbe_meta (key, value) VALUES ('last_sync_count', ?)",
                (str(inserted),),
            )
            conn.execute(
                "INSERT OR REPLACE INTO wonbe_meta (key, value) VALUES ('last_sync_fetched', ?)",
                (str(len(data_rows)),),
            )
            conn.commit()
            return {"ok": True, "fetched": len(data_rows), "inserted": inserted, "synced_at": synced_at}
        finally:
            conn.close()

    @router.post("/lookup-prices")
    def wonbe_lookup_prices(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        codes = [str(c).strip() for c in (payload.get("codes") or []) if str(c).strip()]
        if not codes:
            return {"ok": True, "prices": {}}
        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            placeholders = ",".join(["?"] * len(codes))
            rows = conn.execute(
                f"SELECT 상품코드, 원가 FROM wonbe WHERE 상품코드 IN ({placeholders})",
                codes,
            ).fetchall()
            prices = {r["상품코드"]: r["원가"] for r in rows}
            return {"ok": True, "prices": prices}
        finally:
            conn.close()

    @router.get("/stats")
    def wonbe_stats(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            total = conn.execute("SELECT COUNT(*) FROM wonbe").fetchone()[0]
            db_exists = WONBE_DB_PATH.exists()
            meta_rows = conn.execute("SELECT key, value FROM wonbe_meta").fetchall()
            meta = {r["key"]: r["value"] for r in meta_rows}
            return {
                "ok": True,
                "total": total,
                "db_exists": db_exists,
                "last_sync_at": meta.get("last_sync_at"),
                "last_sync_count": meta.get("last_sync_count"),
                "last_sync_fetched": meta.get("last_sync_fetched"),
            }
        finally:
            conn.close()

    @router.post("/freshness-check")
    async def wonbe_freshness_check(user: str = Depends(get_current_user)):
        async with httpx.AsyncClient(timeout=15.0) as client:
            res = await client.post(
                f"{_ABLY_BASE}/seller/login/",
                json={"email": _ABLY_EMAIL, "password": _ABLY_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "Origin": "https://seller-admin.a-bly.com",
                    "Referer": "https://seller-admin.a-bly.com/",
                    "User-Agent": "Mozilla/5.0",
                },
            )
            if not res.is_success:
                raise HTTPException(status_code=502, detail="에이블리 로그인 실패")
        token = res.json().get("token")
        if not token:
            raise HTTPException(status_code=502, detail="에이블리 로그인 실패: 토큰 없음")

        ably_headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Origin": "https://seller-admin.a-bly.com",
            "Referer": "https://seller-admin.a-bly.com/",
            "User-Agent": "Mozilla/5.0",
        }

        latest_dt: datetime | None = None
        checked_goods = 0
        checked_pages = 0
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                for page in range(1, 3):
                    res = await client.post(
                        f"{_ABLY_BASE}/seller/goods/search/",
                        headers=ably_headers,
                        json={"page": page, "per_page": 30},
                    )
                    res.raise_for_status()
                    data = res.json()
                    goods = data.get("goods", [])
                    checked_pages += 1
                    if not goods:
                        break
                    for g in goods:
                        checked_goods += 1
                        dt = _parse_ably_datetime(g.get("created_at"))
                        if dt and (latest_dt is None or dt > latest_dt):
                            latest_dt = dt
                    if page >= data.get("max_page_number", 1):
                        break
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"상품 목록 조회 실패: {exc}")

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            row = conn.execute("SELECT value FROM wonbe_meta WHERE key = 'last_sync_at'").fetchone()
            last_sync_at_str = row["value"] if row else None

            last_sync_dt = None
            if last_sync_at_str:
                try:
                    last_sync_dt = datetime.strptime(last_sync_at_str, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    last_sync_dt = None

            if latest_dt is None:
                status = "blue"
            elif last_sync_dt is None or last_sync_dt < latest_dt:
                status = "red"
            else:
                status = "blue"

            ingodaegi_added = _sync_ingodaegi_from_wonbe(conn)
            ablystock_added = _sync_ably_stock_from_wonbe(conn)

            checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            latest_created_at = latest_dt.strftime("%Y-%m-%d %H:%M:%S") if latest_dt else None
            meta_values = {
                "freshness_status": status,
                "freshness_checked_at": checked_at,
                "freshness_latest_created_at": latest_created_at or "",
                "freshness_checked_goods": str(checked_goods),
                "freshness_checked_pages": str(checked_pages),
                "freshness_ingodaegi_added": str(ingodaegi_added),
                "freshness_ablystock_added": str(ablystock_added),
            }
            conn.executemany(
                "INSERT OR REPLACE INTO wonbe_meta (key, value) VALUES (?, ?)",
                list(meta_values.items()),
            )
            conn.commit()
        finally:
            conn.close()

        return {
            "ok": True,
            "status": status,
            "latest_created_at": latest_created_at,
            "last_sync_at": last_sync_at_str,
            "checked_goods": checked_goods,
            "checked_pages": checked_pages,
            "checked_at": checked_at,
            "ingodaegi_added": ingodaegi_added,
            "ablystock_added": ablystock_added,
        }

    @router.get("/freshness-status")
    def wonbe_freshness_status(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            keys = [
                "freshness_status",
                "freshness_checked_at",
                "freshness_latest_created_at",
                "freshness_checked_goods",
                "freshness_checked_pages",
                "freshness_ingodaegi_added",
                "freshness_ablystock_added",
                "last_sync_at",
            ]
            placeholders = ", ".join("?" for _ in keys)
            rows = conn.execute(
                f"SELECT key, value FROM wonbe_meta WHERE key IN ({placeholders})",
                keys,
            ).fetchall()
            meta = {r["key"]: r["value"] for r in rows}
        finally:
            conn.close()

        def _int_or_zero(v):
            try:
                return int(v)
            except (TypeError, ValueError):
                return 0

        return {
            "ok": True,
            "status": meta.get("freshness_status"),
            "latest_created_at": meta.get("freshness_latest_created_at") or None,
            "last_sync_at": meta.get("last_sync_at"),
            "checked_goods": _int_or_zero(meta.get("freshness_checked_goods")),
            "checked_pages": _int_or_zero(meta.get("freshness_checked_pages")),
            "checked_at": meta.get("freshness_checked_at"),
            "ingodaegi_added": _int_or_zero(meta.get("freshness_ingodaegi_added")),
            "ablystock_added": _int_or_zero(meta.get("freshness_ablystock_added")),
        }

    @router.post("/sync-ably-sno")
    async def wonbe_sync_ably_sno(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        force = bool(payload.get("force"))

        async with httpx.AsyncClient(timeout=15.0) as client:
            res = await client.post(
                f"{_ABLY_BASE}/seller/login/",
                json={"email": _ABLY_EMAIL, "password": _ABLY_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "Origin": "https://seller-admin.a-bly.com",
                    "Referer": "https://seller-admin.a-bly.com/",
                    "User-Agent": "Mozilla/5.0",
                },
            )
            if not res.is_success:
                raise HTTPException(status_code=502, detail="에이블리 로그인 실패")
        token = res.json().get("token")
        if not token:
            raise HTTPException(status_code=502, detail="에이블리 로그인 실패: 토큰 없음")

        ably_headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Origin": "https://seller-admin.a-bly.com",
            "Referer": "https://seller-admin.a-bly.com/",
            "User-Agent": "Mozilla/5.0",
        }

        # 상품명(대괄호 태그 제거) → sno 매핑을 전체 카탈로그를 훑어서 구성
        name_to_sno: dict[str, int] = {}
        fetched_goods = 0
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                page = 1
                max_page = 1
                while page <= max_page:
                    res = await client.post(
                        f"{_ABLY_BASE}/seller/goods/search/",
                        headers=ably_headers,
                        json={"page": page, "per_page": 1000},
                    )
                    res.raise_for_status()
                    data = res.json()
                    goods = data.get("goods", [])
                    max_page = data.get("max_page_number", 1) or 1
                    for g in goods:
                        fetched_goods += 1
                        sno = g.get("sno")
                        raw_name = str(g.get("name") or "")
                        stripped = re.sub(r"^(\[[^\]]*\]\s*)+", "", raw_name).strip()
                        key = _normalize_cost_base_key(stripped)
                        if key and sno:
                            name_to_sno[key] = sno
                    page += 1
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"상품 목록 조회 실패: {exc}")

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            if force:
                rows = conn.execute("SELECT 상품코드, 상품명 FROM wonbe").fetchall()
            else:
                rows = conn.execute(
                    "SELECT 상품코드, 상품명 FROM wonbe WHERE 에이블리상품번호 IS NULL OR 에이블리상품번호 = ''"
                ).fetchall()

            updates = []
            unmatched = 0
            for r in rows:
                key = _normalize_cost_base_key(r["상품명"])
                sno = name_to_sno.get(key)
                if sno:
                    updates.append((str(sno), r["상품코드"]))
                else:
                    unmatched += 1

            if updates:
                conn.executemany(
                    "UPDATE wonbe SET 에이블리상품번호 = ? WHERE 상품코드 = ?",
                    updates,
                )
                conn.commit()
        finally:
            conn.close()

        return {
            "ok": True,
            "fetched_goods": fetched_goods,
            "considered": len(rows),
            "matched": len(updates),
            "unmatched": unmatched,
        }

    @router.post("/sync-registration-date")
    async def wonbe_sync_registration_date(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        async with httpx.AsyncClient(timeout=15.0) as client:
            res = await client.post(
                f"{_ABLY_BASE}/seller/login/",
                json={"email": _ABLY_EMAIL, "password": _ABLY_PASSWORD},
                headers={
                    "Content-Type": "application/json",
                    "Origin": "https://seller-admin.a-bly.com",
                    "Referer": "https://seller-admin.a-bly.com/",
                    "User-Agent": "Mozilla/5.0",
                },
            )
            if not res.is_success:
                raise HTTPException(status_code=502, detail="에이블리 로그인 실패")
        token = res.json().get("token")
        if not token:
            raise HTTPException(status_code=502, detail="에이블리 로그인 실패: 토큰 없음")

        ably_headers = {
            "Authorization": f"JWT {token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Origin": "https://seller-admin.a-bly.com",
            "Referer": "https://seller-admin.a-bly.com/",
            "User-Agent": "Mozilla/5.0",
        }

        # sno → (등록일, 진열상태, 품절상태) 매핑을 전체 카탈로그를 훑어서 구성
        sno_to_info: dict[str, tuple[str, str, str]] = {}
        fetched_goods = 0
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                page = 1
                max_page = 1
                while page <= max_page:
                    res = await client.post(
                        f"{_ABLY_BASE}/seller/goods/search/",
                        headers=ably_headers,
                        json={"page": page, "per_page": 1000},
                    )
                    res.raise_for_status()
                    data = res.json()
                    goods = data.get("goods", [])
                    max_page = data.get("max_page_number", 1) or 1
                    for g in goods:
                        fetched_goods += 1
                        sno = g.get("sno")
                        if not sno:
                            continue
                        dt = _parse_ably_datetime(g.get("created_at"))
                        created = dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""
                        display_status = "진열" if g.get("is_open") else "미진열"
                        soldout_status = "품절" if g.get("is_soldout") else "정상"
                        sno_to_info[str(sno)] = (created, display_status, soldout_status)
                    page += 1
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"상품 목록 조회 실패: {exc}")

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            rows = conn.execute(
                "SELECT 상품코드, 에이블리상품번호 FROM wonbe WHERE 에이블리상품번호 IS NOT NULL AND 에이블리상품번호 != ''"
            ).fetchall()

            updates = []
            unmatched = 0
            for r in rows:
                info = sno_to_info.get(str(r["에이블리상품번호"]))
                if info:
                    created, display_status, soldout_status = info
                    updates.append((created, display_status, soldout_status, r["상품코드"]))
                else:
                    unmatched += 1

            if updates:
                conn.executemany(
                    "UPDATE wonbe SET 등록일 = ?, 진열상태 = ?, 품절상태 = ? WHERE 상품코드 = ?",
                    updates,
                )
                conn.commit()
        finally:
            conn.close()

        return {
            "ok": True,
            "fetched_goods": fetched_goods,
            "considered": len(rows),
            "matched": len(updates),
            "unmatched": unmatched,
        }

    @router.get("/sync-country/progress")
    def wonbe_sync_country_progress(user: str = Depends(get_current_user)):
        return _country_sync_progress.get(user) or {
            "running": False, "total": 0, "done": 0, "matched": 0,
        }

    @router.post("/sync-country")
    async def wonbe_sync_country(user: str = Depends(get_current_user)):
        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            rows = conn.execute(
                "SELECT 상품코드, 에이블리상품번호 FROM wonbe WHERE 에이블리상품번호 IS NOT NULL AND 에이블리상품번호 != ''"
            ).fetchall()
        finally:
            conn.close()

        if not rows:
            return {"ok": True, "considered": 0, "matched": 0, "unmatched": 0}

        ably_client = AblyClient()
        try:
            await ably_client.login()
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {exc}")

        # 색상/사이즈별로 상품코드는 다르지만 에이블리상품번호(sno)는 같은 경우가 많음 → sno 기준으로 중복 제거 후 1회씩만 조회
        unique_snos = sorted({str(r["에이블리상품번호"]) for r in rows})

        progress = {"running": True, "total": len(unique_snos), "done": 0, "matched": 0}
        _country_sync_progress[user] = progress
        progress_lock = asyncio.Lock()

        # 상품 상세는 상품별 개별 조회라서 동시 요청 수를 제한한다
        semaphore = asyncio.Semaphore(8)

        async def _fetch(sno: str):
            async with semaphore:
                try:
                    detail = await ably_client.get_goods_detail(sno)
                    c = str(detail.get("country") or "").strip()
                except Exception:
                    c = ""
            async with progress_lock:
                progress["done"] += 1
                if c:
                    progress["matched"] += 1
            return sno, c

        try:
            results = await asyncio.gather(*[_fetch(sno) for sno in unique_snos])
        finally:
            progress["running"] = False

        sno_to_country = {sno: c for sno, c in results if c}

        updates = []
        unmatched = 0
        for r in rows:
            c = sno_to_country.get(str(r["에이블리상품번호"]))
            if c:
                updates.append((c, r["상품코드"]))
            else:
                unmatched += 1

        conn = _get_wonbe_db()
        try:
            _init_wonbe_table(conn)
            if updates:
                conn.executemany("UPDATE wonbe SET 제조국 = ? WHERE 상품코드 = ?", updates)
                conn.commit()
        finally:
            conn.close()

        return {
            "ok": True,
            "considered": len(rows),
            "unique_snos": len(unique_snos),
            "matched": len(updates),
            "unmatched": unmatched,
        }

    @router.post("/janggi/save")
    def janggi_save(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        rows = payload.get("rows") or []
        if not rows:
            raise HTTPException(status_code=400, detail="저장할 데이터가 없습니다.")
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            data_rows = [
                (
                    str(r.get("거래처") or ""),
                    str(r.get("거래처상품명") or ""),
                    str(r.get("가격") or ""),
                    str(r.get("옵션") or ""),
                    str(r.get("사이즈") or ""),
                    str(r.get("개수") or ""),
                    str(r.get("날짜") or ""),
                    str(r.get("미송체크") or ""),
                    str(r.get("상품코드") or ""),
                    str(r.get("메모") or ""),
                    str(r.get("거래처합산") or ""),
                )
                for r in rows
            ]
            conn.executemany(
                f"INSERT INTO 날짜별장끼정리 ({', '.join(JANGGI_COLUMNS)}) VALUES ({', '.join(['?'] * len(JANGGI_COLUMNS))})",
                data_rows,
            )
            conn.commit()
            return {"ok": True, "saved": len(data_rows)}
        finally:
            conn.close()

    @router.get("/janggi/search")
    def janggi_search(
        q: str = "",
        date: str = "",
        misong_filter: str = "",
        ilgwal_only: str = "",
        offset: int = 0,
        limit: int = 50,
        user: str = Depends(get_current_user),
    ):
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            q = q.strip()
            date = date.strip()
            misong_filter = misong_filter.strip()
            order = "ORDER BY 날짜 DESC, 거래처 DESC"

            conditions = []
            params: list = []
            if date:
                conditions.append("날짜 = ?")
                params.append(date)
            if q:
                like = f"%{q}%"
                conditions.append("(거래처 LIKE ? OR 거래처상품명 LIKE ? OR 상품코드 LIKE ?)")
                params.extend([like, like, like])
            if misong_filter == "미송":
                conditions.append("(미송체크 LIKE '%미송%' AND 미송체크 NOT LIKE '%픽업%')")
            elif misong_filter == "미송픽업":
                conditions.append("(미송체크 LIKE '%미송픽업%' OR (미송체크 LIKE '%미송%' AND 미송체크 LIKE '%픽업%'))")
            elif misong_filter == "매입차감":
                conditions.append("미송체크 LIKE '%매입차감%'")
            if ilgwal_only == "Y":
                conditions.append("일괄이체 = 'Y'")

            where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

            rows = conn.execute(
                f"SELECT * FROM 날짜별장끼정리 {where} {order} LIMIT ? OFFSET ?",
                params + [limit, offset],
            ).fetchall()
            total = conn.execute(
                f"SELECT COUNT(*) FROM 날짜별장끼정리 {where}",
                params,
            ).fetchone()[0]

            return {
                "ok": True,
                "rows": [dict(r) for r in rows],
                "total": total,
                "offset": offset,
                "limit": limit,
            }
        finally:
            conn.close()

    @router.patch("/janggi/row")
    def janggi_update_row(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        row_id = payload.get("id")
        col = str(payload.get("column") or "").strip()
        value = str(payload.get("value") or "").strip()
        if row_id is None or col not in JANGGI_COLUMNS:
            raise HTTPException(status_code=400, detail="id와 유효한 column이 필요합니다.")
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            cur = conn.execute(
                f"UPDATE 날짜별장끼정리 SET {col} = ? WHERE id = ?",
                (value, row_id),
            )
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 id 없음")
            row = dict(conn.execute("SELECT * FROM 날짜별장끼정리 WHERE id = ?", (row_id,)).fetchone())

            recalc_rows: list[dict] = []
            if col == "가격":
                거래처 = row.get("거래처") or ""
                날짜 = row.get("날짜") or ""
                if 거래처 and 날짜:
                    group = conn.execute(
                        "SELECT id, 가격 FROM 날짜별장끼정리 WHERE 거래처 = ? AND 날짜 = ?",
                        (거래처, 날짜),
                    ).fetchall()
                    total = sum(_parse_amount_janggi(r["가격"]) for r in group)
                    total_str = str(int(total)) if total == int(total) else str(total)
                    ids = [r["id"] for r in group]
                    conn.execute(
                        f"UPDATE 날짜별장끼정리 SET 거래처합산 = ? WHERE 거래처 = ? AND 날짜 = ?",
                        (total_str, 거래처, 날짜),
                    )
                    conn.commit()
                    row["거래처합산"] = total_str
                    recalc_rows = [{"id": rid, "거래처합산": total_str} for rid in ids if rid != row_id]

            return {"ok": True, "row": row, "recalc_rows": recalc_rows}
        finally:
            conn.close()

    @router.post("/janggi/row")
    def janggi_add_row(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            values = tuple(str(payload.get(col) or "") for col in JANGGI_COLUMNS)
            cur = conn.execute(
                f"INSERT INTO 날짜별장끼정리 ({', '.join(JANGGI_COLUMNS)}) VALUES ({', '.join(['?'] * len(JANGGI_COLUMNS))})",
                values,
            )
            conn.commit()
            row = dict(conn.execute("SELECT * FROM 날짜별장끼정리 WHERE id = ?", (cur.lastrowid,)).fetchone())

            recalc_rows: list[dict] = []
            거래처 = row.get("거래처") or ""
            날짜 = row.get("날짜") or ""
            if 거래처 and 날짜:
                group = conn.execute(
                    "SELECT id, 가격 FROM 날짜별장끼정리 WHERE 거래처 = ? AND 날짜 = ?",
                    (거래처, 날짜),
                ).fetchall()
                total = sum(_parse_amount_janggi(r["가격"]) for r in group)
                total_str = str(int(total)) if total == int(total) else str(total)
                conn.execute(
                    "UPDATE 날짜별장끼정리 SET 거래처합산 = ? WHERE 거래처 = ? AND 날짜 = ?",
                    (total_str, 거래처, 날짜),
                )
                conn.commit()
                row["거래처합산"] = total_str
                recalc_rows = [{"id": r["id"], "거래처합산": total_str} for r in group if r["id"] != cur.lastrowid]

            return {"ok": True, "row": row, "recalc_rows": recalc_rows}
        finally:
            conn.close()

    @router.delete("/janggi/row")
    def janggi_delete_row(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        row_id = payload.get("id")
        if row_id is None:
            raise HTTPException(status_code=400, detail="id 필요")
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            cur = conn.execute("DELETE FROM 날짜별장끼정리 WHERE id = ?", (row_id,))
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 id 없음")
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    @router.delete("/janggi/by-date")
    def janggi_delete_by_date(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        date_str = str(payload.get("날짜") or "").strip()
        if not date_str:
            raise HTTPException(status_code=400, detail="날짜 필요")
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            cur = conn.execute("DELETE FROM 날짜별장끼정리 WHERE 날짜 = ?", (date_str,))
            conn.commit()
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    @router.get("/janggi/aliases")
    def janggi_get_aliases(user: str = Depends(get_current_user)):
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            rows = conn.execute("SELECT key, value FROM janggi_aliases").fetchall()
            return {"ok": True, "aliases": {r["key"]: r["value"] for r in rows}}
        finally:
            conn.close()

    @router.put("/janggi/aliases")
    def janggi_save_aliases(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        aliases = payload.get("aliases") or {}
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            conn.execute("DELETE FROM janggi_aliases")
            if aliases:
                conn.executemany(
                    "INSERT INTO janggi_aliases (key, value) VALUES (?, ?)",
                    [(str(k), str(v)) for k, v in aliases.items() if k and v],
                )
            conn.commit()
            return {"ok": True, "count": len(aliases)}
        finally:
            conn.close()

    @router.get("/janggi/top-col-widths")
    def janggi_get_col_widths(user: str = Depends(get_current_user)):
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            rows = conn.execute(
                "SELECT col_key, width FROM janggi_top_col_widths WHERE username = ?", (user,)
            ).fetchall()
            return {"ok": True, "widths": {r["col_key"]: r["width"] for r in rows}}
        finally:
            conn.close()

    @router.put("/janggi/top-col-widths")
    def janggi_save_col_widths(
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        widths = payload.get("widths") or {}
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            conn.execute("DELETE FROM janggi_top_col_widths WHERE username = ?", (user,))
            if widths:
                conn.executemany(
                    "INSERT INTO janggi_top_col_widths (username, col_key, width) VALUES (?, ?, ?)",
                    [(user, str(k), int(round(float(v)))) for k, v in widths.items() if k and v],
                )
            conn.commit()
            return {"ok": True, "count": len(widths)}
        finally:
            conn.close()

    @router.get("/janggi/recent-summary")
    def janggi_recent_summary(user: str = Depends(get_current_user)):
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            latest = conn.execute(
                "SELECT MAX(날짜) AS d FROM 날짜별장끼정리"
            ).fetchone()["d"]
            if not latest:
                return {"ok": True, "date": None, "rows": []}
            raw_rows = conn.execute(
                """SELECT 거래처, 가격
                   FROM 날짜별장끼정리
                   WHERE 날짜 = ? AND 거래처 != ''""",
                (latest,),
            ).fetchall()
            totals: dict[str, float] = {}
            for r in raw_rows:
                supplier = r["거래처"]
                totals[supplier] = totals.get(supplier, 0.0) + _parse_amount_janggi(r["가격"])
            rows = sorted(
                ({"거래처": k, "합산": v} for k, v in totals.items()),
                key=lambda x: x["거래처"],
                reverse=True,
            )
            return {
                "ok": True,
                "date": latest,
                "rows": rows,
            }
        finally:
            conn.close()

    @router.get("/janggi/top-shops")
    async def janggi_top_shops(user: str = Depends(get_current_user)):
        if not _BS4_OK:
            raise HTTPException(status_code=500, detail="beautifulsoup4가 설치되지 않았습니다.")
        today = datetime.now().strftime("%Y-%m-%d")
        try:
            async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
                await client.get(f"{_TOP90_BASE}/login",
                                 headers={"User-Agent": "Mozilla/5.0"})
                await client.post(
                    f"{_TOP90_BASE}/login/auth",
                    data={
                        "login_email": _TOP90_EMAIL,
                        "login_password": _TOP90_PASSWORD,
                        "login_remember": "on",
                    },
                    headers={"User-Agent": "Mozilla/5.0"},
                )
                search_res = await client.post(
                    f"{_TOP90_BASE}/inquiry/done_search",
                    data={"obj[start]": today, "obj[end]": today},
                    headers={
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": f"{_TOP90_BASE}/inquiry/done/",
                        "User-Agent": "Mozilla/5.0",
                    },
                )
                try:
                    items = search_res.json()
                except Exception:
                    raise HTTPException(status_code=502, detail="TOP90 응답 파싱 실패 (로그인 확인)")

                today_items = [
                    it for it in (items if isinstance(items, list) else [])
                    if str(it.get("l_timestamp", "")).startswith(today)
                ]

                seen: set[str] = set()
                shops: list[str] = []
                for it in today_items:
                    detail_res = await client.get(
                        f"{_TOP90_BASE}/inquiry/done_detail/{it['c_idx']}/{it['l_idx']}",
                        headers={"Referer": f"{_TOP90_BASE}/inquiry/done/",
                                 "User-Agent": "Mozilla/5.0"},
                    )
                    soup = _BS4(detail_res.text, "html.parser")
                    for tr in soup.select("table tbody tr"):
                        state_td = tr.select_one("td.i-type")
                        if not state_td or "완료" not in state_td.get_text(" ", strip=True):
                            continue
                        shop_tag = tr.select_one("h2.shopname")
                        if shop_tag:
                            name = shop_tag.get_text(strip=True)
                            if name and name not in seen:
                                seen.add(name)
                                shops.append(name)

                return {"ok": True, "date": today, "shops": shops}
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"TOP90 조회 실패: {exc}")

    # ── 거래처계좌데이터 CRUD ─────────────────────────────────────

    @router.get("/account/rows")
    def account_get_rows(user: str = Depends(get_current_user)):
        conn = _get_janggi_db()
        try:
            _init_account_table(conn)
            rows = conn.execute("SELECT * FROM 거래처계좌데이터 ORDER BY id").fetchall()
            return {"ok": True, "rows": [dict(r) for r in rows]}
        finally:
            conn.close()

    @router.post("/account/save-all")
    def account_save_all(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw_rows = payload.get("rows")
        if not isinstance(raw_rows, list):
            raise HTTPException(status_code=400, detail="rows 형식 오류")
        conn = _get_janggi_db()
        try:
            _init_account_table(conn)
            conn.execute("DELETE FROM 거래처계좌데이터")
            data = []
            for r in raw_rows:
                if not isinstance(r, dict):
                    continue
                data.append(tuple(str(r.get(c) or "").strip() for c in ACCOUNT_COLS))
            if data:
                conn.executemany(
                    "INSERT INTO 거래처계좌데이터 (A, B, C, D, E, F) VALUES (?, ?, ?, ?, ?, ?)",
                    data,
                )
            conn.commit()
            rows = conn.execute("SELECT * FROM 거래처계좌데이터 ORDER BY id").fetchall()
            return {"ok": True, "saved_count": len(data), "rows": [dict(r) for r in rows]}
        finally:
            conn.close()

    @router.post("/account/row")
    def account_add_row(payload: dict = Body(default={}), user: str = Depends(get_current_user)):
        conn = _get_janggi_db()
        try:
            _init_account_table(conn)
            values = {c: str(payload.get(c) or "").strip() for c in ACCOUNT_COLS}
            if not values["B"] and values["D"]:
                values["B"] = _resolve_bank_code(values["D"])
            cur = conn.execute(
                "INSERT INTO 거래처계좌데이터 (A, B, C, D, E, F) VALUES (?, ?, ?, ?, ?, ?)",
                tuple(values[c] for c in ACCOUNT_COLS),
            )
            conn.commit()
            row = conn.execute("SELECT * FROM 거래처계좌데이터 WHERE id = ?", (cur.lastrowid,)).fetchone()
            return {"ok": True, "row": dict(row)}
        finally:
            conn.close()

    @router.patch("/account/row")
    def account_update_row(payload: dict = Body(...), user: str = Depends(get_current_user)):
        row_id = payload.get("id")
        col = str(payload.get("col") or "").strip().upper()
        value = str(payload.get("value") or "").strip()
        if row_id is None or col not in ACCOUNT_COLS:
            raise HTTPException(status_code=400, detail="id와 유효한 col 필요")
        conn = _get_janggi_db()
        try:
            _init_account_table(conn)
            if col == "D":
                # 은행명(D)을 수정하면 은행코드(B)를 자동으로 다시 계산한다
                cur = conn.execute(
                    "UPDATE 거래처계좌데이터 SET D = ?, B = ? WHERE id = ?",
                    (value, _resolve_bank_code(value), row_id),
                )
            else:
                cur = conn.execute(
                    f"UPDATE 거래처계좌데이터 SET {col} = ? WHERE id = ?",
                    (value, row_id),
                )
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 id 없음")
            row = conn.execute("SELECT * FROM 거래처계좌데이터 WHERE id = ?", (row_id,)).fetchone()
            return {"ok": True, "row": dict(row)}
        finally:
            conn.close()

    @router.delete("/account/row")
    def account_delete_row(payload: dict = Body(...), user: str = Depends(get_current_user)):
        row_id = payload.get("id")
        if row_id is None:
            raise HTTPException(status_code=400, detail="id 필요")
        conn = _get_janggi_db()
        try:
            _init_account_table(conn)
            cur = conn.execute("DELETE FROM 거래처계좌데이터 WHERE id = ?", (row_id,))
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 id 없음")
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    @router.post("/account/import-default-xlsx")
    def account_import_default_xlsx(user: str = Depends(get_current_user)):
        default_path = Path(r"C:\Users\ksh29\OneDrive\Desktop\원베\거래처계좌데이터.xlsx")
        if not default_path.exists():
            raise HTTPException(status_code=404, detail=f"파일 없음: {default_path}")
        try:
            import pandas as _pd
            df = _pd.read_excel(default_path, header=None, dtype=object)
            xlsx_rows = df.fillna("").values.tolist()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"파일 읽기 오류: {exc}")
        conn = _get_janggi_db()
        try:
            _init_account_table(conn)
            conn.execute("DELETE FROM 거래처계좌데이터")
            data = []
            for row in xlsx_rows:
                padded = list(row[:6]) + [""] * max(0, 6 - len(list(row[:6])))
                vals = tuple(str(v or "").strip() for v in padded[:6])
                if any(vals):
                    data.append(vals)
            if data:
                conn.executemany(
                    "INSERT INTO 거래처계좌데이터 (A, B, C, D, E, F) VALUES (?, ?, ?, ?, ?, ?)",
                    data,
                )
            conn.commit()
            return {"ok": True, "imported_count": len(data)}
        finally:
            conn.close()

    # ── 이체파일 조회/삭제 ────────────────────────────────────────

    @router.get("/ichae/dates")
    def ichae_dates(user: str = Depends(get_current_user)):
        conn = _get_janggi_db()
        try:
            _init_ichae_table(conn)
            rows = conn.execute(
                "SELECT DISTINCT 날짜 FROM 이체파일 ORDER BY 날짜 DESC LIMIT 60"
            ).fetchall()
            return {"ok": True, "dates": [r["날짜"] for r in rows]}
        finally:
            conn.close()

    @router.get("/ichae/search")
    def ichae_search(
        date: str = "",
        offset: int = 0,
        limit: int = 300,
        user: str = Depends(get_current_user),
    ):
        conn = _get_janggi_db()
        try:
            _init_ichae_table(conn)
            date = date.strip()
            if date:
                rows = conn.execute(
                    "SELECT * FROM 이체파일 WHERE 날짜 = ? ORDER BY 상태 ASC, D ASC LIMIT ? OFFSET ?",
                    (date, limit, offset),
                ).fetchall()
                total = conn.execute(
                    "SELECT COUNT(*) FROM 이체파일 WHERE 날짜 = ?", (date,)
                ).fetchone()[0]
            else:
                rows = conn.execute(
                    "SELECT * FROM 이체파일 ORDER BY 날짜 DESC, 상태 ASC, D ASC LIMIT ? OFFSET ?",
                    (limit, offset),
                ).fetchall()
                total = conn.execute("SELECT COUNT(*) FROM 이체파일").fetchone()[0]
            return {"ok": True, "rows": [dict(r) for r in rows], "total": total}
        finally:
            conn.close()

    @router.patch("/ichae/row")
    def ichae_update_row(payload: dict = Body(...), user: str = Depends(get_current_user)):
        row_id = payload.get("id")
        col = str(payload.get("col") or "").strip().upper()
        value = str(payload.get("value") or "")
        if row_id is None or col not in ("E", "F"):
            raise HTTPException(status_code=400, detail="id와 E 또는 F 컬럼만 수정 가능합니다.")
        conn = _get_janggi_db()
        try:
            _init_ichae_table(conn)
            cur = conn.execute(
                f"UPDATE 이체파일 SET {col} = ? WHERE id = ?", (value, row_id)
            )
            conn.commit()
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="해당 id 없음")
            row = conn.execute("SELECT * FROM 이체파일 WHERE id = ?", (row_id,)).fetchone()
            return {"ok": True, "row": dict(row)}
        finally:
            conn.close()

    @router.delete("/ichae/by-date")
    def ichae_delete_by_date(payload: dict = Body(...), user: str = Depends(get_current_user)):
        date_str = str(payload.get("날짜") or "").strip()
        if not date_str:
            raise HTTPException(status_code=400, detail="날짜 필요")
        conn = _get_janggi_db()
        try:
            _init_ichae_table(conn)
            cur = conn.execute("DELETE FROM 이체파일 WHERE 날짜 = ?", (date_str,))
            conn.commit()
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    @router.get("/ichae/export")
    def ichae_export(month: str = "", user: str = Depends(get_current_user)):
        month = month.strip()
        if month and not re.fullmatch(r"\d{4}-\d{2}", month):
            raise HTTPException(status_code=400, detail="month는 YYYY-MM 형식이어야 합니다.")
        conn = _get_janggi_db()
        try:
            _init_ichae_table(conn)
            if month:
                rows = conn.execute(
                    "SELECT * FROM 이체파일 WHERE 날짜 LIKE ? ORDER BY 날짜 ASC, 상태 ASC, D ASC",
                    (f"{month}%",),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM 이체파일 ORDER BY 날짜 ASC, 상태 ASC, D ASC"
                ).fetchall()
        finally:
            conn.close()

        export_cols = ["날짜", "A", "B", "C", "D", "E", "F", "상태"]
        export_labels = ["날짜", "은행코드", "계좌번호", "입금금액", "거래처명", "거래처가 보는 메모", "우리가 보는 메모", "상태"]

        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet1")
        for ci, label in enumerate(export_labels):
            sheet.write(0, ci, label)
        for ri, row in enumerate(rows, start=1):
            for ci, col in enumerate(export_cols):
                sheet.write(ri, ci, row[col] if row[col] is not None else "")

        buf = io.BytesIO()
        book.save(buf)
        filename = f"이체파일_{month}.xls" if month else "이체파일_전체.xls"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": _content_disposition(filename)},
        )

    # ── 일괄이체목록 마킹 ────────────────────────────────────────

    @router.post("/janggi/bulk-ichae-mark")
    def bulk_ichae_mark(payload: dict = Body(...), user: str = Depends(get_current_user)):
        date_str = str(payload.get("날짜") or "").strip()
        if not date_str:
            raise HTTPException(status_code=400, detail="날짜 필요")
        q = str(payload.get("q") or "").strip()
        misong_filter = str(payload.get("misong_filter") or "").strip()
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            conditions = ["날짜 = ?"]
            params: list = [date_str]
            if q:
                like = f"%{q}%"
                conditions.append("(거래처 LIKE ? OR 거래처상품명 LIKE ? OR 상품코드 LIKE ?)")
                params.extend([like, like, like])
            if misong_filter == "미송":
                conditions.append("(미송체크 LIKE '%미송%' AND 미송체크 NOT LIKE '%픽업%')")
            elif misong_filter == "미송픽업":
                conditions.append("(미송체크 LIKE '%미송픽업%' OR (미송체크 LIKE '%미송%' AND 미송체크 LIKE '%픽업%'))")
            elif misong_filter == "매입차감":
                conditions.append("미송체크 LIKE '%매입차감%'")
            where = f"WHERE {' AND '.join(conditions)}"
            cur = conn.execute(f"UPDATE 날짜별장끼정리 SET 일괄이체 = 'Y' {where}", params)
            conn.commit()
            return {"ok": True, "marked": cur.rowcount}
        finally:
            conn.close()

    @router.delete("/janggi/bulk-ichae-mark")
    def bulk_ichae_clear(payload: dict = Body(...), user: str = Depends(get_current_user)):
        date_str = str(payload.get("날짜") or "").strip()
        row_id = payload.get("id")
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            if row_id is not None:
                cur = conn.execute(
                    "UPDATE 날짜별장끼정리 SET 일괄이체 = '' WHERE id = ?",
                    (row_id,)
                )
            elif date_str:
                cur = conn.execute(
                    "UPDATE 날짜별장끼정리 SET 일괄이체 = '' WHERE 날짜 = ? AND 일괄이체 = 'Y'",
                    (date_str,)
                )
            else:
                raise HTTPException(status_code=400, detail="날짜 또는 id 필요")
            conn.commit()
            return {"ok": True, "cleared": cur.rowcount}
        finally:
            conn.close()

    # ── 이체파일 전환 ─────────────────────────────────────────────

    @router.post("/janggi/to-ichae")
    def janggi_to_ichae(payload: dict = Body(...), user: str = Depends(get_current_user)):
        date_str = str(payload.get("날짜") or "").strip()
        if not date_str:
            raise HTTPException(status_code=400, detail="날짜 필요")
        conn = _get_janggi_db()
        try:
            _init_janggi_table(conn)
            _init_account_table(conn)
            _init_ichae_table(conn)

            include_bulk = bool(payload.get("include_bulk", False))
            if include_bulk:
                janggi_rows = conn.execute(
                    "SELECT * FROM 날짜별장끼정리 WHERE 날짜 = ?", (date_str,)
                ).fetchall()
            else:
                janggi_rows = conn.execute(
                    "SELECT * FROM 날짜별장끼정리 WHERE 날짜 = ? AND (일괄이체 IS NULL OR 일괄이체 != 'Y')",
                    (date_str,)
                ).fetchall()
            if not janggi_rows:
                raise HTTPException(status_code=404, detail=f"{date_str} 날짜의 데이터가 없습니다.")

            # 거래처별 가격 합산 (SUMIF)
            supplier_totals: dict[str, float] = {}
            for row in janggi_rows:
                supplier = str(row["거래처"] or "").strip()
                if not supplier:
                    continue
                supplier_totals[supplier] = supplier_totals.get(supplier, 0.0) + _parse_amount_janggi(row["가격"])

            # 거래처계좌데이터 로드
            account_rows = conn.execute("SELECT * FROM 거래처계좌데이터 ORDER BY id").fetchall()
            account_map: dict[str, dict] = {}
            for ar in account_rows:
                key = str(ar["A"] or "").strip()
                if key and key not in account_map:
                    account_map[key] = dict(ar)

            # 거래처 유니크 행 생성
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            class _ResultRows(list):
                def append(self, row):
                    # 구 계좌 데이터 행은 계좌번호가 빠진 8열 튜플일 수 있어 새 C열을 보정한다.
                    if len(row) == 8:
                        account = account_map.get(row[3]) or {}
                        row = tuple(row[:2]) + (account.get("C", ""),) + tuple(row[2:])
                    super().append(row)

            result_rows = _ResultRows()
            for supplier in sorted(supplier_totals.keys()):
                total = supplier_totals[supplier]
                ar = account_map.get(supplier)
                if ar:
                    result_rows.append((
                        date_str,
                        ar.get("B", "") or _resolve_bank_code(ar.get("D", "")),   # A: 은행코드
                        ar.get("C", ""),   # B: 계좌번호
                        total,             # C: 입금금액
                        ar.get("A", ""),   # D: 거래처명
                        "주식회사 유색",    # E: 거래처가 보는 메모 (기본값)
                        "",                # F: 우리가 보는 메모
                        "매칭",
                        now_str,
                    ))
                else:
                    result_rows.append((
                        date_str, "", "", total, supplier,
                        "주식회사 유색", "",
                        "미등록", now_str,
                    ))

            # 해당 날짜 교체
            conn.execute("DELETE FROM 이체파일 WHERE 날짜 = ?", (date_str,))
            conn.executemany(
                """INSERT INTO 이체파일 (날짜, A, B, C, D, E, F, 상태, 생성일시)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                result_rows,
            )
            conn.commit()

            matched = sum(1 for r in result_rows if r[7] == "매칭")
            return {
                "ok": True,
                "날짜": date_str,
                "총거래처": len(supplier_totals),
                "매칭": matched,
                "미등록": len(result_rows) - matched,
                "저장행수": len(result_rows),
            }
        finally:
            conn.close()

    return router
