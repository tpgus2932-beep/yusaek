from pathlib import Path
from datetime import datetime, timedelta
import asyncio
import json
import tempfile
import io
import re
import urllib.parse
import uuid

import httpx
import pandas as pd
from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from zipfile import BadZipFile
from openpyxl import load_workbook, Workbook
import xlwt

from services.top90_client import execute_main_orders, Top90Error
from services.order_history_store import init_order_history_table, record_order_history
from sdk.ably import AblyClient
from sdk.ezadmin import EzAdminClient, EzAdminSessionExpired
from api.wonbe_routes import (
    load_wonbe_option_sno_map,
    load_wonbe_client_info_by_code,
    load_wonbe_product_name_map,
)
from services.misong_lookup import load_misong_qty_by_code

_EZADMIN_BASE = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_JINMONEY_TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "assets" / "jinmoney_order_template.xlsx"
_JINMONEY_RECEIPT_RE = re.compile(r"org_value=['\"]([^'\"]*)['\"]", re.IGNORECASE)


def build_order_router(
    *,
    require_admin,
    get_db,
    order_cost_base_path: Path,
    get_setting,
    get_shared_db,
    get_user_display,
    is_render: bool = False,
):
    router = APIRouter()
    COST_BASE_CODE_COL = 0
    COST_BASE_MATCH_COL = 8
    COST_BASE_REQUIRED_COLS = COST_BASE_MATCH_COL + 1

    def _safe_str(x):
        if pd.isna(x):
            return ""
        return str(x).strip()

    def _norm_code(v: str) -> str:
        return _safe_str(v).upper()

    def _to_int(x, default=0):
        try:
            if pd.isna(x):
                return default
            return int(float(x))
        except Exception:
            return default

    def _ez_val(html_str):
        s = str(html_str or "")
        # <input ... value='X' ...> → X
        m = re.search(r"<input[^>]+\bvalue=['\"]([^'\"]*)['\"]", s, re.IGNORECASE)
        if m:
            return m.group(1)
        # <a ...>TEXT</a> → TEXT
        m = re.search(r">([^<]+)</a>", s)
        if m:
            return m.group(1).strip()
        # 태그 없으면 그대로
        return re.sub(r"<[^>]+>", "", s).strip()

    def _content_disposition(filename: str) -> str:
        safe_name = (filename or "download.xlsx").replace('"', "")
        ascii_name = "".join(ch if ord(ch) < 128 else "_" for ch in safe_name)
        quoted = urllib.parse.quote(safe_name)
        return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted}"

    def _jinmoney_receipt_qty(value) -> int:
        raw = str(value or "")
        match = _JINMONEY_RECEIPT_RE.search(raw)
        text = match.group(1) if match else _ez_val(raw)
        number = re.search(r"-?\d[\d,]*", text)
        return int(number.group(0).replace(",", "")) if number else 0

    def _jinmoney_size(option_text: str) -> str:
        normalized = re.sub(r"[\[\](){}]", "-", option_text.upper())
        match = re.search(r"(?:^|[-\s/])(2XL|XL|L|M|S)(?=$|[-\s/])", normalized)
        return match.group(1) if match else ""

    def _jinmoney_map_item(supply_product_name: str, option_text: str):
        match = re.match(r"^\s*진머니\s+(.+?)\s*$", supply_product_name, re.IGNORECASE)
        if not match:
            return None, "공급처상품명에서 '진머니 제품명' 형식을 확인할 수 없습니다."

        source_product = re.sub(r"\s+", "", match.group(1)).lower()
        option = re.sub(r"\s+", "", option_text).lower()
        size = _jinmoney_size(option_text)
        if not size:
            return None, "옵션에서 S/M/L/XL 사이즈를 찾을 수 없습니다."

        if source_product in {"1004", "w1004"}:
            if "연청" in option:
                product_code = "w1004-1"
            elif "흑청" in option:
                product_code = "w1004-2"
            else:
                return None, "1004 옵션에서 연청/흑청을 찾을 수 없습니다."
        elif source_product in {"1007", "w1007"}:
            if "흑청" in option:
                product_code = "W1007"
            elif "중청" in option:
                product_code = "w1007-1"
            elif "연청" in option:
                product_code = "w1007-2"
            else:
                return None, "w1007 옵션에서 흑청/중청/연청을 찾을 수 없습니다."
        elif source_product in {"1011", "w1011"}:
            if "슬림" in option:
                product_code = "1011"
            elif "베이직" in option:
                product_code = "1011-1"
            else:
                return None, "w1011 옵션에서 슬림/베이직을 찾을 수 없습니다."
        else:
            return None, f"아직 등록되지 않은 진머니 제품명입니다: {match.group(1)}"

        return {"productCode": product_code, "size": size}, None

    def _clean_product_name(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        while True:
            ns = re.sub(r"^\s*(\[[^\]]*\]|\([^)]*\)|\{[^}]*\})\s*", "", s)
            if ns == s:
                break
            s = ns.strip()
        while True:
            ns = re.sub(r"\s*\([^)]*\)\s*$", "", s)
            if ns == s:
                break
            s = ns.strip()
        return re.sub(r"\s+", " ", s).strip()

    def _clean_option_text(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        parts = [p.strip() for p in s.split("-") if p.strip()]
        return " ".join(parts).strip()

    def _normalize_daily_sales(v) -> str:
        s = _clean_product_name(v)
        s = s.replace("-", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s.lower()

    def _build_daily_sales_cost_map() -> dict[str, object]:
        out: dict[str, object] = {}
        for code, match_name in load_wonbe_product_name_map().items():
            if match_name is None:
                continue
            key = _normalize_daily_sales(match_name)
            if key:
                out[key] = code
        return out

    def _build_daily_sales_rows(ws, skip_header: bool = True) -> list[dict]:
        start_row = 2 if skip_header else 1
        rows: list[dict] = []
        for r in range(start_row, ws.max_row + 1):
            raw_a = ws.cell(row=r, column=1).value
            raw_b = ws.cell(row=r, column=2).value
            raw_d = ws.cell(row=r, column=4).value

            prod = _clean_product_name(raw_a)
            opt = _clean_option_text(raw_b)
            out_a = (f"{prod} {opt}").strip() if opt else prod
            rows.append({"name": out_a, "qty": raw_d})
        return rows

    def _pick_header(value: str | None, fallback: str) -> str:
        v = _safe_str(value)
        return v if v else fallback

    def _build_daily_sales_xls_bytes(
        rows: list[dict],
        headers: list[str],
        include_cols: list[int],
    ) -> bytes:
        book = xlwt.Workbook()
        sheet = book.add_sheet("결과")
        selected_headers = [headers[idx - 1] for idx in include_cols]
        for idx, header in enumerate(selected_headers):
            sheet.write(0, idx, header)

        for i, row in enumerate(rows, start=1):
            selected_values: list[object] = []
            for col_no in include_cols:
                if col_no == 1:
                    selected_values.append(row.get("name", ""))
                elif col_no == 2:
                    selected_values.append(row.get("code", ""))
                elif col_no == 3:
                    selected_values.append(row.get("qty", ""))
            for j, cell_val in enumerate(selected_values):
                sheet.write(i, j, cell_val if cell_val is not None else "")

        buf = io.BytesIO()
        book.save(buf)
        return buf.getvalue()

    def _find_registered_code_variants(conn, normalized_code: str) -> list[dict]:
        rows = conn.execute(
            "SELECT code, qty, created_at, updated_at FROM order_registered_codes"
        ).fetchall()
        out = []
        for r in rows:
            raw_code = _safe_str(r["code"])
            if _norm_code(raw_code) == normalized_code:
                out.append(
                    {
                        "code": raw_code,
                        "qty": int(r["qty"]) if r["qty"] is not None else 0,
                        "created_at": r["created_at"],
                        "updated_at": r["updated_at"],
                    }
                )
        return out

    def _read_order_excel_df(path: Path) -> pd.DataFrame:
        ext = path.suffix.lower()
        last_error = None
        if ext in (".xlsx", ".xlsm"):
            # Some files are mislabeled .xlsx but actually old .xls.
            try:
                return pd.read_excel(path, engine="openpyxl")
            except Exception as e:
                last_error = e
                try:
                    return pd.read_excel(path, engine="xlrd")
                except Exception as e2:
                    last_error = e2
        if ext == ".xls":
            try:
                return pd.read_excel(path, engine="xlrd")
            except Exception as e:
                last_error = e
                try:
                    return pd.read_excel(path, engine="openpyxl")
                except Exception as e2:
                    last_error = e2
        else:
            try:
                return pd.read_excel(path)
            except Exception as e:
                last_error = e

        # Legacy case: files saved as .xls but actual content is HTML table.
        try:
            html_tables = pd.read_html(path)
            if html_tables:
                return html_tables[0]
        except Exception as e:
            last_error = e

        raise ValueError(f"엑셀 형식을 읽을 수 없습니다. 파일 형식 또는 손상 여부를 확인하세요. ({last_error})")

    def _load_cost_base_items() -> list[dict]:
        return [
            {"name": name, "code": code}
            for code, name in load_wonbe_product_name_map().items()
        ]

    def _load_cost_base_name_map() -> dict[str, str]:
        items = _load_cost_base_items()
        m: dict[str, str] = {}
        for it in items:
            code = _norm_code(it.get("code", ""))
            name = it.get("name", "")
            if code and code not in m:
                m[code] = name
        return m

    @router.get("/order/status")
    def order_status(admin: str = Depends(require_admin)):
        conn = get_db()
        count_row = conn.execute("SELECT COUNT(*) AS cnt FROM order_registered_codes").fetchone()
        conn.close()
        return {
            "ok": True,
            "registered_count": int(count_row["cnt"]) if count_row else 0,
            "cost_base_path": str(order_cost_base_path),
            "cost_base_exists": order_cost_base_path.exists(),
        }

    @router.get("/order/cost-base/search")
    def order_cost_base_search(
        q: str = "",
        limit: int = 20,
        admin: str = Depends(require_admin),
    ):
        q_norm = (q or "").strip().lower()
        if limit <= 0 or limit > 100:
            limit = 20
        items = _load_cost_base_items()
        if q_norm:
            items = [
                it
                for it in items
                if q_norm in (it.get("name", "").lower()) or q_norm in (it.get("code", "").lower())
            ]
        # code 기준 dedupe
        seen = set()
        out = []
        for it in items:
            code = it.get("code", "")
            if not code or code in seen:
                continue
            seen.add(code)
            out.append(it)
            if len(out) >= limit:
                break
        return {"ok": True, "items": out}

    @router.get("/order/registered/list")
    def order_registered_list(admin: str = Depends(require_admin)):
        name_map = _load_cost_base_name_map()
        conn = get_db()
        rows = conn.execute(
            "SELECT code, qty, created_at, updated_at FROM order_registered_codes ORDER BY code ASC"
        ).fetchall()
        conn.close()
        merged: dict[str, dict] = {}
        for r in rows:
            code = _norm_code(r["code"])
            if not code:
                continue
            item = {
                "code": code,
                "name": name_map.get(code, ""),
                "qty": int(r["qty"]) if r["qty"] is not None else 0,
                "created_at": r["created_at"],
                "updated_at": r["updated_at"],
            }
            prev = merged.get(code)
            if not prev:
                merged[code] = item
                continue
            prev_updated = _safe_str(prev.get("updated_at"))
            cur_updated = _safe_str(item.get("updated_at"))
            if cur_updated >= prev_updated:
                merged[code] = item
        items = [merged[k] for k in sorted(merged.keys())]
        return {"ok": True, "items": items}

    @router.post("/order/registered/upsert")
    def order_registered_upsert(payload: dict = Body(...), admin: str = Depends(require_admin)):
        code = _norm_code(payload.get("code") or "")
        qty = payload.get("qty")
        if not code:
            raise HTTPException(status_code=400, detail="code가 필요합니다.")
        try:
            qty_i = int(float(str(qty).strip()))
        except Exception:
            raise HTTPException(status_code=400, detail="qty는 숫자여야 합니다.")
        if qty_i < 0:
            raise HTTPException(status_code=400, detail="qty는 0 이상이어야 합니다.")

        now = pd.Timestamp.utcnow().isoformat()
        conn = get_db()
        variants = _find_registered_code_variants(conn, code)
        created_at = now
        if variants:
            created_at_candidates = [_safe_str(v.get("created_at")) for v in variants if _safe_str(v.get("created_at"))]
            if created_at_candidates:
                created_at = min(created_at_candidates)
            for v in variants:
                conn.execute("DELETE FROM order_registered_codes WHERE code = ?", (_safe_str(v["code"]),))
        conn.execute(
            """
            INSERT INTO order_registered_codes (code, qty, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                qty = excluded.qty,
                updated_at = excluded.updated_at
            """,
            (code, qty_i, created_at, now),
        )
        conn.commit()
        conn.close()
        return {"ok": True, "code": code, "qty": qty_i}

    @router.post("/order/registered/add")
    def order_registered_add(payload: dict = Body(...), admin: str = Depends(require_admin)):
        code = _norm_code(payload.get("code") or "")
        qty = payload.get("qty")
        if not code:
            raise HTTPException(status_code=400, detail="code가 필요합니다.")
        try:
            qty_i = int(float(str(qty).strip()))
        except Exception:
            raise HTTPException(status_code=400, detail="qty는 숫자여야 합니다.")
        if qty_i < 0:
            raise HTTPException(status_code=400, detail="qty는 0 이상이어야 합니다.")

        now = pd.Timestamp.utcnow().isoformat()
        conn = get_db()
        variants = _find_registered_code_variants(conn, code)
        if variants:
            conn.close()
            raise HTTPException(status_code=409, detail="이미 등록된 코드입니다.")

        conn.execute(
            "INSERT INTO order_registered_codes (code, qty, created_at, updated_at) VALUES (?, ?, ?, ?)",
            (code, qty_i, now, now),
        )
        conn.commit()
        conn.close()
        return {"ok": True, "code": code, "qty": qty_i}

    @router.delete("/order/registered/{code}")
    def order_registered_delete(code: str, admin: str = Depends(require_admin)):
        code = _norm_code(code or "")
        if not code:
            raise HTTPException(status_code=400, detail="code가 비어있음")
        conn = get_db()
        variants = _find_registered_code_variants(conn, code)
        if variants:
            for v in variants:
                conn.execute("DELETE FROM order_registered_codes WHERE code = ?", (_safe_str(v["code"]),))
        else:
            conn.execute("DELETE FROM order_registered_codes WHERE code = ?", (code,))
        conn.commit()
        conn.close()
        return {"ok": True}

    @router.delete("/order/registered")
    def order_registered_delete_query(code: str = "", admin: str = Depends(require_admin)):
        code = _norm_code(code or "")
        if not code:
            raise HTTPException(status_code=400, detail="code가 비어있음")
        conn = get_db()
        variants = _find_registered_code_variants(conn, code)
        if variants:
            for v in variants:
                conn.execute("DELETE FROM order_registered_codes WHERE code = ?", (_safe_str(v["code"]),))
        else:
            conn.execute("DELETE FROM order_registered_codes WHERE code = ?", (code,))
        conn.commit()
        conn.close()
        return {"ok": True}

    @router.post("/order/check")
    async def order_check(file: UploadFile = File(...), admin: str = Depends(require_admin)):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xls", ".xlsm"}:
            raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능합니다.")

        data = await file.read()
        if not data:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
        tmp = Path(tempfile.gettempdir()) / f"__order_check_{uuid.uuid4().hex}.xlsx"
        tmp.write_bytes(data)
        try:
            try:
                df = _read_order_excel_df(tmp)
            except Exception as e:
                raise HTTPException(status_code=400, detail=str(e))
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

        if df.shape[1] < 10:
            raise HTTPException(status_code=400, detail="업로드 엑셀에 J열까지(최소 10개 컬럼) 필요합니다.")

        conn = get_db()
        rows = conn.execute("SELECT code, qty FROM order_registered_codes").fetchall()
        conn.close()
        reg = {_norm_code(r["code"]): int(r["qty"]) for r in rows if _norm_code(r["code"])}

        col_b = df.iloc[:, 1]
        col_c = df.iloc[:, 2]
        col_e = df.iloc[:, 4]
        col_f = df.iloc[:, 5]
        col_j = df.iloc[:, 9]

        results = []
        for i in range(len(df)):
            code = _norm_code(col_j.iloc[i])
            if not code:
                continue
            if code not in reg:
                continue

            actual_qty = _to_int(col_e.iloc[i], 0)
            received_qty = _to_int(col_f.iloc[i], 0)
            need_qty = int(reg[code])
            shortage = (received_qty - actual_qty) + need_qty
            if shortage == 0 and actual_qty < need_qty:
                name = f"{_safe_str(col_b.iloc[i])} {_safe_str(col_c.iloc[i])}".strip()
                results.append(
                    {
                        "code": code,
                        "name": name,
                        "received_qty": received_qty,
                        "need_qty": need_qty,
                        "actual_qty": actual_qty,
                        "shortage": shortage,
                        "text": f"{name} {shortage}  (부족)",
                    }
                )

        return {"ok": True, "count": len(results), "results": results}

    @router.post("/order/daily-sales/preview")
    async def order_daily_sales_preview(
        file: UploadFile = File(...),
        skip_header: bool = Form(True),
        admin: str = Depends(require_admin),
    ):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xlsm"}:
            raise HTTPException(status_code=400, detail="xlsx/xlsm만 업로드 가능합니다.")
        if not order_cost_base_path.exists():
            raise HTTPException(status_code=400, detail=f"원가베이스 DB가 없습니다: {order_cost_base_path}")

        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")

        tmp = Path(tempfile.gettempdir()) / f"__order_daily_sales_{uuid.uuid4().hex}{ext}"
        tmp.write_bytes(raw)
        try:
            try:
                in_wb = load_workbook(tmp, data_only=True)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"입력 파일 로드 실패: {e}")
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

        in_ws = in_wb.active
        cost_map = _build_daily_sales_cost_map()
        base_rows = _build_daily_sales_rows(in_ws, skip_header=skip_header)

        rows: list[dict] = []
        unmatched_products: list[str] = []
        unmatched_rows = 0
        seen_unmatched: set[str] = set()
        for row in base_rows:
            code = cost_map.get(_normalize_daily_sales(row.get("name", "")), "")
            code_text = _safe_str(code)
            enriched = {"name": row.get("name", ""), "code": code_text, "qty": row.get("qty")}
            rows.append(enriched)
            if code_text:
                continue
            unmatched_rows += 1
            name = _safe_str(row.get("name"))
            if name and name not in seen_unmatched:
                seen_unmatched.add(name)
                unmatched_products.append(name)

        return {
            "ok": True,
            "count": len(rows),
            "unmatched_row_count": unmatched_rows,
            "unmatched_product_count": len(unmatched_products),
            "unmatched_products": unmatched_products,
            "rows": rows,
        }

    @router.post("/order/daily-sales/export")
    async def order_daily_sales_export(
        file: UploadFile = File(...),
        skip_header: bool = Form(True),
        header_col1: str = Form("상품명"),
        header_col2: str = Form("상품코드"),
        header_col3: str = Form("옵션추가항목8"),
        include_col1: bool = Form(True),
        include_col2: bool = Form(True),
        include_col3: bool = Form(True),
        admin: str = Depends(require_admin),
    ):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xlsm"}:
            raise HTTPException(status_code=400, detail="xlsx/xlsm만 업로드 가능합니다.")
        if not order_cost_base_path.exists():
            raise HTTPException(status_code=400, detail=f"원가베이스 DB가 없습니다: {order_cost_base_path}")

        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")

        tmp = Path(tempfile.gettempdir()) / f"__order_daily_sales_export_{uuid.uuid4().hex}{ext}"
        tmp.write_bytes(raw)
        try:
            try:
                in_wb = load_workbook(tmp, data_only=True)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"입력 파일 로드 실패: {e}")
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

        in_ws = in_wb.active
        cost_map = _build_daily_sales_cost_map()
        base_rows = _build_daily_sales_rows(in_ws, skip_header=skip_header)
        rows: list[dict] = []
        for row in base_rows:
            code = cost_map.get(_normalize_daily_sales(row.get("name", "")), "")
            rows.append(
                {
                    "name": row.get("name", ""),
                    "code": _safe_str(code),
                    "qty": row.get("qty"),
                }
            )

        headers = [
            _pick_header(header_col1, "상품명"),
            _pick_header(header_col2, "상품코드"),
            _pick_header(header_col3, "옵션추가항목8"),
        ]
        include_cols: list[int] = []
        if include_col1:
            include_cols.append(1)
        if include_col2:
            include_cols.append(2)
        if include_col3:
            include_cols.append(3)
        if not include_cols:
            raise HTTPException(status_code=400, detail="다운로드할 열을 최소 1개 선택하세요.")

        content = _build_daily_sales_xls_bytes(rows, headers, include_cols)
        src_name = Path(file.filename or "daily_sales.xlsx").stem
        filename = f"{src_name}_일일판매량_가공본.xls"
        headers_resp = {"Content-Disposition": _content_disposition(filename)}
        return Response(
            content=content,
            media_type="application/vnd.ms-excel",
            headers=headers_resp,
        )

    @router.post("/order/daily-sales/process")
    async def order_daily_sales_process(
        file: UploadFile = File(...),
        skip_header: bool = Form(True),
        admin: str = Depends(require_admin),
    ):
        return await order_daily_sales_export(
            file=file,
            skip_header=skip_header,
            header_col1="상품명",
            header_col2="상품코드",
            header_col3="옵션추가항목8",
            include_col1=True,
            include_col2=True,
            include_col3=True,
            admin=admin,
        )

    # ── 인기재고 (에이블리 상품별 판매 통계 랭킹) ────────────────────────────────
    POPULAR_GOODS_SNAPSHOT_KEY = "order_popular_goods"

    def _init_popular_goods_snapshot_table(conn):
        conn.execute("""
            CREATE TABLE IF NOT EXISTS order_popular_goods_snapshot (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                items_json TEXT NOT NULL,
                need_ezadmin_session INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_by TEXT NOT NULL DEFAULT ''
            )
        """)
        conn.commit()

    def _save_popular_goods_snapshot(*, start_date: str, end_date: str, items: list[dict], need_ezadmin_session: bool, updated_by: str):
        conn = get_shared_db()
        try:
            _init_popular_goods_snapshot_table(conn)
            conn.execute("""
                INSERT INTO order_popular_goods_snapshot (id, start_date, end_date, items_json, need_ezadmin_session, updated_at, updated_by)
                VALUES (1, ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                ON CONFLICT(id) DO UPDATE SET
                    start_date = excluded.start_date,
                    end_date = excluded.end_date,
                    items_json = excluded.items_json,
                    need_ezadmin_session = excluded.need_ezadmin_session,
                    updated_at = CURRENT_TIMESTAMP,
                    updated_by = excluded.updated_by
            """, (start_date, end_date, json.dumps(items, ensure_ascii=False), 1 if need_ezadmin_session else 0, updated_by))
            conn.commit()
        finally:
            conn.close()

    @router.get("/order/popular-goods/saved")
    def order_popular_goods_saved(admin: str = Depends(require_admin)):
        """마지막으로 로컬에서 조회해 저장해둔 인기재고 스냅샷을 읽기만 한다 (외부 API 호출 없음 - 배포 환경에서도 동작)."""
        conn = get_shared_db()
        try:
            _init_popular_goods_snapshot_table(conn)
            row = conn.execute(
                "SELECT start_date, end_date, items_json, need_ezadmin_session, updated_at FROM order_popular_goods_snapshot WHERE id = 1"
            ).fetchone()
        finally:
            conn.close()

        if not row:
            return {"ok": True, "saved": False, "live_fetch_enabled": not is_render}

        try:
            items = json.loads(row["items_json"])
        except (TypeError, json.JSONDecodeError):
            items = []
        return {
            "ok": True,
            "saved": True,
            "start_date": row["start_date"],
            "end_date": row["end_date"],
            "count": len(items),
            "items": items,
            "need_ezadmin_session": bool(row["need_ezadmin_session"]),
            "updated_at": row["updated_at"],
            "live_fetch_enabled": not is_render,
        }

    @router.get("/order/popular-goods")
    async def order_popular_goods(
        start_date: str | None = None,
        end_date: str | None = None,
        limit: int = 50,
        admin: str = Depends(require_admin),
    ):
        if is_render:
            raise HTTPException(
                status_code=403,
                detail="배포 환경에서는 실시간 조회를 지원하지 않습니다. 저장된 데이터를 확인하세요.",
            )

        now = datetime.now()
        if not end_date:
            end_date = now.strftime("%Y-%m-%d")
        if not start_date:
            start_date = (now - timedelta(days=4)).strftime("%Y-%m-%d")
        if limit <= 0 or limit > 200:
            limit = 50

        try:
            days = (datetime.strptime(end_date, "%Y-%m-%d") - datetime.strptime(start_date, "%Y-%m-%d")).days + 1
        except ValueError:
            days = 1
        days = max(days, 1)

        try:
            goods = await AblyClient().get_goods_statistics(start_date=start_date, end_date=end_date)
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"에이블리 판매 통계 조회 실패 (HTTP {e.response.status_code})")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 판매 통계 조회 실패: {e}")

        # 옵션(재고) 단위로 한 개의 상품처럼 취급 - 같은 상품이라도 옵션마다 별도 랭킹.
        options: list[dict] = []
        for g in goods:
            for opt in g.get("goods_options") or []:
                options.append(
                    {
                        "goods_sno": g.get("goods_sno"),
                        "goods_name": g.get("goods_name"),
                        "goods_option_sno": opt.get("goods_option_sno"),
                        "goods_option_name": opt.get("goods_option_name"),
                        "order_count": opt.get("order_count") or 0,
                        "order_amount": opt.get("order_amount") or 0,
                        "cart_count": opt.get("cart_count") or 0,
                    }
                )

        # 랭킹은 기간 전체 판매수량 기준(기간이 모든 항목에 동일해 순서는 바뀌지 않음),
        # 화면 표시는 일평균 판매수량(판매수량 / 기간 일수)으로 보여준다.
        items = sorted(options, key=lambda o: o["order_count"], reverse=True)[:limit]
        for rank, item in enumerate(items, start=1):
            item["rank"] = rank
            item["avg_order_count"] = round(item["order_count"] / days, 1)

        # 같은 상품(goods_sno)의 옵션들을 묶어서, 그 상품의 최고 순위 옵션 바로 밑에 붙인다.
        # items가 이미 판매수량 내림차순이라 그룹 내부 순서도 그대로 유지된다.
        grouped: dict[int, list[dict]] = {}
        group_order: list[int] = []
        for item in items:
            sno = item["goods_sno"]
            if sno not in grouped:
                grouped[sno] = []
                group_order.append(sno)
            grouped[sno].append(item)
        items = [item for sno in group_order for item in grouped[sno]]

        # 옵션번호 → 상품코드(DB관리 원가베이스유) → 재고/접수(EZAdmin I100) 매칭.
        option_to_product_id = load_wonbe_option_sno_map()
        client_info_by_code = load_wonbe_client_info_by_code()
        for item in items:
            item["product_id"] = option_to_product_id.get(str(item.get("goods_option_sno") or ""))
            client_info = client_info_by_code.get(item["product_id"] or "", {})
            item["client"] = client_info.get("거래처", "")
            item["client_product_name"] = client_info.get("거래처상품명", "")

        # 상품코드 → 미송수량(노예김승일 미송관리). 없으면 0.
        misong_qty_by_code = load_misong_qty_by_code(get_shared_db)
        for item in items:
            item["misong_qty"] = misong_qty_by_code.get(item.get("product_id") or "", 0)

        unique_product_ids = sorted({item["product_id"] for item in items if item.get("product_id")})
        need_ezadmin_session = False
        stock_cache: dict[str, dict] = {}
        if unique_product_ids:
            ez = EzAdminClient(get_setting)
            sem = asyncio.Semaphore(6)

            async def _lookup(product_id: str) -> tuple[str, dict]:
                async with sem:
                    try:
                        result = await ez.get_stock_and_pending(product_id)
                        return product_id, {**result, "_session_expired": False}
                    except EzAdminSessionExpired:
                        return product_id, {"stock": None, "pending": None, "_session_expired": True}
                    except Exception:
                        return product_id, {"stock": None, "pending": None, "_session_expired": False}

            results = await asyncio.gather(*[_lookup(pid) for pid in unique_product_ids])
            for product_id, result in results:
                if result.pop("_session_expired", False):
                    need_ezadmin_session = True
                stock_cache[product_id] = result

        for item in items:
            product_id = item.get("product_id")
            result = stock_cache.get(product_id) if product_id else None
            item["stock"] = result["stock"] if result else None
            item["pending"] = result["pending"] if result else None

        try:
            _save_popular_goods_snapshot(
                start_date=start_date, end_date=end_date, items=items,
                need_ezadmin_session=need_ezadmin_session, updated_by=admin,
            )
        except Exception:
            pass  # 스냅샷 저장 실패는 이번 조회 결과 응답을 막지 않는다.

        return {
            "ok": True,
            "start_date": start_date,
            "end_date": end_date,
            "days": days,
            "count": len(items),
            "items": items,
            "need_ezadmin_session": need_ezadmin_session,
        }
    # ── 진머니 발주서 (EZAdmin I100 접수수량 조회) ─────────────────────────────
    @router.post("/order/jinmoney/export")
    async def jinmoney_order_export(admin: str = Depends(require_admin)):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}
        if not _JINMONEY_TEMPLATE_PATH.exists():
            raise HTTPException(status_code=500, detail="진머니 주문서 템플릿이 없습니다.")

        today = datetime.now().strftime("%Y-%m-%d")
        jinmoney_query = urllib.parse.quote("진머니")
        par = (
            "auto_search=&search_all_product=&multi_supply_group=&multi_supply=&str_supply_code=0"
            f"&tags_string=&product_tag_include_type=1&query_type=brand&query_str={jinmoney_query}"
            "&stock_type=0&stock_start=&stock_end=&notrans_day=&notrans_cnt=&notrans_status=0&stock_status=0"
            f"&start_date={today}&start_hour=00%3A00%3A00&end_date={today}&end_hour=23%3A59%3A59"
            "&date_period_sel=1&work_type=stockin&work_start=&work_end=&inout_type=0&product_date="
            f"&start_date2={today}&end_date2={today}&date_period_sel2=1&products_sort=1&category=0"
            "&except_soldout=0&temp_soldout=0&location=0"
        )
        headers = {
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": f"{_EZADMIN_BASE}/template40.htm?template=I100",
        }

        raw_items = []
        try:
            async with httpx.AsyncClient(timeout=120.0, verify=False, follow_redirects=True) as client:
                page = 1
                while True:
                    response = await client.post(
                        f"{_EZADMIN_BASE}/function.htm",
                        data={
                            "_search": "false",
                            "nd": str(int(datetime.now().timestamp() * 1000)),
                            "rows": "5000",
                            "page": str(page),
                            "sidx": "",
                            "sord": "asc",
                            "template": "I100",
                            "action": "search",
                            "page_code": "I100",
                            "par": par,
                        },
                        cookies={"PHPSESSID": phpsessid},
                        headers=headers,
                    )
                    if response.status_code >= 400:
                        raise HTTPException(status_code=502, detail=f"이지어드민 조회 실패 (HTTP {response.status_code})")
                    try:
                        data = response.json()
                    except Exception:
                        return {"ok": False, "need_session": True}
                    if not isinstance(data, dict) or "rows" not in data:
                        return {"ok": False, "need_session": True}

                    for row in data.get("rows") or []:
                        cell = row.get("cell", row) or {}
                        supply_product_name = _ez_val(
                            cell.get("brand") or cell.get("supply_product_name")
                        )
                        if not re.match(r"^\s*진머니(?:\s|$)", supply_product_name, re.IGNORECASE):
                            continue
                        # I100 응답 기준:
                        # before_trans = 화면의 "접수", stock_in_standby = "입고대기"
                        receipt_qty = _jinmoney_receipt_qty(cell.get("before_trans"))
                        if receipt_qty <= 0:
                            continue
                        raw_items.append({
                            "supplyProductName": supply_product_name,
                            "option": _ez_val(cell.get("options")),
                            "receiptQty": receipt_qty,
                        })

                    total_pages = int(data.get("total") or 1)
                    if page >= total_pages or page >= 50:
                        break
                    page += 1
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"이지어드민 연결 실패: {type(exc).__name__}")

        if not raw_items:
            raise HTTPException(status_code=404, detail="접수 수량이 있는 진머니 상품이 없습니다.")

        aggregated = {}
        unmatched = []
        for item in raw_items:
            mapped, reason = _jinmoney_map_item(item["supplyProductName"], item["option"])
            if not mapped:
                unmatched.append({**item, "reason": reason})
                continue
            key = (mapped["productCode"], mapped["size"])
            aggregated[key] = aggregated.get(key, 0) + item["receiptQty"]

        if unmatched:
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "매칭되지 않은 진머니 상품이 있어 발주서 생성을 중단했습니다.",
                    "unmatched": unmatched,
                },
            )

        output_rows = [
            {"productCode": code, "size": size, "qty": qty}
            for (code, size), qty in sorted(aggregated.items())
        ]
        if len(output_rows) > 28:
            raise HTTPException(status_code=400, detail="주문서 입력 가능 행(28개)을 초과했습니다.")

        workbook = load_workbook(_JINMONEY_TEMPLATE_PATH)
        sheet = workbook["진머니 주문서양식"]
        for row_number in range(6, 34):
            sheet.cell(row_number, 1).value = row_number - 5
            for column in range(2, 5):
                sheet.cell(row_number, column).value = None
        for index, item in enumerate(output_rows, start=6):
            sheet.cell(index, 2).value = item["productCode"]
            sheet.cell(index, 3).value = item["size"]
            sheet.cell(index, 4).value = item["qty"]
        sheet["D34"] = "=SUM(D6:D33)"

        buffer = io.BytesIO()
        workbook.save(buffer)
        now = datetime.now()
        filename = f"{now.month}월 {now.day}일 진머니 발주서.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": _content_disposition(filename),
                "X-Jinmoney-Items": str(len(output_rows)),
            },
        )

    # ── 메인발주 목록 (EZAdmin IO30 미출고/부족 상품 조회) ───────────────────────
    @router.post("/order/main-order/list")
    async def main_order_list(admin: str = Depends(require_admin)):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        today = datetime.now()
        start = (today - timedelta(days=90)).strftime("%Y-%m-%d")
        end = today.strftime("%Y-%m-%d")

        par = (
            "template=IO30&action=&page_code=IO00&search=1&now_page=&is_sort=&"
            "_sort=supply_options&sort_order=1&product_qty_list=&bill_seq=&"
            "offset_top=&work_no=&location_str=&date_type=collect_date&"
            f"start_date={start}&start_hour=00%3A00%3A00&"
            f"end_date={end}&end_hour=23%3A59%3A59&"
            "date_period_sel=9&multi_shop_group=&multi_shop=&str_shop_code=0&"
            "multi_supply_group=&multi_supply=&str_supply_code=0&"
            "supply_name_search=&brand=&supply_options=&tags_string=&"
            "product_tag_include_type=1&product_id=&name=&options=&"
            "search_keyword_type=origin&search_keyword=&enable_stock_type=2&"
            "order_status=3&except_soldout=1&sel_reserve_qty=none&"
            "sel_return_qty=none&sel_lack_qty=none&sel_req_qty=none&category=0"
        )

        cookies = {"PHPSESSID": phpsessid}
        ez_headers = {
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://ga80.ezadmin.co.kr/template40.htm?template=IO30",
        }
        base_url = f"{_EZADMIN_BASE}/function.htm"

        items = []
        try:
            async with httpx.AsyncClient(timeout=60.0, verify=False, follow_redirects=True) as client:
                page = 1
                while True:
                    nd = str(int(datetime.now().timestamp() * 1000))
                    resp = await client.post(
                        base_url,
                        data={
                            "_search": "false", "nd": nd, "rows": "1000", "page": str(page),
                            "sidx": "", "sord": "asc", "template": "IO30", "action": "search_IO30",
                            "par": par,
                        },
                        cookies=cookies, headers=ez_headers,
                    )
                    if resp.status_code >= 400:
                        return {"ok": False, "error": f"EZAdmin 조회 실패 (HTTP {resp.status_code})"}
                    try:
                        data = resp.json()
                    except Exception:
                        return {"ok": False, "need_session": True}
                    if "rows" not in data:
                        return {"ok": False, "need_session": True}

                    for row in data.get("rows") or []:
                        cell = row.get("cell", row)
                        items.append({
                            "code": _ez_val(cell.get("product_id")),
                            "name": _ez_val(cell.get("product_name")),
                            "options": _ez_val(cell.get("options")),
                            "supplyProductName": _ez_val(cell.get("supply_product_name")),
                            "stock": _to_int(_ez_val(cell.get("stock"))),
                            "notYetDeliv": _to_int(_ez_val(cell.get("not_yet_deliv"))),
                            "lackQty": _to_int(_ez_val(cell.get("lack_qty"))),
                            "requestQty": _to_int(_ez_val(cell.get("request_qty"))),
                        })

                    total_pages = int(data.get("total") or 1)
                    if page >= total_pages or page >= 20:
                        break
                    page += 1
        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        return {"ok": True, "count": len(items), "items": items}

    # ── 메인발주 실행 (top90 발주 등록) ──────────────────────────────────────────
    @router.post("/order/main-order/execute")
    async def main_order_execute(payload: dict = Body(...), admin: str = Depends(require_admin)):
        items = payload.get("items") or []
        groups: dict[str, list[str]] = {}

        for item in items:
            qty = _to_int(item.get("requestQty"), 0)
            if qty <= 0:
                continue
            supply_product_name = str(item.get("supplyProductName") or "").strip()
            if not supply_product_name:
                continue
            parts = supply_product_name.split(" ", 1)
            store_name = parts[0]
            rest = parts[1].strip() if len(parts) > 1 else ""
            options = str(item.get("options") or "").strip()
            pname_parts = [p for p in [rest, options, str(qty)] if p]
            pname = " ".join(pname_parts)
            groups.setdefault(store_name, []).append(pname)

        if not groups:
            return {"ok": False, "error": "발주할 항목이 없습니다 (요청수량이 0보다 큰 항목이 없음)."}

        try:
            result = await execute_main_orders(groups)
        except Top90Error as exc:
            return {"ok": False, "error": str(exc)}
        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        result_by_store = {}
        for s in result.get("success") or []:
            result_by_store[s["store_name"]] = {"result_status": "success"}
        for f in result.get("failed") or []:
            result_by_store[f["store_name"]] = {"result_status": "failed", "reason": f.get("reason", "")}
        record_order_history(
            get_db,
            execution_id=str(uuid.uuid4()),
            recorded_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            action_type="order_execute",
            items=items,
            username=admin,
            display_name=get_user_display(admin),
            result_by_store=result_by_store,
        )

        return {"ok": True, **result}

    # ── 발주내역 (TSV 복사 / 발주 실행 기록) ──────────────────────────────────
    @router.post("/order/main-order/record-tsv-copy")
    async def main_order_record_tsv_copy(payload: dict = Body(...), admin: str = Depends(require_admin)):
        items = payload.get("items") or []
        record_order_history(
            get_db,
            execution_id=str(uuid.uuid4()),
            recorded_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            action_type="tsv_copy",
            items=items,
            username=admin,
            display_name=get_user_display(admin),
        )
        return {"ok": True}

    @router.get("/order/main-order/history")
    def main_order_history(
        offset: int = 0,
        limit: int = 50,
        date_from: str = "",
        date_to: str = "",
        action_type: str = "",
        q: str = "",
        admin: str = Depends(require_admin),
    ):
        offset = max(0, offset)
        limit = min(max(1, limit), 200)
        clauses = []
        params = []
        if date_from:
            clauses.append("recorded_at >= ?")
            params.append(f"{date_from} 00:00:00")
        if date_to:
            clauses.append("recorded_at <= ?")
            params.append(f"{date_to} 23:59:59")
        if action_type:
            clauses.append("action_type = ?")
            params.append(action_type)
        if q.strip():
            like = f"%{q.strip()}%"
            clauses.append(
                "(product_code LIKE ? OR product_name LIKE ? OR store_name LIKE ? "
                "OR recorded_by_username LIKE ? OR recorded_by_display_name LIKE ?)"
            )
            params.extend([like, like, like, like, like])
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        init_order_history_table(get_db)
        conn = get_db()
        try:
            total = conn.execute(f"SELECT COUNT(*) FROM order_history{where}", params).fetchone()[0]
            rows = conn.execute(
                f"SELECT * FROM order_history{where} ORDER BY recorded_at DESC, id DESC LIMIT ? OFFSET ?",
                [*params, limit, offset],
            ).fetchall()
            return {"ok": True, "total": total, "rows": [dict(row) for row in rows]}
        finally:
            conn.close()

    @router.delete("/order/main-order/history")
    def main_order_history_delete(payload: dict = Body(...), admin: str = Depends(require_admin)):
        ids = [int(i) for i in (payload.get("ids") or []) if str(i).strip().lstrip("-").isdigit()]
        if not ids:
            raise HTTPException(status_code=400, detail="삭제할 항목 필요")
        init_order_history_table(get_db)
        conn = get_db()
        try:
            placeholders = ", ".join("?" for _ in ids)
            cur = conn.execute(f"DELETE FROM order_history WHERE id IN ({placeholders})", ids)
            conn.commit()
            return {"ok": True, "deleted": cur.rowcount}
        finally:
            conn.close()

    return router
