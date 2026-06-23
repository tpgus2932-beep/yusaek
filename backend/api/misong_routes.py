from urllib.parse import quote
import io
import warnings
from pathlib import Path

import httpx
import openpyxl
import xlwt
from fastapi import APIRouter, Depends, HTTPException, Body
from fastapi.responses import Response
from datetime import datetime, timezone
from api.amood_hapbae import SHARED_COST_BASE_PATH

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

WAITING_BASE_PATH = SHARED_COST_BASE_PATH
_EZADMIN_BASE = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_INGODAEGI_PATH = Path(r"C:\Users\ksh29\OneDrive\Desktop\원베\입고대기.xlsx")


def build_misong_router(*, get_current_user, get_db, get_setting):
    router = APIRouter(prefix="/noye-kimsungil/misong")

    # ── DB 초기화 ──────────────────────────────────────────────────────────────
    def _init(conn):
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS misong_items (
                id          TEXT PRIMARY KEY,
                A           TEXT NOT NULL DEFAULT '',
                B           TEXT NOT NULL DEFAULT '',
                C           TEXT NOT NULL DEFAULT '',
                D           TEXT NOT NULL DEFAULT '',
                E           TEXT NOT NULL DEFAULT '',
                F           INTEGER NOT NULL DEFAULT 0,
                G           TEXT NOT NULL DEFAULT '',
                original_f  TEXT NOT NULL DEFAULT '',
                added_at    TEXT NOT NULL,
                owner       TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_misong_items_original_f ON misong_items(original_f)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS misong_logs (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                item_id       TEXT NOT NULL,
                type          TEXT NOT NULL,
                qty           INTEGER NOT NULL,
                remaining_qty INTEGER NOT NULL,
                work_date     TEXT NOT NULL DEFAULT '',
                memo          TEXT NOT NULL DEFAULT '',
                supplier_name TEXT NOT NULL DEFAULT '',
                product_name  TEXT NOT NULL DEFAULT '',
                product_code  TEXT NOT NULL DEFAULT '',
                color         TEXT NOT NULL DEFAULT '',
                size          TEXT NOT NULL DEFAULT '',
                ts            TEXT NOT NULL
            )
            """
        )
        log_cols = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(misong_logs)").fetchall()
        }
        if "product_name" not in log_cols:
            conn.execute("ALTER TABLE misong_logs ADD COLUMN product_name TEXT NOT NULL DEFAULT ''")
        if "product_code" not in log_cols:
            conn.execute("ALTER TABLE misong_logs ADD COLUMN product_code TEXT NOT NULL DEFAULT ''")
        if "supplier_name" not in log_cols:
            conn.execute("ALTER TABLE misong_logs ADD COLUMN supplier_name TEXT NOT NULL DEFAULT ''")
        if "color" not in log_cols:
            conn.execute("ALTER TABLE misong_logs ADD COLUMN color TEXT NOT NULL DEFAULT ''")
        if "size" not in log_cols:
            conn.execute("ALTER TABLE misong_logs ADD COLUMN size TEXT NOT NULL DEFAULT ''")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_misong_logs_item_id ON misong_logs(item_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_misong_logs_product_name ON misong_logs(product_name)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_misong_logs_supplier_name ON misong_logs(supplier_name)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_misong_logs_product_code ON misong_logs(product_code)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_misong_logs_ts ON misong_logs(ts)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS misong_alerts (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                type         TEXT NOT NULL,
                product_code TEXT NOT NULL DEFAULT '',
                detail       TEXT NOT NULL DEFAULT '',
                row_info     TEXT NOT NULL DEFAULT '',
                h_value      TEXT NOT NULL DEFAULT '',
                qty          INTEGER NOT NULL DEFAULT 0,
                work_date    TEXT NOT NULL DEFAULT '',
                supplier_name TEXT NOT NULL DEFAULT '',
                product_name TEXT NOT NULL DEFAULT '',
                color        TEXT NOT NULL DEFAULT '',
                size         TEXT NOT NULL DEFAULT '',
                ts           TEXT NOT NULL
            )
            """
        )
        alert_cols = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(misong_alerts)").fetchall()
        }
        alert_column_defs = {
            "h_value": "TEXT NOT NULL DEFAULT ''",
            "qty": "INTEGER NOT NULL DEFAULT 0",
            "work_date": "TEXT NOT NULL DEFAULT ''",
            "supplier_name": "TEXT NOT NULL DEFAULT ''",
            "product_name": "TEXT NOT NULL DEFAULT ''",
            "color": "TEXT NOT NULL DEFAULT ''",
            "size": "TEXT NOT NULL DEFAULT ''",
        }
        for col, col_def in alert_column_defs.items():
            if col not in alert_cols:
                conn.execute(f"ALTER TABLE misong_alerts ADD COLUMN {col} {col_def}")
        conn.commit()

    # ── 헬퍼 ──────────────────────────────────────────────────────────────────
    def _now():
        return datetime.now(timezone.utc).isoformat()

    def _today_date():
        return datetime.now().strftime("%Y-%m-%d")

    def _content_disposition(filename: str) -> str:
        ascii_name = filename.encode("ascii", "ignore").decode() or "download.xlsx"
        return f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(filename)}'

    def _normalize_code(value) -> str:
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return " ".join(str(value or "").split()).strip()

    def _normalize_match_text(value) -> str:
        return " ".join(str(value or "").split()).strip().casefold()

    def _parse_tsv_first_column(text: str) -> list[str]:
        values = []
        for line in str(text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            first_cell = line.split("\t", 1)[0].strip()
            if first_cell:
                values.append(first_cell)
        return values

    def _lookup_waiting_base_product(product_code: str) -> dict | None:
        code = _normalize_code(product_code)
        if not code:
            return None
        if not WAITING_BASE_PATH.exists():
            raise HTTPException(404, "원가베이스유 파일을 찾을 수 없습니다.")

        wb = openpyxl.load_workbook(WAITING_BASE_PATH, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_code = _normalize_code(row[0] if len(row) > 0 else "")
                if row_code == code:
                    return {
                        "A": str(row[5] if len(row) > 5 and row[5] is not None else "").strip(),
                        "B": str(row[6] if len(row) > 6 and row[6] is not None else "").strip(),
                        "D": str(row[2] if len(row) > 2 and row[2] is not None else "").strip(),
                        "E": str(row[3] if len(row) > 3 and row[3] is not None else "").strip(),
                        "originalF": row_code,
                    }
        finally:
            wb.close()
        return None

    def _match_waiting_base_product_code(product_name: str, color: str, size: str) -> str | None:
        name_key = _normalize_match_text(product_name)
        color_key = _normalize_match_text(color)
        size_key = _normalize_match_text(size)
        if not name_key:
            return None
        if not WAITING_BASE_PATH.exists():
            raise HTTPException(404, "원가베이스유 파일을 찾을 수 없습니다.")

        wb = openpyxl.load_workbook(WAITING_BASE_PATH, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            for row in ws.iter_rows(min_row=2, values_only=True):
                base_name = _normalize_match_text(row[6] if len(row) > 6 else "")
                base_color = _normalize_match_text(row[2] if len(row) > 2 else "")
                base_size = _normalize_match_text(row[3] if len(row) > 3 else "")
                if base_name == name_key and base_color == color_key and base_size == size_key:
                    code = _normalize_code(row[0] if len(row) > 0 else "")
                    if code:
                        return code
        finally:
            wb.close()
        return None

    def _get_waiting_base_sheet1(wb):
        if "Sheet2" in wb.sheetnames:
            return wb["Sheet2"]
        if len(wb.worksheets) >= 2:
            return wb.worksheets[1]
        raise HTTPException(400, "원가베이스유 Sheet2를 찾을 수 없습니다.")

    def _item_row_to_dict(row):
        keys = row.keys()
        latest_add_date = (
            row["latest_add_date"]
            if "latest_add_date" in keys and row["latest_add_date"]
            else row["G"]
        )
        today_added_qty = (
            row["today_added_qty"]
            if "today_added_qty" in keys and row["today_added_qty"] is not None
            else 0
        )
        return {
            "id": row["id"],
            "A": row["A"], "B": row["B"], "C": row["C"],
            "D": row["D"], "E": row["E"],
            "F": row["F"],
            "G": latest_add_date,
            "originalF": row["original_f"],
            "addedAt": row["added_at"],
            "latestAddDate": latest_add_date,
            "todayAddedQty": today_added_qty,
        }

    def _fetch_item_rows(conn, item_id: str | None = None, today: str | None = None):
        today = today or _today_date()
        where = ""
        params = [today]
        if item_id:
            where = "WHERE i.id = ?"
            params.append(item_id)
        return conn.execute(
            f"""
            SELECT
                i.*,
                latest_add.latest_add_date,
                COALESCE(today_add.today_added_qty, 0) AS today_added_qty
            FROM misong_items i
            LEFT JOIN (
                SELECT item_id, MAX(work_date) AS latest_add_date
                FROM misong_logs
                WHERE type = 'add' AND work_date <> ''
                GROUP BY item_id
            ) latest_add ON latest_add.item_id = i.id
            LEFT JOIN (
                SELECT item_id, SUM(qty) AS today_added_qty
                FROM misong_logs
                WHERE type = 'add' AND work_date = ?
                GROUP BY item_id
            ) today_add ON today_add.item_id = i.id
            {where}
            ORDER BY i.added_at ASC
            """,
            params,
        ).fetchall()

    def _fetch_item_row(conn, item_id: str, today: str | None = None):
        rows = _fetch_item_rows(conn, item_id, today)
        return rows[0] if rows else None

    def _log_row_to_dict(row):
        return {
            "id": row["id"],
            "item_id": row["item_id"],
            "type": row["type"],
            "qty": row["qty"],
            "remaining_qty": row["remaining_qty"],
            "work_date": row["work_date"],
            "memo": row["memo"],
            "supplier_name": row["supplier_name"] if "supplier_name" in row.keys() else "",
            "product_name": row["product_name"] if "product_name" in row.keys() else "",
            "product_code": row["product_code"] if "product_code" in row.keys() else "",
            "color": row["color"] if "color" in row.keys() else "",
            "size": row["size"] if "size" in row.keys() else "",
            "ts": row["ts"],
        }

    def _alert_row_to_dict(row):
        return {
            "id": row["id"],
            "type": row["type"],
            "productCode": row["product_code"],
            "detail": row["detail"],
            "rowInfo": row["row_info"],
            "hValue": row["h_value"] if "h_value" in row.keys() else "",
            "qty": row["qty"] if "qty" in row.keys() else 0,
            "workDate": row["work_date"] if "work_date" in row.keys() else "",
            "supplierName": row["supplier_name"] if "supplier_name" in row.keys() else "",
            "productName": row["product_name"] if "product_name" in row.keys() else "",
            "color": row["color"] if "color" in row.keys() else "",
            "size": row["size"] if "size" in row.keys() else "",
            "timestamp": row["ts"],
        }

    # ── 아이템 전체 조회 ───────────────────────────────────────────────────────
    @router.get("/items")
    def list_items(today: str = "", user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            rows = _fetch_item_rows(conn, today=today or None)
            return {"ok": True, "items": [_item_row_to_dict(r) for r in rows]}
        finally:
            conn.close()

    # ── 단일 아이템 추가 (수동) ───────────────────────────────────────────────
    @router.post("/items")
    def add_item(payload: dict = Body(...), user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            import uuid as _uuid
            item_id = str(_uuid.uuid4())
            now = _now()
            product_code = _normalize_code(payload.get("originalF", ""))
            if product_code:
                matched = _lookup_waiting_base_product(product_code)
                if matched is None and not all(payload.get(k) for k in ("A", "B", "D", "E")):
                    raise HTTPException(404, f"원가베이스유 Sheet1에서 상품코드 {product_code}를 찾을 수 없습니다.")
                if matched:
                    for key in ("A", "B", "D", "E", "originalF"):
                        if not str(payload.get(key, "") or "").strip():
                            payload[key] = matched[key]
            if not str(payload.get("G", "") or "").strip():
                payload["G"] = _today_date()
            qty = int(payload.get("F", 0))

            if product_code:
                existing = conn.execute(
                    "SELECT * FROM misong_items WHERE original_f = ?",
                    (product_code,),
                ).fetchone()
                if existing:
                    new_qty = int(existing["F"] or 0) + qty
                    work_date = payload.get("G", "") or _today_date()
                    conn.execute(
                        "UPDATE misong_items SET F = ?, G = ? WHERE id = ?",
                        (new_qty, work_date, existing["id"]),
                    )
                    conn.execute(
                        """
                        INSERT INTO misong_logs
                          (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                        VALUES (?, 'add', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            existing["id"],
                            qty,
                            new_qty,
                            work_date,
                            payload.get("memo", "수동 추가"),
                            existing["A"],
                            existing["B"],
                            existing["original_f"],
                            existing["D"],
                            existing["E"],
                            now,
                        ),
                    )
                    conn.commit()
                    row = _fetch_item_row(conn, existing["id"])
                    return {"ok": True, "item": _item_row_to_dict(row), "merged": True}

            conn.execute(
                """
                INSERT INTO misong_items (id, A, B, C, D, E, F, G, original_f, added_at, owner)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item_id,
                    payload.get("A", ""), payload.get("B", ""), payload.get("C", ""),
                    payload.get("D", ""), payload.get("E", ""),
                    qty,
                    payload.get("G", ""),
                    payload.get("originalF", ""),
                    now, user,
                ),
            )
            conn.execute(
                """
                INSERT INTO misong_logs
                  (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                VALUES (?, 'add', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item_id,
                    qty,
                    qty,
                    payload.get("G", ""),
                    payload.get("memo", "수동 추가"),
                    payload.get("A", ""),
                    payload.get("B", ""),
                    payload.get("originalF", ""),
                    payload.get("D", ""),
                    payload.get("E", ""),
                    now,
                ),
            )
            conn.commit()
            row = _fetch_item_row(conn, item_id)
            return {"ok": True, "item": _item_row_to_dict(row)}
        finally:
            conn.close()

    # ── 단일 아이템 수정 ──────────────────────────────────────────────────────
    @router.put("/items/{item_id}")
    def update_item(item_id: str, payload: dict = Body(...), user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            existing = conn.execute("SELECT * FROM misong_items WHERE id = ?", (item_id,)).fetchone()
            if not existing:
                raise HTTPException(404, "항목을 찾을 수 없습니다.")
            conn.execute(
                """
                UPDATE misong_items
                SET A=?, B=?, C=?, D=?, E=?, F=?, G=?, original_f=?
                WHERE id=?
                """,
                (
                    payload.get("A", existing["A"]),
                    payload.get("B", existing["B"]),
                    payload.get("C", existing["C"]),
                    payload.get("D", existing["D"]),
                    payload.get("E", existing["E"]),
                    int(payload.get("F", existing["F"])),
                    payload.get("G", existing["G"]),
                    payload.get("originalF", existing["original_f"]),
                    item_id,
                ),
            )
            conn.commit()
            row = _fetch_item_row(conn, item_id)
            return {"ok": True, "item": _item_row_to_dict(row)}
        finally:
            conn.close()

    # ── 단일 아이템 삭제 ──────────────────────────────────────────────────────
    @router.delete("/items/{item_id}")
    def delete_item(item_id: str, user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            existing = conn.execute("SELECT id FROM misong_items WHERE id = ?", (item_id,)).fetchone()
            if not existing:
                raise HTTPException(404, "항목을 찾을 수 없습니다.")
            conn.execute("DELETE FROM misong_logs WHERE item_id = ?", (item_id,))
            conn.execute("DELETE FROM misong_items WHERE id = ?", (item_id,))
            conn.commit()
            return {"ok": True}
        finally:
            conn.close()

    # ── 전체 아이템 삭제 ──────────────────────────────────────────────────────
    @router.delete("/items")
    def delete_all_items(user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            conn.execute("DELETE FROM misong_logs")
            conn.execute("DELETE FROM misong_items")
            conn.commit()
            return {"ok": True}
        finally:
            conn.close()

    @router.get("/waiting-base/search")
    def search_waiting_base(q: str = "", user: str = Depends(get_current_user)):
        q = q.strip()
        if not q:
            return {"results": []}
        if not WAITING_BASE_PATH.exists():
            raise HTTPException(404, "원가베이스유 파일을 찾을 수 없습니다.")
        query = q.lower()
        wb = openpyxl.load_workbook(WAITING_BASE_PATH, data_only=True, read_only=True)
        try:
            ws = wb.worksheets[0]
            results = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                g_val = str(row[6] if len(row) > 6 and row[6] is not None else "").strip()
                if query in g_val.lower():
                    results.append({
                        "A": str(row[5] if len(row) > 5 and row[5] is not None else "").strip(),
                        "B": g_val,
                        "D": str(row[2] if len(row) > 2 and row[2] is not None else "").strip(),
                        "E": str(row[3] if len(row) > 3 and row[3] is not None else "").strip(),
                        "originalF": str(row[0] if row[0] is not None else "").strip(),
                    })
                    if len(results) >= 30:
                        break
        finally:
            wb.close()
        return {"results": results}

    @router.get("/waiting-base/download")
    def download_waiting_base(user: str = Depends(get_current_user)):
        if not WAITING_BASE_PATH.exists():
            raise HTTPException(
                status_code=404,
                detail=f"원가베이스유 파일을 찾을 수 없습니다: {WAITING_BASE_PATH}",
            )

        conn = get_db()
        try:
            _init(conn)
            rows = conn.execute(
                """
                SELECT original_f, SUM(F) AS qty
                FROM misong_items
                WHERE TRIM(original_f) != ''
                GROUP BY original_f
                """
            ).fetchall()
            qty_by_code = {
                _normalize_code(row["original_f"]): int(row["qty"] or 0)
                for row in rows
                if _normalize_code(row["original_f"])
            }
        finally:
            conn.close()

        try:
            wb = openpyxl.load_workbook(WAITING_BASE_PATH)
            ws = _get_waiting_base_sheet1(wb)
        except Exception as exc:
            if isinstance(exc, HTTPException):
                raise
            raise HTTPException(
                status_code=500,
                detail=f"원가베이스유 Sheet2를 읽지 못했습니다: {exc}",
            ) from exc

        # Sheet2에 존재하는 코드 수집
        matched_codes = set()
        buf = io.BytesIO()
        out_wb = xlwt.Workbook()
        out_ws = out_wb.add_sheet(ws.title[:31] or "Sheet1")
        for row_idx in range(1, ws.max_row + 1):
            code = _normalize_code(ws.cell(row=row_idx, column=1).value)
            if code and row_idx > 1:
                matched_codes.add(code)
            for col_idx in range(1, ws.max_column + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if col_idx == 2:
                    value = "작업수량" if row_idx == 1 and code not in qty_by_code else qty_by_code.get(code, "ZERO")
                out_ws.write(row_idx - 1, col_idx - 1, "" if value is None else value)
        out_wb.save(buf)
        buf.seek(0)

        # 미매칭: misong에는 있지만 Sheet2에 코드가 없는 항목
        unmatched = [code for code in qty_by_code if code not in matched_codes]
        unmatched_count = len(unmatched)
        # 헤더에 미매칭 코드 목록 전달 (쉼표 구분, 최대 500자)
        unmatched_str = ",".join(unmatched)[:500]

        filename = "입고대기_미송수량.xls"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={
                "Content-Disposition": _content_disposition(filename),
                "X-Unmatched-Count": str(unmatched_count),
                "X-Unmatched-Codes": unmatched_str,
                "Access-Control-Expose-Headers": "X-Unmatched-Count, X-Unmatched-Codes",
            },
        )

    @router.post("/waiting-base/append")
    def append_waiting_base(payload: dict = Body(...), user: str = Depends(get_current_user)):
        if not WAITING_BASE_PATH.exists():
            raise HTTPException(
                status_code=404,
                detail=f"원가베이스유 파일을 찾을 수 없습니다: {WAITING_BASE_PATH}",
            )

        values = _parse_tsv_first_column(payload.get("text", ""))
        if not values:
            raise HTTPException(status_code=400, detail="추가할 A열 데이터가 없습니다.")

        try:
            wb = openpyxl.load_workbook(WAITING_BASE_PATH)
            ws = _get_waiting_base_sheet1(wb)
        except Exception as exc:
            if isinstance(exc, HTTPException):
                raise
            raise HTTPException(
                status_code=500,
                detail=f"원가베이스유 Sheet2를 읽지 못했습니다: {exc}",
            ) from exc

        start_row = ws.max_row + 1
        for offset, value in enumerate(values):
            row_idx = start_row + offset
            ws.cell(row=row_idx, column=1).value = value
            ws.cell(row=row_idx, column=2).value = "ZERO"

        try:
            wb.save(WAITING_BASE_PATH)
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail=f"원가베이스유 Sheet2 저장 실패: {exc}",
            ) from exc

        return {"ok": True, "appended": len(values), "start_row": start_row}

    # ── 엑셀 변환 결과 일괄 처리 (미송/미송픽업) ─────────────────────────────
    @router.post("/items/process-rows")
    def process_rows(payload: dict = Body(...), user: str = Depends(get_current_user)):
        """
        H열 기준으로 미송 항목을 일괄 처리.
        payload: { rows: [ { A,B,C,D,E,F,G,H,I,originalF } ] }
        """
        import uuid as _uuid

        rows = payload.get("rows", [])
        today = (payload.get("today") or "").strip() or None
        conn = get_db()
        try:
            _init(conn)
            new_alerts = []
            now = _now()

            for row in rows:
                h_val = (row.get("H") or "").strip()
                original_f = (row.get("originalF") or row.get("I") or "").strip()
                qty = int(row.get("F") or 0)
                work_date = row.get("G", "")
                row_info = f"{row.get('A','')} {row.get('B','')} {row.get('D','')}/{row.get('E','')}"

                if h_val in ("미송", "미송픽업") and not original_f:
                    new_alerts.append({
                        "type": "missing_code",
                        "product_code": "",
                        "detail": "I열 상품코드 없음 - 미송관리 반영 안 됨",
                        "row_info": row_info,
                        "ts": now,
                    })
                    conn.execute(
                        """
                        INSERT INTO misong_alerts
                          (type, product_code, detail, row_info, h_value, qty, work_date, supplier_name, product_name, color, size, ts)
                        VALUES ('missing_code', '', 'I열 상품코드 없음 - 미송관리 반영 안 됨', ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            row_info,
                            h_val,
                            qty,
                            work_date,
                            row.get("A", ""),
                            row.get("B", ""),
                            row.get("D", ""),
                            row.get("E", ""),
                            now,
                        ),
                    )
                    continue

                if h_val == "미송":
                    existing = conn.execute(
                        "SELECT * FROM misong_items WHERE original_f = ?", (original_f,)
                    ).fetchone()

                    if existing:
                        new_f = existing["F"] + qty
                        conn.execute(
                            """
                            INSERT INTO misong_logs
                              (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                            VALUES (?, 'add', ?, ?, ?, '', ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                existing["id"], qty, new_f, work_date,
                                existing["A"], existing["B"], existing["original_f"],
                                existing["D"], existing["E"], now,
                            ),
                        )
                        if new_f == 0:
                            conn.execute("DELETE FROM misong_items WHERE id = ?", (existing["id"],))
                        else:
                            conn.execute(
                                "UPDATE misong_items SET F = ? WHERE id = ?",
                                (new_f, existing["id"]),
                            )
                    else:
                        new_alerts.append({
                            "type": "unmatched_add",
                            "product_code": original_f,
                            "detail": "기존 미송 상품코드와 매칭 안 됨 - 신규 항목으로 추가",
                            "row_info": row_info,
                            "ts": now,
                        })
                        conn.execute(
                            """
                            INSERT INTO misong_alerts
                              (type, product_code, detail, row_info, h_value, qty, work_date, supplier_name, product_name, color, size, ts)
                            VALUES ('unmatched_add', ?, '기존 미송 상품코드와 매칭 안 됨 - 신규 항목으로 추가', ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                original_f,
                                row_info,
                                h_val,
                                qty,
                                work_date,
                                row.get("A", ""),
                                row.get("B", ""),
                                row.get("D", ""),
                                row.get("E", ""),
                                now,
                            ),
                        )
                        item_id = str(_uuid.uuid4())
                        conn.execute(
                            """
                            INSERT INTO misong_items
                              (id, A, B, C, D, E, F, G, original_f, added_at, owner)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                item_id,
                                row.get("A", ""), row.get("B", ""), row.get("C", ""),
                                row.get("D", ""), row.get("E", ""),
                                qty,
                                row.get("G", ""),
                                original_f, now, user,
                            ),
                        )
                        conn.execute(
                            """
                            INSERT INTO misong_logs
                              (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                            VALUES (?, 'add', ?, ?, ?, '', ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                item_id, qty, qty, work_date,
                                row.get("A", ""), row.get("B", ""), original_f,
                                row.get("D", ""), row.get("E", ""), now,
                            ),
                        )

                elif h_val == "미송픽업":
                    existing = conn.execute(
                        "SELECT * FROM misong_items WHERE original_f = ?", (original_f,)
                    ).fetchone()

                    if existing is None:
                        new_alerts.append({
                            "type": "not_found",
                            "product_code": original_f,
                            "detail": "항목 없음",
                            "row_info": row_info,
                            "ts": now,
                        })
                        conn.execute(
                            """
                            INSERT INTO misong_alerts (type, product_code, detail, row_info, ts)
                            VALUES ('not_found', ?, '항목 없음', ?, ?)
                            """,
                            (
                                original_f,
                                row_info,
                                now,
                            ),
                        )
                    else:
                        new_f = existing["F"] - qty

                        if new_f < 0:
                            alert_detail = f"수량 부족 (보유: {existing['F']}, 차감: {qty}) → 음수 처리됨"
                            new_alerts.append({
                                "type": "negative",
                                "product_code": original_f,
                                "detail": alert_detail,
                                "row_info": row_info,
                                "ts": now,
                            })
                            conn.execute(
                                "INSERT INTO misong_alerts (type, product_code, detail, row_info, ts) VALUES ('negative', ?, ?, ?, ?)",
                                (original_f, alert_detail, row_info, now),
                            )

                        # 로그는 항상 기록
                        conn.execute(
                            """
                            INSERT INTO misong_logs
                              (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                            VALUES (?, 'subtract', ?, ?, ?, '', ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                existing["id"], qty, new_f, work_date,
                                existing["A"], existing["B"], existing["original_f"],
                                existing["D"], existing["E"], now,
                            ),
                        )

                        if new_f == 0:
                            # 0이 되면 항목 삭제
                            conn.execute("DELETE FROM misong_items WHERE id = ?", (existing["id"],))
                        else:
                            # 양수든 음수든 수량 업데이트 후 유지
                            conn.execute(
                                "UPDATE misong_items SET F = ? WHERE id = ?",
                                (new_f, existing["id"]),
                            )

            conn.commit()

            items = _fetch_item_rows(conn, today=today)
            alerts = conn.execute(
                "SELECT * FROM misong_alerts ORDER BY ts DESC LIMIT 200"
            ).fetchall()

            return {
                "ok": True,
                "items": [_item_row_to_dict(r) for r in items],
                "alerts": [_alert_row_to_dict(r) for r in alerts],
                "new_alert_count": len(new_alerts),
            }
        finally:
            conn.close()

    # ── 로그 조회 (아이템별 + 날짜 범위 필터) ────────────────────────────────
    @router.get("/logs/{item_id}")
    def get_logs(
        item_id: str,
        date_from: str = "",
        date_to: str = "",
        user: str = Depends(get_current_user),
    ):
        conn = get_db()
        try:
            _init(conn)
            params: list = [item_id]
            where = "item_id = ?"
            if date_from:
                where += " AND work_date >= ?"
                params.append(date_from)
            if date_to:
                where += " AND work_date <= ?"
                params.append(date_to)
            rows = conn.execute(
                f"""
                SELECT
                    l.id, l.item_id, l.type, l.qty, l.remaining_qty, l.work_date, l.memo,
                    COALESCE(NULLIF(l.supplier_name, ''), i.A, '') AS supplier_name,
                    COALESCE(NULLIF(l.product_name, ''), i.B, '') AS product_name,
                    COALESCE(NULLIF(l.product_code, ''), i.original_f, '') AS product_code,
                    COALESCE(NULLIF(l.color, ''), i.D, '') AS color,
                    COALESCE(NULLIF(l.size, ''), i.E, '') AS size,
                    l.ts
                FROM misong_logs l
                LEFT JOIN misong_items i ON i.id = l.item_id
                WHERE {where}
                ORDER BY l.ts DESC
                """,
                params,
            ).fetchall()
            return {"ok": True, "logs": [_log_row_to_dict(r) for r in rows]}
        finally:
            conn.close()

    @router.get("/logs-search")
    def search_logs(
        query: str = "",
        date_from: str = "",
        date_to: str = "",
        limit: int = 100,
        user: str = Depends(get_current_user),
    ):
        conn = get_db()
        try:
            _init(conn)
            keyword = query.strip()
            if not keyword:
                return {"ok": True, "logs": []}

            params: list = []
            where_parts = []
            supplier_expr = "COALESCE(NULLIF(l.supplier_name, ''), i.A, '')"
            name_expr = "COALESCE(NULLIF(l.product_name, ''), i.B, '')"
            code_expr = "COALESCE(NULLIF(l.product_code, ''), i.original_f, '')"
            color_expr = "COALESCE(NULLIF(l.color, ''), i.D, '')"
            size_expr = "COALESCE(NULLIF(l.size, ''), i.E, '')"
            where_parts.append(f"({name_expr} LIKE ? OR {supplier_expr} LIKE ?)")
            params.extend([f"%{keyword}%", f"%{keyword}%"])
            if date_from:
                where_parts.append("l.work_date >= ?")
                params.append(date_from)
            if date_to:
                where_parts.append("l.work_date <= ?")
                params.append(date_to)
            params.append(max(1, min(int(limit or 100), 500)))

            rows = conn.execute(
                f"""
                SELECT
                    l.id, l.item_id, l.type, l.qty, l.remaining_qty, l.work_date, l.memo,
                    {supplier_expr} AS supplier_name,
                    {name_expr} AS product_name,
                    {code_expr} AS product_code,
                    {color_expr} AS color,
                    {size_expr} AS size,
                    l.ts
                FROM misong_logs l
                LEFT JOIN misong_items i ON i.id = l.item_id
                WHERE {' AND '.join(where_parts)}
                ORDER BY l.ts DESC
                LIMIT ?
                """,
                params,
            ).fetchall()
            return {"ok": True, "logs": [_log_row_to_dict(r) for r in rows]}
        finally:
            conn.close()

    # ── 알림 조회 ─────────────────────────────────────────────────────────────
    @router.get("/disappeared")
    def list_disappeared_items(
        limit: int = 100,
        user: str = Depends(get_current_user),
    ):
        conn = get_db()
        try:
            _init(conn)
            safe_limit = max(1, min(int(limit or 100), 500))
            rows = conn.execute(
                """
                SELECT
                    l.id, l.item_id, l.type, l.qty, l.remaining_qty, l.work_date, l.memo,
                    COALESCE(NULLIF(l.supplier_name, ''), i.A, '') AS supplier_name,
                    COALESCE(NULLIF(l.product_name, ''), i.B, '') AS product_name,
                    COALESCE(NULLIF(l.product_code, ''), i.original_f, '') AS product_code,
                    COALESCE(NULLIF(l.color, ''), i.D, '') AS color,
                    COALESCE(NULLIF(l.size, ''), i.E, '') AS size,
                    l.ts
                FROM misong_logs l
                LEFT JOIN misong_items i ON i.id = l.item_id
                WHERE l.remaining_qty = 0
                ORDER BY l.ts DESC
                LIMIT ?
                """,
                (safe_limit,),
            ).fetchall()
            return {"ok": True, "items": [_log_row_to_dict(r) for r in rows]}
        finally:
            conn.close()

    @router.get("/alerts")
    def get_alerts(user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            rows = conn.execute(
                "SELECT * FROM misong_alerts ORDER BY ts DESC LIMIT 200"
            ).fetchall()
            return {"ok": True, "alerts": [_alert_row_to_dict(r) for r in rows]}
        finally:
            conn.close()

    # ── 알림 전체 삭제 ────────────────────────────────────────────────────────
    @router.delete("/alerts")
    def clear_alerts(user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init(conn)
            conn.execute("DELETE FROM misong_alerts")
            conn.commit()
            return {"ok": True}
        finally:
            conn.close()

    @router.post("/alerts/{alert_id}/match-code")
    def match_missing_code_alert(alert_id: int, user: str = Depends(get_current_user)):
        import uuid as _uuid

        conn = get_db()
        try:
            _init(conn)
            alert = conn.execute(
                "SELECT * FROM misong_alerts WHERE id = ?",
                (alert_id,),
            ).fetchone()
            if not alert:
                raise HTTPException(404, "알림을 찾을 수 없습니다.")
            if alert["type"] != "missing_code":
                raise HTTPException(400, "코드없음 알림만 매칭할 수 있습니다.")

            product_name = alert["product_name"] if "product_name" in alert.keys() else ""
            color = alert["color"] if "color" in alert.keys() else ""
            size = alert["size"] if "size" in alert.keys() else ""
            if not _normalize_match_text(product_name):
                raise HTTPException(400, "매칭에 필요한 업로드 엑셀 B열 데이터가 없습니다.")

            matched_code = _match_waiting_base_product_code(product_name, color, size)
            if not matched_code:
                raise HTTPException(
                    404,
                    f"원가베이스유 Sheet1에서 상품명/색상/사이즈가 일치하는 상품코드를 찾지 못했습니다: {product_name} {color}/{size}",
                )

            h_val = alert["h_value"] if "h_value" in alert.keys() else ""
            qty = int(alert["qty"] or 0)
            work_date = alert["work_date"] if "work_date" in alert.keys() else ""
            row_info = alert["row_info"] if "row_info" in alert.keys() else ""
            now = _now()
            new_alerts = []

            if h_val == "미송":
                existing = conn.execute(
                    "SELECT * FROM misong_items WHERE original_f = ?",
                    (matched_code,),
                ).fetchone()
                if existing:
                    new_f = existing["F"] + qty
                    conn.execute(
                        """
                        INSERT INTO misong_logs
                          (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                        VALUES (?, 'add', ?, ?, ?, '코드없음 알림 매칭', ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            existing["id"],
                            qty,
                            new_f,
                            work_date,
                            existing["A"],
                            existing["B"],
                            existing["original_f"],
                            existing["D"],
                            existing["E"],
                            now,
                        ),
                    )
                    conn.execute("UPDATE misong_items SET F = ? WHERE id = ?", (new_f, existing["id"]))
                else:
                    item_id = str(_uuid.uuid4())
                    conn.execute(
                        """
                        INSERT INTO misong_items
                          (id, A, B, C, D, E, F, G, original_f, added_at, owner)
                        VALUES (?, ?, ?, '', ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item_id,
                            alert["supplier_name"],
                            product_name,
                            color,
                            size,
                            qty,
                            work_date,
                            matched_code,
                            now,
                            user,
                        ),
                    )
                    conn.execute(
                        """
                        INSERT INTO misong_logs
                          (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                        VALUES (?, 'add', ?, ?, ?, '코드없음 알림 매칭', ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item_id,
                            qty,
                            qty,
                            work_date,
                            alert["supplier_name"],
                            product_name,
                            matched_code,
                            color,
                            size,
                            now,
                        ),
                    )

            elif h_val == "미송픽업":
                existing = conn.execute(
                    "SELECT * FROM misong_items WHERE original_f = ?",
                    (matched_code,),
                ).fetchone()
                if existing is None:
                    new_alerts.append("not_found")
                    conn.execute(
                        """
                        INSERT INTO misong_alerts
                          (type, product_code, detail, row_info, h_value, qty, work_date, supplier_name, product_name, color, size, ts)
                        VALUES ('not_found', ?, '매칭된 상품코드의 미송 항목 없음', ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            matched_code,
                            row_info,
                            h_val,
                            qty,
                            work_date,
                            alert["supplier_name"],
                            product_name,
                            color,
                            size,
                            now,
                        ),
                    )
                else:
                    new_f = existing["F"] - qty
                    if new_f < 0:
                        alert_detail = f"수량 부족(보유: {existing['F']}, 차감: {qty}) - 매칭 후 픽업 처리"
                        new_alerts.append("negative")
                        conn.execute(
                            """
                            INSERT INTO misong_alerts
                              (type, product_code, detail, row_info, h_value, qty, work_date, supplier_name, product_name, color, size, ts)
                            VALUES ('negative', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                matched_code,
                                alert_detail,
                                row_info,
                                h_val,
                                qty,
                                work_date,
                                existing["A"],
                                existing["B"],
                                existing["D"],
                                existing["E"],
                                now,
                            ),
                        )
                    conn.execute(
                        """
                        INSERT INTO misong_logs
                          (item_id, type, qty, remaining_qty, work_date, memo, supplier_name, product_name, product_code, color, size, ts)
                        VALUES (?, 'subtract', ?, ?, ?, '코드없음 알림 매칭', ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            existing["id"],
                            qty,
                            new_f,
                            work_date,
                            existing["A"],
                            existing["B"],
                            existing["original_f"],
                            existing["D"],
                            existing["E"],
                            now,
                        ),
                    )
                    if new_f == 0:
                        conn.execute("DELETE FROM misong_items WHERE id = ?", (existing["id"],))
                    else:
                        conn.execute("UPDATE misong_items SET F = ? WHERE id = ?", (new_f, existing["id"]))
            else:
                raise HTTPException(400, "알림의 미송 구분값이 올바르지 않습니다.")

            conn.execute("DELETE FROM misong_alerts WHERE id = ?", (alert_id,))
            conn.commit()

            items = _fetch_item_rows(conn)
            alerts = conn.execute(
                "SELECT * FROM misong_alerts ORDER BY ts DESC LIMIT 200"
            ).fetchall()
            return {
                "ok": True,
                "matchedCode": matched_code,
                "items": [_item_row_to_dict(r) for r in items],
                "alerts": [_alert_row_to_dict(r) for r in alerts],
                "new_alert_count": len(new_alerts),
            }
        finally:
            conn.close()

    # ── 로그 메모 수정 ────────────────────────────────────────────────────────
    @router.patch("/logs/{log_id}/memo")
    def update_log_memo(
        log_id: int,
        payload: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        conn = get_db()
        try:
            _init(conn)
            conn.execute(
                "UPDATE misong_logs SET memo = ? WHERE id = ?",
                (payload.get("memo", ""), log_id),
            )
            conn.commit()
            return {"ok": True}
        finally:
            conn.close()

    # ── 입고대기설정 (EZAdmin I200 입고처리) ──────────────────────────────────
    @router.post("/waiting-base/export-to-ezadmin")
    async def waiting_base_export_to_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        conn = get_db()
        try:
            _init(conn)
            rows = conn.execute(
                "SELECT original_f, SUM(F) AS qty FROM misong_items "
                "WHERE TRIM(original_f) != '' GROUP BY original_f"
            ).fetchall()
            qty_by_code = {
                _normalize_code(r["original_f"]): int(r["qty"] or 0)
                for r in rows
                if _normalize_code(r["original_f"])
            }
        finally:
            conn.close()

        if not qty_by_code:
            return {"ok": False, "error": "미송목록이 비어 있습니다."}

        if not _INGODAEGI_PATH.exists():
            return {"ok": False, "error": f"입고대기.xlsx 파일을 찾을 수 없습니다: {_INGODAEGI_PATH}"}

        wb_in = openpyxl.load_workbook(_INGODAEGI_PATH)
        ws_in = wb_in.active

        out_wb = xlwt.Workbook()
        out_ws = out_wb.add_sheet("Sheet1")
        for col_idx in range(1, ws_in.max_column + 1):
            out_ws.write(0, col_idx - 1, ws_in.cell(row=1, column=col_idx).value or "")

        matched_count = 0
        out_row = 1
        for row_idx in range(2, ws_in.max_row + 1):
            code = _normalize_code(ws_in.cell(row=row_idx, column=1).value)
            for col_idx in range(1, ws_in.max_column + 1):
                cell_val = ws_in.cell(row=row_idx, column=col_idx).value
                if col_idx == 2 and code and code in qty_by_code:
                    cell_val = qty_by_code[code]
                out_ws.write(out_row, col_idx - 1, "" if cell_val is None else cell_val)
            out_row += 1
            if code and code in qty_by_code:
                matched_count += 1

        if out_row == 1:
            return {"ok": False, "error": "입고대기.xlsx에 데이터가 없습니다."}

        buf = io.BytesIO()
        out_wb.save(buf)
        xls_bytes = buf.getvalue()

        cookies = {"PHPSESSID": phpsessid}
        ez_headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://ga80.ezadmin.co.kr/template40.htm?template=I210",
            "X-Requested-With": "XMLHttpRequest",
        }
        base_url = f"{_EZADMIN_BASE}/function.htm"
        ts_ms = str(int(datetime.now().timestamp() * 1000))

        try:
            async with httpx.AsyncClient(timeout=300.0, verify=False, follow_redirects=True) as client:
                upload_r = await client.post(
                    base_url,
                    data={"template": "I200", "action": "upload_new"},
                    files={"_file": (f"ingodaegi_{ts_ms}.xls", xls_bytes, "application/vnd.ms-excel")},
                    cookies=cookies, headers=ez_headers,
                )
                if upload_r.status_code >= 400:
                    return {"ok": False, "error": f"업로드 실패 (HTTP {upload_r.status_code})"}

                preview_r = await client.post(
                    base_url,
                    data={"_search": "false", "nd": ts_ms, "rows": "2000", "page": "1",
                          "sidx": "", "sord": "asc", "template": "I200",
                          "action": "load_template_data_new"},
                    cookies=cookies, headers=ez_headers,
                )
                try:
                    preview_r.json()
                except Exception:
                    return {"ok": False, "need_session": True}

                time_flag = datetime.now().strftime("%a %b %d %Y %H:%M:%S GMT+0900 (한국 표준시)")
                apply_r = await client.post(
                    base_url,
                    data={"template": "I200", "action": "apply_new",
                          "bad": "reserve_qty", "type": "arrange",
                          "move_warehouse": "0", "save_stock": "0",
                          "stock_tag": "", "timeFlag": time_flag},
                    cookies=cookies, headers=ez_headers,
                )
                try:
                    apply_data = apply_r.json()
                except Exception:
                    return {"ok": False, "error": f"입고처리 응답 파싱 실패: {apply_r.text[:500]}"}

        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        return {"ok": True, "count": matched_count, "apply_response": apply_data}

    return router
