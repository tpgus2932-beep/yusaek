from collections import defaultdict
from datetime import datetime
from pathlib import Path
import io
import json
import os
import re
import shutil
import sqlite3
import tempfile
import uuid

import httpx
import openpyxl
import pandas as pd
import urllib.parse
import xlwt
from fastapi import APIRouter, Body, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, Response
from services.cost_base_append import append_tsv_rows_to_excel

router = APIRouter()

_EZADMIN_BASE = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_AH_DB_PATH = Path(os.environ.get("APP_DB_PATH") or Path(__file__).resolve().parent.parent / "app.db")
_amood_ezadmin_log: list[dict] = []


def _ah_get_ezadmin_phpsessid() -> str:
    try:
        conn = sqlite3.connect(str(_AH_DB_PATH))
        row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (_EZADMIN_SESSION_KEY,)).fetchone()
        conn.close()
        return (row[0] if row else "") or ""
    except Exception:
        return ""

AMOOD_HAPBAE_ALLOWED_EXCEL = {".xlsx", ".xlsm"}
AMOOD_HAPBAE_ALLOWED_COST_BASE = {".xlsx", ".xls", ".xlsm"}
COST_BASE_CODE_COL = 0
COST_BASE_MATCH_COL = 8
COST_BASE_REQUIRED_COLS = COST_BASE_MATCH_COL + 1
SHARED_COST_BASE_PATH = Path(
    os.environ.get("SHARED_COST_BASE_PATH")
    or os.environ.get("AMOOD_HAPBAE_COST_BASE_PATH")
    or os.environ.get("RETURN_COST_BASE_PATH")
    or r"C:\Users\ksh29\OneDrive\Desktop\원베\원가베이스유.xlsx"
)
AMOOD_HAPBAE_COST_BASE_CACHE: dict[str, object] = {"df": None, "mtime": None, "path": None}

_LAST_FILE_DIR = Path(os.environ.get("AMOOD_HAPBAE_LAST_FILE_DIR") or Path(__file__).resolve().parent.parent / "uploads" / "amood_hapbae")
_LAST_FILE_PATH = _LAST_FILE_DIR / "last_upload.xlsx"
_LAST_META_PATH = _LAST_FILE_DIR / "last_upload_meta.json"


def _ah_save_last_file(data: bytes, filename: str) -> None:
    _LAST_FILE_DIR.mkdir(parents=True, exist_ok=True)
    _LAST_FILE_PATH.write_bytes(data)
    _LAST_META_PATH.write_text(
        json.dumps({"filename": filename, "uploaded_at": datetime.now().isoformat()}, ensure_ascii=False),
        encoding="utf-8",
    )


def _ah_load_last_file() -> tuple[bytes | None, str | None]:
    if not _LAST_FILE_PATH.exists() or not _LAST_META_PATH.exists():
        return None, None
    try:
        meta = json.loads(_LAST_META_PATH.read_text(encoding="utf-8"))
        return _LAST_FILE_PATH.read_bytes(), meta.get("filename")
    except Exception:
        return None, None


def _content_disposition(filename: str) -> str:
    safe_name = (filename or "download").replace('"', "")
    ascii_name = "".join(ch if ord(ch) < 128 else "_" for ch in safe_name)
    ascii_name = re.sub(r"_+", "_", ascii_name).strip("_")
    if not ascii_name:
        ascii_name = "download"
    quoted = urllib.parse.quote(safe_name)
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted}"


def _ah_normalize(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _ah_normalize_match_key(v) -> str:
    return _ah_normalize(v).casefold()


def _ah_remove_leading_bracket_tag(text: str) -> str:
    s = _ah_normalize(text)
    if not s:
        return ""

    lead_patterns = [
        r"^\[[^\]]*\]\s*",
        r"^\([^\)]*\)\s*",
        r"^\{[^}]*\}\s*",
    ]
    tail_patterns = [
        r"\s*\[[^\]]*\]$",
        r"\s*\([^\)]*\)$",
        r"\s*\{[^}]*\}$",
    ]

    changed = True
    while changed:
        changed = False
        for pat in lead_patterns:
            new_s = re.sub(pat, "", s)
            if new_s != s:
                s = new_s.strip()
                changed = True

        for pat in tail_patterns:
            new_s = re.sub(pat, "", s)
            if new_s != s:
                s = new_s.strip()
                changed = True

    return s.strip()


def _ah_merge_j_by_slash(text: str) -> str:
    s = _ah_normalize(text)
    if not s:
        return ""
    parts = [p.strip() for p in s.split("/") if p.strip() != ""]
    parts = ["".join(p.split()) for p in parts]
    return " ".join(parts).strip()


def _ah_get_second_sheet(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if len(wb.worksheets) < 2:
        raise ValueError("엑셀에 두 번째 시트가 없습니다.")
    return wb.worksheets[1]


def _ah_find_conflicts_xlsx(path: Path, skip_header: bool = True):
    ws = _ah_get_second_sheet(path)
    start_row = 2 if skip_header else 1
    c_to_dset: dict[str, set[str]] = defaultdict(set)
    for r in range(start_row, ws.max_row + 1):
        c_val = _ah_normalize(ws.cell(row=r, column=3).value)
        d_val = _ah_normalize(ws.cell(row=r, column=4).value)
        if c_val == "":
            continue
        c_to_dset[c_val].add(d_val)

    conflicts: list[tuple[str, set[str]]] = []
    for c_val, d_set in c_to_dset.items():
        if len(d_set) >= 2:
            conflicts.append((c_val, d_set))
    conflicts.sort(key=lambda x: str(x[0]))
    return ws.title, conflicts


def _ah_build_output_rows_from_hj(path: Path, skip_header: bool = True):
    ws = _ah_get_second_sheet(path)
    start_row = 2 if skip_header else 1
    c_counts = defaultdict(int)
    for r in range(start_row, ws.max_row + 1):
        c_val = _ah_normalize(ws.cell(row=r, column=3).value)
        if c_val != "":
            c_counts[c_val] += 1

    out: list[tuple[str, object]] = []
    for r in range(start_row, ws.max_row + 1):
        c_val = _ah_normalize(ws.cell(row=r, column=3).value)
        if c_val == "" or c_counts.get(c_val, 0) < 2:
            continue

        h_val = _ah_normalize(ws.cell(row=r, column=8).value)
        j_val = _ah_normalize(ws.cell(row=r, column=10).value)
        k_raw = ws.cell(row=r, column=11).value
        try:
            k_qty = float(k_raw) if k_raw is not None and str(k_raw).strip() != "" else 0.0
        except Exception:
            k_qty = 0.0
        if k_qty <= 0:
            k_qty = 1

        if h_val == "" and j_val == "":
            continue

        h_clean = _ah_remove_leading_bracket_tag(h_val)
        j_clean = _ah_merge_j_by_slash(j_val)

        if h_clean and j_clean:
            result = f"{h_clean} {j_clean}"
        elif h_clean:
            result = h_clean
        else:
            result = j_clean

        result = re.sub(r"\s+", " ", result).strip()
        if result:
            out.append((result, k_qty))

    return ws.title, out


def _ah_load_base_cost_map(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    cost_map: dict[str, object] = {}
    for r in range(1, ws.max_row + 1):
        key = _ah_normalize_match_key(ws.cell(row=r, column=COST_BASE_MATCH_COL + 1).value)
        val = ws.cell(row=r, column=COST_BASE_CODE_COL + 1).value
        if key == "":
            continue
        if key not in cost_map:
            cost_map[key] = val
    return cost_map


def _ah_read_cost_base_df(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    if ext == ".xls":
        try:
            return pd.read_excel(path, dtype=str, engine="xlrd")
        except Exception:
            return pd.read_excel(path, dtype=str)
    return pd.read_excel(path, dtype=str)


def _ah_load_cost_base_df():
    path = SHARED_COST_BASE_PATH
    if not path.exists():
        raise FileNotFoundError(f"원가베이스 파일을 찾지 못했습니다: {path}")
    mtime = path.stat().st_mtime
    cached_path = AMOOD_HAPBAE_COST_BASE_CACHE.get("path")
    cached_mtime = AMOOD_HAPBAE_COST_BASE_CACHE.get("mtime")
    if AMOOD_HAPBAE_COST_BASE_CACHE.get("df") is not None and cached_path == str(path) and cached_mtime == mtime:
        return AMOOD_HAPBAE_COST_BASE_CACHE["df"]
    df = _ah_read_cost_base_df(path)
    AMOOD_HAPBAE_COST_BASE_CACHE["df"] = df
    AMOOD_HAPBAE_COST_BASE_CACHE["mtime"] = mtime
    AMOOD_HAPBAE_COST_BASE_CACHE["path"] = str(path)
    return df


def _ah_save_cost_base_df(df: pd.DataFrame):
    path = SHARED_COST_BASE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    AMOOD_HAPBAE_COST_BASE_CACHE["df"] = df
    AMOOD_HAPBAE_COST_BASE_CACHE["mtime"] = path.stat().st_mtime
    AMOOD_HAPBAE_COST_BASE_CACHE["path"] = str(path)


def _ah_cost_base_status() -> dict:
    path = SHARED_COST_BASE_PATH
    exists = path.exists()
    mtime = None
    if exists:
        try:
            mtime = datetime.fromtimestamp(path.stat().st_mtime).isoformat()
        except Exception:
            mtime = None
    return {"path": str(path), "exists": exists, "mtime": mtime}


def _ah_pick_header(value: str | None, fallback: str) -> str:
    header = _ah_normalize(value)
    return header if header else fallback


def _ah_find_unmatched_products(path: Path, cost_map: dict[str, object], skip_header: bool = True):
    _, rows = _ah_build_output_rows_from_hj(path, skip_header=skip_header)
    unmatched_rows = 0
    unique_unmatched: list[str] = []
    seen: set[str] = set()
    for val, _ in rows:
        key = _ah_normalize_match_key(val)
        if key in cost_map:
            continue
        unmatched_rows += 1
        if key and key not in seen:
            seen.add(key)
            unique_unmatched.append(val)
    return unique_unmatched, unmatched_rows


def _ah_build_xls_bytes(
    rows: list[tuple[str, object]],
    cost_map: dict[str, object],
    headers: list[str],
    include_cols: list[int],
) -> bytes:
    merged_rows: list[tuple[str, object, object]] = []
    code_index: dict[str, int] = {}

    for val, qty in rows:
        product_code = cost_map.get(_ah_normalize_match_key(val), "")
        code_key = _ah_normalize_match_key(product_code)

        if code_key == "":
            merged_rows.append((val, qty, product_code))
            continue

        if code_key not in code_index:
            code_index[code_key] = len(merged_rows)
            merged_rows.append((val, qty, product_code))
            continue

        idx = code_index[code_key]
        prev_val, prev_qty, prev_code = merged_rows[idx]
        try:
            prev_num = float(prev_qty) if prev_qty is not None and str(prev_qty).strip() != "" else 0.0
        except Exception:
            prev_num = 0.0
        try:
            add_num = float(qty) if qty is not None and str(qty).strip() != "" else 0.0
        except Exception:
            add_num = 0.0
        total = prev_num + add_num
        merged_qty: object = int(total) if total.is_integer() else total
        merged_rows[idx] = (prev_val, merged_qty, prev_code)

    book = xlwt.Workbook()
    sheet = book.add_sheet("결과")

    selected_headers = [headers[idx - 1] for idx in include_cols]
    for idx, header in enumerate(selected_headers):
        sheet.write(0, idx, header)

    for i, row_data in enumerate(merged_rows, start=1):
        val, qty, product_code = row_data
        selected_values: list[object] = []
        for col_no in include_cols:
            if col_no == 1:
                selected_values.append(val)
            elif col_no == 2:
                selected_values.append(product_code)
            elif col_no == 3:
                selected_values.append(qty if qty is not None else "")
            elif col_no == 4:
                selected_values.append("아무드합배")
        for j, cell_val in enumerate(selected_values):
            sheet.write(i, j, cell_val)

    buf = io.BytesIO()
    book.save(buf)
    return buf.getvalue()


@router.get("/amood-hapbae/cost-base/status")
def amood_hapbae_cost_base_status():
    return {"ok": True, "status": _ah_cost_base_status()}


@router.post("/amood-hapbae/cost-base/reload")
def amood_hapbae_cost_base_reload():
    try:
        df = _ah_load_cost_base_df()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")
    return {"ok": True, "rows": len(df), "status": _ah_cost_base_status()}


@router.post("/amood-hapbae/cost-base/upload")
async def amood_hapbae_cost_base_upload(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in AMOOD_HAPBAE_ALLOWED_COST_BASE:
        raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능")

    SHARED_COST_BASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = Path(tempfile.gettempdir()) / f"amood_hapbae_cost_base_{uuid.uuid4().hex}{ext}"
    data = await file.read()
    tmp_path.write_bytes(data)

    try:
        df = _ah_read_cost_base_df(tmp_path)
        if df.shape[1] < COST_BASE_REQUIRED_COLS:
            raise HTTPException(status_code=400, detail="원가베이스는 최소 A~I열이 필요합니다.")
        shutil.move(str(tmp_path), str(SHARED_COST_BASE_PATH))
        _ah_load_cost_base_df()
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    return {"ok": True, "status": _ah_cost_base_status()}


@router.post("/amood-hapbae/cost-base/append-upload")
async def amood_hapbae_cost_base_append_upload(file: UploadFile = File(...)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in AMOOD_HAPBAE_ALLOWED_COST_BASE:
        raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능")

    if not SHARED_COST_BASE_PATH.exists():
        raise HTTPException(status_code=404, detail=f"원가베이스 파일이 없습니다: {SHARED_COST_BASE_PATH}")

    tmp_path = Path(tempfile.gettempdir()) / f"amood_hapbae_cost_base_append_{uuid.uuid4().hex}{ext}"
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
    tmp_path.write_bytes(data)

    try:
        src_df = _ah_read_cost_base_df(tmp_path)
        if src_df.shape[1] < COST_BASE_REQUIRED_COLS:
            raise HTTPException(status_code=400, detail="업로드 엑셀은 최소 A~I열이 필요합니다.")

        dst_df = _ah_load_cost_base_df().copy()
        if dst_df.shape[1] < COST_BASE_REQUIRED_COLS:
            raise HTTPException(status_code=400, detail="기존 원가베이스는 최소 A~I열이 필요합니다.")

        append_df = src_df.iloc[:, :COST_BASE_REQUIRED_COLS].copy()
        append_df.columns = list(dst_df.columns[:COST_BASE_REQUIRED_COLS])
        append_df = append_df.fillna("")
        append_df = append_df[
            (append_df.iloc[:, COST_BASE_CODE_COL].astype(str).str.strip() != "") | (append_df.iloc[:, COST_BASE_MATCH_COL].astype(str).str.strip() != "")
        ].reset_index(drop=True)

        if append_df.empty:
            raise HTTPException(status_code=400, detail="추가할 데이터가 없습니다. (A/I열 확인)")

        if len(dst_df.columns) > COST_BASE_REQUIRED_COLS:
            for col in dst_df.columns[COST_BASE_REQUIRED_COLS:]:
                append_df[col] = ""
            append_df = append_df.reindex(columns=list(dst_df.columns))

        merged_df = pd.concat([dst_df, append_df], ignore_index=True)
        _ah_save_cost_base_df(merged_df)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    return {
        "ok": True,
        "appended_count": int(len(append_df)),
        "total_rows": int(len(merged_df)),
        "status": _ah_cost_base_status(),
    }


@router.post("/amood-hapbae/cost-base/append-tsv")
def amood_hapbae_cost_base_append_tsv(payload: dict = Body(...)):
    raw_text = str(payload.get("text") or "").strip()
    skip_header = bool(payload.get("skip_header"))

    try:
        result = append_tsv_rows_to_excel(
            SHARED_COST_BASE_PATH,
            raw_text,
            read_df=_ah_read_cost_base_df,
            save_df=_ah_save_cost_base_df,
            required_columns=COST_BASE_REQUIRED_COLS,
            append_columns=2,
            target_column_indices=[COST_BASE_CODE_COL, COST_BASE_MATCH_COL],
            skip_header=skip_header,
        )
        status = _ah_cost_base_status()
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="원가베이스 파일이 없습니다.")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"원가베이스 데이터 추가 실패: {e}")

    return {"ok": True, "status": status, **result}


@router.get("/amood-hapbae/cost-base/download")
def amood_hapbae_cost_base_download():
    path = SHARED_COST_BASE_PATH
    if not path.exists():
        raise HTTPException(status_code=404, detail="원가베이스 파일이 없습니다.")
    return FileResponse(path, filename=path.name)


@router.get("/amood-hapbae/cost-base/preview")
def amood_hapbae_cost_base_preview(offset: int = 0, limit: int = 50, q: str | None = None):
    if offset < 0 or limit <= 0 or limit > 200:
        raise HTTPException(status_code=400, detail="offset/limit 값이 올바르지 않습니다.")
    try:
        df = _ah_load_cost_base_df()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")

    if df.shape[1] < COST_BASE_REQUIRED_COLS:
        raise HTTPException(status_code=400, detail="Cost base requires columns A through I.")

    q_norm = str(q).strip() if q else ""
    if q_norm:
        df_view = df.fillna("").astype(str)
        mask = df_view.apply(lambda row: row.str.contains(q_norm, case=False, na=False)).any(axis=1)
        df_filtered = df[mask].copy()
    else:
        df_filtered = df

    total = len(df_filtered)
    col_names = ["A열 상품코드", "I열 상품명 색상 사이즈"]
    end = min(offset + limit, total)
    rows = []
    for i in range(offset, end):
        r = df_filtered.iloc[i]
        row = []
        for v in [r.iloc[COST_BASE_CODE_COL], r.iloc[COST_BASE_MATCH_COL]]:
            if pd.isna(v):
                row.append("")
            else:
                row.append(v)
        rows.append({"row_index": int(r.name), "values": row})
    return {"ok": True, "columns": col_names, "rows": rows, "total": total, "status": _ah_cost_base_status()}


@router.post("/amood-hapbae/cost-base/edit-batch")
def amood_hapbae_cost_base_edit_batch(payload: dict = Body(...)):
    edits = payload.get("edits")
    if not isinstance(edits, list) or not edits:
        raise HTTPException(status_code=400, detail="edits 값이 올바르지 않습니다.")

    try:
        df = _ah_load_cost_base_df()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")

    _PREVIEW_COL_MAP = [COST_BASE_CODE_COL, COST_BASE_MATCH_COL]  # display 0→A, 1→I

    for item in edits:
        row_index = item.get("row_index")
        column = item.get("column")
        value = item.get("value")
        if row_index is None or not isinstance(row_index, int) or row_index < 0:
            continue
        if row_index >= len(df):
            continue
        if isinstance(column, int):
            if column < 0 or column >= len(_PREVIEW_COL_MAP):
                continue
            actual_idx = _PREVIEW_COL_MAP[column]
            col_name = df.columns[actual_idx]
        elif isinstance(column, str) and column in df.columns:
            col_name = column
        else:
            continue
        df.at[row_index, col_name] = "" if value is None else value

    try:
        _ah_save_cost_base_df(df)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"원가베이스 저장 실패: {e}")

    return {"ok": True, "status": _ah_cost_base_status()}


@router.post("/amood-hapbae/cost-base/add-row")
def amood_hapbae_cost_base_add_row(payload: dict = Body(...)):
    name = _ah_normalize(payload.get("name"))
    code = _ah_normalize(payload.get("code"))
    if not name and not code:
        raise HTTPException(status_code=400, detail="A열 상품코드 또는 I열 상품명 색상 사이즈를 입력하세요.")

    try:
        df = _ah_load_cost_base_df().copy()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")

    if df.shape[1] < COST_BASE_REQUIRED_COLS:
        raise HTTPException(status_code=400, detail="원가베이스는 최소 A~I열이 필요합니다.")

    row_data: dict[str, object] = {}
    row_data[df.columns[COST_BASE_CODE_COL]] = code
    row_data[df.columns[COST_BASE_MATCH_COL]] = name
    for index, col in enumerate(list(df.columns)):
        if index not in (COST_BASE_CODE_COL, COST_BASE_MATCH_COL):
            row_data[col] = ""

    df = pd.concat([df, pd.DataFrame([row_data], columns=list(df.columns))], ignore_index=True)

    try:
        _ah_save_cost_base_df(df)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"원가베이스 저장 실패: {e}")

    return {"ok": True, "status": _ah_cost_base_status(), "row_added": {"name": name, "code": code}}


@router.post("/amood-hapbae/conflicts")
async def amood_hapbae_conflicts(
    file: UploadFile | None = File(None),
    skip_header: bool = Form(True),
):
    if file is not None:
        name = file.filename or ""
        ext = Path(name).suffix.lower()
        if ext not in AMOOD_HAPBAE_ALLOWED_EXCEL:
            raise HTTPException(status_code=400, detail="xlsx/xlsm 파일만 업로드 가능합니다.")
        data = await file.read()
        _ah_save_last_file(data, name)
    else:
        data, saved_name = _ah_load_last_file()
        if data is None:
            raise HTTPException(status_code=400, detail="파일을 선택하거나 이전에 업로드된 파일이 있어야 합니다.")
        name = saved_name or "last_upload.xlsx"
        ext = Path(name).suffix.lower()

    tmp_path = Path(tempfile.gettempdir()) / f"amood_hapbae_conflicts_{uuid.uuid4().hex}{ext}"
    tmp_path.write_bytes(data)

    try:
        sheet, conflicts = _ah_find_conflicts_xlsx(tmp_path, skip_header=skip_header)
        unmatched_products: list[str] = []
        unmatched_rows = 0
        cost_base_exists = SHARED_COST_BASE_PATH.exists()
        if cost_base_exists:
            try:
                cost_map = _ah_load_base_cost_map(SHARED_COST_BASE_PATH)
                unmatched_products, unmatched_rows = _ah_find_unmatched_products(
                    tmp_path,
                    cost_map,
                    skip_header=skip_header,
                )
            except Exception:
                unmatched_products = []
                unmatched_rows = 0

        return {
            "ok": True,
            "sheet": sheet,
            "conflict_count": len(conflicts),
            "cost_base_exists": cost_base_exists,
            "unmatched_product_count": len(unmatched_products),
            "unmatched_row_count": unmatched_rows,
            "unmatched_products": unmatched_products,
            "conflicts": [
                {"c": c_val, "d_values": sorted(list(d_set), key=lambda x: str(x))}
                for c_val, d_set in conflicts
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


@router.post("/amood-hapbae/export")
async def amood_hapbae_export(
    file: UploadFile | None = File(None),
    skip_header: bool = Form(True),
    header_col1: str = Form("상품명"),
    header_col2: str = Form("상품코드"),
    header_col3: str = Form("작업수량"),
    header_col4: str = Form("메모"),
    include_col1: bool = Form(True),
    include_col2: bool = Form(True),
    include_col3: bool = Form(True),
    include_col4: bool = Form(True),
):
    if file is not None:
        name = file.filename or "amood_hapbae.xlsx"
        ext = Path(name).suffix.lower()
        if ext not in AMOOD_HAPBAE_ALLOWED_EXCEL:
            raise HTTPException(status_code=400, detail="xlsx/xlsm 파일만 업로드 가능합니다.")
        data = await file.read()
        _ah_save_last_file(data, name)
    else:
        data, saved_name = _ah_load_last_file()
        if data is None:
            raise HTTPException(status_code=400, detail="파일을 선택하거나 이전에 업로드된 파일이 있어야 합니다.")
        name = saved_name or "amood_hapbae.xlsx"
        ext = Path(name).suffix.lower()

    if not SHARED_COST_BASE_PATH.exists():
        raise HTTPException(
            status_code=400,
            detail=f"원가베이스 파일을 읽을 수 없습니다: {SHARED_COST_BASE_PATH}",
        )

    tmp_path = Path(tempfile.gettempdir()) / f"amood_hapbae_export_{uuid.uuid4().hex}{ext}"
    tmp_path.write_bytes(data)

    try:
        _, rows = _ah_build_output_rows_from_hj(tmp_path, skip_header=skip_header)
        if not rows:
            raise HTTPException(status_code=400, detail="가공할 데이터(H/J)가 없습니다.")

        cost_map = _ah_load_base_cost_map(SHARED_COST_BASE_PATH)

        headers = [
            _ah_pick_header(header_col1, "상품명"),
            _ah_pick_header(header_col2, "상품코드"),
            _ah_pick_header(header_col3, "작업수량"),
            _ah_pick_header(header_col4, "메모"),
        ]
        include_cols: list[int] = []
        if include_col1:
            include_cols.append(1)
        if include_col2:
            include_cols.append(2)
        if include_col3:
            include_cols.append(3)
        if include_col4:
            include_cols.append(4)
        if not include_cols:
            raise HTTPException(status_code=400, detail="다운로드할 열을 최소 1개 선택하세요.")

        content = _ah_build_xls_bytes(rows, cost_map, headers, include_cols)
        filename = f"{Path(name).stem}_가공본.xls"
        headers = {"Content-Disposition": _content_disposition(filename)}
        return Response(
            content=content,
            media_type="application/vnd.ms-excel",
            headers=headers,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


def _ah_build_ezadmin_xls(rows: list[tuple[str, object]], cost_map: dict) -> bytes:
    merged_rows: list[tuple[str, object, object]] = []
    code_index: dict[str, int] = {}
    for val, qty in rows:
        product_code = cost_map.get(_ah_normalize_match_key(val), "")
        code_key = _ah_normalize_match_key(product_code)
        if not code_key:
            merged_rows.append((val, qty, product_code))
            continue
        if code_key not in code_index:
            code_index[code_key] = len(merged_rows)
            merged_rows.append((val, qty, product_code))
            continue
        idx = code_index[code_key]
        prev_val, prev_qty, prev_code = merged_rows[idx]
        try:
            prev_num = float(prev_qty) if prev_qty is not None and str(prev_qty).strip() != "" else 0.0
        except Exception:
            prev_num = 0.0
        try:
            add_num = float(qty) if qty is not None and str(qty).strip() != "" else 0.0
        except Exception:
            add_num = 0.0
        total = prev_num + add_num
        merged_rows[idx] = (prev_val, int(total) if float(total).is_integer() else total, prev_code)

    book = xlwt.Workbook()
    sheet = book.add_sheet("작업")
    sheet.write(0, 0, "상품코드")
    sheet.write(0, 1, "작업수량")
    sheet.write(0, 2, "메모")
    row_idx = 1
    for _, qty, product_code in merged_rows:
        if not _ah_normalize(str(product_code or "")):
            continue
        sheet.write(row_idx, 0, str(product_code))
        sheet.write(row_idx, 1, qty if isinstance(qty, (int, float)) else 0)
        sheet.write(row_idx, 2, "아무드합배")
        row_idx += 1
    buf = io.BytesIO()
    book.save(buf)
    return buf.getvalue(), row_idx - 1


async def _ah_send_to_ezadmin_impl(xls_bytes: bytes, direction: str, phpsessid: str) -> dict:
    cookies = {"PHPSESSID": phpsessid}
    ez_headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": f"{_EZADMIN_BASE}/template40.htm?template=I210",
        "X-Requested-With": "XMLHttpRequest",
    }
    base_url = f"{_EZADMIN_BASE}/function.htm"
    ts_ms = str(int(datetime.now().timestamp() * 1000))
    diag: dict = {}

    async with httpx.AsyncClient(timeout=120.0, verify=False, follow_redirects=True) as client:
        # 템플릿 페이지 먼저 방문 — EZAdmin 서버 사이드 세션 변수 초기화
        await client.get(
            f"{_EZADMIN_BASE}/template40.htm",
            params={"template": "I210"},
            cookies=cookies,
            headers={"User-Agent": "Mozilla/5.0"},
        )

        upload_r = await client.post(
            base_url,
            data={"template": "I200", "action": "upload_new"},
            files={"_file": (f"amood_hapbae_{ts_ms}.xls", xls_bytes, "application/vnd.ms-excel")},
            cookies=cookies,
            headers=ez_headers,
        )
        diag["upload_status"] = upload_r.status_code
        diag["upload_body"] = upload_r.text[:200]
        if upload_r.status_code >= 400:
            return {"ok": False, "error": f"업로드 실패 (HTTP {upload_r.status_code})", "_diag": diag}

        preview_r = await client.post(
            base_url,
            data={
                "_search": "false", "nd": ts_ms,
                "rows": "99999", "page": "1", "sidx": "", "sord": "asc",
                "template": "I200", "action": "load_template_data_new",
            },
            cookies=cookies,
            headers=ez_headers,
        )
        diag["preview_status"] = preview_r.status_code
        try:
            diag["preview_data"] = preview_r.json()
        except Exception:
            diag["preview_body"] = preview_r.text[:200]
            return {"ok": False, "need_session": True, "_diag": diag}

        time_flag = datetime.now().strftime("%a %b %d %Y %H:%M:%S GMT+0900 (한국 표준시)")
        apply_r = await client.post(
            base_url,
            data={
                "template": "I200", "action": "apply_new",
                "bad": "0", "type": direction,
                "move_warehouse": "0", "save_stock": "0",
                "stock_tag": "", "timeFlag": time_flag,
            },
            cookies=cookies,
            headers=ez_headers,
        )
        diag["apply_status"] = apply_r.status_code
        try:
            apply_data = apply_r.json()
            diag["apply_data"] = apply_data
        except Exception:
            diag["apply_body"] = apply_r.text[:300]
            label = "출고" if direction == "out" else "입고"
            return {"ok": False, "error": f"{label}처리 응답 파싱 실패", "_diag": diag}

        result_code = str(apply_data.get("result", "")).strip()
        if result_code == "0" or result_code == "false":
            return {"ok": False, "error": apply_data.get("message") or apply_data.get("msg") or f"EZAdmin 처리 거부 (result={result_code})", "_diag": diag}

    return {"ok": True, "_diag": diag}


@router.get("/amood-hapbae/last-file/info")
def amood_hapbae_last_file_info():
    if not _LAST_META_PATH.exists():
        return {"ok": True, "file": None}
    try:
        meta = json.loads(_LAST_META_PATH.read_text(encoding="utf-8"))
        return {"ok": True, "file": meta}
    except Exception:
        return {"ok": True, "file": None}


@router.get("/amood-hapbae/ezadmin-log")
def amood_hapbae_ezadmin_log():
    return {"log": list(reversed(_amood_ezadmin_log))}


@router.post("/amood-hapbae/send-to-ezadmin")
async def amood_hapbae_send_to_ezadmin(
    file: UploadFile | None = File(None),
    direction: str = Form("out"),
    skip_header: bool = Form(True),
):
    phpsessid = _ah_get_ezadmin_phpsessid()
    if not phpsessid:
        return {"ok": False, "need_session": True}

    if file is not None:
        name = file.filename or "amood_hapbae.xlsx"
        ext = Path(name).suffix.lower()
        if ext not in AMOOD_HAPBAE_ALLOWED_EXCEL:
            raise HTTPException(status_code=400, detail="xlsx/xlsm 파일만 가능합니다.")
        data = await file.read()
        if not data:
            raise HTTPException(status_code=400, detail="파일이 비어 있습니다.")
        _ah_save_last_file(data, name)
    else:
        data, saved_name = _ah_load_last_file()
        if data is None:
            raise HTTPException(status_code=400, detail="파일을 선택하거나 이전에 업로드된 파일이 있어야 합니다.")
        name = saved_name or "amood_hapbae.xlsx"
        ext = Path(name).suffix.lower()

    tmp_path = Path(tempfile.gettempdir()) / f"amood_ez_{uuid.uuid4().hex}{ext}"
    tmp_path.write_bytes(data)

    try:
        _, rows = _ah_build_output_rows_from_hj(tmp_path, skip_header=skip_header)
        if not rows:
            raise HTTPException(status_code=400, detail="가공할 데이터(H/J)가 없습니다.")
        cost_map = _ah_load_base_cost_map(SHARED_COST_BASE_PATH) if SHARED_COST_BASE_PATH.exists() else {}
        xls_bytes, count = _ah_build_ezadmin_xls(rows, cost_map)
        result = await _ah_send_to_ezadmin_impl(xls_bytes, direction, phpsessid)
        if result.get("ok"):
            result["count"] = count
        label = "출고" if direction == "out" else "입고"
        _amood_ezadmin_log.append({
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "label": label,
            "ok": bool(result.get("ok")),
            "count": count if result.get("ok") else 0,
            "error": result.get("error") or "",
            "diag": result.get("_diag"),
        })
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
