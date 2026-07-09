import io
import os
import re
import shutil
import sqlite3
import tempfile
import traceback
import uuid
from collections import Counter
from datetime import datetime, timedelta
from io import StringIO
from pathlib import Path

import httpx
import pandas as pd
from fastapi import APIRouter, Body, Depends, File, HTTPException, Response, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook, load_workbook

from services.pastelco_utils import (
    pastelco_fetch_all_orders,
    pastelco_fetch_shipping_processing_today,
    pastelco_login,
    pastelco_today_kst,
)

_AMOOD_EZADMIN_BASE = "https://ga80.ezadmin.co.kr"
_AMOOD_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_AMOOD_DB_PATH = Path(os.environ.get("APP_DB_PATH") or Path(__file__).resolve().parent.parent / "app.db")

_EZADMIN_COL_ORDER = [
    "상태", "관리번호", "발주일", "판매처", "주문번호",
    "수령자", "주소", "상품코드", "상품명", "옵션",
    "수량", "택배사", "송장번호",
]

_BR_SEP = "\x01"


def _expand_combined_rows(df: "pd.DataFrame", sep: str = _BR_SEP) -> "pd.DataFrame":
    """합배송으로 <br> 구분된 셀을 개별 행으로 분리한다."""
    new_rows = []
    for _, row in df.iterrows():
        max_n = 1
        for val in row:
            if isinstance(val, str) and sep in val:
                max_n = max(max_n, len([p for p in val.split(sep) if p.strip()]))
        if max_n <= 1:
            new_rows.append(row.to_dict())
            continue
        for i in range(max_n):
            new_row = {}
            for col in df.columns:
                val = row[col]
                if isinstance(val, str) and sep in val:
                    parts = [p.strip() for p in val.split(sep) if p.strip()]
                    new_row[col] = parts[i] if i < len(parts) else (parts[-1] if parts else "")
                else:
                    new_row[col] = val
            new_rows.append(new_row)
    if not new_rows:
        return df.iloc[0:0].copy()
    return pd.DataFrame(new_rows, columns=df.columns).reset_index(drop=True)


def _amood_get_ezadmin_phpsessid() -> str:
    try:
        conn = sqlite3.connect(str(_AMOOD_DB_PATH))
        row = conn.execute(
            "SELECT value FROM app_settings WHERE key = ?",
            (_AMOOD_EZADMIN_SESSION_KEY,),
        ).fetchone()
        conn.close()
        return (row[0] if row else "") or ""
    except Exception:
        return ""


def build_amood_router(
    *,
    get_current_user,
    get_amood_state,
    amood_status,
    amood_load_workbooks,
    amood_norm_barcode,
    amood_to_int_qty,
    amood_ws_cell,
    amood_strip_any_brackets,
    amood_fill_down_merged_column,
    amood_norm_key,
    amood_build_output_text,
    amood_collect_rows_by_value,
    load_excel_any,
    content_disposition,
    amood_allowed_excel1,
    amood_allowed_excel2,
    AMOOD_COL1_NAME_RAW,
    AMOOD_COL1_SCAN_BARCODE,
    AMOOD_COL1_ORDER_KEY,
    AMOOD_COL1_NUM_B,
    AMOOD_COL1_NUM_C,
    AMOOD_COL2_ORDER_KEY,
    AMOOD_COL2_NAME,
    AMOOD_COL2_OPTION,
    AMOOD_COL2_QTY,
    AMOOD_COL2_BARCODE,
    AMOOD_COL2_OUTPUT,
    get_shared_incoming_counts,
    set_shared_incoming_counts,
    get_shared_defect_counts,
    set_shared_amood_ezadmin_file,
):
    router = APIRouter()

    def load_pastelco_items_into_state(user: str, items: list):
        wb = Workbook()
        ws_placeholder = wb.active
        ws_placeholder.title = "Sheet1"
        ws_data = wb.create_sheet("주문데이터")

        ws_data.cell(1, 2).value = "주문번호(외부)"
        ws_data.cell(1, 3).value = "주문번호(내부)"
        ws_data.cell(1, 4).value = "선적바코드"
        ws_data.cell(1, 8).value = "상품명"
        ws_data.cell(1, 10).value = "옵션"
        ws_data.cell(1, 11).value = "수량"

        for i, item in enumerate(items, start=2):
            hbl = item.get("ably_pantos_hbl") or {}
            ws_data.cell(i, 2).value = item.get("id")
            ws_data.cell(i, 3).value = (item.get("order") or {}).get("id")
            ws_data.cell(i, 4).value = hbl.get("hbl_no") or ""
            ws_data.cell(i, 8).value = (item.get("product") or {}).get("name_origin") or ""
            opt_vals = (item.get("product_option") or {}).get("option_values_origin") or []
            ws_data.cell(i, 10).value = "/".join(str(v) for v in opt_vals if v)
            ws_data.cell(i, 11).value = item.get("quantity") or 1

        tmp_path = Path(tempfile.gettempdir()) / f"amood_pastelco_{uuid.uuid4().hex}.xlsx"
        wb.save(tmp_path)

        state = get_amood_state(user)
        if state.file1_path and isinstance(state.file1_path, Path) and state.file1_path.exists():
            try:
                state.file1_path.unlink(missing_ok=True)
            except Exception:
                pass
        state.file1_path = tmp_path
        state.file1_name = "pastelco_orders.xlsx"
        state.processed1_path = None
        state.wb1 = None
        state.ws1 = None
        state.current_invoice = None
        state.pending_items = []
        state.waiting_for_items = False
        return state

    def _amood_defect_count(code: str) -> int:
        normalized = amood_norm_barcode(code)
        return int((get_shared_defect_counts() or {}).get(normalized, 0))

    def _amood_item_has_defect(item: dict) -> bool:
        return _amood_defect_count(item.get("barcode", "")) > 0

    def _amood_invoice_has_defect(state) -> bool:
        return any(_amood_item_has_defect(it) for it in state.pending_items)

    def _sync_shared_incoming(state):
        state.incoming_counts = dict(get_shared_incoming_counts() or {})
        return state.incoming_counts

    def _amood_items_view(state) -> list[dict]:
        items = []
        incoming_counts = _sync_shared_incoming(state)
        for it in state.pending_items:
            code = it.get("barcode", "")
            incoming_n = incoming_counts.get(amood_norm_barcode(code), 0)
            items.append(
                {
                    "code": code,
                    "name": it.get("name", "") or "",
                    "option": it.get("option", "") or "",
                    "remain": it.get("remaining", 0),
                    "incoming": incoming_n,
                    "defect": _amood_defect_count(code),
                }
            )
        return items

    def _amood_first_remaining(state):
        incoming_counts = _sync_shared_incoming(state)
        for it in state.pending_items:
            if it.get("remaining", 0) > 0:
                code = it.get("barcode", "")
                incoming_n = incoming_counts.get(amood_norm_barcode(code), 0)
                return {
                    "code": code,
                    "name": it.get("name", "") or "",
                    "option": it.get("option", "") or "",
                    "remain": it.get("remaining", 0),
                    "incoming": incoming_n,
                    "defect": _amood_defect_count(code),
                }
        return None

    def _amood_reset_state(state):
        for path in [state.file1_path, state.file2_path, state.processed1_path, state.processed2_path]:
            if path and isinstance(path, Path) and path.exists():
                try:
                    path.unlink(missing_ok=True)
                except Exception:
                    pass
        set_shared_amood_ezadmin_file(None)
        state.file1_path = None
        state.file2_path = None
        state.file1_name = None
        state.file2_name = None
        state.processed1_path = None
        state.processed2_path = None
        state.wb1 = None
        state.ws1 = None
        state.wb2 = None
        state.ws2 = None
        state.current_invoice = None
        state.pending_items = []
        state.waiting_for_items = False
        state.completed_mgmt_numbers = set()
        state.incoming_counts = {}

    @router.get("/amood/status")
    def amood_status_api(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        _sync_shared_incoming(state)
        return {
            "ok": True,
            "status": amood_status(state),
            "items": state.pending_items,
        }

    @router.post("/amood/incoming/upload")
    async def amood_incoming_upload(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        name = (file.filename or "").lower()
        if not (name.endswith(".xls") or name.endswith(".xlsx")):
            raise HTTPException(status_code=400, detail="xls/xlsx files only")

        suffix = ".xlsx" if name.endswith(".xlsx") else ".xls"
        tmp_path = Path(tempfile.gettempdir()) / f"amood_incoming_{uuid.uuid4().hex}{suffix}"
        data = await file.read()
        tmp_path.write_bytes(data)

        try:
            wb, ws = load_excel_any(tmp_path)
            counts = Counter()
            for r in range(1, ws.max_row + 1):
                code_raw = ws.cell(r, 1).value
                qty_raw = ws.cell(r, 2).value
                code = amood_norm_barcode(code_raw)
                if not code:
                    continue
                qty = amood_to_int_qty(qty_raw)
                if qty > 0:
                    counts[code] += qty
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"incoming load failed: {e}")

        set_shared_incoming_counts(dict(counts))
        state = get_amood_state(user)
        _sync_shared_incoming(state)
        return {
            "ok": True,
            "codes": len(counts),
            "total_qty": sum(counts.values()),
            "status": amood_status(state),
        }

    @router.post("/amood/excel1")
    def amood_upload_excel1(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        name = file.filename or ""
        ext = Path(name).suffix.lower()
        if ext not in amood_allowed_excel1:
            raise HTTPException(status_code=400, detail="excel1은 .xlsx/.xlsm만 가능합니다.")
        tmp_path = Path(tempfile.gettempdir()) / f"amood_excel1_{uuid.uuid4().hex}{ext}"
        with tmp_path.open("wb") as out:
            shutil.copyfileobj(file.file, out)
        state = get_amood_state(user)
        state.file1_path = tmp_path
        state.file1_name = name or tmp_path.name
        state.processed1_path = None
        state.processed2_path = None
        state.wb1 = None
        state.ws1 = None
        state.current_invoice = None
        state.pending_items = []
        state.waiting_for_items = False
        return {"ok": True, "status": amood_status(state)}

    @router.post("/amood/excel2")
    def amood_upload_excel2(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        name = file.filename or ""
        ext = Path(name).suffix.lower()
        if ext not in amood_allowed_excel2:
            raise HTTPException(status_code=400, detail="excel2는 .xls/.xlsx/.xlsm/.htm/.html만 가능합니다.")
        tmp_path = Path(tempfile.gettempdir()) / f"amood_excel2_{uuid.uuid4().hex}{ext}"
        with tmp_path.open("wb") as out:
            shutil.copyfileobj(file.file, out)
        state = get_amood_state(user)
        state.file2_path = tmp_path
        state.file2_name = name or tmp_path.name
        state.processed1_path = None
        state.processed2_path = None
        state.wb2 = None
        state.ws2 = None
        state.current_invoice = None
        state.pending_items = []
        state.waiting_for_items = False
        return {"ok": True, "status": amood_status(state)}

    @router.post("/amood/load-from-ezadmin")
    async def amood_load_from_ezadmin(
        start_date: str = None,
        end_date: str = None,
        user: str = Depends(get_current_user),
    ):
        phpsessid = _amood_get_ezadmin_phpsessid()
        if not phpsessid:
            raise HTTPException(status_code=400, detail="이지어드민 세션(PHPSESSID)이 없습니다.")

        today = datetime.now().strftime("%Y-%m-%d")
        if not end_date:
            end_date = today
        if not start_date:
            start_date = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")

        ez_headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": _AMOOD_EZADMIN_BASE,
            "Referer": f"{_AMOOD_EZADMIN_BASE}/template35.htm?template=S700",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36",
            "X-Requested-With": "XMLHttpRequest",
        }
        cookies = {"PHPSESSID": phpsessid}
        par = (
            f"cs=0&status=1&type=0&dateType=0"
            f"&start_date2={start_date}&start_hour2=00"
            f"&end_date2={end_date}&end_hour2=23"
            f"&date_period_sel2=9"
            f"&search_product_type=name&txt_search_product="
            f"&tag_string=&priorityDeliv=0"
            f"&multi_shop_group=&multi_shop=&str_shop_code=10080&shop_group=0"
            f"&partCount=&price_type=part&part_deliv_price="
            f"&orderReserve=1&multi_supply_group=&multi_supply=&str_supply_code=0"
            f"&s_group_id=&part_deliv_cnt=&check_expect_date=&user_area="
            f"&labels_string=&include_type=1&is_download=0"
            f"&orderType=1&agencyTemplate=20004&check_address=2"
            f"&teamf_transType=tonight&kakaoT_type=0&deliv_today_type=0&hanjin_deliv_type=GN"
        )
        form_data = {
            "template": "S700",
            "action": "get_success_order",
            "_search": "false",
            "nd": str(int(datetime.now().timestamp() * 1000)),
            "rows": "999999",
            "page": "1",
            "sidx": "",
            "sord": "asc",
            "readonly": "T",
            "par": par,
        }

        async with httpx.AsyncClient(timeout=120) as client:
            res = await client.post(
                f"{_AMOOD_EZADMIN_BASE}/function.htm",
                headers=ez_headers,
                cookies=cookies,
                data=form_data,
            )

        try:
            result = res.json()
        except Exception:
            raise HTTPException(status_code=502, detail="이지어드민 응답 파싱 실패 (세션 만료?)")

        api_rows = result.get("rows", [])
        if not api_rows:
            raise HTTPException(status_code=400, detail="불러온 데이터가 없습니다.")

        # [코드]상품명-[옵션] N개 형태 파싱
        _product_re = re.compile(r'\[([A-Z]\d+)\](.*?)(?:-\[([^\]]*)\])?\s+(\d+)개')

        expanded_rows: list[dict] = []
        seen_seq: list[str] = []
        seen_seq_set: set[str] = set()

        for order in api_rows:
            cell = order.get("cell", {})
            seq = str(cell.get("seq", "")).strip()
            status = str(cell.get("status", "")).strip()
            collect_date = str(cell.get("collect_date", "")).strip()
            shop_name = str(cell.get("shop_name", "")).strip()
            order_id = str(cell.get("order_id", "")).strip()
            recv_name = str(cell.get("recv_name", "")).strip()
            recv_address = str(cell.get("recv_address", "")).strip()
            products_str = str(cell.get("products", "")).strip()

            if seq and seq not in seen_seq_set:
                seen_seq_set.add(seq)
                seen_seq.append(seq)

            for m in _product_re.finditer(products_str):
                barcode = m.group(1).strip()
                name = m.group(2).strip().rstrip("-").strip()
                option = m.group(3).strip() if m.group(3) else ""
                qty = int(m.group(4))
                expanded_rows.append({
                    "상태": status,
                    "관리번호": seq,
                    "발주일": collect_date,
                    "판매처": shop_name,
                    "주문번호": order_id,
                    "수령자": recv_name,
                    "주소": recv_address,
                    "상품코드": barcode,
                    "상품명": name,
                    "옵션": option,
                    "수량": qty,
                    "택배사": "",
                    "송장번호": "",
                })

        if not expanded_rows:
            raise HTTPException(status_code=400, detail="상품 데이터를 파싱할 수 없습니다.")

        wb = Workbook()
        ws = wb.active
        for col_idx, header in enumerate(_EZADMIN_COL_ORDER, 1):
            ws.cell(1, col_idx).value = header
        for row_idx, row_data in enumerate(expanded_rows, 2):
            for col_idx, col_name in enumerate(_EZADMIN_COL_ORDER, 1):
                val = row_data.get(col_name)
                if val is not None:
                    ws.cell(row_idx, col_idx).value = val

        tmp_path = Path(tempfile.gettempdir()) / f"amood_excel2_ezadmin_{uuid.uuid4().hex}.xlsx"
        wb.save(tmp_path)

        set_shared_amood_ezadmin_file({
            "file2_path": tmp_path,
            "file2_name": f"이지어드민_{start_date}_{end_date}.xlsx",
        })
        state = get_amood_state(user)

        return {
            "ok": True,
            "count": len(expanded_rows),
            "management_numbers": seen_seq,
            "status": amood_status(state),
        }

    @router.post("/amood/hapbae-pack")
    async def amood_hapbae_pack(
        body: dict = Body(...),
        user: str = Depends(get_current_user),
    ):
        management_numbers = [str(n).strip() for n in body.get("management_numbers", []) if str(n).strip()]
        if len(management_numbers) < 2:
            raise HTTPException(status_code=400, detail="합포할 번호가 2개 이상 필요합니다.")

        phpsessid = _amood_get_ezadmin_phpsessid()
        if not phpsessid:
            raise HTTPException(status_code=400, detail="이지어드민 세션(PHPSESSID)이 없습니다.")

        seq = management_numbers[0]
        pack_seq = ",".join(management_numbers[1:])

        ez_headers = {
            "Accept": "*/*",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Origin": _AMOOD_EZADMIN_BASE,
            "Referer": f"{_AMOOD_EZADMIN_BASE}/template35.htm?template=E900",
            "User-Agent": "Mozilla/5.0",
            "X-Requested-With": "XMLHttpRequest",
        }
        cookies = {"PHPSESSID": phpsessid}
        form_data = {
            "template": "E900",
            "action": "add_pack",
            "seq": seq,
            "pack_seq": pack_seq,
            "cs_type": "",
            "content": "",
            "timeFlag": "0",
        }

        async with httpx.AsyncClient(timeout=30) as client:
            res = await client.post(
                f"{_AMOOD_EZADMIN_BASE}/function.htm",
                headers=ez_headers,
                cookies=cookies,
                data=form_data,
            )

        try:
            result = res.json()
        except Exception:
            result = {"raw": res.text[:500]}

        ok = result.get("result") not in ("0", "false", False, 0) if isinstance(result, dict) else True
        return {"ok": ok, "seq": seq, "pack_seq": pack_seq, "result": result}

    @router.post("/amood/preprocess")
    def amood_preprocess(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        if not state.file1_path or not state.file2_path:
            raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
        try:
            amood_load_workbooks(state)
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="엑셀 로드 실패")
        wb1, ws1 = state.wb1, state.ws1
        wb2, ws2 = state.wb2, state.ws2

        for r in range(2, ws1.max_row + 1):
            v = amood_ws_cell(ws1, AMOOD_COL1_NAME_RAW, r).value
            if v is None:
                continue
            amood_ws_cell(ws1, AMOOD_COL1_NAME_RAW, r).value = amood_strip_any_brackets(str(v))

        amood_fill_down_merged_column(ws2, AMOOD_COL2_ORDER_KEY, start_row=2)
        for r in range(2, ws2.max_row + 1):
            cell = amood_ws_cell(ws2, AMOOD_COL2_ORDER_KEY, r)
            original = cell.value
            cell.value = None
            cell.value = amood_norm_key(original)

        for r in range(2, ws2.max_row + 1):
            name = amood_ws_cell(ws2, AMOOD_COL2_NAME, r).value
            opt = amood_ws_cell(ws2, AMOOD_COL2_OPTION, r).value
            qty = amood_ws_cell(ws2, AMOOD_COL2_QTY, r).value
            out = amood_build_output_text(name, opt, qty)
            amood_ws_cell(ws2, AMOOD_COL2_OUTPUT, r).value = out

        out1 = Path(tempfile.gettempdir()) / f"amood_excel1_processed_{uuid.uuid4().hex}.xlsx"
        out2 = Path(tempfile.gettempdir()) / f"amood_excel2_processed_{uuid.uuid4().hex}.xlsx"
        wb1.save(out1)
        wb2.save(out2)

        state.processed1_path = out1
        state.processed2_path = out2
        return {"ok": True, "status": amood_status(state)}

    @router.get("/amood/download/1")
    def amood_download_excel1(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        if not state.processed1_path or not state.processed1_path.exists():
            raise HTTPException(status_code=404, detail="전처리 결과가 없습니다.")
        name = state.file1_name or "amood_excel1"
        filename = f"{Path(name).stem}_processed.xlsx"
        return FileResponse(state.processed1_path, filename=filename)

    @router.get("/amood/download/2")
    def amood_download_excel2(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        if not state.processed2_path or not state.processed2_path.exists():
            raise HTTPException(status_code=404, detail="전처리 결과가 없습니다.")
        name = state.file2_name or "amood_excel2"
        filename = f"{Path(name).stem}_processed.xlsx"
        return FileResponse(state.processed2_path, filename=filename)

    @router.post("/amood/hapbae-remaining")
    def amood_hapbae_remaining(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        if not state.file1_path or not state.file2_path:
            raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
        try:
            amood_load_workbooks(state)
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="엑셀 로드 실패")

        source_path = state.processed1_path if state.processed1_path and state.processed1_path.exists() else state.file1_path
        wb1 = load_workbook(source_path)
        if len(wb1.worksheets) < 2:
            wb1.close()
            raise HTTPException(status_code=400, detail="아무드 엑셀에 두번째 시트가 없습니다.")
        ws1 = wb1.worksheets[1]
        ws2 = state.ws2

        amood_fill_down_merged_column(ws2, AMOOD_COL2_ORDER_KEY, start_row=2)
        matched_keys = {
            amood_norm_key(amood_ws_cell(ws2, AMOOD_COL2_ORDER_KEY, r).value)
            for r in range(2, ws2.max_row + 1)
        }
        matched_keys.discard("")

        delete_rows = []
        for r in range(2, ws1.max_row + 1):
            order_key = amood_norm_key(amood_ws_cell(ws1, AMOOD_COL1_ORDER_KEY, r).value)
            if order_key and order_key in matched_keys:
                delete_rows.append(r)

        for row_idx in reversed(delete_rows):
            ws1.delete_rows(row_idx, 1)

        remaining_rows = max(ws1.max_row - 1, 0)
        buf = io.BytesIO()
        wb1.save(buf)
        wb1.close()
        buf.seek(0)
        base = Path(state.file1_name or "amood").stem
        filename = f"{base}_합배송.xlsx"
        headers = {
            "Content-Disposition": content_disposition(filename),
            "X-Deleted-Rows": str(len(delete_rows)),
            "X-Remaining-Rows": str(remaining_rows),
        }
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @router.get("/amood/scan/status")
    def amood_scan_status(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        _sync_shared_incoming(state)
        return {
            "ok": True,
            "current_invoice": state.current_invoice,
            "items": _amood_items_view(state),
            "current_next": _amood_first_remaining(state),
            "invoice_has_defect": _amood_invoice_has_defect(state),
        }

    @router.post("/amood/reset")
    def amood_reset(user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        _amood_reset_state(state)
        _sync_shared_incoming(state)
        return {"ok": True, "status": amood_status(state)}

    @router.post("/amood/scan/invoice")
    def amood_scan_invoice(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        _sync_shared_incoming(state)
        amood_load_workbooks(state)

        invoice = (payload.get("invoice") or "").strip()
        if not invoice:
            raise HTTPException(status_code=400, detail="invoice 값이 비어있음")

        amood_fill_down_merged_column(state.ws2, AMOOD_COL2_ORDER_KEY, start_row=2)

        r1_list = []
        target = amood_norm_barcode(invoice).upper()
        for r in range(2, state.ws1.max_row + 1):
            v = amood_ws_cell(state.ws1, AMOOD_COL1_SCAN_BARCODE, r).value
            if not v:
                continue
            if amood_norm_barcode(v).upper() == target:
                r1_list.append(r)

        if not r1_list:
            return {"ok": False, "type": "invoice", "result": "NOT_FOUND", "invoice": invoice}

        order_keys = []
        seen = set()
        for r1 in r1_list:
            ok = amood_ws_cell(state.ws1, AMOOD_COL1_ORDER_KEY, r1).value
            ok = amood_norm_key(ok)
            if ok and ok not in seen:
                seen.add(ok)
                order_keys.append(ok)

        if not order_keys:
            return {"ok": False, "type": "invoice", "result": "NO_ORDER_KEY", "invoice": invoice}

        pending: list[dict] = []
        for order_key in order_keys:
            rows2 = amood_collect_rows_by_value(state.ws2, AMOOD_COL2_ORDER_KEY, order_key, start_row=2)
            last_name = None
            for r in rows2:
                qty_raw = amood_ws_cell(state.ws2, AMOOD_COL2_QTY, r).value
                qty = amood_to_int_qty(qty_raw)
                bc = amood_ws_cell(state.ws2, AMOOD_COL2_BARCODE, r).value
                bc = str(bc).strip() if bc is not None else ""
                name_raw = amood_ws_cell(state.ws2, AMOOD_COL2_NAME, r).value
                name_str = str(name_raw).strip() if name_raw is not None else ""
                # 합배송: 이름이 비어있으면(병합셀) 이전 행 이름으로 채움
                if not name_str and last_name:
                    name_str = last_name
                elif name_str:
                    last_name = name_str
                option = amood_ws_cell(state.ws2, AMOOD_COL2_OPTION, r).value
                # 합배송: qty 셀이 None(병합셀)이고 바코드가 있으면 1로 설정
                if qty_raw is None and bc:
                    qty = 1
                disp = amood_ws_cell(state.ws2, AMOOD_COL2_OUTPUT, r).value
                if disp is None or str(disp).strip() == "":
                    disp = amood_build_output_text(name_str, option, qty)
                pending.append(
                    {
                        "row": r,
                        "barcode": bc,
                        "name": name_str,
                        "option": str(option).strip() if option is not None else "",
                        "display": str(disp).strip() if disp is not None else "",
                        "remaining": qty if qty > 0 else 0,
                    }
                )

        if not pending:
            return {"ok": False, "type": "invoice", "result": "NO_ITEMS", "invoice": invoice}

        state.current_invoice = invoice
        state.pending_items = pending
        all_done = all(it.get("remaining", 0) <= 0 for it in pending)
        state.waiting_for_items = not all_done

        return {
            "ok": True,
            "type": "invoice",
            "result": "SET",
            "invoice": invoice,
            "invoice_done": all_done,
            "items": _amood_items_view(state),
            "current_next": _amood_first_remaining(state),
            "invoice_has_defect": _amood_invoice_has_defect(state),
        }

    @router.post("/amood/scan/item")
    def amood_scan_item(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        _sync_shared_incoming(state)
        if not state.waiting_for_items or not state.pending_items:
            return {"ok": False, "type": "item", "result": "NO_INVOICE"}

        raw = (payload.get("code") or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="code 값이 비어있음")

        scan = amood_norm_barcode(raw)
        matched = None
        for it in state.pending_items:
            if it.get("remaining", 0) <= 0:
                continue
            target = amood_norm_barcode(it.get("barcode", ""))
            if target and (target == scan or scan in target or target in scan):
                matched = it
                break

        if matched is None:
            return {
                "ok": True,
                "type": "item",
                "result": "FALSE",
                "code": raw,
                "remain": 0,
                "items": _amood_items_view(state),
            }

        matched["remaining"] = int(matched.get("remaining", 0)) - 1
        item_has_defect = _amood_item_has_defect(matched)
        try:
            amood_ws_cell(state.ws2, AMOOD_COL2_QTY, matched["row"]).value = matched["remaining"]
        except Exception as e:
            traceback.print_exc()
            print(f"[amood] 셀 쓰기 실패 (row={matched['row']}): {e}")

        all_done = all(it.get("remaining", 0) <= 0 for it in state.pending_items)
        if all_done:
            state.waiting_for_items = False

        return {
            "ok": True,
            "type": "item",
            "result": "TRUE",
            "code": matched.get("barcode", "") or raw,
            "remain": matched.get("remaining", 0),
            "invoice_done": all_done,
            "items": _amood_items_view(state),
            "current_next": _amood_first_remaining(state),
            "item_has_defect": item_has_defect,
            "invoice_has_defect": _amood_invoice_has_defect(state),
        }

    @router.post("/amood/export-shipping")
    def amood_export_shipping(format: str = "xlsx", user: str = Depends(get_current_user)):
        state = get_amood_state(user)
        if not state.file1_path or not state.file2_path:
            raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
        try:
            amood_load_workbooks(state)
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="엑셀 로드 실패")
        wb1, ws1 = state.wb1, state.ws1
        wb2, ws2 = state.wb2, state.ws2

        amood_fill_down_merged_column(ws2, AMOOD_COL2_ORDER_KEY, start_row=2)

        rows: list[dict] = []
        seen_codes: set[str] = set()
        for r in range(2, ws1.max_row + 1):
            order_key = amood_ws_cell(ws1, AMOOD_COL1_ORDER_KEY, r).value
            if not order_key:
                continue
            order_key = str(order_key).strip()
            matched_rows = amood_collect_rows_by_value(ws2, AMOOD_COL2_ORDER_KEY, order_key, start_row=2)
            if not matched_rows:
                continue
            outputs: list[str] = []
            for r2 in matched_rows:
                out_val = amood_ws_cell(ws2, AMOOD_COL2_OUTPUT, r2).value
                if out_val is None or str(out_val).strip() == "":
                    name = amood_ws_cell(ws2, AMOOD_COL2_NAME, r2).value
                    option = amood_ws_cell(ws2, AMOOD_COL2_OPTION, r2).value
                    qty = amood_ws_cell(ws2, AMOOD_COL2_QTY, r2).value
                    out_val = amood_build_output_text(name, option, qty)
                out_text = str(out_val).strip() if out_val is not None else ""
                if out_text and out_text not in outputs:
                    outputs.append(out_text)
            description = " / ".join(outputs)
            code = amood_ws_cell(ws1, AMOOD_COL1_SCAN_BARCODE, r).value
            code = str(code).strip() if code else ""
            if not code:
                continue
            if code in seen_codes:
                continue
            seen_codes.add(code)
            b_val = amood_ws_cell(ws1, AMOOD_COL1_NUM_B, r).value
            c_val = amood_ws_cell(ws1, AMOOD_COL1_NUM_C, r).value
            try:
                b_val = int(b_val)
                c_val = int(c_val)
            except Exception:
                continue
            title = f"{c_val}-{b_val}"
            rows.append({"Title": title, "Description": description, "Code": code})

        if not rows:
            raise HTTPException(status_code=400, detail="추출할 데이터가 없습니다.")

        rows.sort(key=lambda row: str(row.get("Description", "")), reverse=True)
        if str(format or "").lower() == "json":
            return {
                "ok": True,
                "rows": rows,
                "filename": "선적바코드_추출.xlsx",
            }
        df = pd.DataFrame(rows, columns=["Title", "Description", "Code"])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        buf.seek(0)
        filename = "선적바코드_추출.xlsx"
        headers = {"Content-Disposition": content_disposition(filename)}
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @router.post("/amood/load-from-pastelco")
    async def amood_load_from_pastelco(user: str = Depends(get_current_user)):
        try:
            token = await pastelco_login()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Pastelco 로그인 실패: {e}")

        try:
            items = await pastelco_fetch_all_orders(token)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"주문 조회 실패: {e}")

        state = load_pastelco_items_into_state(user, items)

        return {"ok": True, "count": len(items), "status": amood_status(state)}

    @router.post("/amood/load-from-pastelco-shipping-processing-today")
    async def amood_load_from_pastelco_shipping_processing_today(user: str = Depends(get_current_user)):
        today = pastelco_today_kst()
        try:
            token = await pastelco_login()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Pastelco 로그인 실패: {e}")

        try:
            items = await pastelco_fetch_shipping_processing_today(token, today)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"주문 조회 실패: {e}")

        state = load_pastelco_items_into_state(user, items)

        return {
            "ok": True,
            "count": len(items),
            "date": today,
            "status": "SHIPPING_PROCESSING",
            "amood_status": amood_status(state),
        }

    return router
