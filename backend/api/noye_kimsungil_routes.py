from __future__ import annotations

import asyncio
import json
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
import io
import warnings
warnings.filterwarnings("ignore", message="Unverified HTTPS request")
import os
import re
import shutil
import tempfile
import textwrap
import urllib.parse
import uuid
import zipfile

import httpx
import openpyxl
import pandas as pd
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, Response
import xlwt

from api.wonbe_routes import _get_wonbe_db, _init_ably_stock_table

ABLY_BASE     = "https://api.a-bly.com"
ABLY_EMAIL    = "eostm1997@naver.com"
ABLY_PASSWORD = "!Glqgkqdldi1126"


async def _ably_login_noye() -> str:
    async with httpx.AsyncClient(timeout=15.0) as client:
        res = await client.post(
            f"{ABLY_BASE}/seller/login/",
            json={"email": ABLY_EMAIL, "password": ABLY_PASSWORD},
            headers={
                "Content-Type": "application/json",
                "Origin": "https://seller-admin.a-bly.com",
                "Referer": "https://seller-admin.a-bly.com/",
                "User-Agent": "Mozilla/5.0",
            },
        )
        res.raise_for_status()
    token = res.json().get("access_token") or res.json().get("token")
    if not token:
        raise HTTPException(status_code=502, detail="에이블리 로그인 실패")
    return token
import sqlite3

try:
    from PIL import Image as _PILImage, ImageDraw as _PILDraw, ImageFont as _PILFont
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

# ── 불량출력 세션 저장소 (메모리) ──────────────────────────────────────────
_bulyang_sessions: dict[str, dict] = {}


def _bul_mm_to_px(mm: float, dpi: int) -> int:
    return max(1, int(round(mm / 25.4 * dpi)))


def _bul_nz(x, default: str = "") -> str:
    if x is None:
        return default
    try:
        if pd.isna(x):
            return default
    except Exception:
        pass
    return str(x).strip()


def _bul_to_int(x, default: int = 1) -> int:
    try:
        if x in ("", None):
            return default
        return int(float(x))
    except Exception:
        return default


def _bul_find_font() -> str | None:
    for p in [
        r"C:\Windows\Fonts\malgunbd.ttf",
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\gulim.ttc",
        "/System/Library/Fonts/AppleSDGothicNeo.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    ]:
        if os.path.exists(p):
            return p
    return None


_BUL_FONT_PATH = _bul_find_font()


def _bul_load_font(size_px: int):
    if not _PIL_OK:
        return None
    try:
        if _BUL_FONT_PATH:
            return _PILFont.truetype(_BUL_FONT_PATH, size=size_px)
    except Exception:
        pass
    return _PILFont.load_default()


def _bul_build_groups(df: pd.DataFrame) -> tuple[list, dict]:
    raw: dict = {}
    for _, r in df.iterrows():
        vendor = _bul_nz(r.iloc[0])
        name   = _bul_nz(r.iloc[1])
        color  = _bul_nz(r.iloc[2])
        qty    = _bul_to_int(r.iloc[3], 1)
        addr   = _bul_nz(r.iloc[4])
        if not vendor and not addr:
            continue
        raw.setdefault((vendor, addr), []).append({"name": name, "color": color, "qty": qty})

    merged: dict = {}
    for key, items in raw.items():
        combo: dict = {}
        for it in items:
            k = (_bul_nz(it["name"]), _bul_nz(it["color"]))
            combo[k] = combo.get(k, 0) + _bul_to_int(it["qty"], 1)
        out = [{"name": k[0], "color": k[1], "qty": v} for k, v in combo.items()]
        out.sort(key=lambda x: (x["name"], x["color"]))
        merged[key] = out

    keys = list(merged.keys())
    return keys, merged


def _bul_render(
    title_text: str,
    vendor_name: str,
    vendor_addr: str,
    items: list[dict],
    *,
    footer_text: str = "",
    dpi: int = 150,
    # 페이지
    page_w_mm: float = 126.0,
    page_h_mm: float = 103.0,
    # 여백/타이틀
    top_mm: float = 4.5,
    side_mm: float = 15.0,
    title_size_mm: float = 14.0,
    title_gap_mm: float = 2.5,
    title_line_thick_mm: float = 0.5,
    # 주소/거래처명
    addr_v_size_mm: float = 7.0,
    vname_v_size_mm: float = 7.0,
    addr_wrap_chars: int = 28,
    info_gap_below_line_mm: float = 6.0,
    info_extra_gap_mm: float = 0.0,
    # 표/행/폰트
    show_table_header: bool = False,
    header_size_mm: float = 8.2,
    row_size_mm: float = 4.0,
    row_h_mm: float = 6.2,
    row_padding_mm: float = 6.8,
    row_line_factor: float = 0.0,
    two_row_extra_gap_mm: float = 0.0,
    col_gap_mm: float = 3.0,
    # 열 비율
    name_ratio: float = 0.62,
    color_ratio: float = 0.23,
    qty_ratio: float = 0.15,
    # 표 시작 오프셋
    table_top_offset_mm: float = 19.76,
    # 하단 각주
    bottom_mm: float = 10.0,
    footer_size_mm: float = 6.0,
    footer_show_date: bool = True,
    footer_date_format: str = "%Y-%m-%d",
) -> tuple[bytes, dict]:
    if not _PIL_OK:
        raise RuntimeError("Pillow가 설치되지 않았습니다.")

    W = _bul_mm_to_px(page_w_mm, dpi)
    H = _bul_mm_to_px(page_h_mm, dpi)
    left = right = _bul_mm_to_px(side_mm, dpi)
    top = _bul_mm_to_px(top_mm, dpi)

    title_font  = _bul_load_font(_bul_mm_to_px(title_size_mm, dpi))
    header_font = _bul_load_font(_bul_mm_to_px(header_size_mm, dpi))
    row_font    = _bul_load_font(_bul_mm_to_px(row_size_mm, dpi))
    addr_font   = _bul_load_font(_bul_mm_to_px(addr_v_size_mm, dpi))
    vname_font  = _bul_load_font(_bul_mm_to_px(vname_v_size_mm, dpi))
    footer_font = _bul_load_font(_bul_mm_to_px(footer_size_mm, dpi))

    img  = _PILImage.new("RGB", (W, H), "white")
    draw = _PILDraw.Draw(img)

    # 제목
    tb = draw.textbbox((0, 0), title_text, font=title_font)
    tw, th = tb[2] - tb[0], tb[3] - tb[1]
    draw.text(((W - tw) // 2, top), title_text, fill="black", font=title_font)

    # 제목 밑 선
    line_y = top + th + _bul_mm_to_px(title_gap_mm, dpi)
    draw.line((left, line_y, W - right, line_y), fill="black",
              width=max(1, _bul_mm_to_px(title_line_thick_mm, dpi)))

    # 주소 / 거래처명
    y = line_y + _bul_mm_to_px(info_gap_below_line_mm + info_extra_gap_mm, dpi)
    wc = max(10, int(addr_wrap_chars * (W / 476)))
    draw.text((left, y), textwrap.fill(_bul_nz(vendor_addr), width=wc),
              fill="black", font=addr_font)
    vb = draw.textbbox((0, 0), vendor_name, font=vname_font)
    draw.text((W - right - (vb[2] - vb[0]), y), vendor_name, fill="black", font=vname_font)

    # 열 배치
    table_start_y = y + _bul_mm_to_px(table_top_offset_mm, dpi)
    cur_y = table_start_y
    usable_w = W - left - right
    total_r = max(1e-6, name_ratio + color_ratio + qty_ratio)
    nr, cr = name_ratio / total_r, color_ratio / total_r
    gap_px = _bul_mm_to_px(col_gap_mm, dpi)
    name_w  = int(usable_w * nr)
    color_w = int(usable_w * cr)
    qty_w   = usable_w - name_w - color_w - 2 * gap_px
    x_name  = left
    x_color = x_name + name_w + gap_px
    x_qty   = x_color + color_w + gap_px

    # 행 스텝
    rf_a, rf_d = row_font.getmetrics()
    row_font_h = rf_a + rf_d
    row_step = max(
        _bul_mm_to_px(row_h_mm, dpi),
        int(row_font_h * row_line_factor) + _bul_mm_to_px(row_padding_mm, dpi),
    )
    hf_a, hf_d = header_font.getmetrics()
    header_font_h = hf_a + hf_d
    header_step = max(
        _bul_mm_to_px(row_h_mm, dpi),
        int(header_font_h * row_line_factor) + _bul_mm_to_px(row_padding_mm, dpi),
    )

    # 헤더
    if show_table_header:
        draw.text((x_name,  cur_y), "거래처 상품명", font=header_font, fill="black")
        draw.text((x_color, cur_y), "색상",         font=header_font, fill="black")
        draw.text((x_qty,   cur_y), "수량",         font=header_font, fill="black")
        cur_y += header_step

    # 행들
    for idx, it in enumerate(items):
        name  = _bul_nz(it.get("name"))
        color = _bul_nz(it.get("color"))
        qty   = str(it.get("qty", 0))

        nr2 = name
        while True:
            nb = draw.textbbox((0, 0), nr2, font=row_font)
            if nb[2] - nb[0] <= name_w or len(nr2) <= 1:
                break
            nr2 = nr2[:-2] + "…"

        draw.text((x_name,  cur_y), nr2,   fill="black", font=row_font)
        draw.text((x_color, cur_y), color, fill="black", font=row_font)
        qb = draw.textbbox((0, 0), qty, font=row_font)
        draw.text((x_qty + max(0, qty_w - (qb[2] - qb[0])), cur_y), qty,
                  fill="black", font=row_font)
        cur_y += row_step
        if len(items) == 2 and idx == 0 and two_row_extra_gap_mm > 0:
            cur_y += _bul_mm_to_px(two_row_extra_gap_mm, dpi)

    # 하단 각주
    fa, fd = footer_font.getmetrics()
    y_foot = H - _bul_mm_to_px(bottom_mm, dpi) - (fa + fd)
    if footer_show_date:
        draw.text((left, y_foot), datetime.now().strftime(footer_date_format or "%Y-%m-%d"),
                  fill="black", font=footer_font)
    if footer_text:
        max_fw = W - left - right
        ft = footer_text
        while ft and draw.textbbox((0, 0), ft, font=footer_font)[2] > max_fw:
            if len(ft) <= 1:
                break
            ft = ft[:-2] + "…"
        ftw = draw.textbbox((0, 0), ft, font=footer_font)[2]
        draw.text((W - right - ftw, y_foot), ft, fill="black", font=footer_font)

    buf = io.BytesIO()
    img.save(buf, "PNG")
    guides = {
        "x_color": x_color,
        "x_qty": x_qty,
        "table_y": table_start_y,
        "img_w": W,
        "img_h": H,
    }
    return buf.getvalue(), guides


def _bul_sanitize(name: str) -> str:
    out = "".join(c if c.isalnum() or c in " ._-" else "_" for c in (name or "거래처"))
    return out.strip()[:80] or "거래처"


_EZADMIN_BASE        = "https://ga80.ezadmin.co.kr"
_EZADMIN_SESSION_KEY = "ezadmin_phpsessid"
_TODAY_SUPPLY_CODE   = "20002"
_KST = timezone(timedelta(hours=9))
_BROWSER_WEEKDAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_BROWSER_MONTHS   = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _kdg_browser_time_flag(now: datetime) -> str:
    return (
        f"{_BROWSER_WEEKDAYS[now.weekday()]} "
        f"{_BROWSER_MONTHS[now.month - 1]} "
        f"{now.day:02d} {now.year} "
        f"{now:%H:%M:%S} GMT+0900 (한국 표준시)"
    )


def _kdg_looks_like_session_error(response: httpx.Response, body: str) -> bool:
    lowered = (body or "").lower()
    if response.url and "login" in str(response.url).lower():
        return True
    if "<html" in lowered or "<!doctype html" in lowered:
        return True
    return any(token in lowered for token in ("login", "phpsessid", "session", "로그인"))


def _kdg_looks_like_login_page(response: httpx.Response, body: str) -> bool:
    lowered = (body or "").lower()
    if response.url and "login" in str(response.url).lower():
        return True
    return "login.htm" in lowered or "login_form" in lowered


async def _kdg_find_sheet_seq(
    client: httpx.AsyncClient,
    *,
    phpsessid: str,
    start_date: str,
    sheet_title: str,
) -> str | None:
    response = await client.post(
        f"{_EZADMIN_BASE}/function.htm",
        data={
            "_search": "false",
            "nd": str(int(datetime.now(_KST).timestamp() * 1000)),
            "rows": "9999",
            "page": "1",
            "sidx": "",
            "sord": "asc",
            "template": "IM00",
            "action": "get_IM00_grid",
            "par": (
                "template=IM00&action=&page_code=IM00&search=1"
                "&_sort=&sort_order=&date_type=crdate"
                f"&start_date={start_date}&end_date={start_date}"
                "&date_period_sel=0&query_option=title&query_str=&req_status=0"
            ),
        },
        cookies={"PHPSESSID": phpsessid},
        headers={"User-Agent": "Mozilla/5.0", "X-Requested-With": "XMLHttpRequest"},
    )
    body = response.text or ""
    if _kdg_looks_like_session_error(response, body):
        return None
    try:
        obj = response.json()
    except Exception:
        return None

    html_tag = re.compile(r"<[^>]+>")
    matches: list[str] = []
    for row in obj.get("rows", []):
        cell = row.get("cell", {}) or {}
        clean = {
            key: html_tag.sub("", str(value)).strip() if isinstance(value, str) else value
            for key, value in cell.items()
        }
        title = str(clean.get("title") or clean.get("sheet_title") or "").strip()
        if title and title != sheet_title:
            continue
        sheet_no = str(cell.get("sheet") or clean.get("sheet") or "").strip()
        for value in cell.values():
            if sheet_no:
                break
            if isinstance(value, str):
                match = re.search(r"sheet=['\"]?(\w+)['\"]?", value, re.IGNORECASE)
                if match:
                    sheet_no = match.group(1)
        if sheet_no:
            matches.append(sheet_no)

    def _sort_key(v: str):
        try:
            return int(v)
        except Exception:
            return 0

    return sorted(matches, key=_sort_key)[-1] if matches else None


def build_noye_kimsungil_router(*, get_current_user, get_setting, set_setting, get_db):
    router = APIRouter(prefix="/noye-kimsungil")

    KDG_DB_PATH = Path(r"C:\Users\ksh29\OneDrive\Desktop\원베\케이디지원가베이스.db")
    KDG_ALLOWED_BASE = {".xlsx", ".xls", ".xlsm"}

    def _get_kdg_db():
        KDG_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(str(KDG_DB_PATH))
        conn.row_factory = sqlite3.Row
        return conn

    def _init_kdg_table(conn):
        conn.execute("""
            CREATE TABLE IF NOT EXISTS kdg (
                변환품명 TEXT PRIMARY KEY,
                상품코드 TEXT
            )
        """)
        conn.commit()

    def _init_purchase_manager_table(conn):
        conn.execute("""
            CREATE TABLE IF NOT EXISTS noye_purchase_manager_state (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                state_json TEXT NOT NULL,
                updated_by TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

    NUM_LINE_RE = re.compile(r"^\s*([\d,]+)\s+([\d,]+)\s+([\d,]+)\s*$")
    ITEM_RE = re.compile(r"^\s*(?:[A-Za-z]\.)*([A-Za-z])\.(\d+)-(\d+)(SH|LO)?/([^/]+)/([^/\s]+)\s*$")

    def _content_disposition(filename: str) -> str:
        safe_name = (filename or "download").replace('"', "")
        ascii_name = "".join(ch if ord(ch) < 128 else "_" for ch in safe_name)
        ascii_name = re.sub(r"_+", "_", ascii_name).strip("_") or "download"
        quoted = urllib.parse.quote(safe_name)
        return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted}"

    def _safe_str(v) -> str:
        if v is None:
            return ""
        if pd.isna(v):
            return ""
        return str(v).strip()

    def _suffix_to_kor(sfx: str | None) -> str:
        if sfx == "SH":
            return "숏"
        if sfx == "LO":
            return "롱"
        return "기본"

    def _parse_kdg_text(raw: str) -> list[dict]:
        lines = [ln.rstrip() for ln in (raw or "").splitlines() if ln.strip()]
        rows = []
        i = 0
        while i < len(lines):
            ln = lines[i].strip()
            if ln.startswith("품명"):
                i += 1
                continue

            m_item = ITEM_RE.match(ln)
            if not m_item:
                i += 1
                continue

            _, model_no, _opt_no, sfx, color, size = m_item.groups()
            suffix = _suffix_to_kor(sfx)
            a_value = f"{model_no} {color} {size} {suffix}"
            qty = ""
            amount = ""
            if i + 1 < len(lines):
                m_num = NUM_LINE_RE.match(lines[i + 1].strip())
                if m_num:
                    _unit_price, qty, amount = m_num.groups()
                    qty = qty.replace(",", "")
                    amount = amount.replace(",", "")
                    i += 2
                else:
                    i += 1
            else:
                i += 1

            rows.append(
                {
                    "A(변환품명)": a_value,
                    "B(옵션번호)": qty,
                    "원본품명": ln,
                    "모델번호": model_no,
                    "색상": color,
                    "사이즈": size,
                    "기장": suffix,
                    "수량": int(qty) if _safe_str(qty) else 0,
                    "금액": int(amount) if _safe_str(amount) else 0,
                    "원베_B": "",
                }
            )

        if not rows:
            raise ValueError("변환할 품명 라인을 찾지 못했음.")
        return rows

    def _read_base_df(path: Path) -> pd.DataFrame:
        ext = path.suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            return pd.read_excel(path, dtype=str, engine="openpyxl")
        if ext == ".xls":
            try:
                return pd.read_excel(path, dtype=str, engine="xlrd")
            except Exception:
                return pd.read_excel(path, dtype=str)
        return pd.read_excel(path, dtype=str)

    def _load_kdg_base_map() -> dict[str, str]:
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            rows = conn.execute("SELECT 변환품명, 상품코드 FROM kdg").fetchall()
            return {str(r["변환품명"] or ""): str(r["상품코드"] or "") for r in rows}
        finally:
            conn.close()

    def _match_kdg_rows(rows: list[dict], base_map: dict[str, str]) -> tuple[list[dict], int]:
        out = []
        missing = 0
        for row in rows:
            key = _safe_str(row.get("A(변환품명)"))
            code = base_map.get(key, "")
            merged = {**row, "원베_B": code}
            if _safe_str(code) == "":
                missing += 1
            out.append(merged)
        return out, missing

    def _read_excel_any(path: Path) -> pd.DataFrame:
        ext = path.suffix.lower()
        if ext in (".xlsx", ".xlsm"):
            try:
                return pd.read_excel(path, engine="openpyxl")
            except Exception:
                return pd.read_excel(path)
        if ext == ".xls":
            try:
                return pd.read_excel(path, engine="xlrd")
            except Exception:
                return pd.read_excel(path)
        return pd.read_excel(path)

    def _left_first_word(text: str) -> str:
        s = _safe_str(text)
        return s.split(" ", 1)[0].strip() if s else ""

    def _mid_after_first_space(text: str) -> str:
        s = _safe_str(text)
        if not s:
            return ""
        parts = s.split(" ", 1)
        return parts[1].strip() if len(parts) == 2 else ""

    def _extract_bracket_inside(text: str) -> str | None:
        s = _safe_str(text)
        m = re.search(r"\[([^\]]+)\]", s)
        return m.group(1).strip() if m else None

    def _d_from_f(f_text: str) -> str:
        inside = _extract_bracket_inside(f_text)
        if not inside:
            return ""
        return inside.split("-", 1)[0].strip() if "-" in inside else inside.strip()

    def _e_from_f(f_text: str) -> str:
        inside = _extract_bracket_inside(f_text)
        if not inside:
            return "free"
        parts = [p.strip() for p in inside.split("-")]
        if len(parts) <= 1:
            return "free"
        merged = " ".join([p for p in parts[1:] if p]).strip()
        return merged if merged else "free"

    def _ensure_min_columns(df: pd.DataFrame, min_cols: int) -> pd.DataFrame:
        if df.shape[1] >= min_cols:
            return df
        for _ in range(min_cols - df.shape[1]):
            df[f"__EMPTY_{df.shape[1] + 1}__"] = ""
        return df

    def _build_date_chunk_output_df(path: Path) -> pd.DataFrame:
        df = _read_excel_any(path)
        df = _ensure_min_columns(df, 7)
        col_a = df.columns[0]
        col_b = df.columns[1]
        col_d = df.columns[3]
        col_e = df.columns[4]
        col_f = df.columns[5]
        col_g = df.columns[6]

        b_src = df[col_b]
        e_src = df[col_e]
        f_src = df[col_f]
        g_src = df[col_g]

        today_str = date.today().strftime("%Y-%m-%d")
        df[col_a] = b_src.map(_left_first_word)
        df[col_b] = b_src.map(_mid_after_first_space)
        df[col_d] = f_src.map(_d_from_f)
        df[col_e] = f_src.map(_e_from_f)
        df[col_f] = pd.to_numeric(g_src, errors="coerce")
        df[col_g] = today_str

        df_out = df.iloc[:, :7].copy()
        if len(df_out) > 0:
            df_out = df_out.iloc[:-1]
        return df_out

    @router.get("/kdg/base/status")
    def kdg_base_status(user: str = Depends(get_current_user)):
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            total = conn.execute("SELECT COUNT(*) FROM kdg").fetchone()[0]
            return {"ok": True, "exists": total > 0, "total": total}
        finally:
            conn.close()

    @router.get("/kdg/base/preview")
    def kdg_base_preview(
        offset: int = 0,
        limit: int = 50,
        q: str | None = None,
        user: str = Depends(get_current_user),
    ):
        if offset < 0 or limit <= 0 or limit > 200:
            raise HTTPException(status_code=400, detail="offset/limit 값이 올바르지 않습니다.")
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            q_norm = _safe_str(q)
            if q_norm:
                like = f"%{q_norm}%"
                rows = conn.execute(
                    "SELECT rowid, 변환품명, 상품코드 FROM kdg WHERE 변환품명 LIKE ? OR 상품코드 LIKE ? LIMIT ? OFFSET ?",
                    (like, like, limit, offset),
                ).fetchall()
                total = conn.execute(
                    "SELECT COUNT(*) FROM kdg WHERE 변환품명 LIKE ? OR 상품코드 LIKE ?",
                    (like, like),
                ).fetchone()[0]
            else:
                rows = conn.execute(
                    "SELECT rowid, 변환품명, 상품코드 FROM kdg ORDER BY rowid LIMIT ? OFFSET ?",
                    (limit, offset),
                ).fetchall()
                total = conn.execute("SELECT COUNT(*) FROM kdg").fetchone()[0]
            return {
                "ok": True,
                "columns": ["변환품명", "상품코드"],
                "rows": [{"row_index": r["rowid"], "values": [r["변환품명"] or "", r["상품코드"] or ""]} for r in rows],
                "total": total,
            }
        finally:
            conn.close()

    @router.post("/kdg/base/upload")
    async def kdg_base_upload(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in KDG_ALLOWED_BASE:
            raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능합니다.")
        data = await file.read()
        if not data:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
        tmp = Path(tempfile.gettempdir()) / f"kdg_base_{uuid.uuid4().hex}{ext}"
        tmp.write_bytes(data)
        try:
            df = _read_base_df(tmp)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"파일 읽기 실패: {e}")
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass
        if df.shape[1] < 2:
            raise HTTPException(status_code=400, detail="원가베이스는 최소 A,B열이 필요합니다.")
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            conn.execute("DELETE FROM kdg")
            pairs = [
                (_safe_str(r.iloc[0]), _safe_str(r.iloc[1]))
                for _, r in df.iterrows()
                if _safe_str(r.iloc[0])
            ]
            conn.executemany("INSERT OR REPLACE INTO kdg (변환품명, 상품코드) VALUES (?, ?)", pairs)
            conn.commit()
            total = conn.execute("SELECT COUNT(*) FROM kdg").fetchone()[0]
            return {"ok": True, "total": total}
        finally:
            conn.close()

    @router.post("/kdg/base/edit-batch")
    def kdg_base_edit_batch(payload: dict = Body(...), user: str = Depends(get_current_user)):
        edits = payload.get("edits")
        if not isinstance(edits, list) or not edits:
            raise HTTPException(status_code=400, detail="edits 값이 올바르지 않습니다.")
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            for item in edits:
                row_index = item.get("row_index")
                column = item.get("column")
                value = _safe_str(item.get("value"))
                if row_index is None:
                    continue
                if column == 0 or column == "변환품명":
                    conn.execute("UPDATE kdg SET 변환품명=? WHERE rowid=?", (value, row_index))
                elif column == 1 or column == "상품코드":
                    conn.execute("UPDATE kdg SET 상품코드=? WHERE rowid=?", (value, row_index))
            conn.commit()
            return {"ok": True}
        finally:
            conn.close()

    @router.post("/kdg/base/append-tsv")
    def kdg_base_append_tsv(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw_text = _safe_str(payload.get("text"))
        skip_header = bool(payload.get("skip_header"))
        lines = [ln for ln in raw_text.splitlines() if ln.strip()]
        if skip_header and lines:
            lines = lines[1:]
        pairs = []
        for ln in lines:
            parts = ln.split("\t", 1)
            key = parts[0].strip() if parts else ""
            val = parts[1].strip() if len(parts) > 1 else ""
            if key:
                pairs.append((key, val))
        if not pairs:
            raise HTTPException(status_code=400, detail="추가할 데이터가 없습니다.")
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            conn.executemany("INSERT OR REPLACE INTO kdg (변환품명, 상품코드) VALUES (?, ?)", pairs)
            conn.commit()
            total = conn.execute("SELECT COUNT(*) FROM kdg").fetchone()[0]
            return {"ok": True, "appended": len(pairs), "total": total}
        finally:
            conn.close()

    @router.get("/kdg/base/download")
    def kdg_base_download(user: str = Depends(get_current_user)):
        conn = _get_kdg_db()
        try:
            _init_kdg_table(conn)
            rows = conn.execute("SELECT 변환품명, 상품코드 FROM kdg ORDER BY rowid").fetchall()
        finally:
            conn.close()
        book = xlwt.Workbook()
        sheet = book.add_sheet("Sheet1")
        sheet.write(0, 0, "변환품명")
        sheet.write(0, 1, "상품코드")
        for i, row in enumerate(rows, start=1):
            sheet.write(i, 0, row["변환품명"] or "")
            sheet.write(i, 1, row["상품코드"] or "")
        buf = io.BytesIO()
        book.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers={"Content-Disposition": _content_disposition("케이디지원가베이스.xls")},
        )

    @router.post("/kdg/convert")
    def kdg_convert(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw = _safe_str(payload.get("text"))
        rows = _parse_kdg_text(raw)
        return {"ok": True, "count": len(rows), "rows": rows}

    @router.post("/kdg/match")
    def kdg_match(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw = _safe_str(payload.get("text"))
        rows = _parse_kdg_text(raw)
        base_map = _load_kdg_base_map()
        matched, missing = _match_kdg_rows(rows, base_map)
        return {"ok": True, "count": len(matched), "missing": missing, "rows": matched}

    @router.post("/kdg/export-xls")
    def kdg_export_xls(payload: dict = Body(...), user: str = Depends(get_current_user)):
        pre_rows = payload.get("rows")
        if pre_rows and isinstance(pre_rows, list):
            rows = pre_rows
        else:
            raw = _safe_str(payload.get("text"))
            rows = _parse_kdg_text(raw)
        if bool(payload.get("with_match", True)) and not (pre_rows and isinstance(pre_rows, list)):
            base_map = _load_kdg_base_map()
            rows, _ = _match_kdg_rows(rows, base_map)

        book = xlwt.Workbook()
        sheet = book.add_sheet("입고업로드")
        headers_row = ["상품코드", "요청수량", "입고수량"]
        for idx, h in enumerate(headers_row):
            sheet.write(0, idx, h)
        for i, row in enumerate(rows, start=1):
            sheet.write(i, 0, _safe_str(row.get("원베_B")))
            sheet.write(i, 1, "0")
            sheet.write(i, 2, _safe_str(row.get("B(옵션번호)")))

        buf = io.BytesIO()
        book.save(buf)
        headers = {"Content-Disposition": _content_disposition("입고업로드.xls")}
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.ms-excel",
            headers=headers,
        )

    @router.post("/kdg/create-ezadmin-sheet")
    async def kdg_create_ezadmin_sheet(payload: dict = Body(default={}), user: str = Depends(get_current_user)):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        rows = payload.get("rows") or []
        upload_rows = [
            (str(r.get("원베_B") or "").strip(), str(r.get("B(옵션번호)") or "").strip())
            for r in rows
            if str(r.get("원베_B") or "").strip()
        ]
        if not upload_rows:
            return {"ok": False, "error": "상품코드가 있는 행이 없습니다."}

        book = xlwt.Workbook()
        sheet = book.add_sheet("상품일괄추가")
        for ci, h in enumerate(["상품코드", "요청수량", "입고수량"]):
            sheet.write(0, ci, h)
        for ri, (code, qty) in enumerate(upload_rows, start=1):
            sheet.write(ri, 0, code)
            sheet.write(ri, 1, "0")
            sheet.write(ri, 2, qty)
        buf = io.BytesIO()
        book.save(buf)
        xls_bytes = buf.getvalue()

        now = datetime.now(_KST)
        start_date = now.strftime("%Y-%m-%d")
        sheet_title = "KDG입고"
        headers = {
            "User-Agent": "Mozilla/5.0",
            "Referer": f"{_EZADMIN_BASE}/template40.htm?template=IM00",
            "X-Requested-With": "XMLHttpRequest",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        }

        try:
            async with httpx.AsyncClient(timeout=30.0, verify=False, follow_redirects=True) as client:
                response = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "template": "IM00",
                        "action": "new_sheet_each",
                        "start_date": start_date,
                        "sheet_title": sheet_title,
                        "timeFlag": _kdg_browser_time_flag(now),
                    },
                    cookies={"PHPSESSID": phpsessid},
                    headers=headers,
                )
                body = (response.text or "").strip()
                if _kdg_looks_like_session_error(response, body):
                    return {"ok": False, "need_session": True}
                if not (200 <= response.status_code < 300):
                    return {"ok": False, "error": f"EZAdmin 전표 생성 실패 (HTTP {response.status_code})"}
                if body:
                    return {"ok": False, "error": f"전표 생성 응답 오류: {body[:300]}"}

                sheet_seq = None
                for _ in range(5):
                    await asyncio.sleep(0.5)
                    sheet_seq = await _kdg_find_sheet_seq(
                        client,
                        phpsessid=phpsessid,
                        start_date=start_date,
                        sheet_title=sheet_title,
                    )
                    if sheet_seq:
                        break
                if not sheet_seq:
                    return {"ok": False, "error": "전표는 생성됐지만 전표번호를 찾지 못해 상품 일괄추가를 진행하지 못했습니다."}

                upload_response = await client.post(
                    f"{_EZADMIN_BASE}/popup_utf8.htm",
                    data={"template": "IM25", "action": "upload", "seq": sheet_seq},
                    files={"_file": ("kdg_products.xls", xls_bytes, "application/vnd.ms-excel")},
                    cookies={"PHPSESSID": phpsessid},
                    headers={
                        "User-Agent": "Mozilla/5.0",
                        "Referer": f"{_EZADMIN_BASE}/popup35.htm?template=IM25&seq={sheet_seq}",
                    },
                )
        except httpx.HTTPError as exc:
            return {"ok": False, "error": f"EZAdmin 요청 실패: {exc}"}

        upload_body = (upload_response.text or "").strip()
        if _kdg_looks_like_login_page(upload_response, upload_body):
            return {"ok": False, "need_session": True}
        if not (200 <= upload_response.status_code < 300):
            return {"ok": False, "error": f"상품 일괄추가 실패 (HTTP {upload_response.status_code})"}

        return {"ok": True, "sheet_seq": sheet_seq, "uploaded_count": len(upload_rows)}

    @router.post("/kdg/export-date")
    def kdg_export_date(payload: dict = Body(...), user: str = Depends(get_current_user)):
        raw = _safe_str(payload.get("text"))
        rows = _parse_kdg_text(raw)
        df = pd.DataFrame(rows)
        g = (
            df.groupby(["모델번호", "색상", "사이즈", "기장"], as_index=False)[["수량", "금액"]]
            .sum()
        )
        out = pd.DataFrame(
            {
                "A": ["케이디지"] * len(g),
                "B": g["모델번호"].astype(str),
                "C": g["금액"].astype(int).astype(str),
                "D": g["색상"].astype(str),
                "E": (g["사이즈"].astype(str) + " " + g["기장"].astype(str)),
                "F": g["수량"].astype(int).astype(str),
            }
        )

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, header=False, sheet_name="결과")
        buf.seek(0)
        headers = {"Content-Disposition": _content_disposition("날짜별_복사_데이터.xlsx")}
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @router.post("/kdg/date-copy-text")
    def kdg_date_copy_text(payload: dict = Body(...), user: str = Depends(get_current_user)):
        pre_rows = payload.get("rows")
        if pre_rows and isinstance(pre_rows, list):
            rows = pre_rows
        else:
            raw = _safe_str(payload.get("text"))
            rows = _parse_kdg_text(raw)
        df = pd.DataFrame(rows)
        g = (
            df.groupby(["모델번호", "색상", "사이즈", "기장"], as_index=False)[["수량", "금액"]]
            .sum()
        )
        out = pd.DataFrame(
            {
                "A": ["케이디지"] * len(g),
                "B": g["모델번호"].astype(str),
                "C": g["금액"].astype(int).astype(str),
                "D": g["색상"].astype(str),
                "E": (g["사이즈"].astype(str) + " " + g["기장"].astype(str)),
                "F": g["수량"].astype(int).astype(str),
            }
        )
        # Keep a trailing newline so repeated copy/paste blocks append cleanly.
        tsv = out.to_csv(sep="\t", index=False, header=False, lineterminator="\n")
        return {"ok": True, "count": len(out), "text": tsv}

    @router.post("/date-chunk/process")
    async def date_chunk_process(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xls", ".xlsm"}:
            raise HTTPException(status_code=400, detail="xlsx/xls/xlsm만 업로드 가능합니다.")
        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")

        tmp = Path(tempfile.gettempdir()) / f"date_chunk_{uuid.uuid4().hex}{ext}"
        tmp.write_bytes(raw)
        try:
            df_out = _build_date_chunk_output_df(tmp)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"엑셀 읽기 실패: {e}")
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_out.to_excel(writer, sheet_name="Sheet1", index=False)
        buf.seek(0)
        stem = Path(file.filename or "가공대상").stem
        headers = {"Content-Disposition": _content_disposition(f"{stem}_가공.xlsx")}
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    # ── 불량출력 엔드포인트 ────────────────────────────────────────────────

    @router.post("/bulyang/upload")
    async def bulyang_upload(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xls", ".xlsm"}:
            raise HTTPException(status_code=400, detail="xlsx/xls/xlsm만 업로드 가능합니다.")
        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")
        tmp = Path(tempfile.gettempdir()) / f"bulyang_{uuid.uuid4().hex}{ext}"
        tmp.write_bytes(raw)
        try:
            df = _read_excel_any(tmp)
            if df.shape[1] < 5:
                raise ValueError("A~E열(5열)이 필요합니다.")
            df = df.iloc[:, :5]
            keys, groups = _bul_build_groups(df)
            session_id = uuid.uuid4().hex
            _bulyang_sessions[session_id] = {"keys": keys, "groups": groups}
            # 세션 최대 30개 유지
            if len(_bulyang_sessions) > 30:
                oldest = next(iter(_bulyang_sessions))
                del _bulyang_sessions[oldest]
            group_list = [
                {"index": i, "vendor": k[0], "addr": k[1], "item_count": len(groups[k])}
                for i, k in enumerate(keys)
            ]
            return {"ok": True, "session_id": session_id, "total": len(keys), "groups": group_list}
        except Exception as e:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass
            raise HTTPException(status_code=400, detail=str(e))

    @router.post("/bulyang/import")
    def bulyang_import(payload: dict = Body(...), user: str = Depends(get_current_user)):
        rows = payload.get("rows")
        if not isinstance(rows, list) or not rows:
            raise HTTPException(status_code=400, detail="전송할 불량 데이터가 없습니다.")

        normalized_rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            vendor = _bul_nz(row.get("vendor"))
            name = _bul_nz(row.get("name"))
            color = _bul_nz(row.get("color"))
            qty = _bul_to_int(row.get("qty"), 1)
            addr = _bul_nz(row.get("addr"))
            if not vendor and not addr and not name:
                continue
            normalized_rows.append([vendor, name, color, qty, addr])

        if not normalized_rows:
            raise HTTPException(status_code=400, detail="유효한 불량 데이터가 없습니다.")

        try:
            df = pd.DataFrame(normalized_rows, columns=["A", "B", "C", "D", "E"])
            keys, groups = _bul_build_groups(df)
            session_id = uuid.uuid4().hex
            _bulyang_sessions[session_id] = {"keys": keys, "groups": groups}
            if len(_bulyang_sessions) > 30:
                oldest = next(iter(_bulyang_sessions))
                del _bulyang_sessions[oldest]
            group_list = [
                {"index": i, "vendor": k[0], "addr": k[1], "item_count": len(groups[k])}
                for i, k in enumerate(keys)
            ]
            return {"ok": True, "session_id": session_id, "total": len(keys), "groups": group_list}
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))

    @router.get("/bulyang/session/{session_id}")
    def bulyang_session(session_id: str, user: str = Depends(get_current_user)):
        sess = _bulyang_sessions.get(session_id)
        if not sess:
            raise HTTPException(status_code=404, detail="세션이 없습니다. 다시 전송해 주세요.")
        keys = sess["keys"]
        groups = sess["groups"]
        group_list = [
            {"index": i, "vendor": k[0], "addr": k[1], "item_count": len(groups[k])}
            for i, k in enumerate(keys)
        ]
        return {"ok": True, "session_id": session_id, "total": len(keys), "groups": group_list}

    def _bul_layout_kwargs(p: dict) -> dict:
        def f(key, default):
            v = p.get(key, default)
            return v if v is not None else default
        return dict(
            page_w_mm=float(f("page_w_mm", 126.0)),
            page_h_mm=float(f("page_h_mm", 103.0)),
            top_mm=float(f("top_mm", 4.5)),
            side_mm=float(f("side_mm", 15.0)),
            title_size_mm=float(f("title_size_mm", 14.0)),
            title_gap_mm=float(f("title_gap_mm", 2.5)),
            title_line_thick_mm=float(f("title_line_thick_mm", 0.5)),
            addr_v_size_mm=float(f("addr_v_size_mm", 7.0)),
            vname_v_size_mm=float(f("vname_v_size_mm", 7.0)),
            addr_wrap_chars=int(f("addr_wrap_chars", 28)),
            info_gap_below_line_mm=float(f("info_gap_below_line_mm", 6.0)),
            info_extra_gap_mm=float(f("info_extra_gap_mm", 0.0)),
            show_table_header=bool(f("show_table_header", False)),
            header_size_mm=float(f("header_size_mm", 8.2)),
            row_size_mm=float(f("row_size_mm", 4.0)),
            row_h_mm=float(f("row_h_mm", 6.2)),
            row_padding_mm=float(f("row_padding_mm", 6.8)),
            row_line_factor=float(f("row_line_factor", 0.0)),
            two_row_extra_gap_mm=float(f("two_row_extra_gap_mm", 0.0)),
            col_gap_mm=float(f("col_gap_mm", 3.0)),
            name_ratio=float(f("name_ratio", 0.62)),
            color_ratio=float(f("color_ratio", 0.23)),
            qty_ratio=float(f("qty_ratio", 0.15)),
            table_top_offset_mm=float(f("table_top_offset_mm", 19.76)),
            bottom_mm=float(f("bottom_mm", 10.0)),
            footer_size_mm=float(f("footer_size_mm", 6.0)),
            footer_show_date=bool(f("footer_show_date", True)),
            footer_date_format=str(f("footer_date_format", "%Y-%m-%d")),
        )

    @router.post("/bulyang/image/{session_id}/{index}")
    def bulyang_image(
        session_id: str,
        index: int,
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        sess = _bulyang_sessions.get(session_id)
        if not sess:
            raise HTTPException(status_code=404, detail="세션 없음. 파일을 다시 업로드하세요.")
        keys = sess["keys"]
        groups = sess["groups"]
        if index < 0 or index >= len(keys):
            raise HTTPException(status_code=400, detail="인덱스 범위 초과")
        vendor, addr = keys[index]
        items = groups[keys[index]]
        dpi = max(72, min(int(payload.get("dpi", 150)), 300))
        try:
            png, guides = _bul_render(
                title_text=str(payload.get("title") or "벨류스"),
                vendor_name=vendor,
                vendor_addr=addr,
                items=items,
                footer_text=str(payload.get("footer_text") or ""),
                dpi=dpi,
                **_bul_layout_kwargs(payload),
            )
        except RuntimeError as e:
            raise HTTPException(status_code=500, detail=str(e))
        import base64
        return {
            "ok": True,
            "image_b64": base64.b64encode(png).decode(),
            "guides": guides,
        }

    @router.post("/bulyang/export/{session_id}")
    def bulyang_export(
        session_id: str,
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        sess = _bulyang_sessions.get(session_id)
        if not sess:
            raise HTTPException(status_code=404, detail="세션 없음. 파일을 다시 업로드하세요.")
        keys = sess["keys"]
        groups = sess["groups"]
        layout_kw = _bul_layout_kwargs(payload)
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, key in enumerate(keys):
                vendor, addr = key
                items = groups[key]
                try:
                    png, _ = _bul_render(
                        title_text=str(payload.get("title") or "벨류스"),
                        vendor_name=vendor,
                        vendor_addr=addr,
                        items=items,
                        footer_text=str(payload.get("footer_text") or ""),
                        dpi=300,
                        **layout_kw,
                    )
                    zf.writestr(f"{i + 1:03d}_{_bul_sanitize(vendor)}.png", png)
                except Exception:
                    continue
        zip_buf.seek(0)
        headers = {"Content-Disposition": _content_disposition("거래처라벨.zip")}
        return Response(content=zip_buf.getvalue(), media_type="application/zip", headers=headers)

    # ── 날짜청크 엔드포인트 ───────────────────────────────────────────────

    @router.post("/today/reset-stock")
    async def today_reset_stock(payload: dict = Body(...), user: str = Depends(get_current_user)):
        rows = payload.get("rows", [])
        if not rows:
            raise HTTPException(status_code=400, detail="가공된 데이터가 없습니다.")

        # 가공 데이터를 딕셔너리로 변환 (A열 옵션번호 → B열 수량)
        update_map = {str(r["A"]).strip(): str(r["B"]).strip() for r in rows if r.get("A")}

        wonbe_conn = _get_wonbe_db()
        try:
            _init_ably_stock_table(wonbe_conn)
            base_rows = wonbe_conn.execute(
                "SELECT 옵션번호, 수량 FROM 에이블리재고변경 ORDER BY 옵션번호"
            ).fetchall()
        finally:
            wonbe_conn.close()
        if not base_rows:
            raise HTTPException(status_code=404, detail="에이블리재고변경 테이블에 데이터가 없습니다.")

        matched = 0
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = "Sheet1"
        out_ws.cell(1, 1, "에이블리 옵션 번호")
        out_ws.cell(1, 2, "재고 수량")
        for ri, r in enumerate(base_rows, start=2):
            option_no = r["옵션번호"]
            value = update_map.get(option_no, r["수량"])
            if option_no in update_map:
                matched += 1
            out_ws.cell(ri, 1, option_no)
            out_ws.cell(ri, 2, value)

        buf = io.BytesIO()
        out_wb.save(buf)
        buf.seek(0)
        file_bytes = buf.getvalue()

        try:
            token = await _ably_login_noye()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                res = await client.put(
                    f"{ABLY_BASE}/seller/goods/options-bulk-update/stock/",
                    headers={
                        "Authorization": f"JWT {token}",
                        "Origin": "https://seller-admin.a-bly.com",
                        "Referer": "https://seller-admin.a-bly.com/",
                        "User-Agent": "Mozilla/5.0",
                    },
                    files={"file": ("에이블리재고변경.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                )
                res.raise_for_status()
        except HTTPException:
            raise
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"에이블리 재고초기화 실패 (HTTP {e.response.status_code}): {e.response.text[:200]}")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 재고초기화 실패: {e}")

        return {"ok": True, "message": f"오늘출발 재고 초기화 완료 ({matched}건 매칭)"}

    @router.post("/today/set-delivery-type")
    async def today_set_delivery_type(payload: dict = Body(...), user: str = Depends(get_current_user)):
        rows = payload.get("rows", [])
        if not rows:
            raise HTTPException(status_code=400, detail="가공된 데이터가 없습니다.")

        wonbe_conn = _get_wonbe_db()
        try:
            _init_ably_stock_table(wonbe_conn)
            base_snos = {
                r["옵션번호"]
                for r in wonbe_conn.execute("SELECT 옵션번호 FROM 에이블리재고변경").fetchall()
            }
        finally:
            wonbe_conn.close()
        if not base_snos:
            raise HTTPException(status_code=404, detail="에이블리재고변경 테이블에 데이터가 없습니다.")

        # 가공결과 중 기준 파일에 매칭된 A열만 추출
        item_list = [
            str(r["A"]).strip()
            for r in rows
            if r.get("A") and str(r["A"]).strip() in base_snos
        ]

        if not item_list:
            raise HTTPException(status_code=400, detail="기준 파일과 매칭된 옵션번호가 없습니다.")

        try:
            token = await _ably_login_noye()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        try:
            async with httpx.AsyncClient(timeout=60.0) as client:
                res = await client.put(
                    f"{ABLY_BASE}/seller/goods/option-delivery-type-in-bulk/",
                    headers={
                        "Authorization": f"JWT {token}",
                        "Content-Type": "application/json",
                        "Origin": "https://seller-admin.a-bly.com",
                        "Referer": "https://seller-admin.a-bly.com/",
                        "User-Agent": "Mozilla/5.0",
                    },
                    json={
                        "column_to_filter": "sno",
                        "item_list": item_list,
                        "delivery_type": "today",
                    },
                )
                res.raise_for_status()
        except HTTPException:
            raise
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"오출로 변경 실패 (HTTP {e.response.status_code}): {e.response.text[:200]}")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"오출로 변경 실패: {e}")

        return {"ok": True, "message": f"오출로 변경 완료 ({len(item_list)}건)"}

    @router.post("/today/check")
    async def today_check_stock(payload: dict = Body(...), user: str = Depends(get_current_user)):
        rows = payload.get("rows", [])
        if not rows:
            raise HTTPException(status_code=400, detail="가공된 데이터가 없습니다.")

        processed_map: dict[str, int] = {}
        for r in rows:
            sno = str(r.get("A") or "").strip()
            if not sno:
                continue
            try:
                processed_map[sno] = int(float(r.get("B") or 0))
            except (TypeError, ValueError):
                processed_map[sno] = 0

        try:
            token = await _ably_login_noye()
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"에이블리 로그인 실패: {e}")

        ably_by_sno: dict[str, dict] = {}
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                page = 1
                while True:
                    res = await client.get(
                        f"{ABLY_BASE}/seller/today-delivery-goods-options/",
                        headers={
                            "Authorization": f"JWT {token}",
                            "Accept": "application/json",
                            "Origin": "https://my.a-bly.com",
                            "Referer": "https://my.a-bly.com/",
                            "User-Agent": "Mozilla/5.0",
                        },
                        params={"keyword_type": "goods_name", "current_page": page, "per_page": 50},
                    )
                    res.raise_for_status()
                    data = res.json()
                    opts = data.get("data") or []
                    for opt in opts:
                        ably_by_sno[str(opt.get("sno"))] = opt
                    max_page = data.get("max_page_number", 1)
                    if not opts or page >= max_page:
                        break
                    page += 1
        except HTTPException:
            raise
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"오출 목록 조회 실패 (HTTP {e.response.status_code}): {e.response.text[:200]}")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"오출 목록 조회 실패: {e}")

        mismatches = []
        for sno, processed_qty in processed_map.items():
            opt = ably_by_sno.get(sno)
            if opt is None:
                mismatches.append({
                    "sno": sno,
                    "goodsName": "",
                    "optionName": "",
                    "processedQty": processed_qty,
                    "ablyStock": None,
                    "diff": None,
                    "reason": "sno_not_found",
                })
                continue
            ably_stock = int(opt.get("stock") or 0)
            if ably_stock != processed_qty:
                mismatches.append({
                    "sno": sno,
                    "goodsName": opt.get("goods_name", ""),
                    "optionName": opt.get("option_name", ""),
                    "processedQty": processed_qty,
                    "ablyStock": ably_stock,
                    "diff": ably_stock - processed_qty,
                    "reason": "qty_mismatch",
                })

        return {
            "ok": True,
            "checked_at": datetime.now(timezone.utc).isoformat(),
            "processed_count": len(processed_map),
            "ably_count": len(ably_by_sno),
            "mismatches": mismatches,
        }

    @router.post("/date-chunk/copy")
    async def date_chunk_copy(file: UploadFile = File(...), user: str = Depends(get_current_user)):
        ext = Path(file.filename or "").suffix.lower()
        if ext not in {".xlsx", ".xls", ".xlsm"}:
            raise HTTPException(status_code=400, detail="xlsx/xls/xlsm만 업로드 가능합니다.")
        raw = await file.read()
        if not raw:
            raise HTTPException(status_code=400, detail="업로드 파일이 비어 있습니다.")

        tmp = Path(tempfile.gettempdir()) / f"date_chunk_copy_{uuid.uuid4().hex}{ext}"
        tmp.write_bytes(raw)
        try:
            df_out = _build_date_chunk_output_df(tmp)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"엑셀 읽기 실패: {e}")
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

        tsv = df_out.to_csv(sep="\t", index=False, header=False, lineterminator="\n").rstrip("\n")
        return {"ok": True, "rows": len(df_out), "text": tsv}

    @router.post("/today/load-from-ezadmin")
    async def today_load_from_ezadmin(
        payload: dict = Body(default={}),
        user: str = Depends(get_current_user),
    ):
        phpsessid = (get_setting(_EZADMIN_SESSION_KEY) or "").strip()
        if not phpsessid:
            return {"ok": False, "need_session": True}

        cookies = {"PHPSESSID": phpsessid}
        today_str = datetime.now().strftime("%Y-%m-%d")

        try:
            async with httpx.AsyncClient(timeout=120.0, verify=False, follow_redirects=True) as client:
                r = await client.post(
                    f"{_EZADMIN_BASE}/function.htm",
                    data={
                        "_search": "false",
                        "nd": str(int(datetime.now().timestamp() * 1000)),
                        "rows": "2000",
                        "page": "1",
                        "sidx": "",
                        "sord": "asc",
                        "template": "I100",
                        "action": "search",
                        "page_code": "I100",
                        "par": (
                            f"auto_search=&search_all_product=&multi_supply_group=&multi_supply="
                            f"&str_supply_code={_TODAY_SUPPLY_CODE}&tags_string=&product_tag_include_type=1"
                            f"&query_type=name&query_str=&stock_type=0&stock_start=1&stock_end="
                            f"&notrans_day=&notrans_cnt=&notrans_status=0&stock_status=0"
                            f"&start_date={today_str}&start_hour=00%3A00%3A00"
                            f"&end_date={today_str}&end_hour=23%3A59%3A59&date_period_sel=0"
                            f"&work_type=stockin&work_start=&work_end=&inout_type=0&product_date="
                            f"&start_date2={today_str}&end_date2={today_str}&date_period_sel2=0"
                            f"&products_sort=1&category=0&except_soldout=0&temp_soldout=0&location=0"
                        ),
                    },
                    cookies=cookies,
                )
        except Exception as exc:
            return {"ok": False, "error": f"{type(exc).__name__}: {exc}"}

        # JSON 파싱 실패 = 로그인 페이지(HTML) 반환 → 세션 만료
        try:
            obj = r.json()
        except Exception:
            return {"ok": False, "need_session": True}

        # rows 키 없으면 세션 오류 응답으로 판단
        if "rows" not in obj:
            return {"ok": False, "need_session": True}

        rows = []
        for row in obj.get("rows", []):
            c = row.get("cell", {})
            option_no = _safe_str(c.get("option_extra_column1"))
            qty = max(0, int(float(c.get("stock_normal", 0) or 0)))
            if option_no:
                rows.append({"A": option_no, "B": str(qty)})

        return {"ok": True, "rows": rows, "count": len(rows)}

    VAT_VENDORS_KEY = "noye_kimsungil_vat_vendors"

    @router.get("/vat-vendors")
    def get_vat_vendors(user: str = Depends(get_current_user)):
        raw = get_setting(VAT_VENDORS_KEY)
        try:
            vendors = json.loads(raw) if raw else []
        except Exception:
            vendors = []
        return {"ok": True, "vendors": vendors}

    @router.post("/vat-vendors")
    def set_vat_vendors(payload: dict = Body(...), user: str = Depends(get_current_user)):
        vendors = payload.get("vendors")
        if not isinstance(vendors, list):
            raise HTTPException(status_code=400, detail="vendors must be a list")
        cleaned = sorted({str(v).strip() for v in vendors if str(v).strip()})
        set_setting(VAT_VENDORS_KEY, json.dumps(cleaned, ensure_ascii=False))
        return {"ok": True, "vendors": cleaned}

    @router.get("/purchase-manager/state")
    def get_purchase_manager_state(user: str = Depends(get_current_user)):
        conn = get_db()
        try:
            _init_purchase_manager_table(conn)
            row = conn.execute(
                "SELECT state_json, updated_at FROM noye_purchase_manager_state WHERE id = 1"
            ).fetchone()
            if not row:
                return {"ok": True, "state": None}
            try:
                state = json.loads(row["state_json"])
            except (TypeError, json.JSONDecodeError):
                state = None
            return {"ok": True, "state": state, "updatedAt": row["updated_at"]}
        finally:
            conn.close()

    @router.put("/purchase-manager/state")
    def save_purchase_manager_state(payload: dict = Body(...), user: str = Depends(get_current_user)):
        state = payload.get("state")
        if not isinstance(state, dict):
            raise HTTPException(status_code=400, detail="state must be an object")
        try:
            state_json = json.dumps(state, ensure_ascii=False, separators=(",", ":"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="state must be JSON serializable")
        if len(state_json.encode("utf-8")) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="purchase manager state is too large")

        conn = get_db()
        try:
            _init_purchase_manager_table(conn)
            conn.execute("""
                INSERT INTO noye_purchase_manager_state (id, state_json, updated_by, updated_at)
                VALUES (1, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(id) DO UPDATE SET
                    state_json = excluded.state_json,
                    updated_by = excluded.updated_by,
                    updated_at = CURRENT_TIMESTAMP
            """, (state_json, user))
            conn.commit()
            return {"ok": True}
        finally:
            conn.close()

    return router
