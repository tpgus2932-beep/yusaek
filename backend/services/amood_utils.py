import re

import openpyxl
from fastapi import HTTPException
from openpyxl.utils.cell import column_index_from_string

from barcode_core import load_excel_any, normalize_to_yusas

AMOOD_COL1_NAME_RAW = "H"
AMOOD_COL1_SCAN_BARCODE = "D"
AMOOD_COL1_ORDER_KEY = "C"
AMOOD_COL1_NUM_B = "B"
AMOOD_COL1_NUM_C = "C"
AMOOD_COL2_ORDER_KEY = "E"
AMOOD_COL2_NAME = "I"
AMOOD_COL2_OPTION = "J"
AMOOD_COL2_QTY = "K"
AMOOD_COL2_BARCODE = "H"
AMOOD_COL2_OUTPUT = "N"

AMOOD_BRACKET_PATTERNS = [
    r"\[[^\]]*\]",
    r"\([^)]*\)",
    r"\{[^}]*\}",
]


class AmoodState:
    def __init__(self):
        self.file1_path = None
        self.file2_path = None
        self.file1_name = None
        self.file2_name = None
        self.processed1_path = None
        self.processed2_path = None
        self.wb1 = None
        self.ws1 = None
        self.wb2 = None
        self.ws2 = None
        self.current_invoice = None
        self.pending_items: list[dict] = []
        self.waiting_for_items: bool = False
        self.completed_mgmt_numbers: set[str] = set()
        self.incoming_counts: dict[str, int] = {}
        self.ezadmin_saved_at: str | None = None


def _amood_status(state: AmoodState) -> dict:
    incoming_counts = state.incoming_counts or {}
    return {
        "excel1_loaded": state.file1_path is not None,
        "excel2_loaded": state.file2_path is not None,
        "processed": state.processed1_path is not None and state.processed2_path is not None,
        "file1_name": state.file1_name,
        "file2_name": state.file2_name,
        "current_invoice": state.current_invoice,
        "waiting_for_items": state.waiting_for_items,
        "incoming_codes": len(incoming_counts),
        "incoming_total": sum(incoming_counts.values()) if incoming_counts else 0,
        "ezadmin_saved_at": state.ezadmin_saved_at,
    }


def _amood_norm_barcode(s: str | None) -> str:
    raw = re.sub(r"\s+", "", str(s or ""))
    normalized = normalize_to_yusas(raw)
    return normalized or raw


def _amood_strip_any_brackets(s: str | None) -> str:
    if s is None:
        return ""
    text = str(s)
    for pat in AMOOD_BRACKET_PATTERNS:
        text = re.sub(pat, "", text)
    return re.sub(r"\s+", " ", text).strip()


def _amood_to_int_qty(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except Exception:
            return 0
    s = str(v).strip()
    if not s:
        return 0
    m = re.search(r"\d+", s)
    return int(m.group()) if m else 0


def _amood_parse_option(opt: str | None) -> tuple[str, str]:
    if opt is None:
        return ("", "")
    t = str(opt).strip()
    t = re.sub(r"^\s*[\[\(\{]\s*", "", t)
    t = re.sub(r"\s*[\]\)\}]\s*$", "", t)
    t = t.strip()
    if not t:
        return ("", "")
    if "-" in t:
        a, b = t.split("-", 1)
    elif "/" in t:
        a, b = t.split("/", 1)
    else:
        parts = t.split()
        a = parts[0]
        b = parts[1] if len(parts) > 1 else ""
    color = a.strip()
    size = b.strip()
    if size.lower() == "free":
        size = "FREE"
    return (color, size)


def _amood_build_output_text(name: str | None, opt: str | None, qty) -> str:
    base = str(name).strip() if name is not None else ""
    opt_str = str(opt or "").strip()
    opt_clean = re.sub(r"^\s*[\[\(\{]\s*", "", opt_str)
    opt_clean = re.sub(r"\s*[\]\)\}]\s*$", "", opt_clean).strip()

    # 합배송: 이지어드민이 " / "(공백+슬래시+공백)로 여러 상품 옵션을 합친 경우
    if " / " in opt_clean:
        opt_parts = [p.strip() for p in opt_clean.split(" / ") if p.strip()]
        # 이름도 "/"로 구분된 경우
        if "/" in base:
            name_parts = [p.strip() for p in base.split("/") if p.strip()]
        else:
            # 이름은 줄바꿈→공백으로 합쳐진 경우: 단어 수가 옵션 수와 맞으면 페어링
            name_words = base.split()
            name_parts = name_words if len(name_words) == len(opt_parts) else []

        if len(name_parts) == len(opt_parts):
            combined = []
            for n, o in zip(name_parts, opt_parts):
                color, size = _amood_parse_option(o)
                pieces = [n]
                if color:
                    pieces.append(color)
                if size:
                    pieces.append(size)
                combined.append(" ".join(pieces))
            return " / ".join(combined)

        # 페어링 불가: 이름 앞에 두고 옵션 파트를 " / "로 유지
        opt_formatted_parts = []
        for o in opt_parts:
            color, size = _amood_parse_option(o)
            o_pieces = []
            if color:
                o_pieces.append(color)
            if size:
                o_pieces.append(size)
            if o_pieces:
                opt_formatted_parts.append(" ".join(o_pieces))
        opt_formatted = " / ".join(opt_formatted_parts) if opt_formatted_parts else opt_clean
        q = _amood_to_int_qty(qty)
        return " ".join(filter(None, [base, opt_formatted, str(q)])).strip()

    # 단일 상품
    color, size = _amood_parse_option(opt)
    q = _amood_to_int_qty(qty)
    pieces = [base]
    if color:
        pieces.append(color)
    if size:
        pieces.append(size)
    pieces.append(str(q))
    return " ".join(pieces).strip()


def _amood_ws_cell(ws, col_letter: str, row: int):
    return ws[f"{col_letter}{row}"]


def _amood_norm_key(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return m.group(1)
    return re.sub(r"\s+", "", s)


def _amood_collect_rows_by_value(ws, col_letter: str, value: str, start_row: int = 2) -> list[int]:
    target = _amood_norm_key(value)
    out: list[int] = []
    if not target:
        return out
    for r in range(start_row, ws.max_row + 1):
        v = _amood_ws_cell(ws, col_letter, r).value
        if _amood_norm_key(v) == target:
            out.append(r)
    return out


def _amood_fill_down_merged_column(ws, col_letter: str, start_row: int = 2):
    col_idx = column_index_from_string(col_letter)
    targets = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col <= col_idx <= merged.max_col:
            top_val = ws.cell(row=merged.min_row, column=merged.min_col).value
            targets.append((str(merged), merged.min_row, merged.max_row, top_val))
    for ref, _, _, _ in targets:
        ws.unmerge_cells(ref)
    for _, r1, r2, top_val in targets:
        for r in range(r1, r2 + 1):
            ws.cell(row=r, column=col_idx).value = top_val
    last = None
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None or str(v).strip() == "":
            if last is not None:
                ws.cell(row=r, column=col_idx).value = last
        else:
            last = v


def _amood_load_workbooks(state: AmoodState):
    if not state.file1_path or not state.file2_path:
        raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
    if state.wb1 is None or state.ws1 is None:
        state.wb1 = openpyxl.load_workbook(state.file1_path)
        if len(state.wb1.worksheets) < 2:
            raise HTTPException(status_code=400, detail="excel1에 두 번째 시트가 없습니다.")
        state.ws1 = state.wb1.worksheets[1]
    if state.wb2 is None or state.ws2 is None:
        state.wb2, state.ws2 = load_excel_any(state.file2_path)
