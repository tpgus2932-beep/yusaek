from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

# 원가베이스유.xlsx 열 인덱스 (0-based) — 헤더:
# 상품코드, 상품명, 색상, 사이즈, 원가, 거래처, 거래처상품명, 거래처합, 상품명합, 거래처주소, 옵션번호
_NAME_COL = 1
_COLOR_COL = 2
_SIZE_COL = 3
_OPTION_CODE_COL = 10
_REQUIRED_COLS = _OPTION_CODE_COL + 1


def search_cost_base_products(path: Path, q: str, limit: int = 20) -> list[dict]:
    """원가베이스유 엑셀에서 상품명(1열) 기준으로 검색해 옵션(색상/사이즈/옵션번호)을 묶어 반환.

    같은 상품명의 색상/사이즈별 행들을 하나의 항목으로 묶고, 그 항목의
    options에 {code, label}(옵션번호, "색상/사이즈" 표시용 라벨)을 순서대로 모은다.
    실행 화면에서 옵션을 개별 선택/해제할 수 있도록 코드만이 아니라 라벨도 필요하다.
    """
    if not path.exists():
        return []

    q_norm = (q or "").strip().lower()
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    groups: dict[str, list[dict]] = {}
    seen_codes: dict[str, set[str]] = {}
    order: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < _REQUIRED_COLS:
            continue
        name = str(row[_NAME_COL] or "").strip()
        option_code = str(row[_OPTION_CODE_COL] or "").strip()
        if not name or not option_code:
            continue
        if q_norm and q_norm not in name.lower():
            continue
        if name not in groups:
            groups[name] = []
            seen_codes[name] = set()
            order.append(name)
        if option_code in seen_codes[name]:
            continue
        seen_codes[name].add(option_code)
        color = str(row[_COLOR_COL] or "").strip()
        size = str(row[_SIZE_COL] or "").strip()
        label = "/".join(part for part in (color, size) if part) or option_code
        groups[name].append({"code": option_code, "label": label})

    return [{"name": name, "options": groups[name]} for name in order[:limit]]


def filter_matching_order_items(order_items: list[dict], option_codes: set[str]) -> list[dict]:
    """order_items 중 option_stock_sync_code가 option_codes에 속하는 것만 남긴다."""
    return [
        item for item in order_items
        if str(item.get("option_stock_sync_code") or "") in option_codes
    ]


def group_items_by_order_sno(items: list[dict]) -> dict[int, list[dict]]:
    """주문상품 리스트를 order_sno 기준으로 그룹핑 (취소 API가 주문당 1회 호출이라 필요)."""
    grouped: dict[int, list[dict]] = {}
    for item in items:
        grouped.setdefault(item["order_sno"], []).append(item)
    return grouped


def build_soldout_message(template_msg: str, product_names: list[str]) -> str:
    """템플릿의 {상품}을 실제 상품명으로 치환. 중복 상품명은 제거하고 쉼표로 나열."""
    unique_names = list(dict.fromkeys(product_names))
    return template_msg.replace("{상품}", ", ".join(unique_names))
