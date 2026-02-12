from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Body, Header, Depends, Response, Request
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
import tempfile
import uuid
import traceback
import os
import sqlite3
import io
import re
import shutil
import mimetypes
import urllib.parse
from collections import Counter
from datetime import datetime, timedelta, timezone

from barcode_core import process_and_load_any, normalize_to_yusas, load_excel_any

import barcode_core
import pandas as pd
import xlwt
import openpyxl
from openpyxl.utils.cell import column_index_from_string

from passlib.context import CryptContext
from jose import jwt, JWTError
from api.amood_hapbae import router as amood_hapbae_router

print("### barcode_core file =", barcode_core.__file__)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

STATE = {
    "loaded": False,
    "mapping": None,
    "details": None,
    "runs": None,
    "invoice_order": None,
    "invoice_seq": None,
    "code_o_text": None,
    "current_invoice": None,
    "last_scanned_code": None,
    "processed_path": None,
    "defect_counts": None,
    "incoming_counts": None,
}

UPLOAD_BASE = Path(__file__).resolve().parent / "uploads" / "requests"
SHARED_UPLOAD_BASE = Path(__file__).resolve().parent / "uploads" / "shared_files"
ALLOWED_REQUEST_EXTS = {
    ".xlsx",
    ".xls",
    ".csv",
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
}
ALLOWED_SHARED_EXTS = {".xlsx", ".xls", ".csv"}
RETURN_ALLOWED_EXTS = {".xlsx", ".xls", ".xlsm"}
AMOOD_ALLOWED_EXCEL1 = {".xlsx", ".xlsm"}
AMOOD_ALLOWED_EXCEL2 = {".xlsx", ".xls", ".xlsm", ".htm", ".html"}
RETURN_COST_BASE_PATH = Path(
    os.environ.get("RETURN_COST_BASE_PATH", r"C:\Users\ksh29\OneDrive\Desktop\원베\원가베이스유.xlsx")
)
RETURN_STATES: dict[str, "ReturnState"] = {}
AMOOD_STATES: dict[str, "AmoodState"] = {}
RETURN_COST_BASE_CACHE: dict[str, object] = {"df": None, "mtime": None, "path": None}

# ---------- EasyAdmin product upload helpers ----------
HEADER_LIST = [
    "상품명","공급처코드 / 공급처명","공급처 상품명","공급처 옵션","원산지","택배비","중량",
    "원가","공급가","판매가","시중가","옵션1","옵션2","옵션3","옵션관리","바코드",
    "대표 이미지","설명 이미지1","설명 이미지2","설명 이미지3","설명 이미지4","설명 이미지5",
    "비고 이미지","상품설명","상품설명2","재고경고수량","재고위험수량","합포불가",
    "동일상품 합포가능 수량","로케이션","메모","제조사","사은품","담당MD","관리자(정)",
    "관리자(부)","무료배송","카테고리","배송타입","매장간이동","판매시작일","입고대기",
    "판매처코드 / 판매처명","상품태그","상품추가항목1","상품추가항목2","상품추가항목3",
    "상품추가항목4","상품추가항목5","상품추가항목6","상품추가항목7","상품추가항목8",
    "상품추가항목9","상품추가항목10","소진시 품절","입고시 품절해제","소진시 일시품절",
    "입고시 일시품절해제","옵션추가항목1","옵션추가항목2","옵션추가항목3","옵션추가항목4",
    "옵션추가항목5","옵션추가금액(원가)","옵션추가금액(판매가)","매칭시 자동취소",
    "유통기한 경고 설정","판매상태","원가메모","재고단위1","재고단위2","재고단위3","재고단위4","재고단위5"
]

_FRONT_BRACKETS = re.compile(r'^(\s*\[[^\]]*\]\s*)+')
_BACK_BRACKETS = re.compile(r'(\s*\[[^\]]*\]\s*)+$')


def _content_disposition(filename: str) -> str:
    safe_name = (filename or "download").replace('"', "")
    ascii_name = "".join(ch if ord(ch) < 128 else "_" for ch in safe_name)
    ascii_name = re.sub(r"_+", "_", ascii_name).strip("_")
    if not ascii_name:
        ascii_name = "download"
    quoted = urllib.parse.quote(safe_name)
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted}"


def _strip_edge_brackets(text):
    if pd.isna(text):
        return text
    s = str(text)
    s = re.sub(_FRONT_BRACKETS, '', s)
    s = re.sub(_BACK_BRACKETS, '', s)
    return s.strip()


def _split_b_to_c_and_h(value):
    if pd.isna(value):
        return pd.NA, pd.NA
    s = str(value).strip()
    if not s:
        return pd.NA, pd.NA
    parts = s.split()
    if len(parts) <= 2:
        return " ".join(parts), pd.NA
    if len(parts) == 3:
        c_part = " ".join(parts[:2])
        try:
            h_part = int(parts[2])
        except ValueError:
            h_part = pd.NA
        return c_part, h_part
    return " ".join(parts), pd.NA


def _split_l_values(value):
    if pd.isna(value):
        return pd.NA, pd.NA, pd.NA
    tokens = [t.strip() for t in str(value).split(',') if t.strip()]
    tokens = [f":{t}" for t in tokens[:3]]
    while len(tokens) < 3:
        tokens.append(pd.NA)
    return tokens[0], tokens[1], tokens[2]


def _col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def _pos0(col: str) -> int:
    return _col_to_num(col) - 1


def _save_as_xls_bytes(df: pd.DataFrame) -> bytes:
    book = xlwt.Workbook()
    sheet = book.add_sheet("Sheet1")
    for j, col in enumerate(df.columns):
        sheet.write(0, j, col)
    for i, row in df.iterrows():
        for j, val in enumerate(row):
            if pd.isna(val):
                sheet.write(i + 1, j, "")
            else:
                sheet.write(i + 1, j, val)
    buf = io.BytesIO()
    book.save(buf)
    return buf.getvalue()


def _process_easyadmin_product_upload(path: Path) -> bytes:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        df = pd.read_excel(path, engine="openpyxl")
    elif ext == ".xls":
        df = pd.read_excel(path)
    elif ext == ".csv":
        try:
            df = pd.read_csv(path, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(path, encoding="cp949")
    else:
        raise ValueError("지원 형식: xlsx, xls, csv")

    if df.shape[1] < 12:
        raise ValueError("원본 파일에 최소 12열(C, H, L 포함)이 필요합니다.")

    series_b = df.iloc[:, 1]
    series_c = df.iloc[:, 2]
    series_h = df.iloc[:, 7]
    series_l = df.iloc[:, 11]

    col_a = series_c.apply(_strip_edge_brackets)
    col_b = pd.Series(["유색"] * len(df), index=df.index)
    ch_df = series_b.apply(lambda v: pd.Series(_split_b_to_c_and_h(v)))
    lmn_df = series_l.apply(lambda v: pd.Series(_split_l_values(v)))

    out = pd.DataFrame("", index=df.index, columns=HEADER_LIST)
    out.iloc[:, _pos0('A')] = col_a
    out.iloc[:, _pos0('B')] = col_b
    out.iloc[:, _pos0('C')] = ch_df.iloc[:, 0]
    out.iloc[:, _pos0('H')] = ch_df.iloc[:, 1]
    out.iloc[:, _pos0('L')] = lmn_df.iloc[:, 0]
    out.iloc[:, _pos0('M')] = lmn_df.iloc[:, 1]
    out.iloc[:, _pos0('N')] = lmn_df.iloc[:, 2]
    out.iloc[:, _pos0('O')] = 1
    out.iloc[:, _pos0('BG')] = series_h

    return _save_as_xls_bytes(out)


# ---------- Return (반품) helpers ----------
def _read_return_excel(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    if ext == ".xls":
        try:
            return pd.read_excel(path, dtype=str, engine="xlrd")
        except Exception:
            return pd.read_excel(path, dtype=str)
    # 확장자가 없거나 판별 실패 시: openpyxl -> xlrd 순으로 시도
    try:
        return pd.read_excel(path, dtype=str, engine="openpyxl")
    except Exception:
        try:
            return pd.read_excel(path, dtype=str, engine="xlrd")
        except Exception as e:
            raise ValueError(f"지원 형식: xlsx, xls, xlsm (읽기 실패: {e})")


def _read_return_excel_with_header(path: Path, header):
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(path, dtype=str, engine="openpyxl", header=header)
    if ext == ".xls":
        try:
            return pd.read_excel(path, dtype=str, engine="xlrd", header=header)
        except Exception:
            return pd.read_excel(path, dtype=str, header=header)
    try:
        return pd.read_excel(path, dtype=str, engine="openpyxl", header=header)
    except Exception:
        return pd.read_excel(path, dtype=str, engine="xlrd", header=header)


def _clean_invoice(value: str) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in ("nan", "none"):
        return ""
    return re.sub(r"\D+", "", s)


def _clean_product_name(text: str) -> str:
    if text is None:
        return ""
    s = str(text)
    s = re.sub(r"\[[^\]]*\]", " ", s)
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _option_slash_to_space(opt: str) -> str:
    if opt is None:
        return ""
    s = str(opt).strip()
    if s.lower() in ("nan", "none"):
        return ""
    parts = [p.strip() for p in s.split("/") if p.strip()]
    return " ".join(parts)


def _normalize_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def _normalize_key(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.lower()


def _reason_type(k_value: str) -> str:
    s = "" if k_value is None else str(k_value).strip()
    if s.lower() in ("nan", "none"):
        return "미매칭"
    s2 = re.sub(r"\([^)]*\)", "", s).strip()
    if s2.startswith("판매자"):
        return "판매자"
    if s2.startswith("고객"):
        return "고객"
    return "미매칭"


def _clean_qty(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() in ("nan", "none", ""):
        return ""
    s = re.sub(r"\.0$", "", s)
    return s


def _lowercase_size_words(text: str) -> str:
    if text is None:
        return ""
    s = str(text)
    size_words = ["FREE", "XS", "S", "M", "L", "XL", "XXL", "XXXL", "SHORT", "LONG"]
    for w in size_words:
        s = re.sub(rf"\b{w}\b", w.lower(), s, flags=re.IGNORECASE)
    return s


class ReturnState:
    def __init__(self):
        self.df1: pd.DataFrame | None = None
        self.df2: pd.DataFrame | None = None
        self.map_d_to_e: dict[str, str] = {}
        self.df2_index: dict[str, list[int]] = {}
        self.queue_seller: list[dict] = []
        self.queue_customer: list[dict] = []
        self.queue_unmatched: list[dict] = []
        self.all_items: list[dict] = []
        self.last_added_ids: list[int] = []
        self.scanned_barcodes: set[str] = set()
        self.cost_map: dict[str, str] = {}
        self.cost_base_path: Path = RETURN_COST_BASE_PATH
        self.customer_export_df: pd.DataFrame = pd.DataFrame()
        self.last_type: str = "-"
        self.next_id: int = 1


class AmoodState:
    def __init__(self):
        self.file1_path: Path | None = None
        self.file2_path: Path | None = None
        self.file1_name: str | None = None
        self.file2_name: str | None = None
        self.processed1_path: Path | None = None
        self.processed2_path: Path | None = None
        self.wb1 = None
        self.ws1 = None
        self.wb2 = None
        self.ws2 = None
        self.current_invoice: str | None = None
        self.pending_items: list[dict] = []
        self.waiting_for_items: bool = False
        self.completed_mgmt_numbers: set[str] = set()
        self.incoming_counts: dict[str, int] = {}


def _get_return_state(user: str) -> ReturnState:
    state = RETURN_STATES.get(user)
    if not state:
        state = ReturnState()
        RETURN_STATES[user] = state
    return state


def _get_amood_state(user: str) -> AmoodState:
    state = AMOOD_STATES.get(user)
    if not state:
        state = AmoodState()
        AMOOD_STATES[user] = state
    return state


def _amood_status(state: AmoodState) -> dict:
    incoming_counts = state.incoming_counts or {}
    return {
        "excel1_loaded": state.file1_path is not None,
        "excel2_loaded": state.file2_path is not None,
        "processed": state.processed1_path is not None and state.processed2_path is not None,
        "file1_name": state.file1_name,
        "file2_name": state.file2_name,
        "current_invoice": state.current_invoice,
        "waiting_for_items": state.waiting_for_items,
        "incoming_codes": len(incoming_counts),
        "incoming_total": sum(incoming_counts.values()) if incoming_counts else 0,
    }


def _return_status(state: ReturnState) -> dict:
    path = state.cost_base_path
    exists = path.exists()
    mtime = None
    if exists:
        try:
            mtime = datetime.fromtimestamp(path.stat().st_mtime).isoformat()
        except Exception:
            mtime = None
    return {
        "excel1_loaded": state.df1 is not None,
        "excel2_loaded": state.df2 is not None,
        "cost_loaded": bool(state.cost_map),
        "map_count": len(state.map_d_to_e),
        "index_count": len(state.df2_index),
        "cost_count": len(state.cost_map),
        "cost_base_path": str(path),
        "cost_base_exists": exists,
        "cost_base_mtime": mtime,
    }


def _return_rows(df: pd.DataFrame) -> list[dict]:
    if df is None or df.empty:
        return []
    rows = []
    for _, r in df.iterrows():
        item = {}
        for k, v in r.items():
            if pd.isna(v):
                item[k] = ""
            else:
                item[k] = v
        rows.append(item)
    return rows


def _load_return_cost_base(state: ReturnState):
    path = state.cost_base_path
    if not path.exists():
        raise FileNotFoundError(f"원가베이스 파일을 찾지 못했습니다: {path}")
    cost_df = pd.read_excel(path, dtype=str)
    if cost_df.shape[1] < 2:
        raise ValueError("원가베이스는 최소 A,B열이 필요합니다.")
    amap: dict[str, str] = {}
    for _, r in cost_df.iterrows():
        key_raw = r.iloc[0] if len(r) > 0 else ""
        val_raw = r.iloc[1] if len(r) > 1 else ""
        key = _normalize_key("" if pd.isna(key_raw) else str(key_raw))
        val = "" if pd.isna(val_raw) else str(val_raw).strip()
        if key and key not in amap:
            amap[key] = val
    state.cost_map = amap


def _load_cost_base_df():
    path = RETURN_COST_BASE_PATH
    if not path.exists():
        raise FileNotFoundError(f"원가베이스 파일을 찾지 못했습니다: {path}")
    mtime = path.stat().st_mtime
    cached_path = RETURN_COST_BASE_CACHE.get("path")
    cached_mtime = RETURN_COST_BASE_CACHE.get("mtime")
    if RETURN_COST_BASE_CACHE.get("df") is not None and cached_path == str(path) and cached_mtime == mtime:
        return RETURN_COST_BASE_CACHE["df"]
    df = _read_return_excel_with_header(path, header=0)
    if df.shape[0] == 0:
        df_raw = _read_return_excel_with_header(path, header=None)
        if df_raw.shape[0] >= 2:
            new_cols = df_raw.iloc[0].fillna("").astype(str).tolist()
            df_raw = df_raw.iloc[1:].reset_index(drop=True)
            df_raw.columns = new_cols
            df = df_raw
    RETURN_COST_BASE_CACHE["df"] = df
    RETURN_COST_BASE_CACHE["mtime"] = mtime
    RETURN_COST_BASE_CACHE["path"] = str(path)
    return df


def _save_cost_base_df(df: pd.DataFrame):
    path = RETURN_COST_BASE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    RETURN_COST_BASE_CACHE["df"] = df
    RETURN_COST_BASE_CACHE["mtime"] = path.stat().st_mtime
    RETURN_COST_BASE_CACHE["path"] = str(path)


# ---- AMOOD processing (from AMOODBOX2.PY) ----
AMOOD_COL1_NAME_RAW = "H"
AMOOD_COL1_SCAN_BARCODE = "D"
AMOOD_COL1_ORDER_KEY = "C"
AMOOD_COL1_NUM_B = "B"
AMOOD_COL1_NUM_C = "C"
AMOOD_COL2_ORDER_KEY = "E"
AMOOD_COL2_NAME = "I"
AMOOD_COL2_OPTION = "J"
AMOOD_COL2_QTY = "K"
AMOOD_COL2_BARCODE = "H"
AMOOD_COL2_OUTPUT = "N"

AMOOD_BRACKET_PATTERNS = [
    r"\[[^\]]*\]",
    r"\([^)]*\)",
    r"\{[^}]*\}",
]


def _amood_norm_barcode(s: str | None) -> str:
    return re.sub(r"\s+", "", str(s or ""))


def _amood_strip_any_brackets(s: str | None) -> str:
    if s is None:
        return ""
    text = str(s)
    for pat in AMOOD_BRACKET_PATTERNS:
        text = re.sub(pat, "", text)
    return re.sub(r"\s+", " ", text).strip()


def _amood_to_int_qty(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except Exception:
            return 0
    s = str(v).strip()
    if not s:
        return 0
    m = re.search(r"\d+", s)
    return int(m.group()) if m else 0


def _amood_parse_option(opt: str | None) -> tuple[str, str]:
    if opt is None:
        return ("", "")
    t = str(opt).strip()
    t = re.sub(r"^\s*[\[\(\{]\s*", "", t)
    t = re.sub(r"\s*[\]\)\}]\s*$", "", t)
    t = t.strip()
    if not t:
        return ("", "")
    if "-" in t:
        a, b = t.split("-", 1)
    elif "/" in t:
        a, b = t.split("/", 1)
    else:
        parts = t.split()
        a = parts[0]
        b = parts[1] if len(parts) > 1 else ""
    color = a.strip()
    size = b.strip()
    if size.lower() == "free":
        size = "FREE"
    return (color, size)


def _amood_build_output_text(name: str | None, opt: str | None, qty) -> str:
    base = str(name).strip() if name is not None else ""
    color, size = _amood_parse_option(opt)
    q = _amood_to_int_qty(qty)
    pieces = [base]
    if color:
        pieces.append(color)
    if size:
        pieces.append(size)
    pieces.append(str(q))
    return " ".join(pieces).strip()


def _amood_ws_cell(ws, col_letter: str, row: int):
    return ws[f"{col_letter}{row}"]


def _amood_norm_key(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v).strip()
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return m.group(1)
    return re.sub(r"\s+", "", s)


def _amood_collect_rows_by_value(ws, col_letter: str, value: str, start_row: int = 2) -> list[int]:
    target = _amood_norm_key(value)
    out: list[int] = []
    if not target:
        return out
    for r in range(start_row, ws.max_row + 1):
        v = _amood_ws_cell(ws, col_letter, r).value
        if _amood_norm_key(v) == target:
            out.append(r)
    return out


def _amood_fill_down_merged_column(ws, col_letter: str, start_row: int = 2):
    col_idx = column_index_from_string(col_letter)
    targets = []
    for merged in list(ws.merged_cells.ranges):
        if merged.min_col <= col_idx <= merged.max_col:
            top_val = ws.cell(row=merged.min_row, column=merged.min_col).value
            targets.append((str(merged), merged.min_row, merged.max_row, top_val))
    for ref, _, _, _ in targets:
        ws.unmerge_cells(ref)
    for _, r1, r2, top_val in targets:
        for r in range(r1, r2 + 1):
            ws.cell(row=r, column=col_idx).value = top_val
    last = None
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v is None or str(v).strip() == "":
            if last is not None:
                ws.cell(row=r, column=col_idx).value = last
        else:
            last = v


def _amood_load_workbooks(state: AmoodState):
    if not state.file1_path or not state.file2_path:
        raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
    if state.wb1 is None or state.ws1 is None:
        state.wb1 = openpyxl.load_workbook(state.file1_path)
        if len(state.wb1.worksheets) < 2:
            raise HTTPException(status_code=400, detail="excel1에 두 번째 시트가 없습니다.")
        state.ws1 = state.wb1.worksheets[1]
    if state.wb2 is None or state.ws2 is None:
        state.wb2, state.ws2 = load_excel_any(state.file2_path)


def _return_queue_payload(state: ReturnState) -> dict:
    return {
        "seller": state.queue_seller,
        "customer": state.queue_customer,
        "unmatched": state.queue_unmatched,
        "all": state.all_items,
    }

DB_PATH = Path(__file__).with_name("app.db")
JWT_SECRET = os.environ.get("JWT_SECRET", "dev-secret-change-me")
JWT_ALG = "HS256"
BOOT_ID = uuid.uuid4().hex
TOKEN_EXPIRE_MINUTES = 60 * 24

pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")


def _get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def _init_db():
    conn = _get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            display_name TEXT NOT NULL DEFAULT '',
            role TEXT NOT NULL DEFAULT 'user',
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()


def _ensure_user_column(column: str, ddl: str):
    conn = _get_db()
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
    if column not in cols:
        conn.execute(ddl)
        conn.commit()
    conn.close()


_init_db()
_ensure_user_column("display_name", "ALTER TABLE users ADD COLUMN display_name TEXT NOT NULL DEFAULT ''")
_ensure_user_column("role", "ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'user'")


def _init_requests():
    conn = _get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requester_username TEXT NOT NULL,
            requester_display TEXT NOT NULL DEFAULT '',
            assignee_username TEXT NOT NULL,
            assignee_display TEXT NOT NULL DEFAULT '',
            text TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'open',
            created_at TEXT NOT NULL,
            completed_at TEXT,
            acknowledged_at TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def _ensure_request_column(column: str, ddl: str):
    conn = _get_db()
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(requests)").fetchall()]
    if column not in cols:
        conn.execute(ddl)
        conn.commit()
    conn.close()


_init_requests()
_ensure_request_column(
    "requester_display",
    "ALTER TABLE requests ADD COLUMN requester_display TEXT NOT NULL DEFAULT ''",
)
_ensure_request_column(
    "assignee_display",
    "ALTER TABLE requests ADD COLUMN assignee_display TEXT NOT NULL DEFAULT ''",
)
_ensure_request_column(
    "acknowledged_at",
    "ALTER TABLE requests ADD COLUMN acknowledged_at TEXT",
)


def _init_company_credentials():
    conn = _get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS company_credentials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT NOT NULL,
            username TEXT,
            password TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
        """
    )
    conn.commit()
    conn.close()


_init_company_credentials()


def _init_app_settings():
    conn = _get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    conn.commit()
    conn.close()


_init_app_settings()


def _ensure_default_company_pin():
    if not _get_setting("company_pin_hash"):
        _set_setting("company_pin_hash", _hash_pin("0000"))


def _init_request_attachments():
    conn = _get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS request_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            request_id INTEGER NOT NULL,
            original_name TEXT NOT NULL,
            stored_name TEXT NOT NULL,
            mime_type TEXT NOT NULL,
            size INTEGER NOT NULL,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()


_init_request_attachments()


def _init_shared_files():
    conn = _get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS shared_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            original_name TEXT NOT NULL,
            stored_name TEXT NOT NULL,
            mime_type TEXT NOT NULL,
            size INTEGER NOT NULL,
            uploader_username TEXT NOT NULL,
            uploader_display TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        )
        """
    )
    conn.commit()
    conn.close()


_init_shared_files()



def _get_user_display(username: str) -> str:
    conn = _get_db()
    row = conn.execute("SELECT display_name FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    return row["display_name"] if row else ""


def _get_user_role(username: str) -> str:
    conn = _get_db()
    row = conn.execute("SELECT role FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    return row["role"] if row and row["role"] else "user"


def _is_admin(username: str) -> bool:
    return _get_user_role(username) == "admin"


def _count_admins() -> int:
    conn = _get_db()
    row = conn.execute("SELECT COUNT(*) as cnt FROM users WHERE role = 'admin'").fetchone()
    conn.close()
    return int(row["cnt"]) if row else 0


def _ensure_bootstrap_admin():
    username = (os.environ.get("BOOTSTRAP_ADMIN_USERNAME") or "ksh2932").strip()
    password = (os.environ.get("BOOTSTRAP_ADMIN_PASSWORD") or "").strip()
    display_name = (os.environ.get("BOOTSTRAP_ADMIN_DISPLAY_NAME") or "관리자").strip()
    if not password:
        return
    conn = _get_db()
    row = conn.execute("SELECT username FROM users WHERE username = ?", (username,)).fetchone()
    if row:
        conn.execute("UPDATE users SET role = 'admin' WHERE username = ?", (username,))
        conn.commit()
        conn.close()
        return
    conn.execute(
        "INSERT INTO users (username, password_hash, display_name, role, created_at) VALUES (?, ?, ?, ?, ?)",
        (username, _hash_password(password), display_name, "admin", datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()
    conn.close()




def _parse_iso_date(value: str | None):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return None


def _is_visible_completed(completed_at: str | None) -> bool:
    if not completed_at:
        return True
    completed_date = _parse_iso_date(completed_at)
    if not completed_date:
        return True
    return completed_date >= datetime.now(timezone.utc).date()


def _row_to_request(row) -> dict:
    return {
        "id": row["id"],
        "requester_username": row["requester_username"],
        "requester_display": row["requester_display"],
        "assignee_username": row["assignee_username"],
        "assignee_display": row["assignee_display"],
        "text": row["text"],
        "status": row["status"],
        "created_at": row["created_at"],
        "completed_at": row["completed_at"],
        "acknowledged_at": row["acknowledged_at"],
    }


def _is_image_mime(mime: str | None) -> bool:
    return bool(mime) and mime.lower().startswith("image/")


def _get_request_attachments(request_ids: list[int]) -> dict[int, list[dict]]:
    if not request_ids:
        return {}
    conn = _get_db()
    placeholders = ",".join(["?"] * len(request_ids))
    rows = conn.execute(
        f"SELECT * FROM request_attachments WHERE request_id IN ({placeholders}) ORDER BY id ASC",
        request_ids,
    ).fetchall()
    conn.close()
    result: dict[int, list[dict]] = {}
    for row in rows:
        item = {
            "id": row["id"],
            "filename": row["original_name"],
            "mime_type": row["mime_type"],
            "size": row["size"],
            "url": f"/requests/{row['request_id']}/attachments/{row['id']}",
            "is_image": _is_image_mime(row["mime_type"]),
        }
        result.setdefault(row["request_id"], []).append(item)
    return result


def _get_current_user_optional(authorization: str | None, token: str | None):
    raw = None
    if authorization and authorization.startswith("Bearer "):
        raw = authorization.split(" ", 1)[1].strip()
    elif token:
        raw = token.strip()
    if not raw:
        raise HTTPException(status_code=401, detail="Unauthorized")
    try:
        payload = jwt.decode(raw, JWT_SECRET, algorithms=[JWT_ALG])
        if payload.get("boot_id") != BOOT_ID:
            raise HTTPException(status_code=401, detail="Unauthorized")
        username = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Unauthorized")
    except JWTError:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return username


def _row_to_shared_file(row) -> dict:
    return {
        "id": row["id"],
        "filename": row["original_name"],
        "mime_type": row["mime_type"],
        "size": row["size"],
        "uploader_username": row["uploader_username"],
        "uploader_display": row["uploader_display"],
        "created_at": row["created_at"],
        "url": f"/shared-files/{row['id']}",
    }


def _hash_password(password: str) -> str:
    return pwd_context.hash(password)


def _verify_password(password: str, password_hash: str) -> bool:
    return pwd_context.verify(password, password_hash)


def _hash_pin(pin: str) -> str:
    return pwd_context.hash(pin)


def _verify_pin(pin: str, pin_hash: str) -> bool:
    return pwd_context.verify(pin, pin_hash)


def _get_setting(key: str) -> str | None:
    conn = _get_db()
    row = conn.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    return row["value"] if row else None


def _set_setting(key: str, value: str | None):
    conn = _get_db()
    if value is None:
        conn.execute("DELETE FROM app_settings WHERE key = ?", (key,))
    else:
        conn.execute(
            "INSERT INTO app_settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
            (key, value),
        )
    conn.commit()
    conn.close()


_ensure_default_company_pin()


def _to_int(value, default=0):
    try:
        if value is None or (isinstance(value, str) and not value.strip()):
            return default
        return int(float(str(value).strip()))
    except Exception:
        return default


_ensure_bootstrap_admin()


def _create_access_token(username: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(minutes=TOKEN_EXPIRE_MINUTES)
    payload = {"sub": username, "exp": expire, "boot_id": BOOT_ID}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


def _get_current_user(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")
    token = authorization.split(" ", 1)[1].strip()
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        if payload.get("boot_id") != BOOT_ID:
            raise HTTPException(status_code=401, detail="Unauthorized")
        username = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Unauthorized")
    except JWTError:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return username


def _require_admin(user: str = Depends(_get_current_user)):
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="admin required")
    return user


app.include_router(
    amood_hapbae_router,
    dependencies=[Depends(_get_current_user)],
)


@app.get("/ping")
def ping():
    return {"status": "ok"}


@app.post("/auth/register")
def register(payload: dict = Body(...)):
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    display_name = (payload.get("display_name") or "").strip()
    if not username or not password or not display_name:
        raise HTTPException(status_code=400, detail="username/password/display_name required")

    conn = _get_db()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, display_name, role, created_at) VALUES (?, ?, ?, ?, ?)",
            (username, _hash_password(password), display_name, "user", datetime.now(timezone.utc).isoformat()),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="username already exists")
    finally:
        conn.close()

    return {"ok": True}


@app.options("/auth/register")
def register_options():
    return {}


@app.post("/auth/login")
def login(payload: dict = Body(...)):
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    if not username or not password:
        raise HTTPException(status_code=400, detail="username/password required")

    conn = _get_db()
    row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    if not row or not _verify_password(password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="invalid credentials")

    token = _create_access_token(username)
    role = row["role"] if row["role"] else "user"
    return {
        "ok": True,
        "token": token,
        "username": username,
        "display_name": row["display_name"],
        "role": role,
        "is_admin": role == "admin",
    }


@app.options("/auth/login")
def login_options():
    return {}


@app.get("/auth/me")
def me(user: str = Depends(_get_current_user)):
    conn = _get_db()
    row = conn.execute("SELECT display_name, role FROM users WHERE username = ?", (user,)).fetchone()
    conn.close()
    display_name = row["display_name"] if row else ""
    role = row["role"] if row and row["role"] else "user"
    return {"ok": True, "username": user, "display_name": display_name, "role": role, "is_admin": role == "admin"}


@app.options("/auth/me")
def me_options():
    return {}


@app.patch("/auth/profile")
def update_profile(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    display_name = (payload.get("display_name") or "").strip()
    if not display_name:
        raise HTTPException(status_code=400, detail="display_name required")
    conn = _get_db()
    conn.execute("UPDATE users SET display_name = ? WHERE username = ?", (display_name, user))
    conn.commit()
    conn.close()
    return {"ok": True, "username": user, "display_name": display_name}


@app.post("/barcode/upload")
async def barcode_upload(file: UploadFile = File(...), user: str = Depends(_get_current_user)):
    name = (file.filename or "").lower()
    if not (name.endswith(".xls") or name.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="xls/xlsx만 업로드 가능")

    suffix = ".xlsx" if name.endswith(".xlsx") else ".xls"

    # 파일명 충돌 방지
    tmp_path = Path(tempfile.gettempdir()) / f"yusaek_upload_{uuid.uuid4().hex}{suffix}"
    data = await file.read()
    tmp_path.write_bytes(data)

    try:
        result = process_and_load_any(tmp_path)
        print("process_and_load_any return len =", len(result))

        if len(result) == 7:
            processed_path, mapping, details, runs, invoice_order, invoice_seq, code_o_text = result
        elif len(result) == 6:
            mapping, details, runs, invoice_order, invoice_seq, code_o_text = result
            processed_path = None
        else:
            raise Exception(f"unexpected return count: {len(result)}")

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"가공 실패: {e}")

    STATE.update(
        {
            "loaded": True,
            "processed_path": str(processed_path) if processed_path else None,
            "mapping": mapping,
            "details": details,
            "runs": runs,
            "invoice_order": invoice_order,
            "invoice_seq": invoice_seq,
            "code_o_text": code_o_text,
            "current_invoice": None,
            "last_scanned_code": None,
            "defect_counts": {},
        }
    )

    return {
        "ok": True,
        "invoices": len(mapping),
        "codes_total": sum(len(v) for v in mapping.values()),
    }


@app.post("/barcode/incoming/upload")
async def incoming_upload(file: UploadFile = File(...), user: str = Depends(_get_current_user)):
    name = (file.filename or "").lower()
    if not (name.endswith(".xls") or name.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="xls/xlsx files only")

    suffix = ".xlsx" if name.endswith(".xlsx") else ".xls"
    tmp_path = Path(tempfile.gettempdir()) / f"yusaek_incoming_{uuid.uuid4().hex}{suffix}"
    data = await file.read()
    tmp_path.write_bytes(data)

    try:
        wb, ws = load_excel_any(tmp_path)
        counts = Counter()
        for r in range(1, ws.max_row + 1):
            code_raw = ws.cell(r, 1).value
            qty_raw = ws.cell(r, 2).value
            code = normalize_to_yusas(code_raw)
            if not code:
                continue
            qty = _to_int(qty_raw, default=0)
            if qty > 0:
                counts[code] += qty
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"incoming load failed: {e}")

    STATE["incoming_counts"] = dict(counts)
    return {"ok": True, "codes": len(counts), "total_qty": sum(counts.values())}


@app.post("/barcode/product/upload")
async def easyadmin_product_upload(file: UploadFile = File(...), user: str = Depends(_get_current_user)):
    name = (file.filename or "").lower()
    if not (name.endswith(".xls") or name.endswith(".xlsx") or name.endswith(".csv")):
        raise HTTPException(status_code=400, detail="xls/xlsx/csv만 업로드 가능")

    suffix = Path(name).suffix or ".xlsx"
    tmp_path = Path(tempfile.gettempdir()) / f"yusaek_easyadmin_{uuid.uuid4().hex}{suffix}"
    data = await file.read()
    tmp_path.write_bytes(data)

    try:
        xls_bytes = _process_easyadmin_product_upload(tmp_path)
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"가공 실패: {e}")

    filename = f"easyadmin_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
    headers = {"Content-Disposition": _content_disposition(filename)}
    return Response(content=xls_bytes, media_type="application/vnd.ms-excel", headers=headers)


@app.get("/barcode/status")
def barcode_status(user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        return {"loaded": False}
    return {
        "loaded": True,
        "current_invoice": STATE["current_invoice"],
        "invoices": len(STATE["mapping"]),
        "processed_path": STATE["processed_path"],
        "items": _get_all_items(STATE["current_invoice"]) if STATE["current_invoice"] else [],
        "current_next": _get_first_remaining_item(STATE["current_invoice"]),
        "next_preview": _get_next_item_preview(STATE["current_invoice"]),
        "defects": _get_defect_list(),
        "invoice_has_defect": _invoice_has_defect(STATE["current_invoice"]),
    }


def _get_all_items(inv: str):
    """해당 송장의 모든 상품 목록(남은 수량 포함)"""
    mapping = STATE["mapping"]
    if inv not in mapping:
        return []

    if STATE["invoice_order"] and inv in STATE["invoice_order"]:
        codes = STATE["invoice_order"][inv]
    else:
        codes = sorted(mapping[inv].keys())

    incoming_counts = STATE.get("incoming_counts") or {}
    items = []
    for code in codes:
        remain = mapping[inv].get(code, 0)
        run_len = (STATE["runs"] or {}).get(inv, {}).get(code, 0)
        defect_n = (STATE["defect_counts"] or {}).get(code, 0)
        incoming_n = incoming_counts.get(code, 0)
        det = (STATE["details"] or {}).get(inv, {}).get(code, {})
        items.append(
            {
                "code": code,
                "name": det.get("name", "") or "",
                "option": det.get("option", "") or "",
                "remain": remain,
                "run_len": run_len,
                "defect": defect_n,
                "incoming": incoming_n,
            }
        )
    return items


def _get_first_remaining_item(inv: str | None):
    if not inv:
        return None
    for item in _get_all_items(inv):
        if item.get("remain", 0) > 0:
            return item
    return None


def _get_next_item_preview(current_invoice: str | None):
    seq = STATE.get("invoice_seq") or []
    if not seq:
        return None

    last_code = STATE.get("last_scanned_code")
    start_idx = seq.index(current_invoice) if current_invoice in seq else -1

    for i in range(start_idx + 1, len(seq)):
        inv = seq[i]
        item = _get_first_remaining_item(inv)
        if not item:
            continue
        run_len = item.get("run_len", 0)
        if run_len and run_len >= 10:
            continue
        if last_code and item.get("code") == last_code:
            continue
        return {"invoice": inv, **item}
    return None


def _invoice_has_defect(inv: str | None):
    if not inv:
        return False
    defect_counts = STATE.get("defect_counts") or {}
    mapping = STATE.get("mapping") or {}
    if inv not in mapping:
        return False
    for code in mapping[inv].keys():
        if defect_counts.get(code, 0) > 0:
            return True
    return False


def _find_item_detail_by_code(code: str):
    details = STATE.get("details") or {}
    for _, codes in details.items():
        det = codes.get(code)
        if det:
            return {
                "name": det.get("name", "") or "",
                "option": det.get("option", "") or "",
            }
    return {"name": "", "option": ""}


def _get_defect_list():
    defect_counts = STATE.get("defect_counts") or {}
    rows = []
    for code, n in sorted(defect_counts.items()):
        det = _find_item_detail_by_code(code)
        rows.append(
            {
                "code": code,
                "count": n,
                "name": det.get("name", ""),
                "option": det.get("option", ""),
            }
        )
    return rows


def _build_defect_csv() -> str:
    defect_counts = STATE.get("defect_counts") or {}
    code_o_text = STATE.get("code_o_text") or {}
    lines = ["A열(O왼쪽),B열(O오른쪽),C열(옵션명),D열(불량수량)"]
    for code, n in sorted(defect_counts.items()):
        det = _find_item_detail_by_code(code)
        opt = det.get("option", "") or ""
        o_text = (code_o_text.get(code) or "").strip()
        if not o_text:
            name = det.get("name", "") or ""
            o_text = f"{code} {name}".strip()
        o_text = str(o_text).strip().replace(",", " ")
        if " " in o_text:
            left, right = o_text.split(" ", 1)
        else:
            left, right = o_text, ""
        opt_clean = (opt or "").replace(",", " ")
        lines.append(f"{left},{right},{opt_clean},{n}")
    return "\n".join(lines) + "\n"


@app.post("/barcode/scan/invoice")
def scan_invoice(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")

    invoice = (payload.get("invoice") or "").strip()
    if not invoice:
        raise HTTPException(status_code=400, detail="invoice 값이 비어있음")

    if invoice not in STATE["mapping"]:
        return {"ok": False, "type": "invoice", "result": "NOT_FOUND", "invoice": invoice}

    STATE["current_invoice"] = invoice

    first_item = _get_first_remaining_item(invoice)
    if first_item:
        STATE["last_scanned_code"] = first_item.get("code")

    items = _get_all_items(invoice)

    return {
        "ok": True,
        "type": "invoice",
        "result": "SET",
        "invoice": invoice,
        "items": items,
        "current_next": first_item,
        "next_preview": _get_next_item_preview(invoice),
        "defects": _get_defect_list(),
        "invoice_has_defect": _invoice_has_defect(invoice),
    }


@app.post("/barcode/scan/item")
def scan_item(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")

    inv = STATE["current_invoice"]
    if not inv:
        return {"ok": False, "type": "item", "result": "NO_INVOICE"}

    raw = (payload.get("code") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="code 값이 비어있음")

    code = normalize_to_yusas(raw) or raw

    if inv not in STATE["mapping"]:
        return {"ok": False, "type": "item", "result": "BAD_INVOICE", "invoice": inv}

    remain = STATE["mapping"][inv].get(code, 0)
    det = (STATE["details"] or {}).get(inv, {}).get(code, {})
    name = det.get("name", "") or ""
    opt = det.get("option", "") or ""

    if remain <= 0:
        return {
            "ok": True,
            "type": "item",
            "result": "FALSE",
            "invoice": inv,
            "raw": raw,
            "code": code,
            "name": name,
            "option": opt,
            "remain": remain,
            "items": _get_all_items(inv),
            "current_next": _get_first_remaining_item(inv),
            "next_preview": _get_next_item_preview(inv),
            "defects": _get_defect_list(),
        }

    # TRUE 처리: -1
    STATE["mapping"][inv][code] = remain - 1
    STATE["last_scanned_code"] = code

    all_done = all(v == 0 for v in STATE["mapping"][inv].values())

    return {
        "ok": True,
        "type": "item",
        "result": "TRUE",
        "invoice": inv,
        "code": code,
        "name": name,
        "option": opt,
        "remain": STATE["mapping"][inv][code],
        "invoice_done": all_done,
        "items": _get_all_items(inv),
        "current_next": _get_first_remaining_item(inv),
        "next_preview": _get_next_item_preview(inv),
        "defects": _get_defect_list(),
    }


@app.post("/barcode/defect/add")
def add_defect(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")

    raw = (payload.get("code") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="code 값이 비어있음")

    code = normalize_to_yusas(raw) or raw
    defect_counts = STATE.get("defect_counts") or {}
    defect_counts[code] = defect_counts.get(code, 0) + 1
    STATE["defect_counts"] = defect_counts

    inv = STATE.get("current_invoice")
    return {
        "ok": True,
        "code": code,
        "defect_count": defect_counts[code],
        "items": _get_all_items(inv) if inv else [],
        "current_next": _get_first_remaining_item(inv),
        "next_preview": _get_next_item_preview(inv),
        "defects": _get_defect_list(),
    }


@app.get("/barcode/defect/list")
def list_defects(user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")
    return {"ok": True, "defects": _get_defect_list()}


@app.get("/barcode/defect/export")
def export_defects(user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")
    if not (STATE.get("defect_counts") or {}):
        raise HTTPException(status_code=400, detail="불량 목록이 비어있습니다")
    csv_text = _build_defect_csv()
    filename = f"defects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    headers = {"Content-Disposition": _content_disposition(filename)}
    csv_bytes = csv_text.encode("utf-8-sig")
    return Response(content=csv_bytes, media_type="text/csv; charset=utf-8", headers=headers)


@app.post("/barcode/defect/dec")
def decrement_defect(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")
    raw = (payload.get("code") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="code 값이 비어있음")
    code = normalize_to_yusas(raw) or raw
    defect_counts = STATE.get("defect_counts") or {}
    if code in defect_counts:
        defect_counts[code] -= 1
        if defect_counts[code] <= 0:
            del defect_counts[code]
    STATE["defect_counts"] = defect_counts
    inv = STATE.get("current_invoice")
    return {
        "ok": True,
        "defects": _get_defect_list(),
        "items": _get_all_items(inv) if inv else [],
        "current_next": _get_first_remaining_item(inv),
        "next_preview": _get_next_item_preview(inv),
    }


@app.post("/barcode/defect/remove")
def remove_defect(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    if not STATE["loaded"]:
        raise HTTPException(status_code=400, detail="먼저 엑셀을 업로드해주세요")
    raw = (payload.get("code") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="code 값이 비어있음")
    code = normalize_to_yusas(raw) or raw
    defect_counts = STATE.get("defect_counts") or {}
    if code in defect_counts:
        del defect_counts[code]
    STATE["defect_counts"] = defect_counts
    inv = STATE.get("current_invoice")
    return {
        "ok": True,
        "defects": _get_defect_list(),
        "items": _get_all_items(inv) if inv else [],
        "current_next": _get_first_remaining_item(inv),
        "next_preview": _get_next_item_preview(inv),
    }


@app.get("/users")
def list_users(user: str = Depends(_get_current_user)):
    conn = _get_db()
    rows = conn.execute("SELECT username, display_name FROM users ORDER BY username ASC").fetchall()
    conn.close()
    return {"ok": True, "users": [{"username": r["username"], "display_name": r["display_name"]} for r in rows]}


@app.get("/admin/users")
def admin_list_users(admin: str = Depends(_require_admin)):
    conn = _get_db()
    rows = conn.execute(
        "SELECT username, display_name, role, created_at FROM users ORDER BY username ASC"
    ).fetchall()
    conn.close()
    return {
        "ok": True,
        "users": [
            {
                "username": r["username"],
                "display_name": r["display_name"],
                "role": r["role"] if r["role"] else "user",
                "created_at": r["created_at"],
            }
            for r in rows
        ],
    }


@app.patch("/admin/users/{target}/role")
def admin_set_role(target: str, payload: dict = Body(...), admin: str = Depends(_require_admin)):
    role = (payload.get("role") or "").strip()
    if role not in ("admin", "user"):
        raise HTTPException(status_code=400, detail="invalid role")

    conn = _get_db()
    row = conn.execute("SELECT role FROM users WHERE username = ?", (target,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="user not found")

    current_role = row["role"] if row["role"] else "user"
    if current_role == "admin" and role != "admin" and _count_admins() <= 1:
        conn.close()
        raise HTTPException(status_code=400, detail="cannot remove last admin")

    conn.execute("UPDATE users SET role = ? WHERE username = ?", (role, target))
    conn.commit()
    conn.close()
    return {"ok": True, "username": target, "role": role}


@app.delete("/admin/users/{target}")
def admin_delete_user(target: str, admin: str = Depends(_require_admin)):
    if target == admin:
        raise HTTPException(status_code=400, detail="cannot delete self")

    conn = _get_db()
    row = conn.execute("SELECT role FROM users WHERE username = ?", (target,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="user not found")

    role = row["role"] if row["role"] else "user"
    if role == "admin" and _count_admins() <= 1:
        conn.close()
        raise HTTPException(status_code=400, detail="cannot delete last admin")

    conn.execute(
        "DELETE FROM requests WHERE requester_username = ? OR assignee_username = ?",
        (target, target),
    )
    conn.execute("DELETE FROM users WHERE username = ?", (target,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/requests")
def create_request(
    assignee: str = Form(...),
    text: str = Form(...),
    files: list[UploadFile] | None = File(None),
    user: str = Depends(_get_current_user),
):
    assignee = (assignee or "").strip()
    text = (text or "").strip()
    if not assignee or not text:
        raise HTTPException(status_code=400, detail="assignee/text required")

    files = files or []
    for f in files:
        ext = Path(f.filename or "").suffix.lower()
        if ext not in ALLOWED_REQUEST_EXTS:
            raise HTTPException(
                status_code=400,
                detail=f"unsupported file type: {ext or 'unknown'}",
            )

    requester_display = _get_user_display(user)
    assignee_display = _get_user_display(assignee)
    created_at = datetime.now(timezone.utc).isoformat()

    conn = _get_db()
    saved_paths: list[Path] = []
    try:
        cursor = conn.execute(
            """
            INSERT INTO requests (
                requester_username, requester_display,
                assignee_username, assignee_display,
                text, status, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (user, requester_display, assignee, assignee_display, text, "open", created_at),
        )
        request_id = cursor.lastrowid

        if files:
            request_dir = UPLOAD_BASE / str(request_id)
            request_dir.mkdir(parents=True, exist_ok=True)
            for f in files:
                ext = Path(f.filename or "").suffix.lower()
                stored_name = f"{uuid.uuid4().hex}{ext}"
                target_path = request_dir / stored_name
                with target_path.open("wb") as out:
                    shutil.copyfileobj(f.file, out)
                saved_paths.append(target_path)

                size = target_path.stat().st_size
                mime = f.content_type or mimetypes.guess_type(f.filename or "")[0] or "application/octet-stream"
                conn.execute(
                    """
                    INSERT INTO request_attachments (
                        request_id, original_name, stored_name, mime_type, size, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (request_id, f.filename or stored_name, stored_name, mime, size, created_at),
                )

        conn.commit()
    except HTTPException:
        conn.rollback()
        for path in saved_paths:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass
        raise
    except Exception:
        conn.rollback()
        for path in saved_paths:
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass
        raise HTTPException(status_code=500, detail="failed to create request")
    finally:
        conn.close()

    return {"ok": True}


@app.get("/requests/{request_id}/attachments/{attachment_id}")
def get_request_attachment(
    request_id: int,
    attachment_id: int,
    token: str | None = None,
    authorization: str | None = Header(None),
):
    user = _get_current_user_optional(authorization, token)
    conn = _get_db()
    req_row = conn.execute(
        "SELECT requester_username, assignee_username FROM requests WHERE id = ?",
        (request_id,),
    ).fetchone()
    if not req_row:
        conn.close()
        raise HTTPException(status_code=404, detail="request not found")
    if not (_is_admin(user) or user in (req_row["requester_username"], req_row["assignee_username"])):
        conn.close()
        raise HTTPException(status_code=403, detail="forbidden")

    file_row = conn.execute(
        "SELECT * FROM request_attachments WHERE id = ? AND request_id = ?",
        (attachment_id, request_id),
    ).fetchone()
    conn.close()
    if not file_row:
        raise HTTPException(status_code=404, detail="attachment not found")

    file_path = UPLOAD_BASE / str(request_id) / file_row["stored_name"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="file missing")

    return FileResponse(
        file_path,
        media_type=file_row["mime_type"],
        filename=file_row["original_name"],
    )


@app.post("/shared-files")
def upload_shared_file(
    file: UploadFile = File(...),
    user: str = Depends(_get_current_user),
):
    ext = Path(file.filename or "").suffix.lower()
    if ALLOWED_SHARED_EXTS and ext not in ALLOWED_SHARED_EXTS:
        raise HTTPException(status_code=400, detail="지원 형식: xlsx, xls, csv")

    created_at = datetime.now(timezone.utc).isoformat()
    uploader_display = _get_user_display(user)
    stored_name = f"{uuid.uuid4().hex}{ext}"
    SHARED_UPLOAD_BASE.mkdir(parents=True, exist_ok=True)
    target_path = SHARED_UPLOAD_BASE / stored_name

    try:
        with target_path.open("wb") as out:
            shutil.copyfileobj(file.file, out)
        size = target_path.stat().st_size
        mime = file.content_type or mimetypes.guess_type(file.filename or "")[0] or "application/octet-stream"
        conn = _get_db()
        conn.execute(
            """
            INSERT INTO shared_files (
                original_name, stored_name, mime_type, size,
                uploader_username, uploader_display, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                file.filename or stored_name,
                stored_name,
                mime,
                size,
                user,
                uploader_display,
                created_at,
            ),
        )
        conn.commit()
        conn.close()
    except Exception:
        try:
            target_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise HTTPException(status_code=500, detail="파일 업로드 실패")

    return {"ok": True}


@app.get("/shared-files")
def list_shared_files(user: str = Depends(_get_current_user)):
    conn = _get_db()
    rows = conn.execute(
        "SELECT * FROM shared_files ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return {"ok": True, "files": [_row_to_shared_file(r) for r in rows]}


@app.get("/shared-files/{file_id}")
def download_shared_file(
    file_id: int,
    token: str | None = None,
    authorization: str | None = Header(None),
):
    _get_current_user_optional(authorization, token)
    conn = _get_db()
    row = conn.execute("SELECT * FROM shared_files WHERE id = ?", (file_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="file not found")

    file_path = SHARED_UPLOAD_BASE / row["stored_name"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="file missing")

    return FileResponse(
        file_path,
        media_type=row["mime_type"],
        filename=row["original_name"],
    )


@app.delete("/shared-files/{file_id}")
def delete_shared_file(file_id: int, admin: str = Depends(_require_admin)):
    conn = _get_db()
    row = conn.execute("SELECT * FROM shared_files WHERE id = ?", (file_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="file not found")
    conn.execute("DELETE FROM shared_files WHERE id = ?", (file_id,))
    conn.commit()
    conn.close()

    file_path = SHARED_UPLOAD_BASE / row["stored_name"]
    try:
        file_path.unlink(missing_ok=True)
    except Exception:
        pass

    return {"ok": True}


@app.get("/requests/assigned")
def get_assigned_requests(user: str = Depends(_get_current_user)):
    target = user.strip()
    if not target:
        raise HTTPException(status_code=400, detail="assignee required")

    conn = _get_db()
    rows = conn.execute(
        "SELECT * FROM requests WHERE assignee_username = ? ORDER BY created_at DESC",
        (target,),
    ).fetchall()
    conn.close()

    visible_rows = [
        row
        for row in rows
        if not (row["status"] == "completed" and not _is_visible_completed(row["completed_at"]))
    ]
    attachments_map = _get_request_attachments([row["id"] for row in visible_rows])

    items = []
    for row in visible_rows:
        item = _row_to_request(row)
        item["can_complete"] = row["status"] == "open" and row["assignee_username"] == user
        item["attachments"] = attachments_map.get(row["id"], [])
        items.append(item)

    return {"ok": True, "assignee": target, "requests": items}


@app.delete("/requests/assigned/clear")
def clear_assigned_requests(user: str = Depends(_get_current_user)):
    conn = _get_db()
    conn.execute(
        "DELETE FROM requests WHERE assignee_username = ? AND status = 'completed'",
        (user,),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/requests/{request_id}/complete")
def complete_request(request_id: int, user: str = Depends(_get_current_user)):
    conn = _get_db()
    row = conn.execute("SELECT * FROM requests WHERE id = ?", (request_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="request not found")
    if row["assignee_username"] != user:
        conn.close()
        raise HTTPException(status_code=403, detail="forbidden")

    if row["status"] != "completed":
        conn.execute(
            "UPDATE requests SET status = ?, completed_at = ? WHERE id = ?",
            ("completed", datetime.now(timezone.utc).isoformat(), request_id),
        )
        conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/requests/resolved")
def get_resolved_requests(user: str = Depends(_get_current_user)):
    conn = _get_db()
    rows = conn.execute(
        """
        SELECT * FROM requests
        WHERE requester_username = ?
        ORDER BY (status = 'completed') DESC,
                 completed_at DESC,
                 created_at DESC
        """,
        (user,),
    ).fetchall()
    conn.close()

    visible_rows = [
        row
        for row in rows
        if not (row["status"] == "completed" and not _is_visible_completed(row["completed_at"]))
    ]
    attachments_map = _get_request_attachments([row["id"] for row in visible_rows])

    items = []
    for row in visible_rows:
        item = _row_to_request(row)
        item["can_ack"] = row["status"] == "completed" and row["acknowledged_at"] is None
        item["attachments"] = attachments_map.get(row["id"], [])
        items.append(item)

    return {"ok": True, "requests": items}


@app.delete("/requests/sent/clear")
def clear_sent_requests(user: str = Depends(_get_current_user)):
    conn = _get_db()
    conn.execute(
        "DELETE FROM requests WHERE requester_username = ? AND status = 'completed'",
        (user,),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/requests/{request_id}/ack")
def acknowledge_request(request_id: int, user: str = Depends(_get_current_user)):
    conn = _get_db()
    row = conn.execute("SELECT * FROM requests WHERE id = ?", (request_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="request not found")
    if row["requester_username"] != user:
        conn.close()
        raise HTTPException(status_code=403, detail="forbidden")

    conn.execute(
        "UPDATE requests SET acknowledged_at = ? WHERE id = ?",
        (datetime.now(timezone.utc).isoformat(), request_id),
    )
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------- Company Credentials ----------
@app.get("/company-credentials")
def list_company_credentials(user: str = Depends(_get_current_user)):
    is_admin = _is_admin(user)
    conn = _get_db()
    rows = conn.execute(
        "SELECT id, label, username, password, updated_at, created_at FROM company_credentials ORDER BY id DESC"
    ).fetchall()
    conn.close()
    items = []
    for row in rows:
        has_credentials = bool((row["username"] or "").strip() or (row["password"] or "").strip())
        item = {
            "id": row["id"],
            "label": row["label"],
            "has_credentials": has_credentials,
            "updated_at": row["updated_at"],
            "created_at": row["created_at"],
        }
        if is_admin:
            item["username"] = row["username"] or ""
            item["password"] = row["password"] or ""
        items.append(item)
    return {"ok": True, "items": items}


@app.get("/company-credentials/pin")
def get_company_pin_status(user: str = Depends(_get_current_user)):
    pin_hash = _get_setting("company_pin_hash")
    return {"ok": True, "has_pin": bool(pin_hash)}


@app.post("/company-credentials/pin")
def set_company_pin(payload: dict = Body(...), admin: str = Depends(_require_admin)):
    pin = (payload.get("pin") or "").strip()
    if not re.fullmatch(r"\d{4}", pin or ""):
        raise HTTPException(status_code=400, detail="4자리 PIN이 필요합니다.")
    _set_setting("company_pin_hash", _hash_pin(pin))
    return {"ok": True}


@app.post("/company-credentials")
def upsert_company_credentials(
    payload: dict = Body(...),
    admin: str = Depends(_require_admin),
):
    label = (payload.get("label") or "").strip()
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    cid = payload.get("id")

    if not label:
        raise HTTPException(status_code=400, detail="label이 필요합니다.")
    now = datetime.now(timezone.utc).isoformat()
    conn = _get_db()
    if cid:
        row = conn.execute("SELECT id FROM company_credentials WHERE id = ?", (cid,)).fetchone()
        if not row:
            conn.close()
            raise HTTPException(status_code=404, detail="not found")
        conn.execute(
            """
            UPDATE company_credentials
            SET label = ?, username = ?, password = ?, updated_at = ?
            WHERE id = ?
            """,
            (label, username or None, password or None, now, cid),
        )
    else:
        conn.execute(
            """
            INSERT INTO company_credentials (label, username, password, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (label, username or None, password or None, now, now),
        )
    conn.commit()
    conn.close()
    return {"ok": True}


@app.delete("/company-credentials/{cred_id}")
def delete_company_credentials(cred_id: int, admin: str = Depends(_require_admin)):
    conn = _get_db()
    conn.execute("DELETE FROM company_credentials WHERE id = ?", (cred_id,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/company-credentials/{cred_id}/view")
def view_company_credentials(
    cred_id: int,
    payload: dict = Body(...),
    user: str = Depends(_get_current_user),
):
    pin = (payload.get("pin") or "").strip()
    if not re.fullmatch(r"\d{4}", pin or ""):
        raise HTTPException(status_code=400, detail="4자리 PIN이 필요합니다.")

    pin_hash = _get_setting("company_pin_hash")
    if not pin_hash or not _verify_pin(pin, pin_hash):
        raise HTTPException(status_code=403, detail="pin mismatch")

    conn = _get_db()
    row = conn.execute("SELECT * FROM company_credentials WHERE id = ?", (cred_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="not found")

    return {
        "ok": True,
        "label": row["label"],
        "username": row["username"] or "",
        "password": row["password"] or "",
        "updated_at": row["updated_at"],
    }


# ---------- Return (반품) API ----------
@app.get("/returns/state")
def returns_state(user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    return {
        "ok": True,
        "status": _return_status(state),
        "queues": _return_queue_payload(state),
        "onebe": {
            "rows": _return_rows(state.customer_export_df),
        },
        "last_type": state.last_type,
        "scanned_count": len(state.scanned_barcodes),
    }


# ---------- AMOOD Excel API ----------
@app.get("/amood/status")
def amood_status(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    return {
        "ok": True,
        "status": _amood_status(state),
        "items": state.pending_items,
    }


@app.post("/amood/incoming/upload")
async def amood_incoming_upload(file: UploadFile = File(...), user: str = Depends(_get_current_user)):
    name = (file.filename or "").lower()
    if not (name.endswith(".xls") or name.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="xls/xlsx files only")

    suffix = ".xlsx" if name.endswith(".xlsx") else ".xls"
    tmp_path = Path(tempfile.gettempdir()) / f"amood_incoming_{uuid.uuid4().hex}{suffix}"
    data = await file.read()
    tmp_path.write_bytes(data)

    try:
        wb, ws = load_excel_any(tmp_path)
        counts = Counter()
        for r in range(1, ws.max_row + 1):
            code_raw = ws.cell(r, 1).value
            qty_raw = ws.cell(r, 2).value
            code = _amood_norm_barcode(code_raw)
            if not code:
                continue
            qty = _amood_to_int_qty(qty_raw)
            if qty > 0:
                counts[code] += qty
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"incoming load failed: {e}")

    state = _get_amood_state(user)
    state.incoming_counts = dict(counts)
    return {
        "ok": True,
        "codes": len(counts),
        "total_qty": sum(counts.values()),
        "status": _amood_status(state),
    }


@app.post("/amood/excel1")
def amood_upload_excel1(file: UploadFile = File(...), user: str = Depends(_get_current_user)):
    name = file.filename or ""
    ext = Path(name).suffix.lower()
    if ext not in AMOOD_ALLOWED_EXCEL1:
        raise HTTPException(status_code=400, detail="excel1은 .xlsx/.xlsm만 가능합니다.")
    tmp_path = Path(tempfile.gettempdir()) / f"amood_excel1_{uuid.uuid4().hex}{ext}"
    with tmp_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)
    state = _get_amood_state(user)
    state.file1_path = tmp_path
    state.file1_name = name or tmp_path.name
    state.processed1_path = None
    state.processed2_path = None
    state.wb1 = None
    state.ws1 = None
    state.current_invoice = None
    state.pending_items = []
    state.waiting_for_items = False
    return {"ok": True, "status": _amood_status(state)}


@app.post("/amood/excel2")
def amood_upload_excel2(file: UploadFile = File(...), user: str = Depends(_get_current_user)):
    name = file.filename or ""
    ext = Path(name).suffix.lower()
    if ext not in AMOOD_ALLOWED_EXCEL2:
        raise HTTPException(status_code=400, detail="excel2는 .xls/.xlsx/.xlsm/.htm/.html만 가능합니다.")
    tmp_path = Path(tempfile.gettempdir()) / f"amood_excel2_{uuid.uuid4().hex}{ext}"
    with tmp_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)
    state = _get_amood_state(user)
    state.file2_path = tmp_path
    state.file2_name = name or tmp_path.name
    state.processed1_path = None
    state.processed2_path = None
    state.wb2 = None
    state.ws2 = None
    state.current_invoice = None
    state.pending_items = []
    state.waiting_for_items = False
    return {"ok": True, "status": _amood_status(state)}


@app.post("/amood/preprocess")
def amood_preprocess(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    if not state.file1_path or not state.file2_path:
        raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
    try:
        _amood_load_workbooks(state)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 로드 실패")
    wb1, ws1 = state.wb1, state.ws1
    wb2, ws2 = state.wb2, state.ws2

    for r in range(2, ws1.max_row + 1):
        v = _amood_ws_cell(ws1, AMOOD_COL1_NAME_RAW, r).value
        if v is None:
            continue
        _amood_ws_cell(ws1, AMOOD_COL1_NAME_RAW, r).value = _amood_strip_any_brackets(str(v))

    _amood_fill_down_merged_column(ws2, AMOOD_COL2_ORDER_KEY, start_row=2)
    # E열 값을 비웠다가 숫자 문자열로 다시 써줌
    for r in range(2, ws2.max_row + 1):
        cell = _amood_ws_cell(ws2, AMOOD_COL2_ORDER_KEY, r)
        original = cell.value
        cell.value = None
        cell.value = _amood_norm_key(original)

    for r in range(2, ws2.max_row + 1):
        name = _amood_ws_cell(ws2, AMOOD_COL2_NAME, r).value
        opt = _amood_ws_cell(ws2, AMOOD_COL2_OPTION, r).value
        qty = _amood_ws_cell(ws2, AMOOD_COL2_QTY, r).value
        out = _amood_build_output_text(name, opt, qty)
        _amood_ws_cell(ws2, AMOOD_COL2_OUTPUT, r).value = out

    out1 = Path(tempfile.gettempdir()) / f"amood_excel1_processed_{uuid.uuid4().hex}.xlsx"
    out2 = Path(tempfile.gettempdir()) / f"amood_excel2_processed_{uuid.uuid4().hex}.xlsx"
    wb1.save(out1)
    wb2.save(out2)

    state.processed1_path = out1
    state.processed2_path = out2
    return {"ok": True, "status": _amood_status(state)}


@app.get("/amood/download/1")
def amood_download_excel1(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    if not state.processed1_path or not state.processed1_path.exists():
        raise HTTPException(status_code=404, detail="전처리 결과가 없습니다.")
    name = state.file1_name or "amood_excel1"
    filename = f"{Path(name).stem}_processed.xlsx"
    return FileResponse(state.processed1_path, filename=filename)


@app.get("/amood/download/2")
def amood_download_excel2(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    if not state.processed2_path or not state.processed2_path.exists():
        raise HTTPException(status_code=404, detail="전처리 결과가 없습니다.")
    name = state.file2_name or "amood_excel2"
    filename = f"{Path(name).stem}_processed.xlsx"
    return FileResponse(state.processed2_path, filename=filename)


def _amood_items_view(state: AmoodState) -> list[dict]:
    items = []
    incoming_counts = state.incoming_counts or {}
    for it in state.pending_items:
        code = it.get("barcode", "")
        incoming_n = incoming_counts.get(_amood_norm_barcode(code), 0)
        items.append(
            {
                "code": code,
                "name": it.get("name", "") or "",
                "option": it.get("option", "") or "",
                "remain": it.get("remaining", 0),
                "incoming": incoming_n,
            }
        )
    return items


def _amood_first_remaining(state: AmoodState):
    incoming_counts = state.incoming_counts or {}
    for it in state.pending_items:
        if it.get("remaining", 0) > 0:
            code = it.get("barcode", "")
            incoming_n = incoming_counts.get(_amood_norm_barcode(code), 0)
            return {
                "code": code,
                "name": it.get("name", "") or "",
                "option": it.get("option", "") or "",
                "remain": it.get("remaining", 0),
                "incoming": incoming_n,
            }
    return None


def _amood_reset_state(state: AmoodState):
    for path in [state.file1_path, state.file2_path, state.processed1_path, state.processed2_path]:
        if path and isinstance(path, Path) and path.exists():
            try:
                path.unlink(missing_ok=True)
            except Exception:
                pass
    state.file1_path = None
    state.file2_path = None
    state.file1_name = None
    state.file2_name = None
    state.processed1_path = None
    state.processed2_path = None
    state.wb1 = None
    state.ws1 = None
    state.wb2 = None
    state.ws2 = None
    state.current_invoice = None
    state.pending_items = []
    state.waiting_for_items = False
    state.completed_mgmt_numbers = set()
    state.incoming_counts = {}


@app.get("/amood/scan/status")
def amood_scan_status(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    return {
        "ok": True,
        "current_invoice": state.current_invoice,
        "items": _amood_items_view(state),
        "current_next": _amood_first_remaining(state),
    }


@app.post("/amood/reset")
def amood_reset(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    _amood_reset_state(state)
    return {"ok": True, "status": _amood_status(state)}


@app.post("/amood/scan/invoice")
def amood_scan_invoice(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    _amood_load_workbooks(state)

    invoice = (payload.get("invoice") or "").strip()
    if not invoice:
        raise HTTPException(status_code=400, detail="invoice 값이 비어있음")

    _amood_fill_down_merged_column(state.ws2, AMOOD_COL2_ORDER_KEY, start_row=2)

    r1_list = []
    target = _amood_norm_barcode(invoice).upper()
    for r in range(2, state.ws1.max_row + 1):
        v = _amood_ws_cell(state.ws1, AMOOD_COL1_SCAN_BARCODE, r).value
        if not v:
            continue
        if _amood_norm_barcode(v).upper() == target:
            r1_list.append(r)

    if not r1_list:
        return {"ok": False, "type": "invoice", "result": "NOT_FOUND", "invoice": invoice}

    order_keys = []
    seen = set()
    for r1 in r1_list:
        ok = _amood_ws_cell(state.ws1, AMOOD_COL1_ORDER_KEY, r1).value
        ok = _amood_norm_key(ok)
        if ok and ok not in seen:
            seen.add(ok)
            order_keys.append(ok)

    if not order_keys:
        return {"ok": False, "type": "invoice", "result": "NO_ORDER_KEY", "invoice": invoice}

    pending: list[dict] = []
    for order_key in order_keys:
        rows2 = _amood_collect_rows_by_value(state.ws2, AMOOD_COL2_ORDER_KEY, order_key, start_row=2)
        for r in rows2:
            qty = _amood_to_int_qty(_amood_ws_cell(state.ws2, AMOOD_COL2_QTY, r).value)
            if qty <= 0:
                continue
            bc = _amood_ws_cell(state.ws2, AMOOD_COL2_BARCODE, r).value
            bc = str(bc).strip() if bc is not None else ""
            name = _amood_ws_cell(state.ws2, AMOOD_COL2_NAME, r).value
            option = _amood_ws_cell(state.ws2, AMOOD_COL2_OPTION, r).value
            disp = _amood_ws_cell(state.ws2, AMOOD_COL2_OUTPUT, r).value
            if disp is None or str(disp).strip() == "":
                disp = _amood_build_output_text(name, option, qty)
            pending.append(
                {
                    "row": r,
                    "barcode": bc,
                    "name": str(name).strip() if name is not None else "",
                    "option": str(option).strip() if option is not None else "",
                    "display": str(disp).strip() if disp is not None else "",
                    "remaining": qty,
                }
            )

    if not pending:
        return {"ok": False, "type": "invoice", "result": "NO_ITEMS", "invoice": invoice}

    state.current_invoice = invoice
    state.pending_items = pending
    state.waiting_for_items = True

    return {
        "ok": True,
        "type": "invoice",
        "result": "SET",
        "invoice": invoice,
        "items": _amood_items_view(state),
        "current_next": _amood_first_remaining(state),
    }


@app.post("/amood/scan/item")
def amood_scan_item(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    if not state.waiting_for_items or not state.pending_items:
        return {"ok": False, "type": "item", "result": "NO_INVOICE"}

    raw = (payload.get("code") or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="code 값이 비어있음")

    scan = _amood_norm_barcode(raw)
    matched = None
    for it in state.pending_items:
        if it.get("remaining", 0) <= 0:
            continue
        target = _amood_norm_barcode(it.get("barcode", ""))
        if target and (target == scan or scan in target or target in scan):
            matched = it
            break

    if matched is None:
        return {
            "ok": True,
            "type": "item",
            "result": "FALSE",
            "code": raw,
            "remain": 0,
            "items": _amood_items_view(state),
        }

    matched["remaining"] = int(matched.get("remaining", 0)) - 1
    try:
        _amood_ws_cell(state.ws2, AMOOD_COL2_QTY, matched["row"]).value = matched["remaining"]
    except Exception:
        pass

    all_done = all(it.get("remaining", 0) <= 0 for it in state.pending_items)
    if all_done:
        state.waiting_for_items = False

    return {
        "ok": True,
        "type": "item",
        "result": "TRUE",
        "code": matched.get("barcode", "") or raw,
        "remain": matched.get("remaining", 0),
        "invoice_done": all_done,
        "items": _amood_items_view(state),
        "current_next": _amood_first_remaining(state),
    }


@app.post("/amood/export-shipping")
def amood_export_shipping(user: str = Depends(_get_current_user)):
    state = _get_amood_state(user)
    if not state.file1_path or not state.file2_path:
        raise HTTPException(status_code=400, detail="excel1/excel2가 모두 필요합니다.")
    try:
        _amood_load_workbooks(state)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 로드 실패")
    wb1, ws1 = state.wb1, state.ws1
    wb2, ws2 = state.wb2, state.ws2

    _amood_fill_down_merged_column(ws2, AMOOD_COL2_ORDER_KEY, start_row=2)

    rows: list[dict] = []
    seen_codes: set[str] = set()
    for r in range(2, ws1.max_row + 1):
        order_key = _amood_ws_cell(ws1, AMOOD_COL1_ORDER_KEY, r).value
        if not order_key:
            continue
        order_key = str(order_key).strip()
        matched_rows = _amood_collect_rows_by_value(ws2, AMOOD_COL2_ORDER_KEY, order_key, start_row=2)
        if not matched_rows:
            continue
        outputs: list[str] = []
        for r2 in matched_rows:
            out_val = _amood_ws_cell(ws2, AMOOD_COL2_OUTPUT, r2).value
            if out_val is None or str(out_val).strip() == "":
                name = _amood_ws_cell(ws2, AMOOD_COL2_NAME, r2).value
                option = _amood_ws_cell(ws2, AMOOD_COL2_OPTION, r2).value
                qty = _amood_ws_cell(ws2, AMOOD_COL2_QTY, r2).value
                out_val = _amood_build_output_text(name, option, qty)
            out_text = str(out_val).strip() if out_val is not None else ""
            if out_text and out_text not in outputs:
                outputs.append(out_text)
        description = " / ".join(outputs)
        code = _amood_ws_cell(ws1, AMOOD_COL1_SCAN_BARCODE, r).value
        code = str(code).strip() if code else ""
        if not code:
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)
        b_val = _amood_ws_cell(ws1, AMOOD_COL1_NUM_B, r).value
        c_val = _amood_ws_cell(ws1, AMOOD_COL1_NUM_C, r).value
        try:
            b_val = int(b_val)
            c_val = int(c_val)
        except Exception:
            continue
        title = f"{c_val}-{b_val}"
        rows.append({"Title": title, "Description": description, "Code": code})

    if not rows:
        raise HTTPException(status_code=400, detail="추출할 데이터가 없습니다.")

    df = pd.DataFrame(rows, columns=["Title", "Description", "Code"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    buf.seek(0)
    filename = "선적바코드_추출.xlsx"
    headers = {"Content-Disposition": _content_disposition(filename)}
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/returns/excel1")
def returns_upload_excel1(
    file: UploadFile = File(...),
    user: str = Depends(_get_current_user),
):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in RETURN_ALLOWED_EXTS:
        raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능")

    tmp_path = Path(tempfile.gettempdir()) / f"returns_excel1_{uuid.uuid4().hex}{ext}"
    with tmp_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)

    try:
        df = _read_return_excel(tmp_path)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    if df.shape[1] < 5:
        raise HTTPException(status_code=400, detail="1번 엑셀에 D/E열이 없습니다. (열 개수가 부족)")

    df["D_clean"] = df.iloc[:, 3].apply(_clean_invoice)
    df["E_clean"] = df.iloc[:, 4].apply(_clean_invoice)

    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        d = row.get("D_clean", "")
        e = row.get("E_clean", "")
        if d and d not in mapping:
            mapping[d] = e

    state = _get_return_state(user)
    state.df1 = df
    state.map_d_to_e = mapping
    return {"ok": True, "map_count": len(mapping), "status": _return_status(state)}


@app.post("/returns/excel2")
def returns_upload_excel2(
    file: UploadFile = File(...),
    user: str = Depends(_get_current_user),
):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in RETURN_ALLOWED_EXTS:
        raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능")

    tmp_path = Path(tempfile.gettempdir()) / f"returns_excel2_{uuid.uuid4().hex}{ext}"
    with tmp_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)

    try:
        df = _read_return_excel(tmp_path)
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    if df.shape[1] < 13:
        raise HTTPException(status_code=400, detail="2번 엑셀에 필요한 열(F,G,H,K,M)이 없습니다. (열 개수가 부족)")

    df["F_name"] = df.iloc[:, 5].apply(_clean_product_name)
    df["G_opt"] = df.iloc[:, 6].apply(_lowercase_size_words).apply(_option_slash_to_space)
    df["QTY"] = df.iloc[:, 7].apply(_clean_qty)
    df["ITEM_TEXT"] = df.apply(lambda r: _normalize_spaces(f"{r.get('F_name','')} {r.get('G_opt','')}"), axis=1)
    df["REASON_TYPE"] = df.iloc[:, 10].apply(_reason_type)
    df["M_clean"] = df.iloc[:, 12].apply(_clean_invoice)

    idx: dict[str, list[int]] = {}
    for i, v in enumerate(df["M_clean"].tolist()):
        if not v:
            continue
        idx.setdefault(v, []).append(i)

    state = _get_return_state(user)
    state.df2 = df
    state.df2_index = idx
    return {"ok": True, "index_count": len(idx), "status": _return_status(state)}


@app.post("/returns/cost-base/reload")
def returns_cost_base_reload(user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    try:
        _load_return_cost_base(state)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")
    return {"ok": True, "cost_count": len(state.cost_map), "status": _return_status(state)}


@app.post("/returns/cost-base/upload")
def returns_cost_base_upload(
    file: UploadFile = File(...),
    admin: str = Depends(_require_admin),
):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in RETURN_ALLOWED_EXTS:
        raise HTTPException(status_code=400, detail="xls/xlsx/xlsm만 업로드 가능")

    RETURN_COST_BASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = Path(tempfile.gettempdir()) / f"returns_cost_base_{uuid.uuid4().hex}{ext}"
    with tmp_path.open("wb") as out:
        shutil.copyfileobj(file.file, out)

    try:
        df = _read_return_excel(tmp_path)
        if df.shape[1] < 2:
            raise HTTPException(status_code=400, detail="원가베이스는 최소 A,B열이 필요합니다.")
        shutil.move(str(tmp_path), str(RETURN_COST_BASE_PATH))
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    # 업데이트된 파일로 다시 로드
    state = _get_return_state(admin)
    state.cost_base_path = RETURN_COST_BASE_PATH
    _load_return_cost_base(state)

    return {"ok": True, "status": _return_status(state)}


@app.get("/returns/cost-base/download")
def returns_cost_base_download(admin: str = Depends(_require_admin)):
    path = RETURN_COST_BASE_PATH
    if not path.exists():
        raise HTTPException(status_code=404, detail="원가베이스 파일이 없습니다.")
    return FileResponse(path, filename=path.name)


@app.get("/returns/cost-base/preview")
def returns_cost_base_preview(
    offset: int = 0,
    limit: int = 50,
    q: str | None = None,
    user: str = Depends(_get_current_user),
):
    if offset < 0 or limit <= 0 or limit > 200:
        raise HTTPException(status_code=400, detail="offset/limit 값이 올바르지 않습니다.")
    try:
        df = _load_cost_base_df()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")

    if q:
        q_norm = str(q).strip()
    else:
        q_norm = ""

    if q_norm:
        df_view = df.fillna("").astype(str)
        mask = df_view.apply(lambda row: row.str.contains(q_norm, case=False, na=False)).any(axis=1)
        df_filtered = df[mask].copy()
    else:
        df_filtered = df

    total = len(df_filtered)
    col_names = ["1열", "2열"]
    end = min(offset + limit, total)
    rows = []
    for i in range(offset, end):
        r = df_filtered.iloc[i]
        row = []
        for v in r.iloc[:2].values.tolist():
            if pd.isna(v):
                row.append("")
            else:
                row.append(v)
        rows.append({"row_index": int(r.name), "values": row})
    return {"ok": True, "columns": col_names, "rows": rows, "total": total}


@app.post("/returns/cost-base/edit")
def returns_cost_base_edit(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    row_index = payload.get("row_index")
    column = payload.get("column")
    value = payload.get("value")

    if row_index is None or not isinstance(row_index, int) or row_index < 0:
        raise HTTPException(status_code=400, detail="row_index가 올바르지 않습니다.")
    if not isinstance(column, (str, int)):
        raise HTTPException(status_code=400, detail="column 값이 올바르지 않습니다.")

    try:
        df = _load_cost_base_df()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")

    if row_index >= len(df):
        raise HTTPException(status_code=400, detail="row_index 범위를 벗어났습니다.")

    if isinstance(column, int):
        if column < 0 or column >= len(df.columns):
            raise HTTPException(status_code=400, detail="column 범위를 벗어났습니다.")
        col_name = df.columns[column]
    else:
        if column not in df.columns:
            raise HTTPException(status_code=400, detail="유효하지 않은 column 입니다.")
        col_name = column

    df.at[row_index, col_name] = "" if value is None else value
    try:
        _save_cost_base_df(df)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"원가베이스 저장 실패: {e}")

    return {"ok": True}


@app.post("/returns/cost-base/edit-batch")
def returns_cost_base_edit_batch(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    edits = payload.get("edits")
    if not isinstance(edits, list) or not edits:
        raise HTTPException(status_code=400, detail="edits 값이 올바르지 않습니다.")

    try:
        df = _load_cost_base_df()
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"원가베이스 로드 실패: {e}")

    for item in edits:
        row_index = item.get("row_index")
        column = item.get("column")
        value = item.get("value")
        if row_index is None or not isinstance(row_index, int) or row_index < 0:
            continue
        if row_index >= len(df):
            continue
        if isinstance(column, int):
            if column < 0 or column >= len(df.columns):
                continue
            col_name = df.columns[column]
        elif isinstance(column, str) and column in df.columns:
            col_name = column
        else:
            continue
        df.at[row_index, col_name] = "" if value is None else value

    try:
        _save_cost_base_df(df)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"원가베이스 저장 실패: {e}")

    return {"ok": True}


@app.post("/returns/scan")
def returns_scan(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    barcode_raw = (payload.get("barcode") or "").strip()
    barcode = _clean_invoice(barcode_raw)
    if not barcode:
        raise HTTPException(status_code=400, detail="barcode 값이 비어있음")

    state = _get_return_state(user)

    if barcode in state.scanned_barcodes:
        state.last_type = "중복"
        return {
            "ok": True,
            "duplicate": True,
            "last_type": state.last_type,
            "queues": _return_queue_payload(state),
        }

    if not state.map_d_to_e:
        raise HTTPException(status_code=400, detail="먼저 1번 엑셀을 불러오세요.")
    if state.df2 is None or state.df2_index is None:
        raise HTTPException(status_code=400, detail="먼저 2번 엑셀을 불러오세요.")

    e_val = state.map_d_to_e.get(barcode, "")
    if not e_val:
        msg = f"[미매칭] 스캔:{barcode} → 1번(D)에서 찾지 못함"
        state.queue_unmatched.append(
            {"id": state.next_id, "scan": barcode, "match": "", "item_text": msg, "qty": "", "type": "미매칭"}
        )
        state.next_id += 1
        state.last_type = "미매칭"
        return {"ok": True, "last_type": state.last_type, "queues": _return_queue_payload(state)}

    row_indexes = state.df2_index.get(e_val, [])
    if not row_indexes:
        msg = f"[미매칭] 스캔:{barcode} → 1번(E):{e_val} → 2번(M)에서 찾지 못함"
        state.queue_unmatched.append(
            {"id": state.next_id, "scan": barcode, "match": e_val, "item_text": msg, "qty": "", "type": "미매칭"}
        )
        state.next_id += 1
        state.last_type = "미매칭"
        return {"ok": True, "last_type": state.last_type, "queues": _return_queue_payload(state)}

    state.last_added_ids = []
    last_types = set()

    for row_i in row_indexes:
        row = state.df2.iloc[row_i]
        item_text = row.get("ITEM_TEXT", "")
        qty = row.get("QTY", "")
        rtype = row.get("REASON_TYPE", "미매칭")
        if rtype not in ("판매자", "고객"):
            rtype = "미매칭"

        item = {
            "id": state.next_id,
            "scan": barcode,
            "match": e_val,
            "item_text": item_text,
            "qty": qty,
            "type": rtype,
        }
        state.next_id += 1
        state.last_added_ids.append(item["id"])

        if rtype == "판매자":
            state.queue_seller.append(item)
        elif rtype == "고객":
            state.queue_customer.append(item)
        else:
            state.queue_unmatched.append(item)

        state.all_items.append(item)
        last_types.add(rtype)

    if len(last_types) == 1:
        state.last_type = next(iter(last_types))
    else:
        state.last_type = "혼합(" + ",".join(sorted(last_types)) + ")"

    state.scanned_barcodes.add(barcode)

    return {"ok": True, "last_type": state.last_type, "queues": _return_queue_payload(state)}


@app.post("/returns/undo")
def returns_undo(user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    if not state.last_added_ids:
        raise HTTPException(status_code=400, detail="삭제할 최근 스캔 기록이 없습니다.")

    remove_ids = set(state.last_added_ids)
    state.queue_seller = [it for it in state.queue_seller if it.get("id") not in remove_ids]
    state.queue_customer = [it for it in state.queue_customer if it.get("id") not in remove_ids]
    state.queue_unmatched = [it for it in state.queue_unmatched if it.get("id") not in remove_ids]
    state.all_items = [it for it in state.all_items if it.get("id") not in remove_ids]
    state.last_added_ids = []
    state.last_type = "-"
    return {"ok": True, "queues": _return_queue_payload(state), "last_type": state.last_type}


@app.post("/returns/reset")
def returns_reset(user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    state.queue_seller.clear()
    state.queue_customer.clear()
    state.queue_unmatched.clear()
    state.all_items.clear()
    state.last_added_ids.clear()
    state.scanned_barcodes.clear()
    state.customer_export_df = pd.DataFrame()
    state.last_type = "-"
    return {"ok": True}


@app.post("/returns/onebe/build")
def returns_build_onebe(payload: dict = Body(None), user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    source = (payload or {}).get("source", "customer")
    if source == "all":
        items = state.all_items
        if not items:
            raise HTTPException(status_code=400, detail="전체 대기 데이터가 없습니다.")
    else:
        items = state.queue_customer
        if not items:
            raise HTTPException(status_code=400, detail="고객 대기 데이터가 없습니다.")

    if not state.cost_map:
        try:
            _load_return_cost_base(state)
        except Exception:
            raise HTTPException(status_code=400, detail="원가베이스를 먼저 불러오세요.")

    rows = []
    for it in items:
        item_text = _normalize_spaces(it.get("item_text", ""))
        match_key = _normalize_key(item_text)
        product_code = state.cost_map.get(match_key, "")
        matched_flag = "O" if product_code else "X"

        rows.append(
            {
                "상품코드": product_code,
                "요청수량": 0,
                "수량": it.get("qty", ""),
                "가공데이터": item_text,
                "스캔송장": it.get("scan", ""),
                "매칭송장": it.get("match", ""),
                "분류": it.get("type", "고객"),
                "원가베이스매칭": matched_flag,
            }
        )

    state.customer_export_df = pd.DataFrame(rows)
    return {"ok": True, "onebe": {"rows": _return_rows(state.customer_export_df)}}


@app.post("/returns/onebe/consolidate")
def returns_consolidate_onebe(user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    if state.customer_export_df is None or state.customer_export_df.empty:
        raise HTTPException(status_code=400, detail="먼저 '고객대기 → 원베양식 생성'을 실행하세요.")

    df = state.customer_export_df.copy()
    has_code = df["상품코드"].fillna("").astype(str).str.strip() != ""
    df_code = df[has_code].copy()
    df_empty = df[~has_code].copy()

    def to_int_safe(x):
        if x is None:
            return 0
        s = str(x).strip()
        if s == "" or s.lower() in ("nan", "none"):
            return 0
        try:
            return int(float(s))
        except Exception:
            return 0

    df_code["_qty_int"] = df_code["수량"].apply(to_int_safe)

    def merge_match_invoices(series):
        seen = set()
        out = []
        for v in series.fillna("").astype(str).tolist():
            v = v.strip()
            if not v or v.lower() in ("nan", "none"):
                continue
            parts = [p.strip() for p in v.split(",") if p.strip()]
            for p in parts:
                if p not in seen:
                    seen.add(p)
                    out.append(p)
        return ",".join(out)

    agg = (
        df_code.groupby("상품코드", as_index=False)
        .agg(
            {
                "요청수량": "first",
                "_qty_int": "sum",
                "가공데이터": "first",
                "스캔송장": "first",
                "매칭송장": merge_match_invoices,
                "분류": "first",
                "원가베이스매칭": "first",
            }
        )
    )

    agg["입고수량"] = agg["_qty_int"].astype(int)
    agg.drop(columns=["_qty_int"], inplace=True)
    agg.rename(columns={"매칭송장": "요청메모"}, inplace=True)

    new_df = pd.concat([agg, df_empty], ignore_index=True)
    state.customer_export_df = new_df
    return {"ok": True, "onebe": {"rows": _return_rows(state.customer_export_df)}}


@app.post("/returns/onebe/edit")
def returns_edit_onebe(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    if state.customer_export_df is None or state.customer_export_df.empty:
        raise HTTPException(status_code=400, detail="원베양식 데이터가 없습니다.")

    row_index = payload.get("row_index")
    column = (payload.get("column") or "").strip()
    value = (payload.get("value") or "").strip()

    if row_index is None or not isinstance(row_index, int):
        raise HTTPException(status_code=400, detail="row_index가 필요합니다.")
    if column not in state.customer_export_df.columns:
        raise HTTPException(status_code=400, detail="유효하지 않은 컬럼입니다.")
    if row_index < 0 or row_index >= len(state.customer_export_df):
        raise HTTPException(status_code=400, detail="유효하지 않은 행입니다.")

    if column in ("요청수량", "수량"):
        if value == "":
            value = "0"
        try:
            value = str(int(float(value)))
        except Exception:
            raise HTTPException(status_code=400, detail="수량은 숫자여야 합니다.")

    state.customer_export_df.at[row_index, column] = value
    return {"ok": True}


@app.post("/returns/download/onebe")
def returns_download_onebe(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    if state.customer_export_df is None or state.customer_export_df.empty:
        raise HTTPException(status_code=400, detail="원베양식 데이터가 없습니다.")

    columns = payload.get("columns") or []
    if not isinstance(columns, list) or not columns:
        columns = ["상품코드", "요청수량", "수량"]

    for c in columns:
        if c not in state.customer_export_df.columns:
            raise HTTPException(status_code=400, detail=f"유효하지 않은 컬럼: {c}")

    fmt = (payload.get("format") or "xlsx").lower().strip()
    if fmt not in ("xlsx", "xls"):
        fmt = "xlsx"

    header_map = payload.get("header_map") or {}
    if not isinstance(header_map, dict):
        header_map = {}

    out = state.customer_export_df.loc[:, columns].copy()
    rename_map = {}
    for c in columns:
        val = header_map.get(c)
        if isinstance(val, str) and val.strip():
            rename_map[c] = val.strip()
    if rename_map:
        out.rename(columns=rename_map, inplace=True)
    buf = io.BytesIO()
    if fmt == "xlsx":
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="원베양식")
        filename = "원베_고객대기_추출.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        try:
            import xlwt  # noqa: F401
        except Exception:
            raise HTTPException(status_code=400, detail="xls 저장을 위해 xlwt 설치가 필요합니다.")
        with pd.ExcelWriter(buf, engine="xlwt") as writer:
            out.to_excel(writer, index=False, sheet_name="원베양식")
        filename = "원베_고객대기_추출.xls"
        media_type = "application/vnd.ms-excel"
    headers = {"Content-Disposition": _content_disposition(filename)}
    return Response(content=buf.getvalue(), media_type=media_type, headers=headers)


@app.post("/returns/download/queues")
def returns_download_queues(payload: dict = Body(...), user: str = Depends(_get_current_user)):
    state = _get_return_state(user)
    if (not state.queue_seller) and (not state.queue_customer) and (not state.queue_unmatched):
        raise HTTPException(status_code=400, detail="추출할 대기 데이터가 없습니다.")

    fmt = (payload.get("format") or "xlsx").lower().strip()
    if fmt not in ("xlsx", "xls"):
        fmt = "xlsx"

    df_seller = pd.DataFrame(state.queue_seller)
    df_customer = pd.DataFrame(state.queue_customer)
    df_unmatched = pd.DataFrame(state.queue_unmatched)

    for dfx in (df_seller, df_customer, df_unmatched):
        if not dfx.empty:
            dfx.drop(columns=["id"], inplace=True, errors="ignore")
            dfx.rename(
                columns={
                    "scan": "스캔송장",
                    "match": "요청메모",
                    "item_text": "가공데이터",
                    "qty": "입고수량",
                    "type": "분류",
                },
                inplace=True,
            )

    buf = io.BytesIO()
    if fmt == "xlsx":
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_seller.to_excel(writer, index=False, sheet_name="판매자")
            df_customer.to_excel(writer, index=False, sheet_name="고객")
            df_unmatched.to_excel(writer, index=False, sheet_name="미매칭")
        filename = "반품대기_추출.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        try:
            import xlwt  # noqa: F401
        except Exception:
            raise HTTPException(status_code=400, detail="xls 저장을 위해 xlwt 설치가 필요합니다.")
        with pd.ExcelWriter(buf, engine="xlwt") as writer:
            df_seller.to_excel(writer, index=False, sheet_name="판매자")
            df_customer.to_excel(writer, index=False, sheet_name="고객")
            df_unmatched.to_excel(writer, index=False, sheet_name="미매칭")
        filename = "반품대기_추출.xls"
        media_type = "application/vnd.ms-excel"

    headers = {"Content-Disposition": _content_disposition(filename)}
    return Response(content=buf.getvalue(), media_type=media_type, headers=headers)
